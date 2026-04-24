"""
NIFTY 750 MOMENTUM SCREENER
-----------------------------
Reads tickers from 'Ticker Data' sheet col A of the Excel,
fetches OHLCV data via yfinance, applies all strategy filters,
ranks qualifying stocks, and writes results back to Excel.

Runs daily — not restricted to Mondays.

Filters (ALL 5 must pass):
  F1  Trend        : Price > 50 EMA > 200 EMA > 0
  F2  52W Prox     : Price >= 52W High * 0.85  (within 15% of 52W high)
  F3  3M Breakout  : Price >= 3M High (65-day high)  — fresh breakout
  F4  6M Ceiling   : Price < 6M High (126-day high)  — not already extended
  F5  Vol Surge    : 1M Median Volume > 3M Median Volume

Logic behind F3 + F4 together:
  Stock must be breaking out of its 3-month range (F3),
  but has NOT yet cleared its 6-month high (F4).
  This targets stocks in the early stage of a larger breakout —
  just cleared recent resistance, with the 6M high still acting as
  the next target / headroom.

Usage:
    pip install yfinance openpyxl pandas numpy
    python fetch_and_screen.py

Schedule daily (optional):
    Mac/Linux — cron : 0 10 * * 1-5 cd /path/to/folder && python fetch_and_screen.py
    Windows   — Task Scheduler: run python fetch_and_screen.py daily at 10:00 AM
"""

import sys
import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import time
import os

try:
    import yfinance as yf
except ImportError:
    print("Installing yfinance..."); import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "yfinance", "-q"])
    import yfinance as yf

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl..."); import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── CONFIG ───────────────────────────────────────────────────────────────────
EXCEL_FILE       = "nifty750_screener.xlsx"
OUTPUT_FILE      = "nifty750_screener_updated.xlsx"
TICKER_SHEET     = "Ticker Data"
RESULTS_SHEET    = "Top 20 Signals"
LOG_DIR          = "run_logs"          # folder to save daily CSV logs

EMA_SHORT        = 50
EMA_LONG         = 200
BREAKOUT_3M_DAYS = 65                  # ~3 months — must be AT or ABOVE this
BREAKOUT_6M_DAYS = 126                 # ~6 months — must be BELOW this
VOL_SHORT_DAYS   = 21                  # ~1 month
VOL_LONG_DAYS    = 63                  # ~3 months
MAX_DIST_52W     = 0.15                # within 15% of 52-week high
MAX_POSITIONS    = 20
TSL_FACTOR       = 0.95
TARGET_FACTOR    = 1.10
HISTORY_DAYS     = 300                 # enough for 200 EMA + 52W high + 6M window

# Ranking weights (must sum to 1.0)
W_MOMENTUM = 0.30
W_VOLUME   = 0.25
W_BREAKOUT = 0.25
W_EMA_SEP  = 0.20

# Risk / Reward
# Risk   = entry price - TSL           = Price * (1 - TSL_FACTOR)       = 5%
# Reward = 6M High - entry price       = natural resistance target
# RR     = Reward / Risk
# Only stocks with RR >= this threshold are kept
RR_MIN_RATIO = 1.5   # e.g. 1.5 means reward must be at least 1.5× the risk

# Colours
DARK_BG   = "1E2A3A"
GOLD_BG   = "FFF3CD"
GREEN_BG  = "D4EDDA"
RED_BG    = "F8D7DA"
AMBER_BG  = "FFF3CD"
HEADER_BG = "2E4057"
WHITE     = "FFFFFF"
GREEN_FG  = "155724"
RED_FG    = "721C24"

# ── HELPERS ──────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def ema(series, span):
    return series.ewm(span=span, adjust=False).mean()

def median_volume(vol_series, days):
    return vol_series.iloc[-days:].median() if len(vol_series) >= days else np.nan

def percentile_rank(series, value):
    """Return 0-100 percentile rank of value within series (ignoring NaN)."""
    clean = series.dropna()
    if len(clean) == 0 or pd.isna(value):
        return np.nan
    return round(float(np.sum(clean <= value)) / len(clean) * 100, 1)

# ── STEP 1: READ TICKERS ─────────────────────────────────────────────────────
def read_tickers(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[TICKER_SHEET]
    tickers = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                            min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            tickers.append(str(val).strip().upper())
    print(f"  Found {len(tickers)} tickers in Excel")
    return tickers

# ── STEP 2: FETCH DATA ───────────────────────────────────────────────────────
def fetch_data(tickers):
    end   = datetime.today()
    start = end - timedelta(days=HISTORY_DAYS + 60)

    results, failed = [], []
    print(f"\n  Fetching data for {len(tickers)} tickers...\n")

    for i, ticker in enumerate(tickers, 1):
        yf_ticker = ticker if ticker.endswith(".NS") else ticker + ".NS"

        try:
            df = yf.download(yf_ticker, start=start, end=end,
                             progress=False, auto_adjust=True)

            if df is None or len(df) < EMA_LONG + 5:
                failed.append(ticker)
                print(f"  [{i:>3}/{len(tickers)}] {ticker:<15} ✘  Insufficient data")
                continue

            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)

            close  = df["Close"].dropna()
            volume = df["Volume"].dropna()
            high   = df["High"].dropna()

            if len(close) < EMA_LONG:
                failed.append(ticker)
                continue

            price   = float(close.iloc[-1])
            e50     = float(ema(close, EMA_SHORT).iloc[-1])
            e200    = float(ema(close, EMA_LONG).iloc[-1])
            h52w    = float(high.iloc[-252:].max()) if len(high) >= 252 else float(high.max())
            h3m     = float(high.iloc[-BREAKOUT_3M_DAYS:].max())
            h6m     = float(high.iloc[-BREAKOUT_6M_DAYS:].max()) \
                      if len(high) >= BREAKOUT_6M_DAYS else float(high.max())
            l3m     = float(close.iloc[-BREAKOUT_3M_DAYS:].min())
            v1m_med = float(median_volume(volume, VOL_SHORT_DAYS))
            v3m_med = float(median_volume(volume, VOL_LONG_DAYS))
            ret_3m  = float((close.iloc[-1] - close.iloc[-BREAKOUT_3M_DAYS])
                            / close.iloc[-BREAKOUT_3M_DAYS]) \
                      if len(close) >= BREAKOUT_3M_DAYS else np.nan

            # Headroom = how far price is below the 6M high (higher = more room)
            headroom_6m = (h6m - price) / h6m * 100 if h6m > 0 else np.nan

            # ── Risk / Reward ─────────────────────────────────────────────
            # Risk   : price → TSL          = price × (1 − TSL_FACTOR)
            # Reward : price → 6M High      = 6M High − price   (natural resistance)
            # RR     : reward / risk
            risk_pts    = price * (1 - TSL_FACTOR)                         # ₹ at risk
            reward_pts  = (h6m - price) if h6m > price else np.nan        # ₹ to target
            rr_ratio    = round(reward_pts / risk_pts, 2) \
                          if (not np.isnan(reward_pts) and risk_pts > 0) else np.nan
            risk_pct    = round((1 - TSL_FACTOR) * 100, 2)                # always 5%
            reward_pct  = round(reward_pts / price * 100, 2) \
                          if not np.isnan(reward_pts) else np.nan

            results.append({
                "Ticker":      ticker,
                "Price":       round(price, 2),
                "EMA50":       round(e50, 2),
                "EMA200":      round(e200, 2),
                "High52W":     round(h52w, 2),
                "High3M":      round(h3m, 2),
                "High6M":      round(h6m, 2),
                "Low3M":       round(l3m, 2),
                "Vol1M":       round(v1m_med),
                "Vol3M":       round(v3m_med),
                "Return3M":    round(ret_3m * 100, 2) if not np.isnan(ret_3m) else np.nan,
                "Headroom6M":  round(headroom_6m, 2)  if not np.isnan(headroom_6m) else np.nan,
                "RiskPts":     round(risk_pts, 2),
                "RewardPts":   round(reward_pts, 2)   if not np.isnan(reward_pts) else np.nan,
                "RiskPct":     risk_pct,
                "RewardPct":   reward_pct,
                "RR":          rr_ratio,
            })

            trend_ok = "✔" if (price > e50 > e200 > 0) else "·"
            bo3m_ok  = "✔" if price >= h3m  else "·"
            h6m_ok   = "✔" if price <  h6m  else "·"   # below 6M high = good
            print(f"  [{i:>3}/{len(tickers)}] {ticker:<15} "
                  f"Trend:{trend_ok} 3MBO:{bo3m_ok} <6MH:{h6m_ok}  "
                  f"₹{price:>9,.2f}  3MH:₹{h3m:>9,.2f}  6MH:₹{h6m:>9,.2f}")

        except Exception as ex:
            failed.append(ticker)
            print(f"  [{i:>3}/{len(tickers)}] {ticker:<15} ✘  Error: {ex}")

        if i % 10 == 0:
            time.sleep(1)

    print(f"\n  ✔ Fetched: {len(results)}   ✘ Failed: {len(failed)}")
    if failed:
        print(f"  Failed tickers: {', '.join(failed)}")

    return pd.DataFrame(results)

# ── STEP 3: APPLY FILTERS ────────────────────────────────────────────────────
def apply_filters(df):
    """
    6 filters — ALL must pass:

      F1  Trend        : Price > 50 EMA > 200 EMA > 0
      F2  52W Prox     : Price >= 52W High × 0.85  (within 15% of 52W high)
      F3  3M Breakout  : Price >= 3M High (65-day high)   — just broke out
      F4  6M Ceiling   : Price <  6M High (126-day high)  — headroom remains
      F5  Vol Surge    : 1M Median Volume > 3M Median Volume
      F6  R/R Ratio    : (6M High − Price) / (Price × 5%) >= RR_MIN_RATIO

    F3 + F4 + F6 together:
      Stock has broken its 3M high (momentum confirmed), still below its
      6M high (headroom to run), AND the distance to 6M high is at least
      RR_MIN_RATIO × the 5% stop — so the trade has mathematical edge.
    """
    df = df.copy()
    df["F_Trend"]      = (df["Price"] > df["EMA50"]) & \
                         (df["EMA50"] > df["EMA200"]) & \
                         (df["EMA200"] > 0)
    df["F_52WProx"]    = df["Price"] >= df["High52W"] * (1 - MAX_DIST_52W)
    df["F_Breakout3M"] = df["Price"] >= df["High3M"]           # AT or ABOVE 3M high
    df["F_Below6M"]    = df["Price"] <  df["High6M"]           # BELOW 6M high (headroom)
    df["F_Volume"]     = df["Vol1M"]  >  df["Vol3M"]
    df["F_RR"]         = df["RR"] >= RR_MIN_RATIO              # R/R must meet minimum
    df["AllPass"]      = (df["F_Trend"]      &
                          df["F_52WProx"]    &
                          df["F_Breakout3M"] &
                          df["F_Below6M"]    &
                          df["F_Volume"]     &
                          df["F_RR"])
    return df

# ── STEP 4: RANK QUALIFYING STOCKS ──────────────────────────────────────────
def rank_stocks(df):
    df = df.copy()
    df["VolRatio"]    = df["Vol1M"] / df["Vol3M"].replace(0, np.nan)
    df["BreakoutStr"] = (df["Price"] - df["High3M"]) / df["High3M"].replace(0, np.nan) * 100
    df["EMASep"]      = (df["Price"] - df["EMA50"])  / df["EMA50"].replace(0, np.nan)  * 100

    passing = df[df["AllPass"]].copy()
    all_q   = df

    if passing.empty:
        print("\n  ⚠  No stocks passed all 6 filters today.")
        return passing

    for col, score_col in [
        ("Return3M",    "S_Momentum"),
        ("VolRatio",    "S_Volume"),
        ("BreakoutStr", "S_Breakout"),
        ("EMASep",      "S_EMA"),
    ]:
        ref = all_q[col].dropna()
        passing[score_col] = passing[col].apply(
            lambda v: percentile_rank(ref, v) if pd.notna(v) else np.nan
        )

    passing["CompositeScore"] = (
        passing["S_Momentum"] * W_MOMENTUM +
        passing["S_Volume"]   * W_VOLUME   +
        passing["S_Breakout"] * W_BREAKOUT +
        passing["S_EMA"]      * W_EMA_SEP
    ).round(1)

    passing = passing.sort_values("CompositeScore", ascending=False).reset_index(drop=True)
    passing["Rank"] = passing.index + 1
    return passing

# ── STEP 5: PRINT RESULTS ────────────────────────────────────────────────────
def print_results(df_all, top20):
    total  = len(df_all)
    n_pass = int(df_all["AllPass"].sum()) if "AllPass" in df_all else 0

    print("\n" + "═"*82)
    print("  SCREENING SUMMARY")
    print("═"*82)
    print(f"  Total tickers screened  : {total}")
    for col, label in [
        ("F_Trend",       "F1  Trend (P>50E>200E)      "),
        ("F_52WProx",     "F2  Within 15% of 52W High  "),
        ("F_Breakout3M",  "F3  3M Breakout (>=65d High) "),
        ("F_Below6M",     "F4  Below 6M High (<126d H)  "),
        ("F_Volume",      "F5  Volume Surge             "),
        ("F_RR",          f"F6  R/R >= {RR_MIN_RATIO}x            "),
    ]:
        if col in df_all:
            n = int(df_all[col].sum())
            bar = "█" * int(n / max(total, 1) * 30)
            print(f"    {label}: {n:>4} / {total}  {bar}")

    print(f"  {'─'*44}")
    print(f"  Passed ALL 5 filters    : {n_pass}")
    print(f"  Top {MAX_POSITIONS} selected         : {min(MAX_POSITIONS, len(top20))}")
    print("═"*82)

    if top20.empty:
        print("\n  No qualifying stocks today.\n")
        return

    print(f"\n  {'Rk':<4} {'Ticker':<13} {'Price':>9}  {'Score':>5}  "
          f"{'3MRet':>6}  {'VolR':>5}  {'RR':>5}  {'Reward%':>7}  {'TSL':>9}  {'Target':>9}")
    print("  " + "─"*84)

    for _, row in top20.head(MAX_POSITIONS).iterrows():
        tsl      = round(row["Price"] * TSL_FACTOR, 2)
        target   = round(row["Price"] * TARGET_FACTOR, 2)
        rr       = row.get("RR", 0)
        reward   = row.get("RewardPct", 0)
        rr_flag  = "🟢" if rr >= 2.0 else ("🟡" if rr >= RR_MIN_RATIO else "🔴")
        print(f"  {int(row['Rank']):<4} {row['Ticker']:<13} "
              f"₹{row['Price']:>8,.2f}  {row['CompositeScore']:>5.1f}  "
              f"{row['Return3M']:>5.1f}%  {row['VolRatio']:>5.2f}x  "
              f"{rr_flag}{rr:>4.1f}x  {reward:>6.1f}%  "
              f"₹{tsl:>8,.2f}  ₹{target:>8,.2f}")
    print(f"\n  R/R legend:  🟢 ≥ 2.0x   🟡 {RR_MIN_RATIO}–1.99x   🔴 < {RR_MIN_RATIO}x (filtered out)")
    print()

# ── STEP 6: SAVE DAILY LOG ───────────────────────────────────────────────────
def save_daily_log(top20):
    """Saves today's top 20 as a timestamped CSV for historical reference."""
    if top20.empty:
        return
    os.makedirs(LOG_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    log_path = os.path.join(LOG_DIR, f"top20_{date_str}.csv")
    cols     = ["Rank","Ticker","Price","CompositeScore","Return3M",
                "VolRatio","EMASep","BreakoutStr","High3M","High6M",
                "Headroom6M","RiskPct","RewardPct","RR"]
    cols     = [c for c in cols if c in top20.columns]
    top20[cols].head(MAX_POSITIONS).to_csv(log_path, index=False)
    print(f"  📁 Daily log saved : {log_path}")

# ── STEP 7: WRITE BACK TO EXCEL ──────────────────────────────────────────────
def write_excel(df_all, top20, input_path, output_path):
    wb = load_workbook(input_path)

    # ── Update Ticker Data sheet ──────────────────────────────────────────────
    ws_td = wb[TICKER_SHEET]
    ticker_to_row = {}
    for row in ws_td.iter_rows(min_row=4, max_row=ws_td.max_row,
                               min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value and str(cell.value).strip():
            ticker_to_row[str(cell.value).strip().upper()] = cell.row

    col_map  = {"Price": 3, "EMA50": 4, "EMA200": 5,
                "High52W": 6, "High3M": 7, "Low3M": 8,
                "Vol1M": 9, "Vol3M": 10, "Return3M": 11}
    num_fmts = {3:"#,##0.00", 4:"#,##0.00", 5:"#,##0.00",
                6:"#,##0.00", 7:"#,##0.00", 8:"#,##0.00",
                9:"#,##0", 10:"#,##0", 11:"0.00"}

    for _, r in df_all.iterrows():
        tk = r["Ticker"]
        if tk not in ticker_to_row:
            continue
        excel_row = ticker_to_row[tk]
        for field, col in col_map.items():
            val  = r.get(field, np.nan)
            cell = ws_td.cell(row=excel_row, column=col)
            cell.value         = float(val) if pd.notna(val) else None
            cell.number_format = num_fmts.get(col, "General")
            cell.font          = Font(name="Arial", color="000000", size=10)
            cell.fill          = PatternFill("solid", start_color="E8F8EF")
            cell.alignment     = Alignment(horizontal="center", vertical="center")
            cell.border        = thin_border()

    # ── Rebuild Top 20 Signals sheet ─────────────────────────────────────────
    if RESULTS_SHEET in wb.sheetnames:
        del wb[RESULTS_SHEET]
    ws_r = wb.create_sheet(RESULTS_SHEET)
    ws_r.sheet_view.showGridLines = False

    run_ts = datetime.now().strftime("%d %b %Y  %H:%M")
    ws_r.row_dimensions[1].height = 35
    ws_r.merge_cells("A1:T1")
    c = ws_r.cell(row=1, column=1,
        value=f"🟢  TOP 20 BUY SIGNALS  —  {run_ts}  "
              f"|  6 Filters: Trend | 52W Prox | >=3MH | <6MH | Vol Surge | R/R≥{RR_MIN_RATIO}x")
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=12)
    c.fill      = PatternFill("solid", start_color=DARK_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers    = ["Rank","Ticker","Price ₹","3M Ret %","Vol Ratio","Score",
                  "Risk ₹","Risk %","Reward ₹","Reward %","R/R Ratio",
                  "6M Headroom %","TSL (−5%)","Target (+10%)",
                  "F1 Trend","F2 52W","F3 >=3MH","F4 <6MH","F5 Vol","F6 R/R"]
    col_widths = [6,12,12,10,10,7, 10,8,10,9,10, 13,12,14, 10,9,10,9,9,9]

    ws_r.row_dimensions[2].height = 25
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws_r.cell(row=2, column=i, value=h)
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        # Colour-code section groups
        bg = HEADER_BG
        if i in (7,8,9,10,11): bg = "5D4037"   # risk/reward = brown
        if i >= 15:             bg = "2E7D32"   # filter flags = dark green
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
        ws_r.column_dimensions[get_column_letter(i)].width = w

    for pos, (_, row) in enumerate(top20.head(MAX_POSITIONS).iterrows(), 1):
        r_idx    = pos + 2
        ws_r.row_dimensions[r_idx].height = 20
        tsl      = round(row["Price"] * TSL_FACTOR, 2)
        target   = round(row["Price"] * TARGET_FACTOR, 2)
        rr       = round(row.get("RR", 0), 2)
        headroom = round(row.get("Headroom6M", 0), 2)
        risk_pts = round(row.get("RiskPts", 0), 2)
        rew_pts  = round(row.get("RewardPts", 0), 2)
        risk_pct = round(row.get("RiskPct", 0), 2)
        rew_pct  = round(row.get("RewardPct", 0), 2)

        # R/R quality label
        rr_label = f"🟢 {rr:.1f}x" if rr >= 2.0 else (f"🟡 {rr:.1f}x" if rr >= RR_MIN_RATIO else f"🔴 {rr:.1f}x")

        values = [
            int(row["Rank"]),
            row["Ticker"],
            round(row["Price"], 2),
            round(row["Return3M"], 2),       # shown as %
            round(row["VolRatio"], 2),
            round(row["CompositeScore"], 1),
            risk_pts,                        # ₹ at risk
            risk_pct,                        # % at risk (always 5%)
            rew_pts,                         # ₹ reward to 6M high
            rew_pct,                         # % reward to 6M high
            rr_label,                        # R/R with emoji
            headroom,                        # % below 6M high
            tsl,
            target,
            "✔" if row["F_Trend"]       else "✘",
            "✔" if row["F_52WProx"]     else "✘",
            "✔" if row["F_Breakout3M"]  else "✘",
            "✔" if row["F_Below6M"]     else "✘",
            "✔" if row["F_Volume"]      else "✘",
            "✔" if row["F_RR"]          else "✘",
        ]
        # col index → number format
        num_fmts_r = {
            3:"#,##0.00",  4:"0.00%",   5:"0.00x",
            7:"#,##0.00",  8:"0.00%",   9:"#,##0.00",  10:"0.00%",
            12:"0.00%",    13:"#,##0.00", 14:"#,##0.00"
        }

        row_bg = GOLD_BG if pos <= 5 else ("E8F8F5" if pos % 2 == 0 else WHITE)

        for c_idx, val in enumerate(values, 1):
            # scale % fields
            if c_idx in (4, 8, 10, 12) and isinstance(val, (int, float)):
                val = val / 100
            is_pass = str(val) == "✔"
            is_fail = str(val) == "✘"
            cell = ws_r.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = Font(name="Arial", size=10, bold=(pos <= 3),
                                  color=GREEN_FG if is_pass else
                                       (RED_FG   if is_fail else "000000"))
            cell.fill      = PatternFill("solid",
                                start_color=GREEN_BG if is_pass else
                                           (RED_BG   if is_fail else row_bg))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
            if c_idx in num_fmts_r:
                cell.number_format = num_fmts_r[c_idx]

    # Footer
    footer_row = MAX_POSITIONS + 4
    n_cols = len(headers)
    ws_r.merge_cells(f"A{footer_row}:{get_column_letter(n_cols)}{footer_row}")
    c = ws_r.cell(row=footer_row, column=1,
        value=f"Risk = Price × 5% (TSL)  |  Reward = 6M High − Price  |  "
              f"R/R = Reward ÷ Risk  |  Min R/R = {RR_MIN_RATIO}x  |  "
              f"🟢 ≥2.0x  🟡 {RR_MIN_RATIO}–1.99x  ·  "
              f"F1:Trend  F2:52WH  F3:3MBO  F4:<6MH  F5:VolSurge  F6:R/R")
    c.font      = Font(name="Arial", italic=True, size=9, color="555555")
    c.fill      = PatternFill("solid", start_color="F0F4F8")
    c.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    print(f"  ✅ Excel written : {output_path}")

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═"*82)
    print("  NIFTY 750 MOMENTUM SCREENER  (Daily)")
    print(f"  Run date : {datetime.now().strftime('%A, %d %b %Y  %H:%M')}")
    print("═"*82)

    if not os.path.exists(EXCEL_FILE):
        print(f"\n  ❌ ERROR: '{EXCEL_FILE}' not found in current folder.")
        print(f"     Current folder: {os.getcwd()}")
        sys.exit(1)

    print(f"\n  Reading tickers from : {EXCEL_FILE}")
    tickers = read_tickers(EXCEL_FILE)
    if not tickers:
        print("  ❌ No tickers found in Ticker Data sheet (col A, from row 4).")
        sys.exit(1)

    print("\n  Fetching market data from Yahoo Finance...")
    df_raw = fetch_data(tickers)
    if df_raw.empty:
        print("  ❌ No data fetched. Check internet connection or ticker symbols.")
        sys.exit(1)

    print("\n  Applying 6 filters...")
    df_filtered = apply_filters(df_raw)

    print("  Ranking qualifying stocks...")
    top20 = rank_stocks(df_filtered)

    print_results(df_filtered, top20)
    save_daily_log(top20)

    print("  Writing results to Excel...")
    write_excel(df_filtered, top20, EXCEL_FILE, OUTPUT_FILE)

    print(f"\n  Done! Open '{OUTPUT_FILE}' → 'Top 20 Signals' tab.")
    print("═"*82 + "\n")

if __name__ == "__main__":
    main()