"""
Alpha & Low Volatility Ranking
==============================
Ranks NSE N500 stocks uniquely combining 12M Jensen's Alpha 
and 12M Volatility. Equal weight composite.
"""

import sys
import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Append the Sharpe Score folder to sys.path to load momentum_lib
LIB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Sharpe Score"))
if LIB_PATH not in sys.path:
    sys.path.insert(0, LIB_PATH)

import momentum_lib as ml

# ── CONFIG ────────────────────────────────────────────────────────────────────
FILE         = "n500.xlsx" if len(sys.argv) < 2 else sys.argv[1]
# Fallback to the original Sharpe Score folder if n500.xlsx isn't in current dir
FILE_PATH    = FILE if os.path.exists(FILE) else os.path.join(LIB_PATH, FILE)

OUTPUT_FILE  = "n500_alphavol.xlsx"
RFR_ANNUAL   = 0.07
TRADING_DAYS = 252
TOP_N        = 20
WINDOW_DAYS  = 252  # 12M

rfr_daily = RFR_ANNUAL / TRADING_DAYS

# ── LOAD ──────────────────────────────────────────────────────────────────────
print(f"Loading {FILE_PATH} ...")
prices_df, nifty_series, stock_tickers, dates = ml.load_prices(FILE_PATH)

valid_days = sum(1 for d in prices_df.columns
                 if prices_df[d].notna().any() and (prices_df[d] != 0).any())
print(f"  {len(prices_df)} stocks  |  {len(dates)} date columns  "
      f"({valid_days} actual trading days)  "
      f"|  {dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}\n")

# ── COMPUTE SCORES ────────────────────────────────────────────────────────────
alpha_vol_df, z_df = ml.compute_alpha_vol(
    prices_df, stock_tickers, nifty_series,
    window=WINDOW_DAYS, rfr_daily=rfr_daily, trading_days=TRADING_DAYS
)

# ── COMBINE & RANK ────────────────────────────────────────────────────────────
result = z_df.join(alpha_vol_df)

# Rank by COMPOSITE Score
result["RANK"] = result["COMPOSITE"].rank(ascending=False, method="first", na_option="bottom")
result = result.sort_values("COMPOSITE", ascending=False)

# ── CONSOLE OUTPUT ────────────────────────────────────────────────────────────
SEP  = "-" * 75
HEAD = (f"{'RNK':>4}  {'TICKER':<12}  {'COMP':>9}  "
        f"{'Z_ALPHA':>9}  {'Z_INV_V':>9}  {'ALPHA_12M':>9}  {'INV_VOL':>9}")

print(f"\n{'':=<75}")
print(f"  ALPHA & LOW VOLATILITY - TOP {TOP_N}  .  Equal Weight Composite")
print(f"  Window: 12M ({WINDOW_DAYS}d)  |  RFR={RFR_ANNUAL*100:.1f}%")
print(f"{'':=<75}")
print(HEAD); print(SEP)

def fs(v, w=7): return f"{v:>{w}.3f}" if pd.notna(v) else f"{'--':>{w}}"
def fp(v, w=7): return f"{v:>{w}.1f}" if pd.notna(v) else f"{'--':>{w}}"

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 1):
    print(f"{i:>4}  {ticker:<12}  "
          f"{fs(row['COMPOSITE'],9)}  "
          f"{fs(row['Z_ALPHA'],9)}  {fs(row['Z_INV_VOL'],9)}  "
          f"{fs(row['ALPHA'],9)}  {fs(row['INV_VOL'],9)}")

print(SEP)
print(f"\n  COMPOSITE = 0.5 * (Z_ALPHA + Z_INV_VOL)\n")

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
print(f"Writing {OUTPUT_FILE} ...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border_all():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# Light Theme Colors
HDR_FILL   = fill("E5F3FF")  # Light Blue
ALT_FILL   = fill("F9F9F9")  # Very Light Gray
EVEN_FILL  = fill("FFFFFF")  # White
HDR_FONT   = Font(name="Calibri", color="003366", bold=True,  size=11)
TEXT_FONT  = Font(name="Calibri", color="000000",             size=11)
BOLD_FONT  = Font(name="Calibri", color="000000", bold=True,  size=11)
BLUE_FONT  = Font(name="Calibri", color="0066CC", bold=True,  size=11)

def set_hdr(cell, value):
    cell.value     = value
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border_all()

def set_cell(cell, value, font=None, bg=None, num_fmt=None, align="right"):
    cell.value     = value
    cell.font      = font or TEXT_FONT
    cell.fill      = bg or EVEN_FILL
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border_all()
    if num_fmt:
        cell.number_format = num_fmt

# ── SHEET 1 — TOP20 ───────────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("TOP20")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C3"

ws1.merge_cells("A1:G1")
tc           = ws1["A1"]
tc.value     = (f"N500 ALPHA & VOLATILITY  .  Top {TOP_N} by COMPOSITE  .  "
                f"RFR={RFR_ANNUAL*100:.1f}%  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}")
tc.font      = Font(name="Calibri", color="003366", bold=True, size=11)
tc.fill      = fill("D0E8FF")
tc.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

top20_cols = [
    ("RNK",        6), 
    ("TICKER",    12), 
    ("COMPOSITE", 12),
    ("Z_ALPHA",   10), 
    ("Z_INV_VOL", 10), 
    ("ALPHA_12M", 12), 
    ("INV_VOL",   12)
]
for c, (col_name, col_w) in enumerate(top20_cols, 1):
    set_hdr(ws1.cell(row=2, column=c), col_name)
    ws1.column_dimensions[get_column_letter(c)].width = col_w
ws1.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 3):
    bg     = ALT_FILL if i % 2 == 0 else EVEN_FILL
    rank_v = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    values = [
        (rank_v,              BOLD_FONT,  bg, "0"),
        (ticker,              BOLD_FONT,  bg, None),
        (row["COMPOSITE"],    BLUE_FONT,  bg, "0.000"),
        (row["Z_ALPHA"],      TEXT_FONT,  bg, "0.000"),
        (row["Z_INV_VOL"],    TEXT_FONT,  bg, "0.000"),
        (row["ALPHA"],        TEXT_FONT,  bg, "0.000"),
        (row["INV_VOL"],      TEXT_FONT,  bg, "0.000")
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws1.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws1.row_dimensions[i].height = 16

# ── SHEET 2 — CALCS ───────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("CALCS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C3"

ws2.merge_cells("A1:F1")
t2           = ws2["A1"]
t2.value     = (f"N500 ALPHA & VOLATILITY  .  All {len(stock_tickers)} stocks  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}")
t2.font      = Font(name="Calibri", color="003366", bold=True, size=11)
t2.fill      = fill("D0E8FF")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

calcs_cols = [
    ("RANK",       6), 
    ("TICKER",    12),
    ("ALPHA_12M", 12), 
    ("Z_ALPHA",   10), 
    ("INV_VOL",   12), 
    ("Z_INV_VOL", 10)
]

for c, (col_name, col_w) in enumerate(calcs_cols, 1):
    set_hdr(ws2.cell(row=2, column=c), col_name)
    ws2.column_dimensions[get_column_letter(c)].width = col_w
ws2.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.iterrows(), 3):
    bg         = ALT_FILL if i % 2 == 0 else EVEN_FILL
    rank_v     = int(row["RANK"]) if pd.notna(row["RANK"]) else None

    values = [
        (rank_v,             BOLD_FONT, bg, "0"),
        (ticker,             BOLD_FONT, bg, None),
        (row["ALPHA"],       TEXT_FONT, bg, "0.000"),
        (row["Z_ALPHA"],     TEXT_FONT, bg, "0.000"),
        (row["INV_VOL"],     TEXT_FONT, bg, "0.000"),
        (row["Z_INV_VOL"],   TEXT_FONT, bg, "0.000"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws2.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws2.row_dimensions[i].height = 16

wb_out.save(OUTPUT_FILE)
print(f"  +  Saved -> {OUTPUT_FILE}")
print(f"     Sheet 'TOP20' : top {TOP_N} stocks")
print(f"     Sheet 'CALCS' : all {len(stock_tickers)} stocks, {len(calcs_cols)} columns")
