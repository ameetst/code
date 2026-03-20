"""
NSE Options Trade Signal Generator  v2.0
==========================================
Reads your existing ETF momentum system outputs and generates
actionable weekly options buy signals with full entry + exit specs.

DATA SOURCE: Local CSV files downloaded from nseindia.com/option-chain
  - No API keys, no scraping, no dependencies other than pandas + openpyxl
  - Download the options chain CSV for each index + expiry you want to trade
  - Put the files in the same folder as this script — do NOT rename them

HOW TO DOWNLOAD:
  1. Go to https://www.nseindia.com/option-chain
  2. Select index (NIFTY / BANKNIFTY / FINNIFTY / MIDCPNIFTY / NIFTYNXT50)
  3. Select expiry date from the dropdown
  4. Click the CSV download button at top-right of the table
  5. Move the downloaded file into this folder
     NSE names them: option-chain-ED-NIFTY-27-Mar-2026.csv
     The script reads index name and expiry date from the filename automatically.

DOWNLOAD ONE FILE PER EXPIRY PER INDEX.
For best results download both the weekly AND monthly expiry for each index —
the script will pick the better one automatically based on IV rank and trend quality.

INPUTS (all in same folder):
  - etf_rankings.xlsx        — from your ETF momentum system (read-only)
  - holdings_log.json        — from your ETF momentum system (read-only)
  - option-chain-ED-*.csv    — downloaded from NSE (one per expiry per index)
  - iv_history.json          — auto-created and maintained by this script
  - signals_log.json         — auto-created and maintained by this script

OUTPUT:
  - options_trades.xlsx      — Trade Signals sheet + Signal History sheet

RUN:  python options_trader.py
DEPS: pip install pandas openpyxl python-dateutil
"""

import csv
import json
import math
import os
import re
import sys
from datetime import datetime, timedelta
from glob import glob

import pandas as pd
import numpy as np


# ─────────────────────────────────────────────────────────────
# CONFIG — edit only this block
# ─────────────────────────────────────────────────────────────
class CONFIG:
    ETF_RANKINGS_FILE = "etf_rankings.xlsx"
    HOLDINGS_LOG_FILE = "holdings_log.json"
    IV_HISTORY_FILE   = "iv_history.json"
    SIGNALS_LOG_FILE  = "signals_log.json"
    OUTPUT_FILE       = "options_trades.xlsx"

    # Pattern to find downloaded NSE option chain CSVs
    CSV_PATTERN       = "option-chain-ED-*.csv"

    # Trade budget
    MAX_PREMIUM       = 25_000    # ₹ hard ceiling per trade (total premium paid)

    # Strike selection — delta targeting via moneyness proxy
    TARGET_DELTA_HIGH = 0.42      # when R² > HIGH_CONF_R2: ~1-2% OTM (near ATM)
    TARGET_DELTA_LOW  = 0.32      # when R² ≤ HIGH_CONF_R2: ~3-4% OTM

    # IV rank thresholds
    IVR_MAX_TO_BUY    = 50        # skip if IVR > 50 (premium too expensive to buy)
    IVR_CHEAP         = 30        # IVR < 30 = clearly cheap — prefer weekly

    # Exit rules
    SL_PCT            = 0.50      # stop-loss: exit when LTP drops to 50% of entry
    TARGET_PCT        = 2.00      # profit target: exit at 2× entry (100% gain)
    TIME_STOP_DAYS    = 2         # exit N calendar days before expiry regardless of P&L

    # Expiry selection
    MIN_DTE_WEEKLY    = 5         # weekly expiry must have ≥ 5 days left to be used
    HIGH_CONF_R2      = 0.85      # R² threshold for near-ATM strikes and weekly preference


# ─────────────────────────────────────────────────────────────
# SECTOR → NSE INDEX MAPPING
# ─────────────────────────────────────────────────────────────
SECTOR_TO_INDEX = {
    "PSU_BANK":        "BANKNIFTY",
    "PRIVATE_BANK":    "BANKNIFTY",
    "BANKING_BROAD":   "BANKNIFTY",
    "FACTOR_VALUE":    "FINNIFTY",
    "FINANCIAL":       "FINNIFTY",
    "MIDCAP":          "MIDCPNIFTY",
    "SMALLCAP":        "MIDCPNIFTY",
    "IT_TECH":         "NIFTY",
    "HEALTHCARE":      "NIFTY",
    "PHARMA":          "NIFTY",
    "BROAD_MARKET":    "NIFTY",
    "FACTOR_MOMENTUM": "NIFTY",
    "FACTOR_QUALITY":  "NIFTYNXT50",
    "FACTOR_LOWVOL":   "NIFTYNXT50",
    "INTERNATIONAL":   "NIFTY",
    "ENERGY":          "NIFTY",
    "METAL":           "NIFTY",
    "GOLD":            "NIFTY",
    "SILVER":          "NIFTY",
    "GOVT_BONDS":      None,      # no liquid options on NSE — skip
}

LOT_SIZES = {
    "NIFTY":       75,
    "BANKNIFTY":   15,
    "FINNIFTY":    40,
    "MIDCPNIFTY":  75,
    "NIFTYNXT50":  25,
}

# Keywords used to identify index name from CSV filename
# Order matters — BANKNIFTY must come before NIFTY
INDEX_KEYWORDS = {
    "BANKNIFTY":  ["BANKNIFTY", "BANK-NIFTY"],
    "FINNIFTY":   ["FINNIFTY", "FIN-NIFTY", "FINSERVICE"],
    "MIDCPNIFTY": ["MIDCPNIFTY", "MIDCP-NIFTY", "MIDCAP-SELECT"],
    "NIFTYNXT50": ["NIFTYNXT50", "NIFTYNXT-50", "NIFTYNEXT50"],
    "NIFTY":      ["NIFTY"],      # checked last — it's a substring of the others
}


# ─────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────
def clean_num(val) -> float | None:
    """Parse NSE number string like '1,23,456.78' or '-' to float."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("-", "", "nan", "N/A"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


# ─────────────────────────────────────────────────────────────
# STEP 1 — READ REGIME FROM holdings_log.json
# ─────────────────────────────────────────────────────────────
def read_regime(log_file: str) -> dict:
    if not os.path.exists(log_file):
        print(f"[ERROR] {log_file} not found. Run etf_momentum_ranking.py first.")
        sys.exit(1)

    with open(log_file) as f:
        log = json.load(f)

    if not log:
        print("[ERROR] holdings_log.json is empty.")
        sys.exit(1)

    # Detect format: month-keyed {"2026-03": {...}} vs flat {"regime": "...", ...}
    first_val = log[list(log.keys())[0]]
    if isinstance(first_val, dict):
        entry = log[sorted(log.keys())[-1]]
    else:
        entry = log

    regime_raw = entry.get("regime", "UNKNOWN")
    if   "BEAR"    in regime_raw.upper(): regime = "BEAR"
    elif "PARTIAL" in regime_raw.upper(): regime = "PARTIAL"
    elif "BULL"    in regime_raw.upper(): regime = "BULL"
    else:                                 regime = regime_raw.upper()

    allocation = entry.get("allocation", entry.get("active_slots", []))
    if not isinstance(allocation, list):
        allocation = []

    active = [s.get("ticker","?") for s in allocation
              if isinstance(s, dict) and s.get("ticker","CASH") != "CASH"]

    print(f"[Regime] '{regime_raw}' → {regime} | Active ETFs: {active}")
    return {"regime": regime, "regime_raw": regime_raw, "slots": allocation}


# ─────────────────────────────────────────────────────────────
# STEP 2 — READ TOP SECTORS FROM etf_rankings.xlsx
# ─────────────────────────────────────────────────────────────
def read_top_sectors(rankings_file: str, n: int = 5) -> pd.DataFrame:
    if not os.path.exists(rankings_file):
        print(f"[ERROR] {rankings_file} not found.")
        sys.exit(1)

    df = pd.read_excel(rankings_file, sheet_name="Rankings", header=1)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if "investable" in cl and "rank" in cl: col_map["inv_rank"]   = col
        if "sector"     in cl:                  col_map.setdefault("sector", col)
        if "ticker"     in cl:                  col_map.setdefault("ticker", col)
        if "clenow"     in cl and "bld" in cl:  col_map["clenow_bld"] = col
        if "clenow"     in cl and "3m"  in cl and "score" in cl: col_map["clenow_3m"] = col

    # Positional fallback per 24-column layout in context doc
    for k, pos in [("inv_rank",0),("ticker",7),("sector",9),("clenow_bld",15),("clenow_3m",14)]:
        if k not in col_map:
            col_map[k] = df.columns[pos]

    df_inv = df[pd.to_numeric(df[col_map["inv_rank"]], errors="coerce").notna()].copy()
    df_inv["_rank"] = pd.to_numeric(df_inv[col_map["inv_rank"]], errors="coerce")
    df_inv = df_inv.sort_values("_rank").head(n)

    result = pd.DataFrame({
        "inv_rank":   df_inv["_rank"].values,
        "ticker":     df_inv[col_map["ticker"]].values,
        "sector":     df_inv[col_map["sector"]].values,
        "clenow_bld": pd.to_numeric(df_inv[col_map["clenow_bld"]], errors="coerce").values,
        "clenow_3m":  pd.to_numeric(df_inv[col_map["clenow_3m"]], errors="coerce").values,
    })

    print(f"\n[Sectors] Top {n} investable ETFs:")
    for _, row in result.iterrows():
        idx = SECTOR_TO_INDEX.get(str(row["sector"]).upper(), "NIFTY")
        print(f"  Rank {int(row['inv_rank'])}: {row['ticker']} | {row['sector']} → {idx or 'SKIP'}")

    return result


# ─────────────────────────────────────────────────────────────
# STEP 3 — PARSE NSE OPTIONS CHAIN CSV FILES
# ─────────────────────────────────────────────────────────────
def detect_index(fname: str) -> str | None:
    name = os.path.basename(fname).upper()
    for index, keywords in INDEX_KEYWORDS.items():
        if any(kw in name for kw in keywords):
            return index
    return None


def detect_expiry(fname: str) -> datetime | None:
    m = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", os.path.basename(fname))
    if m:
        try:
            return datetime.strptime(m.group(1), "%d-%b-%Y")
        except ValueError:
            pass
    return None


def parse_nse_csv(filepath: str) -> tuple[list[dict], float, datetime | None]:
    """
    Parses one NSE options chain CSV file.

    NSE CSV layout (23 columns, rows 0-1 are headers):
      Cols 0-10:  CALL data  (blank, OI, chng_OI, vol, IV, LTP, chng, bidQty, bid, ask, askQty)
      Col  11:    STRIKE
      Cols 12-22: PUT data   (bidQty, bid, ask, askQty, chng, LTP, IV, vol, chng_OI, OI, blank)

    Spot estimated via put-call parity: spot ≈ strike + call_LTP - put_LTP
    at the strike where |call_LTP - put_LTP| is minimised.
    """
    expiry = detect_expiry(filepath)
    rows   = []
    with open(filepath, encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            rows.append(row)

    records = []
    for row in rows[2:]:
        if len(row) < 22:
            continue
        strike = clean_num(row[11])
        if not strike or strike <= 0:
            continue
        records.append({
            "strike":   strike, "expiry": expiry,
            "call_ltp": clean_num(row[5])  or 0,
            "call_iv":  clean_num(row[4])  or 0,
            "call_oi":  clean_num(row[1])  or 0,
            "call_vol": clean_num(row[3])  or 0,
            "call_bid": clean_num(row[8])  or 0,
            "call_ask": clean_num(row[9])  or 0,
            "put_ltp":  clean_num(row[17]) or 0,
            "put_iv":   clean_num(row[18]) or 0,
            "put_oi":   clean_num(row[21]) or 0,
            "put_vol":  clean_num(row[19]) or 0,
            "put_bid":  clean_num(row[13]) or 0,
            "put_ask":  clean_num(row[14]) or 0,
        })

    spot  = 0.0
    valid = [r for r in records if r["call_ltp"] > 0.5 and r["put_ltp"] > 0.5]
    if valid:
        atm  = min(valid, key=lambda r: abs(r["call_ltp"] - r["put_ltp"]))
        spot = round(atm["strike"] + atm["call_ltp"] - atm["put_ltp"], 1)

    return records, spot, expiry


def load_all_chains(pattern: str) -> dict:
    """
    Finds all CSVs matching pattern and returns:
      { "NIFTY": [(expiry, records, spot), ...], "BANKNIFTY": [...], ... }
    Each list sorted nearest expiry first.
    """
    chains   = {}
    csvfiles = sorted(glob(pattern))

    if not csvfiles:
        print(f"\n[CSV] No files found matching '{pattern}'")
        print("      Download option chain CSVs from nseindia.com/option-chain")
        return chains

    print(f"\n[CSV] Found {len(csvfiles)} file(s):")
    for fp in csvfiles:
        index  = detect_index(fp)
        expiry = detect_expiry(fp)
        fname  = os.path.basename(fp)

        if not index:
            print(f"  SKIP {fname} — index not recognised"); continue
        if not expiry:
            print(f"  SKIP {fname} — expiry not found in filename"); continue

        records, spot, _ = parse_nse_csv(fp)
        if not records:
            print(f"  SKIP {fname} — no data rows"); continue

        print(f"  {fname}")
        print(f"    {index} | Expiry {expiry.strftime('%d-%b-%Y')} | "
              f"{len(records)} strikes | Spot ₹{spot:,.1f}")

        chains.setdefault(index, []).append((expiry, records, spot))

    for idx in chains:
        chains[idx].sort(key=lambda x: x[0])

    return chains


# ─────────────────────────────────────────────────────────────
# STEP 4 — IV RANK
# ─────────────────────────────────────────────────────────────
def load_iv_history(iv_file: str) -> dict:
    return json.load(open(iv_file)) if os.path.exists(iv_file) else {}


def save_iv_history(iv_file: str, history: dict):
    json.dump(history, open(iv_file, "w"), indent=2)


def compute_atm_iv(records: list[dict], spot: float) -> float:
    nearest3 = sorted(records, key=lambda r: abs(r["strike"] - spot))[:3]
    ivs = [r["call_iv"] for r in nearest3 if r["call_iv"] > 0] + \
          [r["put_iv"]  for r in nearest3 if r["put_iv"]  > 0]
    return round(float(np.mean(ivs)), 2) if ivs else 0.0


def record_iv(history: dict, index: str, iv: float, date_str: str) -> dict:
    if iv > 0:
        history.setdefault(index, {})[date_str] = round(iv, 2)
        entries = sorted(history[index].items())
        if len(entries) > 260:
            history[index] = dict(entries[-260:])
    return history


def compute_iv_rank(history: dict, index: str, current_iv: float) -> float:
    data = history.get(index, {})
    if len(data) < 20:
        print(f"  [IVR] {index}: {len(data)} weeks of history so far "
              f"(IVR stabilises after ~5 weeks). Using neutral 40.")
        return 40.0
    vals = list(data.values())
    lo, hi = min(vals), max(vals)
    if hi == lo:
        return 50.0
    return round(max(0.0, min(100.0, (current_iv - lo) / (hi - lo) * 100)), 1)


# ─────────────────────────────────────────────────────────────
# STEP 5 — EXPIRY SELECTION
# ─────────────────────────────────────────────────────────────
def _is_last_thursday(dt: datetime) -> bool:
    return dt.weekday() == 3 and (dt + timedelta(days=7)).month != dt.month


def select_expiry(expiry_list: list[datetime], today: datetime,
                  ivr: float, r2: float) -> tuple[datetime | None, str]:
    if not expiry_list:
        return None, "No CSVs downloaded for this index"

    nearest     = expiry_list[0]
    dte_nearest = (nearest - today).days

    # Find best monthly candidate
    monthly = expiry_list[-1]
    for exp in expiry_list:
        if _is_last_thursday(exp):
            monthly = exp
            break

    if len(expiry_list) == 1:
        label = "monthly" if _is_last_thursday(nearest) else "only available expiry"
        return nearest, (f"{label.capitalize()} — only one CSV downloaded "
                         f"({nearest.strftime('%d-%b-%Y')}, {dte_nearest}d). "
                         f"Download both weekly + monthly CSVs for full selection.")

    if dte_nearest < CONFIG.MIN_DTE_WEEKLY:
        mdte = (monthly - today).days
        return monthly, (f"Monthly selected: weekly {nearest.strftime('%d-%b-%Y')} "
                         f"too close ({dte_nearest}d < min {CONFIG.MIN_DTE_WEEKLY}d). "
                         f"Using {monthly.strftime('%d-%b-%Y')} ({mdte}d).")

    if ivr < CONFIG.IVR_CHEAP and r2 > CONFIG.HIGH_CONF_R2:
        return nearest, (f"Weekly selected: IVR={ivr:.0f} (cheap) + "
                         f"R²={r2:.2f} (high confidence). "
                         f"Expiry {nearest.strftime('%d-%b-%Y')} ({dte_nearest}d).")

    mdte = (monthly - today).days
    why  = []
    if ivr >= CONFIG.IVR_CHEAP:
        why.append(f"IVR={ivr:.0f} not cheap enough for weekly")
    if r2 <= CONFIG.HIGH_CONF_R2:
        why.append(f"R²={r2:.2f} below threshold")
    return monthly, (f"Monthly selected: {', '.join(why)}. "
                     f"Expiry {monthly.strftime('%d-%b-%Y')} ({mdte}d).")


# ─────────────────────────────────────────────────────────────
# STEP 6 — STRIKE SELECTION
# ─────────────────────────────────────────────────────────────
def select_strike(records: list[dict], spot: float,
                  r2: float, option_type: str) -> dict | None:
    if not records or spot <= 0:
        return None

    target_delta     = CONFIG.TARGET_DELTA_HIGH if r2 > CONFIG.HIGH_CONF_R2 \
                       else CONFIG.TARGET_DELTA_LOW
    target_moneyness = max(0.005, (0.5 - target_delta) * 0.33)

    ltp_k = "call_ltp" if option_type == "CE" else "put_ltp"
    iv_k  = "call_iv"  if option_type == "CE" else "put_iv"
    oi_k  = "call_oi"  if option_type == "CE" else "put_oi"
    bid_k = "call_bid" if option_type == "CE" else "put_bid"
    ask_k = "call_ask" if option_type == "CE" else "put_ask"

    for r in records:
        r["_mono"] = abs(r["strike"] - spot) / spot

    side = [r for r in records
            if r[ltp_k] > 0
            and (r["strike"] >= spot if option_type == "CE" else r["strike"] <= spot)]

    if not side:
        return None

    best   = min(side, key=lambda r: abs(r["_mono"] - target_moneyness))
    delta  = round(max(0.10, min(0.50, 0.5 - best["_mono"] / 0.33)), 2)

    return {
        "strike":        int(best["strike"]),
        "ltp":           best[ltp_k],
        "iv":            best[iv_k],
        "oi":            best[oi_k],
        "bid":           best[bid_k],
        "ask":           best[ask_k],
        "moneyness_pct": round(best["_mono"] * 100, 2),
        "approx_delta":  delta,
    }


# ─────────────────────────────────────────────────────────────
# STEP 7 — LOT SIZING + EXIT LEVELS
# ─────────────────────────────────────────────────────────────
def prev_trading_day(dt: datetime) -> datetime:
    """
    Returns dt itself if it is a weekday (Mon-Fri).
    If dt falls on Saturday, returns the preceding Friday.
    If dt falls on Sunday, returns the preceding Friday.
    Does not account for public holidays — NSE holiday calendar
    is not fetched here. Handle holidays manually if needed.
    """
    while dt.weekday() >= 5:   # 5=Saturday, 6=Sunday
        dt -= timedelta(days=1)
    return dt


def compute_trade(index: str, sd: dict, expiry: datetime, today: datetime) -> dict | None:
    lot_size     = LOT_SIZES.get(index, 75)
    ltp          = sd["ltp"]
    if ltp <= 0:
        print("  [SKIP] LTP = ₹0"); return None

    cost_per_lot = lot_size * ltp
    if cost_per_lot > CONFIG.MAX_PREMIUM:
        print(f"  [SKIP] 1 lot = ₹{cost_per_lot:,.0f} > budget ₹{CONFIG.MAX_PREMIUM:,}")
        return None

    lots          = max(1, math.floor(CONFIG.MAX_PREMIUM / cost_per_lot))
    total         = round(lots * cost_per_lot, 0)

    return {
        "lot_size":      lot_size,
        "lots":          lots,
        "cost_per_lot":  round(cost_per_lot, 0),
        "total_premium": total,
        "max_loss":      total,
        "sl_ltp":        round(ltp * CONFIG.SL_PCT, 2),
        "sl_pnl":        round(-total * (1 - CONFIG.SL_PCT), 0),
        "target_ltp":    round(ltp * CONFIG.TARGET_PCT, 2),
        "target_pnl":    round(total * (CONFIG.TARGET_PCT - 1), 0),
        "time_stop":     prev_trading_day(expiry - timedelta(days=CONFIG.TIME_STOP_DAYS)),
    }


# ─────────────────────────────────────────────────────────────
# STEP 8 — GENERATE ALL SIGNALS
# ─────────────────────────────────────────────────────────────
def _blank_signal(**kwargs) -> dict:
    base = {
        "run_date":"","index":"","sector":"","inv_rank":0,"regime":"",
        "spot":0,"atm_iv":0,"iv_rank":0,"strategy":"","strike":0,
        "strike_label":"","strike_iv":0,"approx_delta":0,"moneyness_pct":0,
        "entry_ltp":0,"expiry":"","dte":0,"expiry_reason":"",
        "lot_size":0,"lots":0,"cost_per_lot":0,"total_premium":0,
        "max_loss":0,"sl_ltp":0,"sl_pnl":0,"target_ltp":0,"target_pnl":0,
        "time_stop":"","r2_proxy":0,"clenow_bld":0,"status":"","skip_reason":"",
    }
    base.update(kwargs)
    return base


def generate_signals(regime_data, top_sectors, chains, iv_history, today) -> list[dict]:
    signals   = []
    regime    = regime_data["regime"]
    today_str = today.strftime("%Y-%m-%d")

    print(f"\n{'='*60}")
    print(f"  Regime: {regime}  |  Generating signals...")
    print(f"{'='*60}")

    if regime == "BEAR":
        print("  BEAR regime — no buy signals. Stay in cash.")
        return []

    seen = set()

    for _, row in top_sectors.iterrows():
        sector = str(row["sector"]).upper()
        index  = SECTOR_TO_INDEX.get(sector, "NIFTY")
        rank   = int(row["inv_rank"])
        base   = dict(run_date=today_str, index=index or "", sector=sector,
                      inv_rank=rank, regime=regime)

        if index is None:
            print(f"\n[Skip] {sector} → no liquid options index (e.g. GOVT_BONDS)"); continue
        if index in seen:
            print(f"\n[Skip] {index} already covered by a higher-ranked sector"); continue

        print(f"\n{'─'*50}")
        print(f"[Signal] {sector} rank {rank} → {index}")

        if index not in chains:
            msg = f"No CSV downloaded for {index} — see instructions at top of script"
            print(f"  [SKIP] {msg}")
            signals.append(_blank_signal(**base, strategy="SKIPPED", status="SKIPPED", skip_reason=msg))
            seen.add(index); continue

        nearest_expiry, nearest_records, spot = chains[index][0]
        print(f"  Spot: ₹{spot:,.1f}  |  {len(nearest_records)} strikes loaded")

        atm_iv     = compute_atm_iv(nearest_records, spot)
        iv_history = record_iv(iv_history, index, atm_iv, today_str)
        ivr        = compute_iv_rank(iv_history, index, atm_iv)
        print(f"  ATM IV: {atm_iv:.1f}%  |  IV Rank: {ivr:.0f}")

        if ivr > CONFIG.IVR_MAX_TO_BUY:
            msg = (f"IVR {ivr:.0f} > {CONFIG.IVR_MAX_TO_BUY} — "
                   f"premium too expensive. Wait for IV to cool.")
            print(f"  [SKIP] {msg}")
            signals.append(_blank_signal(**base, spot=spot, atm_iv=atm_iv, iv_rank=ivr,
                                         strategy="SKIPPED", status="SKIPPED", skip_reason=msg))
            seen.add(index); continue

        clenow = float(row["clenow_bld"]) if not math.isnan(float(row["clenow_bld"] or 0)) else 0.5
        r2     = min(0.99, max(0.50, 0.50 + clenow * 0.5))

        expiry_list = [exp for (exp, _, _) in chains[index]]
        expiry, expiry_reason = select_expiry(expiry_list, today, ivr, r2)

        if expiry is None:
            signals.append(_blank_signal(**base, spot=spot, atm_iv=atm_iv, iv_rank=ivr,
                                         strategy="SKIPPED", status="SKIPPED", skip_reason=expiry_reason))
            seen.add(index); continue

        dte = (expiry - today).days
        print(f"  Expiry: {expiry.strftime('%d-%b-%Y')} ({dte}d) | {expiry_reason}")

        # Use records for the selected expiry
        sel_records, sel_spot = nearest_records, spot
        for (exp, recs, sp) in chains[index]:
            if exp == expiry:
                sel_records, sel_spot = recs, sp; break

        sd = select_strike(sel_records, sel_spot, r2, "CE")
        if sd is None:
            msg = f"No suitable strike found in chain for {expiry.strftime('%d-%b-%Y')}"
            print(f"  [SKIP] {msg}")
            signals.append(_blank_signal(**base, spot=sel_spot, atm_iv=atm_iv, iv_rank=ivr,
                                         strategy="SKIPPED", status="SKIPPED", skip_reason=msg))
            seen.add(index); continue

        print(f"  Strike: {sd['strike']} CE  |  LTP ₹{sd['ltp']:.2f}  |  "
              f"IV {sd['iv']:.1f}%  |  Delta ≈ {sd['approx_delta']}  |  OTM {sd['moneyness_pct']:.1f}%")

        trade = compute_trade(index, sd, expiry, today)
        if trade is None:
            signals.append(_blank_signal(**base, spot=sel_spot, atm_iv=atm_iv, iv_rank=ivr,
                                         strategy="SKIPPED", status="SKIPPED",
                                         skip_reason=f"1 lot cost ₹{sd['ltp']*LOT_SIZES.get(index,75):,.0f} > ₹{CONFIG.MAX_PREMIUM:,} budget"))
            seen.add(index); continue

        sig = _blank_signal(
            **base,
            spot=sel_spot, atm_iv=atm_iv, iv_rank=ivr,
            strategy="Buy Call", strike=sd["strike"],
            strike_label=f"{sd['strike']} CE", strike_iv=sd["iv"],
            approx_delta=sd["approx_delta"], moneyness_pct=sd["moneyness_pct"],
            entry_ltp=sd["ltp"],
            expiry=expiry.strftime("%d-%b-%Y"), dte=dte,
            expiry_reason=expiry_reason,
            lot_size=trade["lot_size"], lots=trade["lots"],
            cost_per_lot=trade["cost_per_lot"],
            total_premium=trade["total_premium"],
            max_loss=trade["max_loss"],
            sl_ltp=trade["sl_ltp"], sl_pnl=trade["sl_pnl"],
            target_ltp=trade["target_ltp"], target_pnl=trade["target_pnl"],
            time_stop=trade["time_stop"].strftime("%d-%b-%Y"),
            r2_proxy=round(r2, 3), clenow_bld=round(clenow, 4),
            status="ACTIVE", skip_reason="",
        )
        signals.append(sig)
        seen.add(index)

        print(f"\n  *** TRADE SIGNAL ***")
        print(f"  {index} {sd['strike']} CE | Entry ₹{sd['ltp']:.2f} | "
              f"{trade['lots']} lot(s) | Total ₹{trade['total_premium']:,.0f}")
        print(f"  Stop-loss  : ₹{trade['sl_ltp']}  (₹{abs(trade['sl_pnl']):,.0f} loss if hit)")
        print(f"  Target     : ₹{trade['target_ltp']}  (₹{trade['target_pnl']:,.0f} gain if hit)")
        print(f"  Time stop  : {trade['time_stop'].strftime('%d-%b-%Y')} (2d before expiry)")

    return signals


# ─────────────────────────────────────────────────────────────
# STEP 9 — SIGNAL HISTORY
# ─────────────────────────────────────────────────────────────
def load_signals_log(log_file: str) -> list:
    return json.load(open(log_file)) if os.path.exists(log_file) else []


def save_signals_log(log_file: str, log: list):
    json.dump(log, open(log_file, "w"), indent=2, default=str)


# ─────────────────────────────────────────────────────────────
# STEP 10 — WRITE options_trades.xlsx
# ─────────────────────────────────────────────────────────────
DISPLAY_COLS = [
    ("Run date",         "run_date"),
    ("Index",            "index"),
    ("Regime",           "regime"),
    ("Sector (driver)",  "sector"),
    ("Inv rank",         "inv_rank"),
    ("Spot (₹)",         "spot"),
    ("ATM IV%",          "atm_iv"),
    ("IV rank",          "iv_rank"),
    ("Strategy",         "strategy"),
    ("Strike",           "strike_label"),
    ("Strike IV%",       "strike_iv"),
    ("Delta (approx)",   "approx_delta"),
    ("OTM %",            "moneyness_pct"),
    ("Entry LTP (₹)",    "entry_ltp"),
    ("Expiry",           "expiry"),
    ("DTE",              "dte"),
    ("Expiry reason",    "expiry_reason"),
    ("Lot size",         "lot_size"),
    ("Lots",             "lots"),
    ("Cost/lot (₹)",     "cost_per_lot"),
    ("Total premium (₹)","total_premium"),
    ("Max loss (₹)",     "max_loss"),
    ("Stop-loss LTP",    "sl_ltp"),
    ("Stop-loss P&L",    "sl_pnl"),
    ("Target LTP",       "target_ltp"),
    ("Target P&L",       "target_pnl"),
    ("Time stop date",   "time_stop"),
    ("Clenow blended",   "clenow_bld"),
    ("R² proxy",         "r2_proxy"),
    ("Status",           "status"),
    ("Skip reason",      "skip_reason"),
]


def save_excel(signals: list[dict], history: list[dict], output_file: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trade Signals"

    H  = PatternFill("solid", fgColor="1F3864")
    A  = PatternFill("solid", fgColor="C6EFCE")
    SK = PatternFill("solid", fgColor="FFEB9C")
    BR = PatternFill("solid", fgColor="FFC7CE")
    SC = PatternFill("solid", fgColor="D9E1F2")

    hf = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
    nf = Font(name="Calibri", size=10)
    bf = Font(name="Calibri", bold=True,  size=10)
    rf = Font(name="Calibri", bold=True,  color="C00000", size=10)
    gf = Font(name="Calibri", bold=True,  color="375623", size=10)
    tf = Font(name="Calibri", bold=True,  color="1F3864", size=13)

    th = Side(style="thin", color="BFBFBF")
    BD = Border(left=th, right=th, top=th, bottom=th)
    CC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LC = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    nc = len(DISPLAY_COLS)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    c = ws.cell(row=1, column=1,
                value=f"NSE Options Trade Signals — {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    c.font = tf; c.alignment = CC
    ws.row_dimensions[1].height = 26

    # Header row
    for ci, (hdr, _) in enumerate(DISPLAY_COLS, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = hf; c.fill = H; c.alignment = CC; c.border = BD
    ws.row_dimensions[2].height = 30

    active  = [s for s in signals if s.get("status") == "ACTIVE"]
    skipped = [s for s in signals if s.get("status") == "SKIPPED"]

    if not signals:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=nc)
        c = ws.cell(row=3, column=1, value="BEAR REGIME — No trade signals. Stay in cash.")
        c.font = rf; c.fill = BR; c.alignment = CC
        ws.row_dimensions[3].height = 22
    else:
        for ri, sig in enumerate(active + skipped, start=3):
            fill = A if sig["status"] == "ACTIVE" else SK
            for ci, (hdr, key) in enumerate(DISPLAY_COLS, 1):
                val = sig.get(key, "")
                c   = ws.cell(row=ri, column=ci, value=val)
                c.font = nf; c.fill = fill; c.alignment = CC; c.border = BD
                if "Stop-loss"  in hdr:            c.font = rf
                elif "Target"   in hdr and "Time" not in hdr: c.font = gf
                elif hdr in ("Total premium (₹)", "Max loss (₹)"): c.font = bf
                elif hdr == "Expiry reason":        c.alignment = LC
            ws.row_dimensions[ri].height = 18

    # Exit rules box
    if active:
        sr = len(signals) + 4
        ws.cell(row=sr, column=1, value="EXIT RULES").font = bf
        ws.cell(row=sr, column=1).fill = SC
        for i, (rule, desc) in enumerate([
            ("Stop-loss (50%)", "Exit immediately when LTP falls to the Stop-loss LTP value."),
            ("Profit target",   "Exit when LTP hits Target LTP (100% gain). Book the profit."),
            ("Time stop",       "Exit on Time Stop Date — 2 days before expiry, regardless of P&L."),
            ("Weekly check",    "Run script every Mon/Fri. Check all three exit levels before placing new trades."),
        ], 1):
            ws.cell(row=sr+i, column=1, value=rule).font = bf
            c = ws.cell(row=sr+i, column=2, value=desc)
            c.font = nf; c.alignment = LC
            ws.merge_cells(start_row=sr+i, start_column=2, end_row=sr+i, end_column=10)

    # Column widths
    for ci, w in enumerate(
        [13,11,9,18,9,12,9,9,12,16,11,13,9,13,14,6,36,10,7,13,18,13,14,14,13,14,14,13,9,10,28],
        1
    ):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    # History sheet
    ws2 = wb.create_sheet("Signal History")
    all_hist = history + signals
    if all_hist:
        hist_df = pd.DataFrame(all_hist)
        for ci, col in enumerate(hist_df.columns, 1):
            c = ws2.cell(row=1, column=ci, value=col)
            c.font = hf; c.fill = H; c.alignment = CC; c.border = BD
        for ri, (_, row) in enumerate(hist_df.iterrows(), 2):
            st   = str(row.get("status",""))
            fill = A if st=="ACTIVE" else (SK if st=="SKIPPED" else PatternFill("solid", fgColor="FFFFFF"))
            for ci, val in enumerate(row.values, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.font = nf; c.fill = fill; c.alignment = CC; c.border = BD
        for ci in range(1, len(hist_df.columns)+1):
            ws2.column_dimensions[get_column_letter(ci)].width = 16
        ws2.freeze_panes = "A2"

    wb.save(output_file)
    print(f"\n[Done] Saved → {output_file}")


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    print("=" * 60)
    print("  NSE Options Trade Signal Generator  v2.0")
    print(f"  {today.strftime('%d %b %Y (%A)')}")
    print("=" * 60)

    regime_data = read_regime(CONFIG.HOLDINGS_LOG_FILE)
    top_sectors = read_top_sectors(CONFIG.ETF_RANKINGS_FILE)
    chains      = load_all_chains(CONFIG.CSV_PATTERN)
    iv_history  = load_iv_history(CONFIG.IV_HISTORY_FILE)

    signals     = generate_signals(regime_data, top_sectors, chains, iv_history, today)

    save_iv_history(CONFIG.IV_HISTORY_FILE, iv_history)

    history     = load_signals_log(CONFIG.SIGNALS_LOG_FILE)
    save_excel(signals, history, CONFIG.OUTPUT_FILE)
    history.extend(signals)
    save_signals_log(CONFIG.SIGNALS_LOG_FILE, history)

    # Summary
    active  = [s for s in signals if s["status"] == "ACTIVE"]
    skipped = [s for s in signals if s["status"] == "SKIPPED"]

    print("\n" + "=" * 60)
    print("  SIGNAL SUMMARY")
    print("=" * 60)
    print(f"  Regime  : {regime_data['regime']}")
    print(f"  Active  : {len(active)}  |  Skipped: {len(skipped)}")
    if not signals:
        print("  → BEAR regime. No trades. Full cash.")
    for s in active:
        print(f"\n  [{s['index']}] {s['strike_label']}  |  Expiry {s['expiry']} ({s['dte']}d)")
        print(f"    Entry  ₹{s['entry_ltp']:.2f}  |  {s['lots']} lot(s)  |  Total ₹{s['total_premium']:,.0f}")
        print(f"    SL     ₹{s['sl_ltp']}  (₹{abs(s['sl_pnl']):,.0f} loss if hit)")
        print(f"    Target ₹{s['target_ltp']}  (₹{s['target_pnl']:,.0f} gain if hit)")
        print(f"    Exit by {s['time_stop']} (time stop)")
    for s in skipped:
        print(f"\n  [SKIP] {s['index']} — {s['skip_reason']}")
    print(f"\n  Output → {CONFIG.OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
