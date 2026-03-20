"""
NSE Options Position Monitor
==============================
Tracks MTM on open positions from options_trader.py signals.
Run this daily (or any time) with fresh downloaded CSVs.

HOW TO USE:
  1. Download fresh option chain CSVs from nseindia.com/option-chain
     (same way you did when generating signals — one per index/expiry)
  2. Put them in this folder
  3. Run: python options_monitor.py

The script reads your open positions from signals_log.json,
looks up the current LTP from the fresh CSVs, and tells you
exactly what to do — HOLD, EXIT (stop-loss hit), EXIT (target hit),
or EXIT (time stop approaching).

DEPS: pip install pandas openpyxl
"""

import csv
import json
import os
import re
from datetime import datetime, timedelta
from glob import glob

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
SIGNALS_LOG   = "signals_log.json"
CSV_PATTERN   = "option-chain-ED-*.csv"
OUTPUT_FILE   = "options_monitor.xlsx"

# Exit alert thresholds
TIME_STOP_WARN_DAYS = 3    # warn when time stop is within N days


# ─────────────────────────────────────────────
# LOAD OPEN POSITIONS FROM signals_log.json
# ─────────────────────────────────────────────
def load_open_positions() -> list[dict]:
    if not os.path.exists(SIGNALS_LOG):
        print(f"[ERROR] {SIGNALS_LOG} not found. Run options_trader.py first.")
        return []

    with open(SIGNALS_LOG) as f:
        log = json.load(f)

    open_pos = [s for s in log if s.get("status") == "ACTIVE"]

    if not open_pos:
        print("[INFO] No open ACTIVE positions found in signals_log.json")
        print("       Run options_trader.py first to generate trade signals.")

    return open_pos


# ─────────────────────────────────────────────
# PARSE CSVs — reused from options_trader.py
# ─────────────────────────────────────────────
INDEX_KEYWORDS = {
    "BANKNIFTY":  ["BANKNIFTY", "BANK-NIFTY"],
    "FINNIFTY":   ["FINNIFTY", "FIN-NIFTY"],
    "MIDCPNIFTY": ["MIDCPNIFTY", "MIDCP-NIFTY"],
    "NIFTYNXT50": ["NIFTYNXT50", "NIFTYNXT-50", "NIFTYNEXT50"],
    "NIFTY":      ["NIFTY"],
}

def prev_trading_day(dt: datetime) -> datetime:
    """Roll back to nearest weekday. Saturday/Sunday → preceding Friday."""
    while dt.weekday() >= 5:
        dt -= timedelta(days=1)
    return dt


def clean_num(val) -> float | None:
    if val is None: return None
    s = str(val).strip()
    if s in ("-", "", "nan"): return None
    try: return float(s.replace(",", ""))
    except: return None

def detect_index(fname: str) -> str | None:
    name = os.path.basename(fname).upper()
    for index, kws in INDEX_KEYWORDS.items():
        if any(k in name for k in kws):
            return index
    return None

def detect_expiry(fname: str) -> datetime | None:
    m = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", os.path.basename(fname))
    if m:
        try: return datetime.strptime(m.group(1), "%d-%b-%Y")
        except: pass
    return None

def parse_csv_for_ltp(filepath: str) -> dict:
    """
    Returns a dict keyed by strike → {call_ltp, put_ltp, call_iv, put_iv}
    from a downloaded NSE options chain CSV.
    """
    data = {}
    rows = []
    with open(filepath, encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            rows.append(row)

    for row in rows[2:]:
        if len(row) < 22: continue
        strike = clean_num(row[11])
        if not strike: continue
        data[int(strike)] = {
            "call_ltp": clean_num(row[5])  or 0,
            "call_iv":  clean_num(row[4])  or 0,
            "put_ltp":  clean_num(row[17]) or 0,
            "put_iv":   clean_num(row[18]) or 0,
        }
    return data


def load_fresh_chains() -> dict:
    """
    Loads all downloaded CSVs.
    Returns: { ("NIFTY", "27-Mar-2026"): {strike: {call_ltp, ...}}, ... }
    """
    chains   = {}
    csvfiles = sorted(glob(CSV_PATTERN))

    if not csvfiles:
        print(f"[WARN] No CSV files found. Download fresh CSVs from nseindia.com/option-chain")
        return chains

    for fp in csvfiles:
        index  = detect_index(fp)
        expiry = detect_expiry(fp)
        if not index or not expiry:
            continue
        key            = (index, expiry.strftime("%d-%b-%Y"))
        chains[key]    = parse_csv_for_ltp(fp)
        print(f"  Loaded: {os.path.basename(fp)} → {index} {expiry.strftime('%d-%b-%Y')} "
              f"({len(chains[key])} strikes)")

    return chains


# ─────────────────────────────────────────────
# LOOKUP CURRENT LTP FOR A POSITION
# ─────────────────────────────────────────────
def get_current_ltp(pos: dict, chains: dict) -> float | None:
    """
    Finds the current LTP for a position from the loaded chain data.
    Tries exact expiry match first, then nearest available expiry.
    """
    index        = pos.get("index", "")
    strike       = int(pos.get("strike", 0))
    expiry_str   = pos.get("expiry", "")
    option_type  = "CE"  # we only buy calls in this system

    if not index or not strike:
        return None

    # Try exact expiry match
    key = (index, expiry_str)
    if key in chains:
        row = chains[key].get(strike)
        if row:
            return row["call_ltp"] if option_type == "CE" else row["put_ltp"]

    # Try any available expiry for this index (in case user downloaded different expiry)
    for (idx, exp_str), strike_data in chains.items():
        if idx == index:
            row = strike_data.get(strike)
            if row:
                ltp = row["call_ltp"] if option_type == "CE" else row["put_ltp"]
                if ltp > 0:
                    print(f"  [Note] Used {exp_str} CSV (exact expiry {expiry_str} not found)")
                    return ltp

    return None


# ─────────────────────────────────────────────
# COMPUTE MTM + EXIT STATUS
# ─────────────────────────────────────────────
def compute_mtm(pos: dict, current_ltp: float | None, today: datetime) -> dict:
    """
    Given a position and current LTP, computes:
      - unrealised P&L
      - % move from entry
      - which exit rule fires (if any)
      - recommended action
    """
    entry_ltp    = float(pos.get("entry_ltp", 0))
    sl_ltp       = float(pos.get("sl_ltp", 0))
    target_ltp   = float(pos.get("target_ltp", 0))
    lots         = int(pos.get("lots", 1))
    lot_size     = int(pos.get("lot_size", 75))
    total_prem   = float(pos.get("total_premium", 0))
    time_stop_s  = pos.get("time_stop", "")

    try:
        time_stop = datetime.strptime(time_stop_s, "%d-%b-%Y")
    except Exception:
        time_stop = None

    if time_stop:
        time_stop    = prev_trading_day(time_stop)
        days_to_ts   = (time_stop - today).days
    else:
        days_to_ts   = 999
    expiry_s     = pos.get("expiry", "")
    try:
        expiry       = datetime.strptime(expiry_s, "%d-%b-%Y")
        days_to_exp  = (expiry - today).days
    except Exception:
        days_to_exp  = 999

    result = {
        "current_ltp":   current_ltp,
        "entry_ltp":     entry_ltp,
        "pnl_per_lot":   None,
        "total_pnl":     None,
        "pnl_pct":       None,
        "days_to_exp":   days_to_exp,
        "days_to_ts":    days_to_ts,
        "action":        "HOLD",
        "exit_reason":   "",
        "alert_level":   "green",   # green / amber / red
    }

    if current_ltp is None:
        result["action"]      = "CHECK MANUALLY"
        result["exit_reason"] = "LTP not found in downloaded CSV — download fresh CSV for this expiry"
        result["alert_level"] = "amber"
        return result

    # P&L calculation
    pnl_per_unit = current_ltp - entry_ltp
    pnl_per_lot  = pnl_per_unit * lot_size
    total_pnl    = pnl_per_lot  * lots
    pnl_pct      = (current_ltp / entry_ltp - 1) * 100 if entry_ltp > 0 else 0

    result["pnl_per_lot"] = round(pnl_per_lot, 0)
    result["total_pnl"]   = round(total_pnl, 0)
    result["pnl_pct"]     = round(pnl_pct, 1)

    # Check exit rules in priority order
    if current_ltp <= sl_ltp:
        result["action"]      = "EXIT NOW — STOP-LOSS HIT"
        result["exit_reason"] = (f"LTP ₹{current_ltp:.2f} ≤ stop-loss ₹{sl_ltp:.2f}. "
                                  f"Exit immediately. Loss: ₹{abs(total_pnl):,.0f}")
        result["alert_level"] = "red"

    elif current_ltp >= target_ltp:
        result["action"]      = "EXIT NOW — TARGET HIT"
        result["exit_reason"] = (f"LTP ₹{current_ltp:.2f} ≥ target ₹{target_ltp:.2f}. "
                                  f"Book profit: ₹{total_pnl:,.0f}")
        result["alert_level"] = "green_exit"

    elif days_to_ts <= 0:
        result["action"]      = "EXIT NOW — TIME STOP"
        result["exit_reason"] = (f"Time stop date {time_stop_s} reached. "
                                  f"Exit today regardless of P&L. "
                                  f"Current P&L: ₹{total_pnl:+,.0f}")
        result["alert_level"] = "red"

    elif days_to_ts <= TIME_STOP_WARN_DAYS:
        result["action"]      = f"PREPARE TO EXIT — {days_to_ts}d to time stop"
        result["exit_reason"] = (f"Time stop is {time_stop_s} ({days_to_ts}d away). "
                                  f"Start monitoring closely. P&L: ₹{total_pnl:+,.0f}")
        result["alert_level"] = "amber"

    elif pnl_pct >= 75:
        result["action"]      = "CONSIDER EXITING — near target"
        result["exit_reason"] = (f"Up {pnl_pct:.0f}% — within 25% of target ₹{target_ltp:.2f}. "
                                  f"Consider partial exit or tighten mental stop.")
        result["alert_level"] = "amber"

    elif pnl_pct <= -40:
        result["action"]      = "WATCH CLOSELY — approaching stop-loss"
        result["exit_reason"] = (f"Down {abs(pnl_pct):.0f}% — approaching stop-loss ₹{sl_ltp:.2f}. "
                                  f"Do not average down. Current P&L: ₹{total_pnl:+,.0f}")
        result["alert_level"] = "amber"

    else:
        result["action"]      = "HOLD"
        result["exit_reason"] = (f"No exit triggered. Next check: "
                                  f"SL ₹{sl_ltp:.2f} | Target ₹{target_ltp:.2f} | "
                                  f"Time stop {time_stop_s} ({days_to_ts}d)")
        result["alert_level"] = "green"

    return result


# ─────────────────────────────────────────────
# WRITE options_monitor.xlsx
# ─────────────────────────────────────────────
def save_monitor_excel(positions: list[dict], mtm_results: list[dict],
                        today: datetime, output_file: str):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Position Monitor"

    # Styles
    HDR  = PatternFill("solid", fgColor="1F3864")
    RED  = PatternFill("solid", fgColor="FFC7CE")
    AMB  = PatternFill("solid", fgColor="FFEB9C")
    GRN  = PatternFill("solid", fgColor="C6EFCE")
    GEX  = PatternFill("solid", fgColor="00B050")   # bright green = exit for profit
    WHT  = PatternFill("solid", fgColor="FFFFFF")
    SEC  = PatternFill("solid", fgColor="D9E1F2")

    hf   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    nf   = Font(name="Calibri", size=10)
    bf   = Font(name="Calibri", bold=True, size=10)
    rf   = Font(name="Calibri", bold=True, color="C00000", size=10)
    gf   = Font(name="Calibri", bold=True, color="375623", size=10)
    wgf  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    tf   = Font(name="Calibri", bold=True, color="1F3864", size=13)

    th   = Side(style="thin", color="BFBFBF")
    BD   = Border(left=th, right=th, top=th, bottom=th)
    CC   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LC   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Title
    title_cols = 18
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_cols)
    t           = ws.cell(row=1, column=1,
                          value=f"NSE Options — Position Monitor  |  {today.strftime('%d %b %Y (%A)')}")
    t.font      = tf
    t.alignment = CC
    ws.row_dimensions[1].height = 26

    # Column headers
    headers = [
        "Index", "Strike", "Expiry", "DTE",
        "Entry LTP", "Current LTP", "Move %",
        "Lots", "Total premium", "Unrealised P&L",
        "Stop-loss", "Target", "Time stop", "Days to TS",
        "Action", "Detail",
        "Entry date", "Sector",
    ]
    for ci, h in enumerate(headers, 1):
        c           = ws.cell(row=2, column=ci, value=h)
        c.font      = hf
        c.fill      = HDR
        c.alignment = CC
        c.border    = BD
    ws.row_dimensions[2].height = 30

    if not positions:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=title_cols)
        c           = ws.cell(row=3, column=1, value="No open positions found in signals_log.json")
        c.font      = bf
        c.alignment = CC
        ws.row_dimensions[3].height = 22
    else:
        for ri, (pos, mtm) in enumerate(zip(positions, mtm_results), start=3):
            # Pick row fill based on alert level
            al   = mtm["alert_level"]
            fill = RED if al == "red" else (AMB if al == "amber" else
                   GEX if al == "green_exit" else GRN)
            action_font = (rf  if al == "red"        else
                           Font(name="Calibri", bold=True, color="FF8C00", size=10) if al == "amber" else
                           wgf if al == "green_exit" else gf)

            pnl     = mtm["total_pnl"]
            pnl_pct = mtm["pnl_pct"]
            cur_ltp = mtm["current_ltp"]

            row_vals = [
                pos.get("index",""),
                pos.get("strike_label", str(pos.get("strike",""))),
                pos.get("expiry",""),
                mtm["days_to_exp"],
                pos.get("entry_ltp", 0),
                cur_ltp if cur_ltp is not None else "No CSV",
                f"{pnl_pct:+.1f}%" if pnl_pct is not None else "—",
                pos.get("lots", 0),
                pos.get("total_premium", 0),
                pnl if pnl is not None else "—",
                pos.get("sl_ltp", 0),
                pos.get("target_ltp", 0),
                pos.get("time_stop",""),
                mtm["days_to_ts"],
                mtm["action"],
                mtm["exit_reason"],
                pos.get("run_date",""),
                pos.get("sector",""),
            ]

            for ci, val in enumerate(row_vals, 1):
                c           = ws.cell(row=ri, column=ci, value=val)
                c.font      = nf
                c.fill      = fill
                c.alignment = CC
                c.border    = BD
                # Overrides
                if ci == 15:   # Action column
                    c.font = action_font
                if ci == 16:   # Detail column
                    c.alignment = LC
                if ci == 10 and pnl is not None:  # P&L
                    c.font = gf if pnl >= 0 else rf

            ws.row_dimensions[ri].height = 20

    # Column widths
    col_widths = [12,16,14,7,12,14,10,7,16,16,12,12,14,12,32,42,13,16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    # ── Legend ──────────────────────────────────────
    legend_row = len(positions) + 5 if positions else 5
    ws.cell(row=legend_row, column=1, value="Legend").font = bf
    ws.cell(row=legend_row, column=1).fill = SEC

    legend_items = [
        (GEX, wgf, "EXIT NOW — target hit (green + bold)"),
        (RED, rf,  "EXIT NOW — stop-loss hit or time stop reached"),
        (AMB, bf,  "Watch closely — approaching an exit level"),
        (GRN, gf,  "HOLD — no exit triggered"),
    ]
    for i, (fill, font, desc) in enumerate(legend_items, 1):
        c1 = ws.cell(row=legend_row+i, column=1, value=desc)
        c1.font = font; c1.fill = fill; c1.alignment = LC
        ws.merge_cells(start_row=legend_row+i, start_column=1,
                        end_row=legend_row+i,   end_column=5)

    # ── How to use box ───────────────────────────────
    hr = legend_row + 7
    ws.cell(row=hr, column=1, value="HOW TO USE THIS MONITOR").font = bf
    ws.cell(row=hr, column=1).fill = SEC
    steps = [
        "1. Download fresh option chain CSVs from nseindia.com/option-chain (same index + expiry as your open trades)",
        "2. Put the CSVs in this folder and re-run: python options_monitor.py",
        "3. Check the Action column — EXIT NOW means exit that day, no waiting",
        "4. HOLD means no action needed. Re-check tomorrow or on Monday/Friday",
        "5. If you exit a position, update its status to CLOSED in signals_log.json (change 'ACTIVE' to 'CLOSED')",
    ]
    for i, step in enumerate(steps, 1):
        c = ws.cell(row=hr+i, column=1, value=step)
        c.font = nf; c.alignment = LC
        ws.merge_cells(start_row=hr+i, start_column=1, end_row=hr+i, end_column=title_cols)

    wb.save(output_file)
    print(f"[Done] Saved → {output_file}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    print("=" * 60)
    print("  NSE Options Position Monitor")
    print(f"  {today.strftime('%d %b %Y (%A)')}")
    print("=" * 60)

    # Load open positions
    positions = load_open_positions()
    if not positions:
        print("\nNothing to monitor. Exiting.")
        return

    print(f"\n[Positions] {len(positions)} open trade(s) found:")
    for p in positions:
        print(f"  {p.get('index')} {p.get('strike_label','')} | "
              f"Expiry {p.get('expiry','')} | "
              f"Entry ₹{p.get('entry_ltp',0):.2f} | "
              f"SL ₹{p.get('sl_ltp',0)} | Target ₹{p.get('target_ltp',0)}")

    # Load fresh CSVs
    print(f"\n[CSVs] Scanning for option chain files...")
    chains = load_fresh_chains()

    # Compute MTM for each position
    print(f"\n[MTM] Computing mark-to-market...")
    print("=" * 60)

    mtm_results = []
    for pos in positions:
        cur_ltp = get_current_ltp(pos, chains)
        mtm     = compute_mtm(pos, cur_ltp, today)
        mtm_results.append(mtm)

        index  = pos.get("index","")
        strike = pos.get("strike_label","")
        pnl    = mtm["total_pnl"]
        action = mtm["action"]

        ltp_str = f"₹{cur_ltp:.2f}" if cur_ltp is not None else "not found"
        pnl_str = f"₹{pnl:+,.0f}" if pnl is not None else "—"

        print(f"\n  [{index}] {strike}")
        print(f"  Current LTP : {ltp_str}  (entry: ₹{pos.get('entry_ltp',0):.2f})")
        print(f"  P&L         : {pnl_str}  ({mtm['pnl_pct']:+.1f}%)" if mtm["pnl_pct"] is not None
              else f"  P&L         : {pnl_str}")
        print(f"  ACTION      : {action}")
        if mtm["exit_reason"]:
            print(f"  Detail      : {mtm['exit_reason']}")

    print("\n" + "=" * 60)

    # Save Excel
    save_monitor_excel(positions, mtm_results, today, "options_monitor.xlsx")

    # Summary
    exits  = [m for m in mtm_results if "EXIT NOW" in m["action"]]
    warns  = [m for m in mtm_results if "PREPARE" in m["action"] or "CONSIDER" in m["action"] or "WATCH" in m["action"]]
    holds  = [m for m in mtm_results if m["action"] == "HOLD"]

    print(f"\n  SUMMARY")
    print(f"  Exit now  : {len(exits)}")
    print(f"  Watch     : {len(warns)}")
    print(f"  Hold      : {len(holds)}")

    if exits:
        print(f"\n  *** ACTION REQUIRED — exit the following TODAY: ***")
        for pos, mtm in zip(positions, mtm_results):
            if "EXIT NOW" in mtm["action"]:
                print(f"  → {pos['index']} {pos.get('strike_label','')} | {mtm['action']}")

    print(f"\n  Output → options_monitor.xlsx")
    print("=" * 60)


if __name__ == "__main__":
    main()
