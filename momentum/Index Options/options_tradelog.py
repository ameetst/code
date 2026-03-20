"""
NSE Options Trade Log
======================
Closes open positions, records exit prices, and maintains a full
P&L workbook with trade-by-trade history and running statistics.

HOW TO USE:
  When you exit a trade (stop-loss / target / time stop), run:
    python options_tradelog.py

  The script will:
    1. Show all open positions
    2. Ask you which one(s) to close and at what exit price
    3. Record the exit in trade_log.json
    4. Write/update trade_log.xlsx with full P&L analysis

  You can also run it with no exits just to refresh the Excel report.

DEPS: pip install pandas openpyxl
"""

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SIGNALS_LOG  = "signals_log.json"
TRADE_LOG    = "trade_log.json"
OUTPUT_FILE  = "trade_log.xlsx"


# ─────────────────────────────────────────────
# LOAD / SAVE
# ─────────────────────────────────────────────
def load_json(path: str, default):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return default


def save_json(path: str, data):
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)


# ─────────────────────────────────────────────
# CLOSE A TRADE INTERACTIVELY
# ─────────────────────────────────────────────
def close_trades_interactively():
    """
    Prompts the user to close open positions.
    Updates signals_log.json status to CLOSED.
    Appends closed trades to trade_log.json.
    """
    signals = load_json(SIGNALS_LOG, [])
    open_trades = [s for s in signals if s.get("status") == "ACTIVE"]

    if not open_trades:
        print("\n[INFO] No open positions to close.")
        return

    print("\n" + "=" * 60)
    print("  CLOSE A POSITION")
    print("=" * 60)
    print("\nOpen positions:")
    for i, t in enumerate(open_trades):
        print(f"  [{i+1}] {t['index']} {t.get('strike_label','')} | "
              f"Expiry {t.get('expiry','')} | "
              f"Entry ₹{t.get('entry_ltp',0):.2f} | "
              f"SL ₹{t.get('sl_ltp',0)} | Target ₹{t.get('target_ltp',0)} | "
              f"Lots {t.get('lots',0)}")

    print(f"\n  [0] Done — no more closures")

    trade_log = load_json(TRADE_LOG, [])

    while True:
        try:
            choice = input("\nEnter position number to close (0 to finish): ").strip()
            if choice == "0":
                break
            idx = int(choice) - 1
            if idx < 0 or idx >= len(open_trades):
                print("  Invalid number. Try again.")
                continue

            trade = open_trades[idx]

            # Ask for exit price
            exit_ltp_str = input(
                f"  Exit LTP for {trade['index']} {trade.get('strike_label','')} "
                f"(entry was ₹{trade.get('entry_ltp',0):.2f}): ₹"
            ).strip()
            exit_ltp = float(exit_ltp_str.replace(",", ""))

            # Ask for exit reason
            print("  Exit reason:")
            print("    [1] Stop-loss hit")
            print("    [2] Target hit")
            print("    [3] Time stop")
            print("    [4] Manual / other")
            reason_choice = input("  Choose [1-4]: ").strip()
            reasons = {
                "1": "Stop-loss", "2": "Target hit",
                "3": "Time stop", "4": "Manual"
            }
            exit_reason = reasons.get(reason_choice, "Manual")

            exit_date = datetime.today().strftime("%Y-%m-%d")

            # Compute P&L
            lot_size     = int(trade.get("lot_size", 75))
            lots         = int(trade.get("lots", 1))
            entry_ltp    = float(trade.get("entry_ltp", 0))
            total_prem   = float(trade.get("total_premium", 0))

            pnl_per_unit = exit_ltp - entry_ltp
            pnl_per_lot  = pnl_per_unit * lot_size
            total_pnl    = pnl_per_lot * lots
            pnl_pct      = (exit_ltp / entry_ltp - 1) * 100 if entry_ltp > 0 else 0

            # Duration
            try:
                entry_date = datetime.strptime(trade.get("run_date","")[:10], "%Y-%m-%d")
                exit_dt    = datetime.strptime(exit_date, "%Y-%m-%d")
                hold_days  = (exit_dt - entry_date).days
            except Exception:
                hold_days  = 0

            closed = {
                **trade,
                "exit_ltp":    round(exit_ltp, 2),
                "exit_date":   exit_date,
                "exit_reason": exit_reason,
                "pnl_per_lot": round(pnl_per_lot, 0),
                "total_pnl":   round(total_pnl, 0),
                "pnl_pct":     round(pnl_pct, 1),
                "hold_days":   hold_days,
                "status":      "CLOSED",
            }

            # Update signals_log: mark as CLOSED
            for s in signals:
                if (s.get("index") == trade.get("index") and
                    s.get("strike") == trade.get("strike") and
                    s.get("run_date") == trade.get("run_date") and
                    s.get("status") == "ACTIVE"):
                    s["status"]      = "CLOSED"
                    s["exit_ltp"]    = closed["exit_ltp"]
                    s["exit_date"]   = exit_date
                    s["exit_reason"] = exit_reason
                    s["total_pnl"]   = closed["total_pnl"]
                    break

            trade_log.append(closed)

            pnl_str = f"₹{total_pnl:+,.0f}"
            print(f"\n  ✓ Closed: {trade['index']} {trade.get('strike_label','')} | "
                  f"Exit ₹{exit_ltp:.2f} | P&L {pnl_str} ({pnl_pct:+.1f}%) | "
                  f"Reason: {exit_reason}")

        except (ValueError, KeyboardInterrupt):
            print("  Invalid input. Skipping.")
            continue

    save_json(SIGNALS_LOG, signals)
    save_json(TRADE_LOG, trade_log)
    print(f"\n[Saved] {SIGNALS_LOG} and {TRADE_LOG} updated.")


# ─────────────────────────────────────────────
# WRITE trade_log.xlsx
# ─────────────────────────────────────────────
def write_trade_log_excel(trade_log: list, output_file: str):
    """
    Writes a 3-sheet workbook:
      Sheet 1: Trade Log        — every closed trade, row by row
      Sheet 2: P&L Summary      — monthly breakdown + running totals
      Sheet 3: Stats            — win rate, avg win/loss, expectancy
    """
    wb  = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────
    HDR   = PatternFill("solid", fgColor="1F3864")
    WIN   = PatternFill("solid", fgColor="C6EFCE")
    LOSS  = PatternFill("solid", fgColor="FFC7CE")
    EVEN  = PatternFill("solid", fgColor="F2F2F2")
    ODD   = PatternFill("solid", fgColor="FFFFFF")
    SEC   = PatternFill("solid", fgColor="D9E1F2")
    DARK  = PatternFill("solid", fgColor="1F3864")
    AMBER = PatternFill("solid", fgColor="FFEB9C")

    hf    = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
    nf    = Font(name="Arial", size=10)
    bf    = Font(name="Arial", bold=True,  size=10)
    rf    = Font(name="Arial", bold=True,  color="C00000", size=10)
    gf    = Font(name="Arial", bold=True,  color="375623", size=10)
    tf    = Font(name="Arial", bold=True,  color="1F3864", size=13)
    wf    = Font(name="Arial", bold=True,  color="FFFFFF", size=11)

    th    = Side(style="thin",   color="BFBFBF")
    BD    = Border(left=th, right=th, top=th, bottom=th)
    CC    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LC    = Alignment(horizontal="left",   vertical="center")
    RC    = Alignment(horizontal="right",  vertical="center")

    # ════════════════════════════════════════════
    # SHEET 1: TRADE LOG
    # ════════════════════════════════════════════
    ws1       = wb.active
    ws1.title = "Trade Log"

    LOG_COLS = [
        ("#",             ""),
        ("Entry date",    "run_date"),
        ("Index",         "index"),
        ("Strike",        "strike_label"),
        ("Expiry",        "expiry"),
        ("Sector",        "sector"),
        ("Regime",        "regime"),
        ("IV rank",       "iv_rank"),
        ("Lots",          "lots"),
        ("Lot size",      "lot_size"),
        ("Entry LTP (₹)", "entry_ltp"),
        ("Exit LTP (₹)",  "exit_ltp"),
        ("Move (₹)",      ""),           # formula
        ("Move %",        ""),           # formula
        ("Total prem (₹)","total_premium"),
        ("P&L (₹)",       "total_pnl"),
        ("P&L %",         "pnl_pct"),
        ("Hold days",     "hold_days"),
        ("Exit reason",   "exit_reason"),
        ("Exit date",     "exit_date"),
    ]
    NC = len(LOG_COLS)

    # Title
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    t = ws1.cell(row=1, column=1,
                 value=f"NSE Options — Trade Log  |  Updated {datetime.now().strftime('%d %b %Y')}")
    t.font = tf; t.alignment = CC
    ws1.row_dimensions[1].height = 26

    # Header
    for ci, (hdr, _) in enumerate(LOG_COLS, 1):
        c = ws1.cell(row=2, column=ci, value=hdr)
        c.font = hf; c.fill = HDR; c.alignment = CC; c.border = BD
    ws1.row_dimensions[2].height = 28

    closed = [t for t in trade_log if t.get("status") == "CLOSED"]

    if not closed:
        ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NC)
        c = ws1.cell(row=3, column=1, value="No closed trades yet.")
        c.font = bf; c.alignment = CC
    else:
        for ri, trade in enumerate(sorted(closed, key=lambda x: x.get("exit_date", "")), start=3):
            pnl  = float(trade.get("total_pnl", 0) or 0)
            fill = WIN if pnl >= 0 else LOSS

            entry_col = 11   # K = Entry LTP
            exit_col  = 12   # L = Exit LTP

            for ci, (hdr, key) in enumerate(LOG_COLS, 1):
                if key:
                    val = trade.get(key, "")
                    # format pnl_pct nicely
                    if key == "pnl_pct" and val != "":
                        try: val = float(val) / 100
                        except: pass
                elif hdr == "Move (₹)":
                    # Exit LTP - Entry LTP
                    ecol = get_column_letter(entry_col)
                    xcol = get_column_letter(exit_col)
                    val  = f"={xcol}{ri}-{ecol}{ri}"
                elif hdr == "Move %":
                    ecol = get_column_letter(entry_col)
                    xcol = get_column_letter(exit_col)
                    val  = f"=IF({ecol}{ri}=0,0,({xcol}{ri}-{ecol}{ri})/{ecol}{ri})"
                elif hdr == "#":
                    val = ri - 2
                else:
                    val = ""

                c = ws1.cell(row=ri, column=ci, value=val)
                c.font = nf; c.fill = fill; c.alignment = CC; c.border = BD

                # Colour P&L column
                if hdr == "P&L (₹)":
                    c.font = gf if pnl >= 0 else rf
                    c.font = Font(name="Arial", bold=True,
                                  color="375623" if pnl >= 0 else "C00000", size=10)
                if hdr in ("P&L %", "Move %"):
                    c.number_format = "0.0%"
                if hdr in ("P&L (₹)", "Total prem (₹)", "Move (₹)",
                           "Entry LTP (₹)", "Exit LTP (₹)"):
                    c.number_format = '#,##0.00'
                if hdr == "Exit reason":
                    c.alignment = LC

            ws1.row_dimensions[ri].height = 18

        # Totals row
        tr = len(closed) + 3
        ws1.cell(row=tr, column=1, value="TOTAL").font = bf
        ws1.cell(row=tr, column=1).fill = SEC

        pnl_col   = get_column_letter(16)
        prem_col  = get_column_letter(15)
        ws1.cell(row=tr, column=15, value=f"=SUM({prem_col}3:{prem_col}{tr-1})")
        ws1.cell(row=tr, column=16, value=f"=SUM({pnl_col}3:{pnl_col}{tr-1})")

        for col in [15, 16]:
            c = ws1.cell(row=tr, column=col)
            c.font = bf; c.fill = SEC; c.alignment = CC; c.border = BD
            c.number_format = '#,##0.00'

    # Column widths
    widths = [5,13,10,16,13,16,9,9,7,10,14,13,12,10,16,14,10,12,14,13]
    for ci, w in enumerate(widths[:NC], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.freeze_panes = "A3"

    # ════════════════════════════════════════════
    # SHEET 2: P&L SUMMARY (monthly)
    # ════════════════════════════════════════════
    ws2       = wb.create_sheet("P&L Summary")

    # Title
    ws2.merge_cells("A1:L1")
    t = ws2.cell(row=1, column=1, value="P&L Summary — by Month")
    t.font = tf; t.alignment = CC
    ws2.row_dimensions[1].height = 26

    # Monthly breakdown headers
    month_hdrs = ["Month","Trades","Winners","Losers","Win rate",
                  "Gross P&L (₹)","Avg win (₹)","Avg loss (₹)",
                  "Best trade (₹)","Worst trade (₹)",
                  "Premium deployed (₹)","Return on premium"]
    for ci, h in enumerate(month_hdrs, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = hf; c.fill = HDR; c.alignment = CC; c.border = BD
    ws2.row_dimensions[2].height = 28

    # Group closed trades by month
    from collections import defaultdict
    monthly: dict[str, list] = defaultdict(list)
    for trade in closed:
        ed = str(trade.get("exit_date",""))[:7]   # "2026-03"
        if ed:
            monthly[ed].append(trade)

    ri = 3
    monthly_pnl_rows = []
    for month in sorted(monthly.keys()):
        trades_m = monthly[month]
        pnls     = [float(t.get("total_pnl", 0) or 0) for t in trades_m]
        prems    = [float(t.get("total_premium", 0) or 0) for t in trades_m]
        wins     = [p for p in pnls if p >= 0]
        losses   = [p for p in pnls if p < 0]
        gross    = sum(pnls)
        win_rate = len(wins) / len(pnls) if pnls else 0
        avg_win  = sum(wins) / len(wins) if wins else 0
        avg_loss = sum(losses) / len(losses) if losses else 0
        best     = max(pnls) if pnls else 0
        worst    = min(pnls) if pnls else 0
        prem     = sum(prems)
        rop      = gross / prem if prem > 0 else 0

        fill = WIN if gross >= 0 else LOSS
        row_vals = [
            datetime.strptime(month, "%Y-%m").strftime("%b %Y"),
            len(pnls), len(wins), len(losses),
            win_rate, gross, avg_win, avg_loss, best, worst, prem, rop
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = nf; c.fill = fill; c.alignment = CC; c.border = BD
            if ci in (5, 12):
                c.number_format = "0.0%"
            if ci in (6, 7, 8, 9, 10, 11):
                c.number_format = "#,##0"
        ws2.row_dimensions[ri].height = 18
        monthly_pnl_rows.append(ri)
        ri += 1

    # Grand total row
    if monthly_pnl_rows:
        tr  = ri
        gr  = f"3:{tr-1}"
        ws2.cell(row=tr, column=1, value="ALL TIME").font = bf
        ws2.cell(row=tr, column=1).fill = SEC

        formulas = {
            2:  f"=SUM(B{gr})",
            3:  f"=SUM(C{gr})",
            4:  f"=SUM(D{gr})",
            5:  f"=IFERROR(C{tr}/B{tr},0)",
            6:  f"=SUM(F{gr})",
            7:  f"=IFERROR(F{tr}/C{tr},0)",
            8:  f"=IFERROR(F{tr}-(C{tr}*G{tr}),0)",
            9:  f"=MAX(I{gr})",
            10: f"=MIN(J{gr})",
            11: f"=SUM(K{gr})",
            12: f"=IFERROR(F{tr}/K{tr},0)",
        }
        for ci, formula in formulas.items():
            c = ws2.cell(row=tr, column=ci, value=formula)
            c.font = bf; c.fill = SEC; c.alignment = CC; c.border = BD
            if ci in (5, 12):
                c.number_format = "0.0%"
            if ci in (6, 7, 8, 9, 10, 11):
                c.number_format = "#,##0"

    mw = [12,9,9,9,10,18,15,15,16,17,22,18]
    for ci, w in enumerate(mw, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = "A3"

    # ════════════════════════════════════════════
    # SHEET 3: STATS DASHBOARD
    # ════════════════════════════════════════════
    ws3       = wb.create_sheet("Stats")

    ws3.merge_cells("A1:F1")
    t = ws3.cell(row=1, column=1, value="Performance Statistics")
    t.font = tf; t.alignment = CC
    ws3.row_dimensions[1].height = 26

    all_pnls  = [float(t.get("total_pnl",0) or 0) for t in closed]
    all_prems = [float(t.get("total_premium",0) or 0) for t in closed]
    wins_     = [p for p in all_pnls if p >= 0]
    losses_   = [p for p in all_pnls if p < 0]
    n         = len(all_pnls)

    def _s(v, fmt="₹"):
        if fmt == "₹":  return f"₹{v:,.0f}"
        if fmt == "%":  return f"{v:.1f}%"
        if fmt == "x":  return f"{v:.2f}x"
        return str(v)

    win_rate    = len(wins_) / n * 100  if n else 0
    avg_win     = sum(wins_)  / len(wins_)   if wins_   else 0
    avg_loss    = sum(losses_)/ len(losses_) if losses_ else 0
    pf          = abs(sum(wins_) / sum(losses_)) if losses_ else float("inf")
    total_pnl   = sum(all_pnls)
    total_prem  = sum(all_prems)
    rop_total   = total_pnl / total_prem * 100 if total_prem > 0 else 0
    expectancy  = (win_rate/100 * avg_win) + ((1 - win_rate/100) * avg_loss)
    best_trade  = max(all_pnls)  if all_pnls else 0
    worst_trade = min(all_pnls)  if all_pnls else 0
    consec_wins = _max_consecutive(all_pnls, positive=True)
    consec_loss = _max_consecutive(all_pnls, positive=False)

    stats_rows = [
        ("OVERVIEW",        None,           None),
        ("Total trades",    n,              ""),
        ("Winners",         len(wins_),     ""),
        ("Losers",          len(losses_),   ""),
        ("Open positions",  len([t for t in trade_log if t.get("status")=="ACTIVE"]), ""),
        ("",                None,           None),
        ("P&L",             None,           None),
        ("Total P&L",       total_pnl,      "₹"),
        ("Total premium deployed", total_prem, "₹"),
        ("Return on premium",  rop_total,   "%"),
        ("Best trade",      best_trade,     "₹"),
        ("Worst trade",     worst_trade,    "₹"),
        ("",                None,           None),
        ("QUALITY",         None,           None),
        ("Win rate",        win_rate,       "%"),
        ("Avg winning trade",  avg_win,     "₹"),
        ("Avg losing trade",   avg_loss,    "₹"),
        ("Profit factor",      pf,          "x"),
        ("Expectancy / trade", expectancy,  "₹"),
        ("Max consec wins",    consec_wins, ""),
        ("Max consec losses",  consec_loss, ""),
        ("",                None,           None),
        ("EXIT ANALYSIS",   None,           None),
    ]

    # Exit breakdown
    exit_counts: dict[str,int] = {}
    for t in closed:
        er = str(t.get("exit_reason","Other"))
        exit_counts[er] = exit_counts.get(er, 0) + 1
    for reason, count in exit_counts.items():
        stats_rows.append((f"  {reason}", count, "trades"))

    # Index breakdown
    stats_rows.append(("", None, None))
    stats_rows.append(("INDEX BREAKDOWN", None, None))
    idx_pnl: dict[str, list] = defaultdict(list)
    for t in closed:
        idx_pnl[t.get("index","?")].append(float(t.get("total_pnl",0) or 0))
    for idx, pnls in sorted(idx_pnl.items()):
        stats_rows.append((f"  {idx}",
                           f"{len(pnls)} trades | P&L ₹{sum(pnls):+,.0f}", ""))

    row_idx = 2
    for (label, value, fmt) in stats_rows:
        if value is None:
            # Section header
            if label:
                ws3.merge_cells(start_row=row_idx, start_column=1,
                                end_row=row_idx,   end_column=4)
                c = ws3.cell(row=row_idx, column=1, value=label)
                c.font = hf; c.fill = HDR; c.alignment = CC
                ws3.row_dimensions[row_idx].height = 20
            else:
                ws3.row_dimensions[row_idx].height = 8
            row_idx += 1
            continue

        c1 = ws3.cell(row=row_idx, column=1, value=label)
        c1.font = nf; c1.alignment = LC
        ws3.merge_cells(start_row=row_idx, start_column=1,
                        end_row=row_idx,   end_column=2)

        # Format value
        if fmt == "₹" and isinstance(value, (int, float)):
            disp = f"₹{value:+,.0f}" if label != "Total premium deployed" else f"₹{value:,.0f}"
        elif fmt == "%" and isinstance(value, (int, float)):
            disp = f"{value:.1f}%"
        elif fmt == "x" and isinstance(value, (int, float)):
            disp = f"{value:.2f}x" if value != float("inf") else "∞"
        else:
            disp = value

        c2 = ws3.cell(row=row_idx, column=3, value=disp)
        c2.alignment = RC

        # Colour code key P&L rows
        if fmt == "₹" and isinstance(value, (int, float)):
            c2.font = gf if value >= 0 else rf
        elif label == "Win rate":
            c2.font = gf if win_rate >= 50 else rf
        else:
            c2.font = bf

        ws3.row_dimensions[row_idx].height = 18
        row_idx += 1

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 10

    wb.save(output_file)
    print(f"[Done] Saved → {output_file}")


def _max_consecutive(pnls: list[float], positive: bool) -> int:
    max_run = cur = 0
    for p in pnls:
        if (p >= 0) == positive:
            cur += 1
            max_run = max(max_run, cur)
        else:
            cur = 0
    return max_run


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    today = datetime.today()

    print("=" * 60)
    print("  NSE Options Trade Log")
    print(f"  {today.strftime('%d %b %Y (%A)')}")
    print("=" * 60)

    trade_log = load_json(TRADE_LOG, [])
    open_cnt  = len([s for s in load_json(SIGNALS_LOG, []) if s.get("status") == "ACTIVE"])
    closed_cnt = len([t for t in trade_log if t.get("status") == "CLOSED"])

    print(f"\n  Open positions : {open_cnt}")
    print(f"  Closed trades  : {closed_cnt}")

    if open_cnt > 0:
        ans = input("\nDo you want to close any positions now? [y/N]: ").strip().lower()
        if ans == "y":
            close_trades_interactively()

    # Refresh trade_log from file (may have been updated above)
    trade_log = load_json(TRADE_LOG, [])

    print(f"\n[Excel] Writing {OUTPUT_FILE}...")
    write_trade_log_excel(trade_log, OUTPUT_FILE)

    # Console P&L summary
    closed = [t for t in trade_log if t.get("status") == "CLOSED"]
    if closed:
        all_pnl  = [float(t.get("total_pnl",0) or 0) for t in closed]
        all_prem = [float(t.get("total_premium",0) or 0) for t in closed]
        wins     = [p for p in all_pnl if p >= 0]
        total    = sum(all_pnl)
        prem     = sum(all_prem)

        print(f"\n  {'─'*40}")
        print(f"  Closed trades  : {len(closed)}")
        print(f"  Win rate       : {len(wins)/len(closed)*100:.0f}%  "
              f"({len(wins)}W / {len(closed)-len(wins)}L)")
        print(f"  Total P&L      : ₹{total:+,.0f}")
        print(f"  Return on prem : {total/prem*100:.1f}%  "
              f"(on ₹{prem:,.0f} deployed)" if prem > 0 else "")
        print(f"  {'─'*40}")

    print(f"\n  Output → {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
