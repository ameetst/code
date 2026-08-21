"""
milt_update_prices_dhan.py
===========================
Dhan HQ v2 equivalent of milt_update_prices.py. Fetches ~2 years of daily
Open/High/Low/Close/Volume for the MILT strategy's NSE 750 universe via
the shared `dhandata` library, and writes the same 5-sheet workbook shape
(DATA/OPEN/HIGH/LOW/VOLUME, tickers-in-rows x dates-in-columns) that
milt_strategy.py / momentum_lib.py already expect. No changes needed
there.

This is a NEW, separate script. milt_update_prices.py (the yfinance
fetcher) is left untouched as a fallback -- run whichever one you want,
independently.

Setup (one-time)
-----------------
    cd <path-to>\\dhan_datahq
    pip install -e .

Then make sure dhan_datahq\\.dhan_token_cache.json has a usable token
(a manually generated 24h token is enough to start) -- see
dhan_datahq/README.md.

Usage
-----
    python milt_update_prices_dhan.py [--tickers-from N750.xlsx]
        [--output MILT_N750_updated_dhan.xlsx] [--lookback-years 2]

NOTE: the default output filename (MILT_N750_updated_dhan.xlsx) is
DIFFERENT from the yfinance fetcher's (MILT_N750_updated.xlsx), so
running this never touches/risks the existing yfinance-sourced workbook.
Once you've compared the two and are happy with the Dhan data, point
--output at MILT_N750_updated.xlsx yourself (or rename the file) to
actually switch milt_strategy.py over to it.

Known differences vs. the yfinance fetcher -- see dhan_datahq/README.md
for the full list (adjustment methodology unconfirmed, NIFTY 500 is the
plain price index not the ^CRSLDX Total Return Index, ~1% of tickers
may not resolve due to renames/demergers/delistings since N750.xlsx was
last built).
"""

import argparse
import datetime
import sys
from pathlib import Path

import openpyxl
import pandas as pd

import dhandata as dh

# ── Config ────────────────────────────────────────────────────────────────
DEFAULT_TICKERS_FROM = "N750.xlsx"
DEFAULT_OUTPUT = "MILT_N750_updated_dhan.xlsx"
DEFAULT_LOOKBACK_YEARS = 2

NIFTY_INDEX_NAME = "NIFTY 500"       # matched by name against Dhan's index list
NIFTY_TICKER_LABEL = "NIFTY500"      # row label used in the output sheets,
                                      # matching what milt_strategy.py / N750.xlsx expect

FIELDS = ["open", "high", "low", "close", "volume"]
SHEET_FOR_FIELD = {
    "open": "OPEN", "high": "HIGH", "low": "LOW",
    "close": "DATA", "volume": "VOLUME",
}

# Tickers whose Dhan UNDERLYING_SYMBOL differs from the name in N750.xlsx
# (renames/demergers since the workbook was last rebuilt). Add entries
# here as you identify them from the "unmatched" report below --
# e.g. "TATAMOTORS": "TMCV"  (if you determine that's the right successor).
ALIAS_MAP = {
    # "OLD_TICKER": "DHAN_TICKER",
}


def read_universe(template_path: str) -> list:
    """Read the ticker list from column A of an existing N750-format DATA sheet."""
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    ws = wb["DATA"]
    tickers = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        val = row[0].value
        if val:
            tickers.append(str(val).strip())
    wb.close()
    return tickers


def main(template_path: str, output_file: str, lookback_years: float):
    print(f"Reading universe from: {template_path}")
    tickers = read_universe(template_path)
    # N750.xlsx already carries a NIFTY500 row as part of its own universe
    # (the yfinance fetcher handled it by remapping that one ticker's
    # Yahoo symbol to ^CRSLDX in place). We fetch the benchmark separately
    # via the index API below, so drop it here -- otherwise it both shows
    # up as a false "unmatched" equity AND ends up duplicated in the
    # output sheets (two rows sharing the NIFTY500 label breaks any
    # downstream .loc[] lookup by date, including momentum_lib.py's own
    # loaders, which silently drop one of the two rows rather than error).
    tickers = [t for t in tickers if t != NIFTY_TICKER_LABEL]
    print(f"  {len(tickers)} tickers (benchmark row excluded here; fetched separately below)\n")

    today = datetime.date.today()
    from_date = (today - datetime.timedelta(days=int(lookback_years * 365.25))).isoformat()
    # Dhan's toDate is non-inclusive -- pad by a day so today's bar (once
    # the market's closed) is actually included.
    to_date = (today + datetime.timedelta(days=1)).isoformat()
    print(f"Date range: {from_date} -> {today.isoformat()} (inclusive)\n")

    print(f"Fetching {len(tickers)} equities' OHLCV via Dhan "
          f"(single-symbol calls, rate-limited to ~5/sec -- this will take a few minutes) ...")
    data, unmatched, failed = dh.get_daily_history_bulk(
        tickers, from_date=from_date, to_date=to_date,
        fields=FIELDS, alias_map=ALIAS_MAP, progress=True,
    )
    print()

    print(f"Fetching benchmark index ({NIFTY_INDEX_NAME}) ...")
    try:
        nifty_df = dh.get_index_history_df(NIFTY_INDEX_NAME, from_date=from_date, to_date=to_date)
        data[NIFTY_TICKER_LABEL] = nifty_df[[c for c in FIELDS if c in nifty_df.columns]]
        print(f"  ok ({len(nifty_df)} rows)")
    except Exception as e:
        print(f"  [ERROR] {e}")
    print()

    if unmatched:
        print(f"{len(unmatched)} ticker(s) had no match in the Dhan instrument master "
              f"(renamed/delisted/demerged since {template_path} was built?):")
        for t in unmatched:
            print(f"   {t}")
        print("  -> add a mapping to ALIAS_MAP at the top of this script if you find the new symbol.\n")

    if failed:
        print(f"{len(failed)} ticker(s) matched but the API call failed:")
        for t, err in failed.items():
            print(f"   {t}: {err}")
        print()

    all_tickers_for_sheet = tickers + [NIFTY_TICKER_LABEL]
    all_dates = sorted({d for df in data.values() for d in df.index})
    if all_dates:
        print(f"Total distinct trading dates: {len(all_dates)} "
              f"({all_dates[0]} -> {all_dates[-1]})")
    else:
        print("No data fetched -- aborting before writing an empty workbook.")
        sys.exit(1)

    print(f"\nWriting {output_file} ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for field, sheet_name in SHEET_FOR_FIELD.items():
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value="TICKER")
        for c, d in enumerate(all_dates, start=2):
            cell = ws.cell(row=1, column=c, value=d)
            cell.number_format = "dd-mmm-yy"
        for r, ticker in enumerate(all_tickers_for_sheet, start=2):
            ws.cell(row=r, column=1, value=ticker)
            df = data.get(ticker)
            if df is None or field not in df.columns:
                continue
            series = df[field]
            for c, d in enumerate(all_dates, start=2):
                if d in series.index:
                    v = series.loc[d]
                    if isinstance(v, pd.Series):
                        # Defensive fallback: dhandata already dedupes same-date
                        # rows, but if a duplicate index ever slips through,
                        # don't crash the whole fetch over one cell -- take
                        # the last value for that date.
                        v = v.iloc[-1]
                    if pd.notna(v):
                        v = int(v) if field == "volume" else round(float(v), 2)
                        ws.cell(row=r, column=c, value=v)

    wb.save(output_file)
    print("Done.")

    n_ok = len([t for t in tickers if t in data])
    print(f"\n{n_ok}/{len(tickers)} equity tickers fetched successfully "
          f"({len(unmatched)} unmatched, {len(failed)} API failures).")

    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Fetch OHLCV history for the MILT strategy universe via Dhan HQ v2.")
    parser.add_argument("--tickers-from", type=str, default=DEFAULT_TICKERS_FROM,
                        help=f"Existing xlsx to read the ticker universe from (default: {DEFAULT_TICKERS_FROM})")
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT,
                        help=f"Output workbook path (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--lookback-years", type=float, default=DEFAULT_LOOKBACK_YEARS,
                        help=f"Years of daily history to fetch (default: {DEFAULT_LOOKBACK_YEARS})")
    args = parser.parse_args()

    if not Path(args.tickers_from).exists():
        print(f"ERROR: {args.tickers_from} not found.")
        sys.exit(1)

    main(args.tickers_from, args.output, args.lookback_years)
