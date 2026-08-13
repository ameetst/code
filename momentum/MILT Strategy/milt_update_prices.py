"""
milt_update_prices.py
======================
Fetches 2 years of daily Open/High/Low/Close/Volume for the MILT strategy's
NSE 750 universe and writes a fresh, self-contained workbook.

Unlike update_stock_price.py (Sharpe's fetcher, Close+Volume only, driven by
a hand-built Excel array-formula template), this script:
  - needs High/Low (for ATR) and Open (for Monday-open execution) in addition
    to Close/Volume, and
  - writes literal date headers directly (no template formula to maintain).

Usage:
    pip install yfinance openpyxl pandas
    python milt_update_prices.py [--tickers-from N750.xlsx] [--period 2y]

Ticker universe is read from the existing N750.xlsx DATA sheet (column A) so
we don't need a second copy of the 750-ticker list.

Output: MILT_N750_updated.xlsx with 5 sheets, each tickers-in-rows x
dates-in-columns (row 1 = literal date values):
  DATA    — Close
  OPEN    — Open
  HIGH    — High
  LOW     — Low
  VOLUME  — Volume
"""

import argparse
import sys
import time
from pathlib import Path

import openpyxl
import pandas as pd
import yfinance as yf

# ── Config ────────────────────────────────────────────────────────────────────
BATCH_SIZE = 50          # tickers per yfinance batch download
SLEEP_SEC  = 2            # pause between batches (avoid rate-limiting)
DEFAULT_PERIOD = "2y"     # need >1y of daily bars so 23-week MA / 20-period
                           # weekly Bollinger Band have a burn-in period before
                           # the first valid signal

FIELDS = ["Open", "High", "Low", "Close", "Volume"]
SHEET_FOR_FIELD = {
    "Open": "OPEN", "High": "HIGH", "Low": "LOW",
    "Close": "DATA", "Volume": "VOLUME",
}

TICKER_OVERRIDES = {
    "NIFTY500": "^CRSLDX",   # Nifty 500 Total Return Index
}
# ─────────────────────────────────────────────────────────────────────────────


def ns(ticker: str) -> str:
    """Map a sheet ticker to its Yahoo Finance symbol."""
    if ticker in TICKER_OVERRIDES:
        return TICKER_OVERRIDES[ticker]
    t = ticker.upper()
    if not t.endswith(".NS") and not t.endswith(".BO") and not t.startswith("^"):
        return t + ".NS"
    return t


def read_universe(template_path: str) -> list[str]:
    """Read the ticker list from column A of an existing N750-format DATA sheet."""
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    ws = wb["DATA"]
    tickers = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        val = row[0].value
        if val:
            tickers.append(str(val).strip())
    wb.close()
    return tickers


def batch_download(tickers_ns: list[str], period: str) -> dict[str, pd.DataFrame]:
    """
    Download OHLCV for a batch of Yahoo symbols.

    Returns dict: field -> DataFrame (index=date, columns=Yahoo symbol)
    for field in Open/High/Low/Close/Volume.
    """
    raw = yf.download(
        tickers_ns,
        period=period,
        auto_adjust=True,      # adjusts Open/High/Low/Close for splits+dividends
        progress=False,
        threads=True,
    )
    if raw.empty:
        return {f: pd.DataFrame() for f in FIELDS}

    out = {}
    if isinstance(raw.columns, pd.MultiIndex):
        top = raw.columns.get_level_values(0)
        for f in FIELDS:
            out[f] = raw[f] if f in top else pd.DataFrame()
    else:
        # single-ticker download collapses the MultiIndex
        for f in FIELDS:
            out[f] = raw[[f]] if f in raw.columns else pd.DataFrame()

    for f, df in out.items():
        if not df.empty:
            df.index = df.index.date
    return out


def main(template_path: str, output_file: str, period: str):
    print(f"Reading universe from: {template_path}")
    tickers = read_universe(template_path)
    total = len(tickers)
    print(f"  {total} tickers")
    print()

    ns_to_original = {ns(t): t for t in tickers}
    ns_tickers = list(ns_to_original.keys())
    index_tickers  = [t for t in ns_tickers if t.startswith("^")]
    equity_tickers = [t for t in ns_tickers if not t.startswith("^")]

    # field -> {original_ticker: {date: value}}
    all_data: dict[str, dict[str, dict]] = {f: {t: {} for t in tickers} for f in FIELDS}

    def store_batch(field_frames: dict[str, pd.DataFrame]):
        for f, df in field_frames.items():
            if df.empty:
                continue
            for col in df.columns:
                orig = ns_to_original.get(col, col.replace(".NS", ""))
                if orig not in all_data[f]:
                    continue
                series = df[col].dropna()
                if f == "Volume":
                    all_data[f][orig] = {d: int(v) for d, v in series.to_dict().items()}
                else:
                    all_data[f][orig] = {d: round(float(v), 2) for d, v in series.to_dict().items()}

    # ── Index tickers individually ─────────────────────────────────────────────
    if index_tickers:
        print(f"Downloading {len(index_tickers)} index ticker(s): {index_tickers}")
        for sym in index_tickers:
            orig = ns_to_original[sym]
            print(f"  {sym} ({orig})", end="", flush=True)
            try:
                frames = batch_download([sym], period)
                store_batch(frames)
                n_rows = len(frames["Close"]) if not frames["Close"].empty else 0
                print(f"  [ok] ({n_rows} rows)")
            except Exception as e:
                print(f"  [ERROR] {e}")
            time.sleep(1)

    # ── Equity tickers in batches ────────────────────────────────────────────
    batches = [equity_tickers[i:i + BATCH_SIZE] for i in range(0, len(equity_tickers), BATCH_SIZE)]
    print(f"\nDownloading {len(equity_tickers)} equities (OHLCV) in {len(batches)} batches of <={BATCH_SIZE}...")

    for idx, batch in enumerate(batches, 1):
        print(f"  Batch {idx}/{len(batches)}: {batch[0]} ... {batch[-1]}", end="", flush=True)
        try:
            frames = batch_download(batch, period)
            store_batch(frames)
            n_rows = len(frames["Close"]) if not frames["Close"].empty else 0
            print(f"  [ok] ({n_rows} rows)")
        except Exception as e:
            print(f"  [ERROR] {e}")
        if idx < len(batches):
            time.sleep(SLEEP_SEC)

    # ── Union of all dates seen across every field/ticker ─────────────────────
    all_dates = set()
    for f in FIELDS:
        for date_map in all_data[f].values():
            all_dates.update(date_map.keys())
    dates_sorted = sorted(all_dates)
    print(f"\nTotal distinct trading dates: {len(dates_sorted)} "
          f"({dates_sorted[0]} -> {dates_sorted[-1]})" if dates_sorted else "\nNo data downloaded.")

    # ── Write output workbook ──────────────────────────────────────────────────
    print(f"\nWriting {output_file} ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for field, sheet_name in SHEET_FOR_FIELD.items():
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value="TICKER")
        for c, d in enumerate(dates_sorted, start=2):
            cell = ws.cell(row=1, column=c, value=d)
            cell.number_format = "dd-mmm-yy"
        for r, ticker in enumerate(tickers, start=2):
            ws.cell(row=r, column=1, value=ticker)
            date_map = all_data[field].get(ticker, {})
            for c, d in enumerate(dates_sorted, start=2):
                v = date_map.get(d)
                if v is not None:
                    ws.cell(row=r, column=c, value=v)

    wb.save(output_file)
    print("Done.")

    # ── Report tickers with no data at all ─────────────────────────────────────
    errors = [t for t in tickers if not all_data["Close"].get(t)]
    if errors:
        print(f"\n{len(errors)} tickers had no Close data:")
        for t in errors[:30]:
            print(f"   {t}")
        if len(errors) > 30:
            print(f"   ... and {len(errors) - 30} more")
    else:
        print("All tickers fetched successfully.")

    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fetch OHLCV history for the MILT strategy universe.")
    parser.add_argument("--tickers-from", type=str, default="N750.xlsx",
                        help="Existing xlsx to read the ticker universe from (default: N750.xlsx)")
    parser.add_argument("--output", type=str, default="MILT_N750_updated.xlsx",
                        help="Output workbook path (default: MILT_N750_updated.xlsx)")
    parser.add_argument("--period", type=str, default=DEFAULT_PERIOD,
                        help="yfinance history period, e.g. 2y, 5y (default: 2y)")
    args = parser.parse_args()

    if not Path(args.tickers_from).exists():
        print(f"ERROR: {args.tickers_from} not found.")
        sys.exit(1)

    main(args.tickers_from, args.output, args.period)
