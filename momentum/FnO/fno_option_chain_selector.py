"""
FNO Option Chain Contract Selector — Phase 2
=============================================
Reads the Top 10 output from fno_top10_picker.py, then processes
NSE option chain CSV files to recommend the specific contract to enter
for each stock.

Usage
-----
1. Run fno_top10_picker.py first to generate FNO_Top10_Picks.xlsx
2. Download NSE option chain CSVs for your top stocks:
     nseindia.com → F&O → Option Chain → select stock → Download CSV
   Save each file as: TICKER_optionchain.csv (e.g. VEDL_optionchain.csv)
   OR use the NSE default filename format: option-chain-ED-TICKER-DD-Mon-YYYY.csv
3. Place all CSVs in the same folder as this script
4. Run:
     python fno_option_chain_selector.py

Outputs
-------
  FNO_Option_Contracts.xlsx   — contract recommendations with verdicts,
                                 warnings, and action guidance

Strike Selection Logic (ADX + R² based)
-----------------------------------------
  ADX > 20  AND  R²_90 > 0.15  →  Slight OTM  (strong clean trend)
  ADX < 15  AND  R²_90 < 0.05  →  ITM         (weak trend, buy delta)
  Everything else               →  ATM         (default)

Verdict logic
--------------
  ENTER NOW         — IDEAL or GOOD entry zone, liquid, within budget
  WAIT FOR PULLBACK — WATCH zone (extended >3% above EMA20)
  CAUTION           — liquidity concern or conflicting signals
  SKIP              — price moved significantly since signal date,
                      or budget exceeded, or no liquid strikes

Dependencies: openpyxl, numpy (pip install openpyxl numpy)
"""

import os
import re
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
TOP10_FILE      = 'FNO_Top10_Picks.xlsx'   # output from fno_top10_picker.py
OUTPUT_FILE     = 'FNO_Option_Contracts.xlsx'
CAPITAL_MIN     = 25000    # minimum budget per trade (₹)
CAPITAL_MAX     = 50000    # maximum budget per trade (₹)
MIN_OI          = 100      # minimum open interest for liquidity
MIN_VOL         = 5        # minimum volume for liquidity

# NSE lot sizes — update if NSE revises these
LOT_SIZES = {
    'AARTIIND':700,'ABB':125,'ABBOTINDIA':50,'ACC':400,'ADANIENT':250,
    'ADANIPORTS':1250,'ALKEM':150,'AMBUJACEM':1500,'ANGELONE':300,
    'APLAPOLLO':700,'APOLLOHOSP':125,'APOLLOTYRE':1750,'ASHOKLEY':3500,
    'ASTRAL':500,'ATGL':1000,'ATUL':100,'AUBANK':1000,'AUROPHARMA':650,
    'AXISBANK':1200,'BAJAJ-AUTO':250,'BAJAJFINSV':500,'BAJFINANCE':125,
    'BALKRISIND':300,'BANDHANBNK':3600,'BANKBARODA':5850,'BANKINDIA':5400,
    'BEL':3700,'BERGEPAINT':1100,'BHARATFORG':500,'BHARTIARTL':500,
    'BHEL':2950,'BIOCON':2600,'BLUESTARCO':300,'BPCL':1800,'BRITANNIA':200,
    'BSE':400,'CANBK':4350,'CANFINHOME':1250,'CASTROLIND':3700,
    'CGPOWER':1250,'CHAMBLFERT':2000,'CHOLAFIN':500,'CIPLA':650,
    'COALINDIA':1350,'COFORGE':150,'COLPAL':700,'CONCOR':1000,
    'CROMPTON':2000,'CUMMINSIND':300,'DALBHARAT':200,'DEEPAKNTR':325,
    'DELHIVERY':2750,'DIVISLAB':200,'DIXON':100,'DLF':1650,'DRREDDY':125,
    'EICHERMOT':175,'ESCORTS':275,'EXIDEIND':2750,'FEDERALBNK':5000,
    'GAIL':3850,'GLENMARK':650,'GMRINFRA':15000,'GODREJCP':500,
    'GODREJPROP':300,'GRANULES':1700,'GRASIM':475,'GUJGASLTD':1250,
    'HAL':175,'HAVELLS':500,'HCLTECH':700,'HDFCAMC':200,'HDFCBANK':550,
    'HDFCLIFE':1100,'HEROMOTOCO':300,'HINDALCO':1075,'HINDCOPPER':2950,
    'HINDPETRO':1850,'HINDUNILVR':300,'HUDCO':3550,'ICICIBANK':700,
    'ICICIGI':400,'ICICIPRULI':750,'IDEA':40000,'IDFCFIRSTB':8000,
    'IEX':3750,'IGL':1375,'INDHOTEL':1500,'INDIANB':2700,'INDIGO':300,
    'INDUSINDBK':500,'INDUSTOWER':2800,'INFY':400,'IOC':2500,'IPCALAB':400,
    'IRCTC':875,'IRFC':4000,'ITC':1600,'JINDALSTEL':500,'JSWENERGY':1250,
    'JSWSTEEL':600,'JUBLFOOD':1250,'KEI':350,'KOTAKBANK':400,
    'LAURUSLABS':800,'LICHSGFIN':1000,'LT':175,'LTIM':150,'LTTS':200,
    'LUPIN':500,'M&M':350,'M&MFIN':2000,'MANAPPURAM':3000,'MARICO':1100,
    'MARUTI':100,'MAXHEALTH':700,'MCX':250,'MFSL':800,'MPHASIS':275,
    'MRF':10,'MUTHOOTFIN':375,'NATIONALUM':3250,'NAUKRI':125,'NBCC':6300,
    'NESTLEIND':40,'NMDC':4350,'NTPC':3375,'NYKAA':3500,'OBEROIRLTY':400,
    'OIL':1925,'ONGC':1925,'PAGEIND':15,'PEL':275,'PERSISTENT':150,
    'PETRONET':3000,'PFC':2700,'PHOENIXLTD':400,'PIDILITIND':400,
    'PIIND':200,'PNB':8000,'POLYCAB':275,'POONAWALLA':1700,
    'POWERINDIA':25,'POWERGRID':2900,'PVRINOX':750,'RAMCOCEM':700,
    'RBLBANK':5000,'RECLTD':3000,'RELIANCE':250,'SAIL':6700,
    'SBICARD':1000,'SBILIFE':750,'SBIN':1500,'SHREECEM':25,
    'SHRIRAMFIN':250,'SIEMENS':275,'SJVN':5000,'SKFINDIA':300,
    'SONACOMS':1400,'SRF':250,'SUNPHARMA':700,'SUNTV':1000,
    'SUPREMEIND':175,'TATACOMM':500,'TATACONSUM':1100,'TATAMOTORS':1400,
    'TATAPOWER':3375,'TATASTEEL':1350,'TCS':175,'TECHM':600,
    'TIINDIA':300,'TITAN':375,'TORNTPHARM':250,'TORNTPOWER':500,
    'TRENT':275,'TVSMOTOR':350,'UBL':400,'ULTRACEMCO':100,
    'UNIONBANK':8000,'UNITDSPR':1200,'UPL':1300,'VEDL':975,
    'VOLTAS':500,'WIPRO':1500,'YESBANK':40000,'ZOMATO':4500,
}


# ══════════════════════════════════════════════════════════════════════════════
# OPTION CHAIN PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_option_chain(filepath):
    """
    Parse NSE option chain CSV into list of strike records.

    NSE CSV layout (row 1 = CALLS,,PUTS, row 2 = headers, row 3+ = data):
      CE side  : cols 1-10
      Strike   : col 11
      PE side  : cols 12-22

    Handles both normal strikes (700) and high-value strikes ("1,800").
    """
    with open(filepath, 'r') as f:
        content = f.read()

    lines = content.strip().split('\n')
    records = []

    for line in lines[2:]:
        # Remove quotes around comma-formatted numbers: "1,800.00" → 1800.00
        line = re.sub(r'"([\d,\.]+)"', lambda m: m.group(1).replace(',', ''), line)
        cols = line.strip().split(',')
        if len(cols) < 12:
            continue

        def clean(val):
            val = val.strip().rstrip('\r')
            if val in ('-', '', 'NA'):
                return None
            try:
                return float(val.replace(',', ''))
            except ValueError:
                return None

        strike = clean(cols[11])
        if strike is None:
            continue

        records.append({
            'strike'    : strike,
            'ce_ltp'    : clean(cols[5]),
            'ce_iv'     : clean(cols[4]),
            'ce_oi'     : clean(cols[1]),
            'ce_vol'    : clean(cols[3]),
            'ce_chg_oi' : clean(cols[2]),
            'ce_bid'    : clean(cols[8]),
            'ce_ask'    : clean(cols[9]),
            'pe_ltp'    : clean(cols[17]),
            'pe_oi'     : clean(cols[21]),
            'pe_vol'    : clean(cols[19]),
        })

    return records


def estimate_spot(records):
    """
    Estimate current spot price from option chain via CE/PE put-call parity.
    The strike where CE LTP ≈ PE LTP is approximately ATM = spot.
    """
    parity = [
        (r, abs((r['ce_ltp'] or 0) - (r['pe_ltp'] or 0)))
        for r in records if r['ce_ltp'] and r['pe_ltp']
    ]
    if not parity:
        return None
    return min(parity, key=lambda x: x[1])[0]['strike']


# ══════════════════════════════════════════════════════════════════════════════
# STRIKE SELECTION
# ══════════════════════════════════════════════════════════════════════════════

def select_strike(records, spot, adx, r2_90):
    """
    Choose the best CE strike based on ADX + R² signal strength.

    Logic:
      Strong trend  (ADX>20, R²_90>0.15) → slight OTM (1 step above ATM)
      Weak trend    (ADX<15, R²_90<0.05) → ITM (1 step below ATM, buy delta)
      Default                             → ATM

    Only considers strikes with sufficient liquidity (OI >= MIN_OI, Vol >= MIN_VOL).
    Falls back to looser filter if no liquid strikes found.
    """
    # Build liquid strike list
    liquid = [
        r for r in records
        if r['ce_ltp'] and r['ce_oi'] and r['ce_oi'] >= MIN_OI
        and r['ce_vol'] and r['ce_vol'] >= MIN_VOL
    ]
    if not liquid:
        liquid = [r for r in records if r['ce_ltp'] and r['ce_oi'] and r['ce_oi'] >= 50]
    if not liquid:
        liquid = [r for r in records if r['ce_ltp']]
    if not liquid:
        return None, None, 'NONE', None

    liquid_strikes = sorted(set(r['strike'] for r in liquid))
    atm_strike     = min(liquid_strikes, key=lambda x: abs(x - spot))
    atm_idx        = liquid_strikes.index(atm_strike)

    # Determine bias
    if adx and adx > 20 and r2_90 and r2_90 > 0.15:
        bias       = 'OTM'
        target_idx = min(atm_idx + 1, len(liquid_strikes) - 1)
    elif (adx is None or adx < 15) and (r2_90 is None or r2_90 < 0.05):
        bias       = 'ITM'
        target_idx = max(atm_idx - 1, 0)
    else:
        bias       = 'ATM'
        target_idx = atm_idx

    chosen_strike = liquid_strikes[target_idx]
    chosen        = next((r for r in liquid    if r['strike'] == chosen_strike), None)
    atm_rec       = next((r for r in records   if r['strike'] == atm_strike),    None)

    return chosen, atm_rec, bias, atm_strike


# ══════════════════════════════════════════════════════════════════════════════
# POSITION SIZING
# ══════════════════════════════════════════════════════════════════════════════

def calc_lots(ltp, lot_size):
    """
    How many lots fit within CAPITAL_MIN–CAPITAL_MAX budget.
    Always at least 1 lot. If 1 lot already exceeds CAPITAL_MAX,
    still returns 1 (flagged as over-budget in warnings).
    """
    if not ltp or ltp <= 0:
        return 1, 0, 0
    cost_per_lot = round(ltp * lot_size, 0)
    if cost_per_lot >= CAPITAL_MIN:
        lots = 1
    else:
        lots = max(1, int(CAPITAL_MAX // cost_per_lot))
    total = round(lots * cost_per_lot, 0)
    return lots, cost_per_lot, total


# ══════════════════════════════════════════════════════════════════════════════
# TOP 10 FILE READER
# ══════════════════════════════════════════════════════════════════════════════

def load_top10(filepath):
    """
    Read the Top 10 Picks sheet from FNO_Top10_Picks.xlsx.
    Returns list of dicts with ticker, price, EMA levels, ADX, R² values.
    """
    wb  = openpyxl.load_workbook(filepath, data_only=True)
    ws  = wb['Top 10 Picks']
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains 'Ticker')
    header_row = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip() == 'Ticker' for c in row if c):
            header_row = i
            break

    if header_row is None:
        raise ValueError(f"Could not find header row in {filepath}")

    headers = [str(c).strip() if c else '' for c in rows[header_row]]

    def col(name):
        for i, h in enumerate(headers):
            if name.lower() in h.lower():
                return i
        return None

    stocks = []
    for row in rows[header_row + 1:]:
        if not row or not row[1]:
            continue
        ticker = str(row[1]).strip() if row[1] else None
        if not ticker or ticker in ('Ticker', '—', ''):
            continue

        def get(name):
            idx = col(name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            if val is None or str(val).strip() in ('—', '', 'None'):
                return None
            try:
                return float(str(val).replace('%','').replace('₹','').replace(',',''))
            except:
                return str(val).strip()

        stocks.append({
            'ticker'     : ticker,
            'price'      : get('Price'),
            'ema20'      : get('EMA20'),
            'ema50'      : get('EMA50'),
            'ema200'     : get('EMA200'),
            'adx'        : get('ADX'),
            'pct_rank'   : get('52W'),
            'r2_90'      : get('R2 (90'),
            'r2_252'     : get('R2 (252'),
            'conviction' : get('Conviction'),
            'entry_zone' : get('Entry'),
            'lot'        : LOT_SIZES.get(ticker, None),
        })

    return stocks


# ══════════════════════════════════════════════════════════════════════════════
# CSV FILE FINDER
# ══════════════════════════════════════════════════════════════════════════════

def find_csv(ticker):
    """
    Look for NSE option chain CSV for a given ticker in the current directory.
    Accepts both naming conventions:
      TICKER_optionchain.csv
      option-chain-ED-TICKER-DD-Mon-YYYY.csv  (NSE default)
    """
    patterns = [
        f"{ticker}_optionchain.csv",
        f"{ticker}_optionchain*.csv",
        f"option-chain-ED-{ticker}-*.csv",
        f"*{ticker}*.csv",
    ]
    for pattern in patterns:
        matches = glob.glob(pattern, recursive=False)
        if matches:
            return matches[0]
    return None


# ══════════════════════════════════════════════════════════════════════════════
# ANALYSIS ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def analyse_stock(stock):
    """
    For one stock: find its CSV, parse chain, select strike, compute all outputs.
    Returns a result dict.
    """
    ticker  = stock['ticker']
    csv_path = find_csv(ticker)

    if not csv_path:
        return {
            'ticker': ticker, 'conviction': stock.get('conviction'),
            'entry_zone': stock.get('entry_zone'), 'valid': False,
            'note': f"No option chain CSV found — expected: {ticker}_optionchain.csv"
        }

    records   = parse_option_chain(csv_path)
    spot_live = estimate_spot(records)
    spot      = spot_live if spot_live else stock.get('price')

    if not spot:
        return {'ticker': ticker, 'conviction': stock.get('conviction'),
                'entry_zone': stock.get('entry_zone'), 'valid': False,
                'note': 'Could not determine spot price'}

    adx    = stock.get('adx')
    r2_90  = stock.get('r2_90')
    r2_252 = stock.get('r2_252')
    lot    = stock.get('lot')

    if not lot:
        return {'ticker': ticker, 'conviction': stock.get('conviction'),
                'entry_zone': stock.get('entry_zone'), 'valid': False,
                'note': f'Lot size not found for {ticker} — add to LOT_SIZES dict'}

    chosen, atm_rec, bias, atm_strike = select_strike(records, spot, adx, r2_90)

    if not chosen:
        return {'ticker': ticker, 'conviction': stock.get('conviction'),
                'entry_zone': stock.get('entry_zone'), 'spot': spot,
                'valid': False, 'note': 'No liquid CE strikes found'}

    ltp  = chosen['ce_ltp']
    iv   = chosen['ce_iv']
    oi   = chosen['ce_oi']
    vol  = chosen['ce_vol']
    bid  = chosen['ce_bid']
    ask  = chosen['ce_ask']

    spread     = round(ask - bid, 2) if bid and ask else None
    spread_pct = round(spread / ltp * 100, 1) if spread and ltp else None

    lots, cost_per_lot, total_cost = calc_lots(ltp, lot)

    # EMA-based spot target
    ema20 = stock.get('ema20') or spot
    ema50 = stock.get('ema50') or spot * 0.97
    move_pct     = max(0.025, (ema20 - ema50) / max(ema50, 1) * 0.6)
    target_spot  = round(spot * (1 + move_pct), 1)

    # Price drift from signal date
    signal_price = stock.get('price') or spot
    price_drift  = round((spot - signal_price) / signal_price * 100, 1) if signal_price else 0

    # ── Warnings ──────────────────────────────────────────────────────────────
    warnings = []
    liq_ok   = bool(oi and oi >= MIN_OI and vol and vol >= MIN_VOL)

    if price_drift > 30:
        warnings.append(f'Price moved +{price_drift}% since signal — trend may be captured')
    if price_drift > 15:
        warnings.append(f'Price up +{price_drift}% since signal date — enter cautiously')
    if str(stock.get('entry_zone','')).upper() in ('WATCH',):
        warnings.append('Extended >3% above EMA20 — consider waiting for pullback')
    if spread_pct and spread_pct > 5:
        warnings.append(f'Wide bid-ask spread ({spread_pct}%) — use limit orders only')
    if not liq_ok:
        warnings.append('Low liquidity — use limit orders, check OI before entering')
    if total_cost > CAPITAL_MAX:
        warnings.append(f'1 lot = ₹{total_cost:,.0f} — exceeds budget of ₹{CAPITAL_MAX:,.0f}')

    # ── Verdict ───────────────────────────────────────────────────────────────
    zone = str(stock.get('entry_zone', '')).upper()
    if price_drift > 30:
        verdict = 'SKIP'
    elif total_cost > CAPITAL_MAX * 1.5:
        verdict = 'SKIP'
    elif not liq_ok:
        verdict = 'CAUTION'
    elif zone == 'IDEAL' and not any('budget' in w for w in warnings):
        verdict = 'ENTER NOW'
    elif zone == 'GOOD' and (spread_pct is None or spread_pct < 4):
        verdict = 'ENTER NOW'
    elif zone == 'WATCH':
        verdict = 'WAIT FOR PULLBACK'
    else:
        verdict = 'ENTER NOW'

    return {
        'ticker'      : ticker,
        'conviction'  : stock.get('conviction'),
        'entry_zone'  : stock.get('entry_zone'),
        'adx'         : adx, 'r2_90': r2_90, 'r2_252': r2_252,
        'ema20'       : ema20, 'ema50': ema50,
        'signal_price': signal_price,
        'spot'        : spot, 'price_drift': price_drift,
        'atm_strike'  : atm_strike,
        'strike'      : chosen['strike'],
        'bias'        : bias,
        'ltp'         : ltp, 'iv': iv,
        'oi'          : int(oi) if oi else None,
        'vol'         : int(vol) if vol else None,
        'bid'         : bid, 'ask': ask,
        'spread'      : spread, 'spread_pct': spread_pct,
        'lot'         : lot, 'lots': lots,
        'cost_per_lot': cost_per_lot,
        'total_cost'  : total_cost,
        'target_spot' : target_spot,
        'liq_ok'      : liq_ok,
        'warnings'    : warnings,
        'verdict'     : verdict,
        'valid'       : True,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def hf(c):
    return PatternFill('solid', start_color=c, fgColor=c)

def ft(bold=False, color='1A1A1A', size=9):
    return Font(name='Arial', bold=bold, color=color, size=size)

def al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = Border(
    left=Side(style='thin', color='D0D7E2'), right=Side(style='thin', color='D0D7E2'),
    top=Side(style='thin', color='D0D7E2'),  bottom=Side(style='thin', color='D0D7E2'),
)
THICK = Border(
    left=Side(style='thin', color='D0D7E2'), right=Side(style='thin', color='D0D7E2'),
    top=Side(style='thin', color='D0D7E2'),  bottom=Side(style='medium', color='2C4A7C'),
)

VERDICT_STYLE = {
    'ENTER NOW'        : ('1B5E20', 'C8E6C9'),
    'WAIT FOR PULLBACK': ('E65100', 'FFF3E0'),
    'CAUTION'          : ('F57F17', 'FFF9C4'),
    'SKIP'             : ('B71C1C', 'FFCDD2'),
}
ZONE_STYLE = {
    'IDEAL': ('1B5E20', 'E8F5E9'),
    'GOOD' : ('33691E', 'F1F8E9'),
    'WATCH': ('E65100', 'FFF3E0'),
    'NOTE' : ('B71C1C', 'FFEBEE'),
}


def build_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contract Recommendations'
    ws.sheet_view.showGridLines = False

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'FNO OPTIONS CONTRACT SELECTOR'
    ws['A1'].font      = Font(name='Arial', bold=True, color='FFFFFF', size=15)
    ws['A1'].fill      = hf('0D1B2A')
    ws['A1'].alignment = al()
    ws.row_dimensions[1].height = 34

    ws.merge_cells('A2:Q2')
    ws['A2'] = (f'Budget ₹{CAPITAL_MIN:,}–₹{CAPITAL_MAX:,} per trade  |  '
                f'Strike logic: ADX+R² based  |  '
                f'Min OI: {MIN_OI}  |  Min Vol: {MIN_VOL}')
    ws['A2'].font      = Font(name='Arial', color='90A4AE', size=8)
    ws['A2'].fill      = hf('0D1B2A')
    ws['A2'].alignment = al()
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 8

    # ── Column headers ────────────────────────────────────────────────────────
    hdrs  = ['#', 'Ticker', 'Verdict', 'Conv%', 'Zone',
             'Spot ₹', 'ATM', 'Strike', 'Bias',
             'LTP ₹', 'IV%', 'OI', 'Vol', 'Spread',
             'Lot', 'Lots', 'Outlay ₹']
    cws   = [4, 13, 16, 7, 8, 10, 9, 9, 7, 8, 7, 8, 8, 10, 6, 5, 12]

    for ci, (h, w) in enumerate(zip(hdrs, cws), 1):
        c = ws.cell(4, ci, h)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill      = hf('1A2D45')
        c.alignment = al('center', 'center', True)
        c.border    = THICK
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for rank, r in enumerate(results, 1):
        row  = rank + 4
        verd = r.get('verdict', '—')
        vft, vbg = VERDICT_STYLE.get(verd, ('555555', 'F5F5F5'))
        zone     = str(r.get('entry_zone') or '—').upper()
        zft, zbg = ZONE_STYLE.get(zone, ('555555', 'F5F5F5'))
        bg       = 'FFFDE7' if rank == 1 else ('F8F8F8' if rank % 2 else 'FFFFFF')

        if not r.get('valid'):
            vals = [rank, r['ticker'], 'NO DATA',
                    r.get('conviction', '—'), zone,
                    '—', '—', '—', '—', '—', '—', '—', '—', '—', '—', '—', '—']
        else:
            spread_str = (f"₹{r['spread']} ({r['spread_pct']}%)"
                          if r['spread'] else '—')
            vals = [
                rank, r['ticker'], verd,
                r.get('conviction', '—'), zone,
                r['spot'], int(r['atm_strike']), int(r['strike']), r['bias'],
                r['ltp'],
                r['iv']  if r['iv']  else '—',
                r['oi']  if r['oi']  else '—',
                r['vol'] if r['vol'] else '—',
                spread_str, r['lot'], r['lots'],
                f"₹{r['total_cost']:,.0f}",
            ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row, ci, val)
            c.border    = THIN
            c.fill      = hf(bg)
            c.font      = ft(False, '1A1A1A', 9)
            c.alignment = al('left' if ci == 2 else 'center')
            if ci == 3:
                c.font = Font(name='Arial', bold=True, color=vft, size=9)
                c.fill = hf(vbg)
            if ci == 5:
                c.font = Font(name='Arial', bold=True, color=zft, size=8)
                c.fill = hf(zbg)
            if ci == 17 and r.get('valid'):
                c.font = Font(name='Arial', bold=(verd == 'ENTER NOW'),
                              color='1A1A1A', size=9)
        ws.row_dimensions[row].height = 15

    # ── Warnings block ────────────────────────────────────────────────────────
    wr = len(results) + 7
    ws.merge_cells(f'A{wr}:Q{wr}')
    ws[f'A{wr}'] = 'WARNINGS & NOTES PER STOCK'
    ws[f'A{wr}'].font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws[f'A{wr}'].fill      = hf('37474F')
    ws[f'A{wr}'].alignment = al('left')
    ws.row_dimensions[wr].height = 20
    wr += 1

    has_warnings = False
    for r in results:
        if r.get('valid') and r.get('warnings'):
            has_warnings = True
            ws.merge_cells(f'A{wr}:B{wr}')
            ws.merge_cells(f'C{wr}:Q{wr}')
            c1 = ws[f'A{wr}']
            c1.value     = r['ticker']
            c1.font      = Font(name='Arial', bold=True, color='0D47A1', size=9)
            c1.fill      = hf('E3F2FD')
            c1.alignment = al('left')
            c1.border    = THIN
            c2 = ws[f'C{wr}']
            c2.value     = ' | '.join(r['warnings'])
            c2.font      = Font(name='Arial', color='374151', size=8)
            c2.fill      = hf('FAFAFA')
            c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c2.border    = THIN
            ws.row_dimensions[wr].height = 22
            wr += 1

    if not has_warnings:
        ws.merge_cells(f'A{wr}:Q{wr}')
        ws[f'A{wr}'] = 'No warnings — all positions within normal parameters.'
        ws[f'A{wr}'].font = Font(name='Arial', italic=True, color='9E9E9E', size=9)
        ws.row_dimensions[wr].height = 18
        wr += 1

    # ── Action guidance ───────────────────────────────────────────────────────
    wr += 1
    ws.merge_cells(f'A{wr}:Q{wr}')
    ws[f'A{wr}'] = 'HOW TO READ THIS TABLE'
    ws[f'A{wr}'].font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws[f'A{wr}'].fill      = hf('1B5E20')
    ws[f'A{wr}'].alignment = al('left')
    ws.row_dimensions[wr].height = 20
    wr += 1

    guidance = [
        ('ENTER NOW',
         'IDEAL or GOOD entry zone + liquid + within budget. Enter at market or limit at Ask price.'),
        ('WAIT FOR PULLBACK',
         'Stock is extended above EMA20. Set price alert at EMA20 level. Enter only on touch.'),
        ('CAUTION',
         'Liquidity concern. Always use limit orders. Check OI/Volume on NSE before placing order.'),
        ('SKIP',
         'Price has moved significantly since signal, or 1 lot exceeds budget. No edge at current levels.'),
        ('Strike — ATM',
         'At-the-money. Balanced delta and premium. Default for moderate ADX/R² situations.'),
        ('Strike — OTM',
         'One step above ATM. Cheaper premium, needs more move. Used when ADX>20 + R²_90>0.15.'),
        ('Strike — ITM',
         'One step below ATM. Higher delta, costs more. Used when ADX<15 + R²_90<0.05 (buy delta safety).'),
    ]

    for label, note in guidance:
        ws.merge_cells(f'A{wr}:D{wr}')
        ws.merge_cells(f'E{wr}:Q{wr}')
        c1 = ws.cell(wr, 1, label)
        c1.font      = Font(name='Arial', bold=True, color='0D47A1', size=9)
        c1.fill      = hf('F0F4F8')
        c1.alignment = al('left')
        c1.border    = THIN
        c2 = ws.cell(wr, 5, note)
        c2.font      = Font(name='Arial', color='374151', size=9)
        c2.fill      = hf('FAFAFA')
        c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c2.border    = THIN
        ws.row_dimensions[wr].height = 20
        wr += 1

    ws.freeze_panes = 'A5'
    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── Load Top 10 list ──────────────────────────────────────────────────────
    if not os.path.exists(TOP10_FILE):
        print(f"ERROR: {TOP10_FILE} not found.")
        print("Run fno_top10_picker.py first to generate this file.")
        return

    print(f"Loading {TOP10_FILE} ...")
    try:
        stocks = load_top10(TOP10_FILE)
    except Exception as e:
        print(f"ERROR reading {TOP10_FILE}: {e}")
        return

    print(f"Found {len(stocks)} stocks in Top 10 list\n")

    # ── Process each stock ────────────────────────────────────────────────────
    results = []
    for stock in stocks:
        ticker = stock['ticker']
        csv_path = find_csv(ticker)
        status = f"  Processing {ticker:<14}"
        if csv_path:
            status += f"→ {os.path.basename(csv_path)}"
        else:
            status += "→ ⚠ No CSV found"
        print(status)

        result = analyse_stock(stock)
        results.append(result)

    # Sort: ENTER NOW first, then by conviction
    order = {'ENTER NOW': 0, 'WAIT FOR PULLBACK': 1, 'CAUTION': 2, 'SKIP': 3}
    results.sort(key=lambda r: (
        order.get(r.get('verdict', 'SKIP'), 3),
        -(r.get('conviction') or 0)
    ))

    # ── Print terminal summary ─────────────────────────────────────────────────
    print(f"\n{'='*100}")
    print(f"{'#':<3} {'TICKER':<13} {'CONV%':>6} {'ZONE':<8} {'SPOT':>9} {'STRIKE':>8} "
          f"{'BIAS':<5} {'LTP':>7} {'OI':>7} {'OUTLAY':>10}  VERDICT")
    print(f"{'='*100}")

    for i, r in enumerate(results, 1):
        if not r.get('valid'):
            print(f"{i:<3} {r['ticker']:<13} {'—':>6} {'—':<8} {'—':>9}  {r.get('note','No CSV')}")
        else:
            print(f"{i:<3} {r['ticker']:<13} {r.get('conviction') or '—':>6} "
                  f"{str(r.get('entry_zone') or '—'):<8} {r['spot']:>9.2f} "
                  f"{int(r['strike']):>8} {r['bias']:<5} {r['ltp']:>7.2f} "
                  f"{str(r['oi'] or '—'):>7} ₹{r['total_cost']:>8,.0f}  {r['verdict']}")
            for w in r.get('warnings', []):
                print(f"    ⚠  {w}")

    # ── Save Excel ────────────────────────────────────────────────────────────
    build_excel(results, OUTPUT_FILE)
    print(f"\nSaved → {OUTPUT_FILE}")

    enter_now = [r for r in results if r.get('verdict') == 'ENTER NOW']
    if enter_now:
        print(f"\n{'─'*60}")
        print(f"  TOP PICKS TO ENTER NOW:")
        for r in enter_now[:3]:
            print(f"  ✓ {r['ticker']} CE {int(r['strike'])}  LTP ₹{r['ltp']}  "
                  f"→ {r['lots']} lot(s) = ₹{r['total_cost']:,.0f}")
    print()


if __name__ == '__main__':
    main()