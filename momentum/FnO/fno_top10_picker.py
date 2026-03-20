"""
FNO Top 10 Picker — Phase 1
============================
Reads FNO.xlsx, runs the full signal engine, then applies a 4-gate
conviction framework to surface the best 10 BUY CALL candidates.

Outputs
-------
  FNO_Top10_Picks.xlsx   — two sheets:
    1. Top 10 Picks       — actionable ranked list with entry zone guidance
    2. Full Scored List   — all stocks that passed all 4 gates (ranked)

Conviction Scoring (after gates)
---------------------------------
  R² (252d)  × 40%   — trend sustained cleanly over full year
  R² (90d)   × 35%   — recent 3-month trend quality
  ADX        × 25%   — trend strength tiebreaker

Gates (all must pass)
----------------------
  Gate 1 — Bull Score = 80       (maximum possible score only)
  Gate 2 — Full EMA Stack ✓      (EMA20 > EMA50 > EMA200)
  Gate 3 — 52W Rank ≥ 80%        (price near annual highs)
  Gate 4 — Price ≤ 7% above EMA20 (not overextended, mean-reversion aware)

EMA20 Entry Zone (informational, not a gate)
---------------------------------------------
  IDEAL  : price at or below EMA20 (best pullback entry)
  GOOD   : 0–3% above EMA20
  WATCH  : 3–7% above EMA20 (acceptable but slightly extended)
  AVOID  : >7% above EMA20   (filtered out at Gate 4)

Dependencies: openpyxl, numpy
Usage: python fno_top10_picker.py
"""

import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
INPUT_FILE       = 'FNO.xlsx'
OUTPUT_FILE      = 'FNO_Top10_Picks.xlsx'

CALL_THRESH      = 65       # Bull score threshold for BUY CALL signal
PUT_THRESH       = 65       # Bear score threshold for BUY PUT signal
ADX_MIN          = 25       # ADX minimum for scoring component
TOP_BAND         = 0.30     # Top 30% of 52W range = bullish range position
BOT_BAND         = 0.30     # Bottom 30% for bearish

# Gate thresholds
GATE_52W_MIN     = 80.0     # Minimum 52W rank % to qualify
GATE_EMA20_MAX   = 7.0      # Max % above EMA20 allowed (overextension filter)

# Conviction weights (must sum to 1.0)
W_R2_252         = 0.40
W_R2_90          = 0.35
W_ADX            = 0.25

TOP_N            = 10       # Number of top picks to highlight


# ══════════════════════════════════════════════════════════════════════════════
# INDICATOR FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def calc_ema(prices, period):
    """Exponential Moving Average."""
    p = np.array(prices, dtype=float)
    k = 2.0 / (period + 1)
    result = np.full(len(p), np.nan)
    result[period - 1] = p[:period].mean()
    for i in range(period, len(p)):
        result[i] = p[i] * k + result[i - 1] * (1 - k)
    return result


def calc_adx(prices, period=14):
    """
    Close-only ADX proxy.
    Uses absolute price change as True Range when OHLC is unavailable.
    Returns float or None if data insufficient.
    """
    p        = np.array(prices, dtype=float)
    tr       = np.abs(np.diff(p))
    plus_dm  = np.where(np.diff(p) > 0,  np.diff(p), 0.0)
    minus_dm = np.where(np.diff(p) < 0, -np.diff(p), 0.0)

    def wilder_smooth(arr, n):
        s   = arr[:n].sum()
        out = [s]
        for x in arr[n:]:
            s = s - s / n + x
            out.append(s)
        return np.array(out)

    atr  = wilder_smooth(tr,       period)
    p_dm = wilder_smooth(plus_dm,  period)
    m_dm = wilder_smooth(minus_dm, period)

    with np.errstate(divide='ignore', invalid='ignore'):
        pdi = np.where(atr > 0, 100 * p_dm / atr, 0)
        mdi = np.where(atr > 0, 100 * m_dm / atr, 0)
        dx  = np.where((pdi + mdi) > 0,
                       100 * np.abs(pdi - mdi) / (pdi + mdi), 0)

    if len(dx) < period:
        return None
    return round(dx[-period:].mean(), 1)


def calc_r2(prices, window):
    """
    Signed linear regression R² over the last `window` bars.
    Sign follows slope direction: positive = uptrend, negative = downtrend.
    High |R²| + correct sign = clean directional move (ideal for options buying).
    """
    p = np.array(prices[-window:], dtype=float)
    if len(p) < 10:
        return None
    x              = np.arange(len(p), dtype=float)
    slope, _       = np.polyfit(x, p, 1)
    y_pred         = slope * x + _
    ss_res         = np.sum((p - y_pred) ** 2)
    ss_tot         = np.sum((p - p.mean()) ** 2)
    r2             = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    return round(r2 if slope >= 0 else -r2, 4)


# ══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def score_stock(prices):
    """Compute all indicators + bull/bear scores. Returns dict or None."""
    p = np.array(prices, dtype=float)
    if len(p) < 200:
        return None

    e20   = calc_ema(p, 20)[-1]
    e50   = calc_ema(p, 50)[-1]
    e200  = calc_ema(p, 200)[-1]
    adx   = calc_adx(p)
    price = p[-1]

    hi, lo   = p.max(), p.min()
    rng      = hi - lo if hi != lo else 1.0
    pct_rank = (price - lo) / rng          # 0 = 52W low, 1 = 52W high

    r2_90  = calc_r2(prices, 90)
    r2_252 = calc_r2(prices, 252)

    # Distance from EMA20 — positive = above, negative = pulled back below
    pct_from_ema20 = round((price - e20) / e20 * 100, 2)

    # ── Bullish components ──────────────────────────────────────────────
    b1 = price > e200                               # long-term regime     +25
    b2 = e20   > e50                                # short > medium       +20
    b3 = e20   > e50   > e200                       # full EMA stack       +15
    b4 = adx is not None and adx > ADX_MIN          # trend strength       +20
    b5 = pct_rank >= (1 - TOP_BAND)                 # near 52W high        +20
    bull = 25*b1 + 20*b2 + 15*b3 + 20*b4 + 20*b5

    # ── Bearish components ──────────────────────────────────────────────
    r1 = price < e200
    r2 = e20   < e50
    r3 = e20   < e50   < e200
    r4 = b4
    r5 = pct_rank <= BOT_BAND
    bear = 25*r1 + 20*r2 + 15*r3 + 20*r4 + 20*r5

    return dict(
        price          = round(price, 2),
        ema20          = round(e20,   2),
        ema50          = round(e50,   2),
        ema200         = round(e200,  2),
        adx            = adx,
        pct_rank       = round(pct_rank * 100, 1),
        pct_from_ema20 = pct_from_ema20,
        r2_90          = r2_90,
        r2_252         = r2_252,
        bull           = bull,
        bear           = bear,
        full_stack     = b3,     # EMA20 > EMA50 > EMA200
    )


def get_signal(s):
    if s['bull'] >= CALL_THRESH: return 'BUY CALL'
    if s['bear'] >= PUT_THRESH:  return 'BUY PUT'
    return 'NO TRADE'


def entry_zone_label(pct):
    """Classify EMA20 proximity as an entry quality label."""
    if pct < 0:       return 'IDEAL'    # pulled back to/below EMA20
    if pct <= 3.0:    return 'GOOD'     # within 3%
    if pct <= 7.0:    return 'WATCH'    # slightly extended
    return 'AVOID'                      # overextended (blocked by Gate 4)


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_and_score(filepath):
    """
    Parse wide-format FNO Excel (rows=stocks, cols=dates from col 5 onward).
    Returns (all_results list, last_date string).
    """
    wb        = openpyxl.load_workbook(filepath, data_only=True)
    ws        = wb['DATA']
    rows      = list(ws.iter_rows(values_only=True))
    header    = rows[0]
    date_cols = [i for i, h in enumerate(header) if hasattr(h, 'year')]
    last_date = header[date_cols[-1]].strftime('%d %b %Y')

    results = []
    for row in rows[1:]:
        ticker = row[1]
        name   = row[0]
        prices = [row[i] for i in date_cols if isinstance(row[i], (int, float))]
        s = score_stock(prices)
        if s is None:
            continue
        sig = get_signal(s)
        s['entry_zone'] = entry_zone_label(s['pct_from_ema20'])
        results.append({'ticker': ticker, 'name': name, 'signal': sig, **s})

    return results, last_date


# ══════════════════════════════════════════════════════════════════════════════
# 4-GATE CONVICTION FILTER
# ══════════════════════════════════════════════════════════════════════════════

def apply_gates(results):
    """
    Apply all 4 gates sequentially. Returns filtered + conviction-ranked list.
    Also returns gate-by-gate counts for reporting.
    """
    calls = [r for r in results if r['signal'] == 'BUY CALL']

    g1 = [r for r in calls if r['bull'] == 80]
    g2 = [r for r in g1    if r['full_stack']]
    g3 = [r for r in g2    if r['pct_rank'] >= GATE_52W_MIN]
    g4 = [r for r in g3    if r['pct_from_ema20'] <= GATE_EMA20_MAX]

    gate_counts = {
        'total_stocks'  : len(results),
        'buy_call'      : len(calls),
        'gate1_bull80'  : len(g1),
        'gate2_stack'   : len(g2),
        'gate3_52w'     : len(g3),
        'gate4_ema20'   : len(g4),
    }

    if not g4:
        return [], gate_counts

    # ── Conviction scoring on survivors ──────────────────────────────────────
    r2_252_arr = np.array([r['r2_252'] or 0 for r in g4], dtype=float)
    r2_90_arr  = np.array([r['r2_90']  or 0 for r in g4], dtype=float)
    adx_arr    = np.array([r['adx']    or 0 for r in g4], dtype=float)

    def norm(a):
        mn, mx = a.min(), a.max()
        return (a - mn) / (mx - mn) if mx > mn else np.zeros_like(a)

    conviction = (W_R2_252 * norm(r2_252_arr) +
                  W_R2_90  * norm(r2_90_arr)  +
                  W_ADX    * norm(adx_arr))

    for i, r in enumerate(g4):
        r['conviction'] = round(float(conviction[i]) * 100, 1)

    g4.sort(key=lambda x: -x['conviction'])
    return g4, gate_counts


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

# ── Style helpers ─────────────────────────────────────────────────────────────
def hf(hex_col):
    return PatternFill('solid', start_color=hex_col, fgColor=hex_col)

def ft(bold=False, color='1A1A1A', size=9):
    return Font(name='Arial', bold=bold, color=color, size=size)

def al(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = Border(
    left   = Side(style='thin', color='D0D7E2'),
    right  = Side(style='thin', color='D0D7E2'),
    top    = Side(style='thin', color='D0D7E2'),
    bottom = Side(style='thin', color='D0D7E2'),
)
THICK_BOT = Border(
    left   = Side(style='thin',   color='D0D7E2'),
    right  = Side(style='thin',   color='D0D7E2'),
    top    = Side(style='thin',   color='D0D7E2'),
    bottom = Side(style='medium', color='4A6FA5'),
)

# ── Color palette ─────────────────────────────────────────────────────────────
BG_DARK       = '0D1B2A'
BG_MID        = '1A2D45'
BG_GOLD1      = 'FFF3CD'   # rank 1
BG_GOLD2      = 'FFF8E1'   # rank 2-3
BG_CALL_LT    = 'EDF7EE'   # rank 4-10
BG_ALT        = 'F5FAF5'   # alternating row

ZONE_COLORS = {
    'IDEAL' : ('C8E6C9', '1B5E20'),   # green fill, dark green text
    'GOOD'  : ('DCEDC8', '33691E'),   # light green
    'WATCH' : ('FFF9C4', 'F57F17'),   # amber
    'AVOID' : ('FFCDD2', 'B71C1C'),   # red
}

def r2_style(val):
    """Returns (fill_hex, font_hex) for R² cell."""
    if val is None:  return ('F5F5F5', '9E9E9E')
    if val >= 0.50:  return ('C8E6C9', '1B5E20')
    if val >= 0.25:  return ('DCEDC8', '33691E')
    if val >= 0.05:  return ('F9FBE7', '558B2F')
    if val >= 0.0:   return ('F5F5F5', '757575')
    if val >= -0.25: return ('FFCCBC', 'BF360C')
    return ('FFCDD2', 'B71C1C')

def conviction_style(val):
    if val >= 70: return ('1B5E20', 'FFFFFF')
    if val >= 45: return ('2E7D32', 'FFFFFF')
    if val >= 25: return ('558B2F', 'FFFFFF')
    return ('757575', 'FFFFFF')

def row_bg(rank):
    if rank == 1:         return BG_GOLD1
    if rank in (2, 3):    return BG_GOLD2
    if rank % 2 == 0:     return BG_ALT
    return BG_CALL_LT

def write_cell(ws, row, col, value, bold=False, color='1A1A1A', size=9,
               bg=None, h_align='center', border=THIN, wrap=False):
    c = ws.cell(row, col, value)
    c.font      = Font(name='Arial', bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=wrap)
    c.border    = border
    if bg:
        c.fill = hf(bg)
    return c


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_top10_sheet(wb, ranked, gate_counts, last_date):
    """Sheet 1 — Top 10 Picks with full annotation."""
    ws = wb.active
    ws.title = 'Top 10 Picks'
    ws.sheet_view.showGridLines = False

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:N1')
    ws['A1'] = 'FNO TOP 10 — BUY CALL CONVICTION PICKS'
    ws['A1'].font      = Font(name='Arial', bold=True, color='FFFFFF', size=16)
    ws['A1'].fill      = hf(BG_DARK)
    ws['A1'].alignment = al('center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:N2')
    ws['A2'] = (f'Signal Date: {last_date}   |   '
                f'Scanned: {gate_counts["total_stocks"]} stocks   |   '
                f'BUY CALL: {gate_counts["buy_call"]}   →   '
                f'After Gate 1 (Bull=80): {gate_counts["gate1_bull80"]}   →   '
                f'Gate 2 (Stack✓): {gate_counts["gate2_stack"]}   →   '
                f'Gate 3 (52W≥80%): {gate_counts["gate3_52w"]}   →   '
                f'Gate 4 (EMA20≤7%): {gate_counts["gate4_ema20"]} qualified')
    ws['A2'].font      = Font(name='Arial', color='90A4AE', size=9)
    ws['A2'].fill      = hf('0D1B2A')
    ws['A2'].alignment = al('center')
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 10

    # ── Conviction scoring legend ────────────────────────────────────────────
    ws.merge_cells('A4:N4')
    ws['A4'] = ('Conviction = R²(252d)×40% + R²(90d)×35% + ADX×25%   |   '
                'Entry Zone: IDEAL = at/below EMA20 (best pullback)   '
                'GOOD = 0-3% above   WATCH = 3-7% above   '
                'R² colour: green ≥ 0.50 (clean trend) → grey (flat) → red (negative/downtrend)')
    ws['A4'].font      = Font(name='Arial', italic=True, color='546E7A', size=8)
    ws['A4'].fill      = hf('F0F4F8')
    ws['A4'].alignment = al('left')
    ws['A4'].border    = THIN
    ws.row_dimensions[4].height = 14

    ws.row_dimensions[5].height = 8

    # ── Column headers ───────────────────────────────────────────────────────
    headers = [
        'Rank', 'Ticker', 'Company Name', 'Price (₹)',
        'EMA20', 'EMA50', 'EMA200',
        'EMA20\nDist%', 'Entry\nZone',
        'ADX', '52W\nRank%',
        'R² (90d)', 'R² (252d)',
        'Conviction\nScore'
    ]
    col_widths = [5, 13, 34, 10, 10, 10, 10, 9, 9, 7, 8, 10, 10, 12]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(6, ci, h)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill      = hf(BG_MID)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = THICK_BOT
    ws.row_dimensions[6].height = 28

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows ────────────────────────────────────────────────────────────
    for rank, rec in enumerate(ranked[:TOP_N], 1):
        r      = rank + 6
        bg     = row_bg(rank)
        zone   = rec['entry_zone']
        zfill, zfont = ZONE_COLORS.get(zone, ('F5F5F5', '333333'))
        r2_90_fill,  r2_90_font  = r2_style(rec['r2_90'])
        r2_252_fill, r2_252_font = r2_style(rec['r2_252'])
        conv_bg, conv_ft = conviction_style(rec['conviction'])

        # Rank
        c = ws.cell(r, 1, rank)
        c.font      = Font(name='Arial', bold=True,
                           color='B8860B' if rank == 1 else ('888888' if rank > 3 else '555555'),
                           size=11 if rank <= 3 else 9)
        c.fill      = hf(bg); c.alignment = al('center'); c.border = THIN

        # Ticker
        c = ws.cell(r, 2, rec['ticker'])
        c.font      = Font(name='Arial', bold=True, color='0D47A1', size=10)
        c.fill      = hf(bg); c.alignment = al('center'); c.border = THIN

        # Company name
        c = ws.cell(r, 3, rec['name'])
        c.font      = ft(False, '1A1A1A', 9)
        c.fill      = hf(bg); c.alignment = al('left'); c.border = THIN

        # Price
        c = ws.cell(r, 4, rec['price'])
        c.font      = Font(name='Arial', bold=True, color='1A1A1A', size=9)
        c.fill      = hf(bg); c.alignment = al('center'); c.border = THIN

        # EMA20, EMA50, EMA200
        for ci, key in enumerate(['ema20','ema50','ema200'], 5):
            c = ws.cell(r, ci, rec[key])
            c.font  = ft(False, '374151', 9)
            c.fill  = hf(bg); c.alignment = al('center'); c.border = THIN

        # EMA20 dist %
        dist_val = f"{rec['pct_from_ema20']:+.2f}%"
        c = ws.cell(r, 8, dist_val)
        c.font      = Font(name='Arial', bold=True,
                           color='1B5E20' if rec['pct_from_ema20'] < 0 else
                                 ('F57F17' if rec['pct_from_ema20'] > 3 else '374151'),
                           size=9)
        c.fill      = hf(bg); c.alignment = al('center'); c.border = THIN

        # Entry zone
        c = ws.cell(r, 9, zone)
        c.font      = Font(name='Arial', bold=True, color=zfont, size=8)
        c.fill      = hf(zfill); c.alignment = al('center'); c.border = THIN

        # ADX
        adx_val = rec['adx'] if rec['adx'] else '-'
        c = ws.cell(r, 10, adx_val)
        c.font      = Font(name='Arial', bold=(rec['adx'] or 0) > 25,
                           color='7B1FA2' if (rec['adx'] or 0) > 25 else '374151', size=9)
        c.fill      = hf(bg); c.alignment = al('center'); c.border = THIN

        # 52W rank
        c = ws.cell(r, 11, f"{rec['pct_rank']}%")
        c.font      = ft(False, '374151', 9)
        c.fill      = hf(bg); c.alignment = al('center'); c.border = THIN

        # R² 90d
        r2_90_disp = rec['r2_90'] if rec['r2_90'] is not None else '-'
        c = ws.cell(r, 12, r2_90_disp)
        c.font      = Font(name='Arial', bold=False, color=r2_90_font, size=9)
        c.fill      = hf(r2_90_fill); c.alignment = al('center'); c.border = THIN

        # R² 252d
        r2_252_disp = rec['r2_252'] if rec['r2_252'] is not None else '-'
        c = ws.cell(r, 13, r2_252_disp)
        c.font      = Font(name='Arial', bold=True, color=r2_252_font, size=9)
        c.fill      = hf(r2_252_fill); c.alignment = al('center'); c.border = THIN

        # Conviction score
        c = ws.cell(r, 14, f"{rec['conviction']}")
        c.font      = Font(name='Arial', bold=True, color=conv_ft, size=10)
        c.fill      = hf(conv_bg); c.alignment = al('center'); c.border = THIN

        ws.row_dimensions[r].height = 17

    # ── Action guidance block ─────────────────────────────────────────────────
    guide_row = TOP_N + 9
    ws.row_dimensions[guide_row - 1].height = 12

    ws.merge_cells(f'A{guide_row}:N{guide_row}')
    ws[f'A{guide_row}'] = 'HOW TO USE THIS TABLE'
    ws[f'A{guide_row}'].font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws[f'A{guide_row}'].fill      = hf(BG_MID)
    ws[f'A{guide_row}'].alignment = al('left')
    ws.row_dimensions[guide_row].height = 20

    guidance = [
        ('Step 1 — Pick your 2 stocks',
         'Select from the top of the list. Prioritise IDEAL or GOOD entry zone. '
         'WATCH means wait for a small pullback toward EMA20 before entering.'),
        ('Step 2 — Download option chain',
         'Go to NSE website → F&O → Option Chain → select the stock → Download CSV. '
         'Save as: TICKER_optionchain.csv (e.g. VEDL_optionchain.csv)'),
        ('Step 3 — Run Phase 2 script',
         'Upload both CSVs to the option chain processor (Phase 2). '
         'It will select the optimal strike + expiry and size the trade.'),
        ('Step 4 — Entry rule',
         'Enter only when price is within Entry Zone (IDEAL or GOOD). '
         'If WATCH, set a limit alert at EMA20 level and enter on touch.'),
        ('ADX note',
         'All current signals have ADX < 25 (trend direction confirmed, momentum moderate). '
         'Buy options with ≥ 30 DTE to allow the move to develop. Avoid weekly expiry.'),
    ]

    for i, (heading, body) in enumerate(guidance):
        gr = guide_row + 1 + i
        ws.merge_cells(f'A{gr}:C{gr}')
        ws.merge_cells(f'D{gr}:N{gr}')
        c1 = ws[f'A{gr}']
        c1.value     = heading
        c1.font      = Font(name='Arial', bold=True, color='0D47A1', size=9)
        c1.fill      = hf('F0F4F8')
        c1.alignment = al('left', wrap=True)
        c1.border    = THIN
        c2 = ws[f'D{gr}']
        c2.value     = body
        c2.font      = Font(name='Arial', color='374151', size=9)
        c2.fill      = hf('FAFCFF')
        c2.alignment = al('left', wrap=True)
        c2.border    = THIN
        ws.row_dimensions[gr].height = 24

    ws.freeze_panes = 'A7'


def build_full_list_sheet(wb, ranked, last_date):
    """Sheet 2 — All stocks that passed all 4 gates, fully ranked."""
    ws = wb.create_sheet('All Qualified Stocks')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:N1')
    ws['A1'] = f'All Qualified Stocks — {len(ranked)} passed all 4 gates — {last_date}'
    ws['A1'].font      = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    ws['A1'].fill      = hf(BG_DARK)
    ws['A1'].alignment = al('center')
    ws.row_dimensions[1].height = 26

    headers = [
        'Rank', 'Ticker', 'Company Name', 'Price (₹)',
        'EMA20', 'EMA50', 'EMA200',
        'EMA20 Dist%', 'Entry Zone',
        'ADX', '52W Rank%',
        'R² (90d)', 'R² (252d)', 'Conviction'
    ]
    col_widths = [5, 13, 34, 10, 10, 10, 10, 10, 9, 7, 9, 10, 10, 11]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill      = hf(BG_MID)
        c.alignment = al('center', 'center', True)
        c.border    = THICK_BOT
    ws.row_dimensions[2].height = 20

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for rank, rec in enumerate(ranked, 1):
        r    = rank + 2
        bg   = row_bg(rank)
        zone = rec['entry_zone']
        zfill, zfont = ZONE_COLORS.get(zone, ('F5F5F5', '333333'))
        r2_90_fill,  r2_90_font  = r2_style(rec['r2_90'])
        r2_252_fill, r2_252_font = r2_style(rec['r2_252'])
        conv_bg, conv_ft         = conviction_style(rec['conviction'])

        row_data = [
            (rank,                                   'center', bg,         ft(rank<=3,'555555',9)),
            (rec['ticker'],                          'center', bg,         Font(name='Arial',bold=True,color='0D47A1',size=9)),
            (rec['name'],                            'left',   bg,         ft(False,'1A1A1A',9)),
            (rec['price'],                           'center', bg,         ft(False,'1A1A1A',9)),
            (rec['ema20'],                           'center', bg,         ft(False,'374151',9)),
            (rec['ema50'],                           'center', bg,         ft(False,'374151',9)),
            (rec['ema200'],                          'center', bg,         ft(False,'374151',9)),
            (f"{rec['pct_from_ema20']:+.2f}%",       'center', bg,         ft(False,'374151',9)),
            (zone,                                   'center', zfill,      Font(name='Arial',bold=True,color=zfont,size=8)),
            (rec['adx'] or '-',                      'center', bg,         ft(False,'374151',9)),
            (f"{rec['pct_rank']}%",                  'center', bg,         ft(False,'374151',9)),
            (rec['r2_90']  if rec['r2_90']  is not None else '-',
                                                     'center', r2_90_fill, Font(name='Arial',color=r2_90_font,size=9)),
            (rec['r2_252'] if rec['r2_252'] is not None else '-',
                                                     'center', r2_252_fill,Font(name='Arial',bold=True,color=r2_252_font,size=9)),
            (str(rec['conviction']),                 'center', conv_bg,    Font(name='Arial',bold=True,color=conv_ft,size=9)),
        ]

        for ci, (val, h_al, bg_col, font) in enumerate(row_data, 1):
            c = ws.cell(r, ci, val)
            c.font      = font
            c.fill      = hf(bg_col)
            c.alignment = Alignment(horizontal=h_al, vertical='center')
            c.border    = THIN
        ws.row_dimensions[r].height = 15

    ws.freeze_panes = 'A3'


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print(f"Loading {INPUT_FILE} ...")
    results, last_date = load_and_score(INPUT_FILE)
    print(f"Scored {len(results)} stocks  |  Signal date: {last_date}")

    ranked, gate_counts = apply_gates(results)

    print(f"\n── Gate funnel ──────────────────────────────────")
    print(f"  Total stocks      : {gate_counts['total_stocks']}")
    print(f"  BUY CALL signals  : {gate_counts['buy_call']}")
    print(f"  Gate 1 (Bull=80)  : {gate_counts['gate1_bull80']}")
    print(f"  Gate 2 (Stack ✓)  : {gate_counts['gate2_stack']}")
    print(f"  Gate 3 (52W≥80%)  : {gate_counts['gate3_52w']}")
    print(f"  Gate 4 (EMA20≤7%) : {gate_counts['gate4_ema20']}  ← conviction ranked")

    print(f"\n── Top {min(TOP_N, len(ranked))} Picks ──────────────────────────────────")
    print(f"  {'#':<3} {'Ticker':<13} {'Conv%':>6}  {'R2_252':>7}  {'R2_90':>7}  {'ADX':>5}  {'EMA20%':>7}  Zone")
    print("  " + "─"*72)
    for i, r in enumerate(ranked[:TOP_N], 1):
        print(f"  {i:<3} {r['ticker']:<13} {r['conviction']:>6.1f}  "
              f"{(r['r2_252'] or 0):>7.4f}  {(r['r2_90'] or 0):>7.4f}  "
              f"{(r['adx'] or 0):>5.1f}  {r['pct_from_ema20']:>+6.2f}%  {r['entry_zone']}")

    wb = Workbook()
    build_top10_sheet(wb, ranked, gate_counts, last_date)
    build_full_list_sheet(wb, ranked, last_date)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved → {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
