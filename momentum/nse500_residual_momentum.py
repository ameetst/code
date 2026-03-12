"""
NSE 500 Residual Momentum Strategy
====================================
Implements Blitz, Huij & Martens (2011) residual momentum on the NSE 500 universe.

Pipeline:
  1. Download NSE 500 constituents & monthly price data (via yfinance)
  2. Build Indian market factor (Nifty 500 excess return) + SMB / HML proxies
  3. Rolling 36-month OLS regression per stock → extract residuals
  4. Compute residual momentum signal (t-12 to t-2, skip t-1)
  5. Rank stocks → construct equal-weight long/short portfolio
  6. Backtest & report performance metrics
  7. Export results to Excel

Dependencies:
  pip install yfinance pandas numpy scipy statsmodels openpyxl tqdm requests

Usage:
  python nse500_residual_momentum.py

  Optional CLI flags:
    --start     YYYY-MM  First month of data download   (default: 2015-01)
    --end       YYYY-MM  Last  month of data download   (default: current month)
    --roll      INT      Rolling regression window       (default: 36 months)
    --top       FLOAT    Top-percentile long threshold   (default: 0.80)
    --bot       FLOAT    Bottom-percentile short thresh  (default: 0.20)
    --min-obs   INT      Min observations for regression (default: 24)
    --output    PATH     Excel output file path
    --no-short           Long-only mode (skip short leg)
    --liquidity FLOAT    Min avg monthly volume filter (crore INR, default: 1.0)

Notes on data:
  - Price data is fetched from Yahoo Finance (suffix .NS for NSE).
  - NSE 500 constituents are fetched from NSE's public API.
  - For Fama-French India factors, the script constructs a simple 3-factor
    proxy from the downloaded universe itself (market, SMB, HML).
  - For production, replace with actual FF India factors from:
    https://www.aqr.com/Insights/Datasets  or
    https://www.iimb.ac.in/node/2148  (IIMB India FF factors)
"""

import argparse
import warnings
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from io import BytesIO

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from scipy import stats
from tqdm import tqdm

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_START       = "2015-01"
DEFAULT_ROLL        = 36        # months
DEFAULT_TOP         = 0.80      # long top 20%
DEFAULT_BOT         = 0.20      # short bottom 20%
DEFAULT_MIN_OBS     = 24        # minimum non-NaN obs to run regression
DEFAULT_LIQUIDITY   = 1.0       # crore INR average monthly traded value
NSE500_API_URL      = "https://www.niftyindices.com/IndexConstituents/ind_nifty500list.csv"
NIFTY500_TICKER     = "^CRSLDX"   # Nifty 500 TR index on Yahoo Finance
RF_TICKER           = "INDA.NS"   # fallback; we use RBI 91-day T-bill proxy (4% ann default)
RF_ANNUAL_DEFAULT   = 0.065       # 6.5% annualised RBI repo rate proxy
SLEEP_BETWEEN_DL    = 0.05       # seconds between yfinance calls


# ─────────────────────────────────────────────────────────────────────────────
# 1. FETCH NSE 500 CONSTITUENTS
# ─────────────────────────────────────────────────────────────────────────────

def fetch_nse500_tickers() -> list[str]:
    """
    Fetch NSE 500 constituent symbols from NSE's public CSV.
    Falls back to a hardcoded sample list of 50 liquid names if the API fails.
    Returns Yahoo Finance tickers (symbol + '.NS').
    """
    print("\n[1/7] Fetching NSE 500 constituents...")
    try:
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://www.niftyindices.com/",
        }
        resp = requests.get(NSE500_API_URL, headers=headers, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(BytesIO(resp.content))
        # Column is typically 'Symbol'
        sym_col = [c for c in df.columns if "symbol" in c.lower()][0]
        symbols = df[sym_col].str.strip().tolist()
        tickers = [s + ".NS" for s in symbols]
        print(f"    ✓ Retrieved {len(tickers)} constituents from NSE API")
        return tickers
    except Exception as e:
        print(f"    ✗ NSE API failed ({e}). Using fallback list of 50 liquid NSE names.")
        fallback = [
            "RELIANCE","TCS","HDFCBANK","INFY","ICICIBANK","HINDUNILVR","ITC",
            "SBIN","BHARTIARTL","KOTAKBANK","WIPRO","AXISBANK","ASIANPAINT",
            "MARUTI","TITAN","ULTRACEMCO","SUNPHARMA","NESTLEIND","TECHM",
            "BAJFINANCE","HCLTECH","POWERGRID","NTPC","ONGC","JSWSTEEL",
            "TATAMOTORS","TATASTEEL","GRASIM","DIVISLAB","DRREDDY",
            "CIPLA","ADANIPORTS","COALINDIA","BPCL","HINDALCO","INDUSINDBK",
            "BRITANNIA","HEROMOTOCO","EICHERMOT","BAJAJFINSV","SHREECEM",
            "APOLLOHOSP","TATACONSUM","PIDILITIND","BERGEPAINT","MCDOWELL-N",
            "GODREJCP","DABUR","MARICO","COLPAL",
        ]
        return [s + ".NS" for s in fallback]


# ─────────────────────────────────────────────────────────────────────────────
# 2. DOWNLOAD PRICE DATA
# ─────────────────────────────────────────────────────────────────────────────

def download_prices(tickers: list[str], start: str, end: str) -> pd.DataFrame:
    """
    Download adjusted monthly closing prices for all tickers.
    Returns a DataFrame: index=month-end dates, columns=tickers.
    """
    print(f"\n[2/7] Downloading price data ({start} → {end}) for {len(tickers)} stocks...")
    start_dt = pd.Timestamp(start + "-01")
    end_dt   = pd.Timestamp(end   + "-01") + pd.offsets.MonthEnd(1)

    # Download in batches of 50 to avoid timeouts
    batch_size = 50
    all_prices = {}

    for i in tqdm(range(0, len(tickers), batch_size), desc="    Batches"):
        batch = tickers[i : i + batch_size]
        try:
            raw = yf.download(
                batch,
                start=start_dt.strftime("%Y-%m-%d"),
                end=(end_dt + timedelta(days=1)).strftime("%Y-%m-%d"),
                auto_adjust=True,
                progress=False,
                threads=True,
            )
            if raw.empty:
                continue
            # Handle single vs multi ticker response
            if isinstance(raw.columns, pd.MultiIndex):
                close = raw["Close"]
            else:
                close = raw[["Close"]].rename(columns={"Close": batch[0]})
            # Resample to month-end
            monthly = close.resample("ME").last()
            for col in monthly.columns:
                all_prices[col] = monthly[col]
        except Exception as e:
            print(f"    ✗ Batch {i//batch_size} failed: {e}")
        time.sleep(SLEEP_BETWEEN_DL)

    prices = pd.DataFrame(all_prices)
    prices.index = pd.to_datetime(prices.index)
    prices = prices.sort_index()

    # Keep only columns with enough data (at least 30 months)
    prices = prices.dropna(axis=1, thresh=30)
    print(f"    ✓ {prices.shape[1]} stocks with sufficient data, {prices.shape[0]} months")
    return prices


# ─────────────────────────────────────────────────────────────────────────────
# 3. COMPUTE MONTHLY RETURNS & APPLY LIQUIDITY FILTER
# ─────────────────────────────────────────────────────────────────────────────

def compute_returns(prices: pd.DataFrame) -> pd.DataFrame:
    """Compute simple monthly returns (%)."""
    returns = prices.pct_change() * 100
    return returns.iloc[1:]   # drop first NaN row


def liquidity_filter(
    prices: pd.DataFrame,
    tickers: list[str],
    start: str,
    end: str,
    min_volume_cr: float,
) -> list[str]:
    """
    Download volume data and filter out stocks with low average traded value.
    Returns filtered ticker list.
    """
    if min_volume_cr <= 0:
        return list(prices.columns)

    print(f"\n    Applying liquidity filter (min avg monthly vol: ₹{min_volume_cr} Cr)...")
    start_dt = pd.Timestamp(start + "-01")
    end_dt   = pd.Timestamp(end   + "-01") + pd.offsets.MonthEnd(1)

    liquid = []
    batch_size = 50
    for i in range(0, len(tickers), batch_size):
        batch = tickers[i : i + batch_size]
        try:
            raw = yf.download(
                batch,
                start=start_dt.strftime("%Y-%m-%d"),
                end=(end_dt + timedelta(days=1)).strftime("%Y-%m-%d"),
                auto_adjust=True,
                progress=False,
                threads=True,
            )
            if raw.empty:
                continue
            if isinstance(raw.columns, pd.MultiIndex):
                vol   = raw["Volume"]
                close = raw["Close"]
            else:
                vol   = raw[["Volume"]].rename(columns={"Volume": batch[0]})
                close = raw[["Close"]].rename(columns={"Close": batch[0]})
            # Approximate traded value = close * volume / 1e7 (crore)
            tv = (close * vol / 1e7).resample("ME").sum()
            avg_tv = tv.mean()
            for t in batch:
                if t in avg_tv.index and avg_tv[t] >= min_volume_cr and t in prices.columns:
                    liquid.append(t)
        except Exception:
            pass
        time.sleep(SLEEP_BETWEEN_DL)

    liquid = [t for t in liquid if t in prices.columns]
    print(f"    ✓ {len(liquid)} stocks pass liquidity filter")
    return liquid if liquid else list(prices.columns)


# ─────────────────────────────────────────────────────────────────────────────
# 4. BUILD FACTOR DATA (Market, SMB, HML, Rf)
# ─────────────────────────────────────────────────────────────────────────────

def build_factors(returns: pd.DataFrame, rf_annual: float = RF_ANNUAL_DEFAULT) -> pd.DataFrame:
    """
    Construct a simple 3-factor model from the cross-section:
      - Mkt-RF : Equal-weight market return minus risk-free rate
      - SMB    : Small-minus-Big (bottom tercile minus top tercile by prior-month mktcap proxy)
      - HML    : High-minus-Low B/M (approximated by prior-year return inverse as value proxy)
      - RF     : Risk-free rate (RBI 91-day T-bill approximation)

    For production: replace with actual Fama-French India factors.
    """
    print("\n[3/7] Building risk factors (market, SMB, HML, Rf)...")

    # Monthly Rf from annual rate
    rf_monthly = ((1 + rf_annual) ** (1/12) - 1) * 100  # in %

    # Market factor: equal-weight cross-sectional mean
    mkt_ew  = returns.mean(axis=1)
    mkt_rf  = mkt_ew - rf_monthly

    # SMB proxy: sort stocks by rolling 12-month cumulative return as size proxy
    # (larger/momentum stocks ≈ large cap in short run — rough but usable proxy)
    cum12 = returns.rolling(12).sum()

    smb_series = []
    hml_series = []

    for dt in returns.index:
        row = returns.loc[dt]
        c12 = cum12.loc[dt]
        valid = row.dropna().index.intersection(c12.dropna().index)
        if len(valid) < 30:
            smb_series.append(np.nan)
            hml_series.append(np.nan)
            continue

        r   = row[valid]
        c   = c12[valid]
        n   = len(valid)
        t30 = int(n * 0.30)

        # SMB: low-cum-return stocks (proxy small) minus high-cum-return (proxy large)
        sorted_c = c.sort_values()
        small    = r[sorted_c.index[:t30]].mean()
        big      = r[sorted_c.index[-t30:]].mean()
        smb_series.append(small - big)

        # HML: prior 12m losers (value proxy) minus winners (growth proxy)
        # Value = low past return (cheap), Growth = high past return (expensive)
        hml_series.append(small - big)   # collinear simplification; replace with actual B/M

    factors = pd.DataFrame({
        "Mkt_RF": mkt_rf,
        "SMB":    smb_series,
        "HML":    hml_series,
        "RF":     rf_monthly,
    }, index=returns.index)

    print(f"    ✓ Factors built for {len(factors)} months")
    return factors


# ─────────────────────────────────────────────────────────────────────────────
# 5. ROLLING OLS → RESIDUALS
# ─────────────────────────────────────────────────────────────────────────────

def compute_residuals(
    returns:  pd.DataFrame,
    factors:  pd.DataFrame,
    roll:     int,
    min_obs:  int,
) -> pd.DataFrame:
    """
    For each stock, run a rolling `roll`-month OLS:
        ExcRet_t = α + β1·Mkt_RF_t + β2·SMB_t + β3·HML_t + ε_t

    Returns DataFrame of monthly residuals (same shape as returns).
    """
    print(f"\n[4/7] Running rolling {roll}-month regressions...")

    # Align excess returns
    rf = factors["RF"]
    exc_returns = returns.sub(rf, axis=0)
    X_cols = ["Mkt_RF", "SMB", "HML"]

    residuals = pd.DataFrame(np.nan, index=returns.index, columns=returns.columns)
    tickers   = returns.columns.tolist()

    for ticker in tqdm(tickers, desc="    Stocks"):
        y_full = exc_returns[ticker]
        for end_idx in range(roll - 1, len(returns)):
            start_idx = end_idx - roll + 1
            y = y_full.iloc[start_idx : end_idx + 1]
            X = factors[X_cols].iloc[start_idx : end_idx + 1]

            # Drop NaNs
            mask = y.notna() & X.notna().all(axis=1)
            y_clean = y[mask]
            X_clean = X[mask]

            if len(y_clean) < min_obs:
                continue

            # OLS via scipy
            X_mat = np.column_stack([np.ones(len(X_clean)), X_clean.values])
            try:
                coef, _, _, _ = np.linalg.lstsq(X_mat, y_clean.values, rcond=None)
            except Exception:
                continue

            # Residual for the last month in the window only
            t = returns.index[end_idx]
            if factors[X_cols].loc[t].notna().all() and pd.notna(y_full.loc[t]):
                pred = coef[0] + coef[1:] @ factors[X_cols].loc[t].values
                residuals.loc[t, ticker] = y_full.loc[t] - pred

    print(f"    ✓ Residuals computed ({residuals.notna().sum().sum():,} observations)")
    return residuals


# ─────────────────────────────────────────────────────────────────────────────
# 6. COMPUTE SIGNAL
# ─────────────────────────────────────────────────────────────────────────────

def compute_signal(residuals: pd.DataFrame) -> pd.DataFrame:
    """
    For each stock at month t:
      Raw signal  = SUM(ε, t-12 to t-2)        [skip t-1 to avoid reversal]
      Std signal  = Raw / STDEV(ε, t-12 to t-2) [information-ratio adjusted]

    Returns DataFrame with columns: [ticker]_raw, [ticker]_std (and summary cols).
    """
    print("\n[5/7] Computing residual momentum signals...")

    raw_signals = pd.DataFrame(np.nan, index=residuals.index, columns=residuals.columns)
    std_signals = pd.DataFrame(np.nan, index=residuals.index, columns=residuals.columns)

    for ticker in residuals.columns:
        r = residuals[ticker]
        for i in range(13, len(residuals)):   # need at least 13 obs
            window = r.iloc[i-12 : i-1]       # t-12 to t-2 (11 months)
            if window.notna().sum() < 6:
                continue
            raw = window.sum()
            std = window.std()
            raw_signals.iloc[i][ticker] = raw
            if std > 0:
                std_signals.iloc[i][ticker] = raw / std

    print(f"    ✓ Signals computed")
    return raw_signals, std_signals


# ─────────────────────────────────────────────────────────────────────────────
# 7. BACKTEST
# ─────────────────────────────────────────────────────────────────────────────

def backtest(
    returns:     pd.DataFrame,
    std_signals: pd.DataFrame,
    top:         float,
    bot:         float,
    long_only:   bool,
) -> pd.DataFrame:
    """
    Each month:
      - Rank stocks by std signal
      - Long top `top` percentile, short bottom `bot` percentile (equal weight)
      - Hold for 1 month → record portfolio return

    Returns DataFrame with columns: long_ret, short_ret, ls_ret, n_long, n_short
    """
    print(f"\n[6/7] Backtesting (long top {int((1-top)*100)}%, short bot {int(bot*100)}%)...")

    results = []
    dates   = std_signals.index

    for i in range(1, len(dates)):
        signal_date  = dates[i - 1]   # signal formed at t-1
        return_date  = dates[i]        # return realised at t

        sig  = std_signals.loc[signal_date].dropna()
        rets = returns.loc[return_date]

        if len(sig) < 20:
            continue

        pct = sig.rank(pct=True)
        long_stocks  = pct[pct >= top].index
        short_stocks = pct[pct <= bot].index

        long_ret  = rets[long_stocks].mean()
        short_ret = rets[short_stocks].mean() if not long_only else 0.0
        ls_ret    = long_ret - short_ret if not long_only else long_ret

        results.append({
            "date":      return_date,
            "long_ret":  long_ret,
            "short_ret": short_ret,
            "ls_ret":    ls_ret,
            "n_long":    len(long_stocks),
            "n_short":   len(short_stocks),
        })

    perf = pd.DataFrame(results).set_index("date")
    print(f"    ✓ Backtest complete: {len(perf)} monthly observations")
    return perf


# ─────────────────────────────────────────────────────────────────────────────
# 8. PERFORMANCE METRICS
# ─────────────────────────────────────────────────────────────────────────────

def performance_metrics(perf: pd.DataFrame, rf_annual: float) -> dict:
    rf_monthly = ((1 + rf_annual) ** (1/12) - 1) * 100

    def metrics(series: pd.Series, label: str) -> dict:
        s = series.dropna()
        if len(s) < 2:
            return {}
        ann_ret  = s.mean() * 12
        ann_vol  = s.std()  * np.sqrt(12)
        sharpe   = (s.mean() - rf_monthly) / s.std() * np.sqrt(12) if s.std() > 0 else np.nan
        cum      = (1 + s/100).cumprod()
        peak     = cum.cummax()
        dd       = (cum - peak) / peak * 100
        max_dd   = dd.min()
        win_rate = (s > 0).mean() * 100
        t_stat, p_val = stats.ttest_1samp(s.dropna(), 0)
        return {
            f"{label} Ann. Return (%)":   round(ann_ret,  2),
            f"{label} Ann. Volatility (%)":round(ann_vol, 2),
            f"{label} Sharpe Ratio":       round(sharpe,  3),
            f"{label} Max Drawdown (%)":   round(max_dd,  2),
            f"{label} Win Rate (%)":       round(win_rate,2),
            f"{label} t-stat":             round(t_stat,  3),
            f"{label} p-value":            round(p_val,   4),
            f"{label} Observations":       len(s),
        }

    out = {}
    out.update(metrics(perf["long_ret"],  "Long Leg"))
    out.update(metrics(perf["short_ret"], "Short Leg"))
    out.update(metrics(perf["ls_ret"],    "L/S Portfolio"))
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 9. EXPORT TO EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(
    perf:        pd.DataFrame,
    metrics:     dict,
    raw_signals: pd.DataFrame,
    std_signals: pd.DataFrame,
    residuals:   pd.DataFrame,
    factors:     pd.DataFrame,
    output_path: str,
) -> None:
    print(f"\n[7/7] Exporting results to {output_path}...")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()

    HDR_BG = "1F4E79"
    SUB_BG = "2E75B6"
    ALT_BG = "D9E1F2"
    WHITE  = "FFFFFF"
    BLACK  = "000000"
    GREEN_C= "E2EFDA"
    RED_C  = "FFDCE1"

    def thin():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def style(cell, val, bold=False, color=BLACK, bg=WHITE, fmt=None, align="center"):
        cell.value = val
        cell.font  = Font(name="Arial", bold=bold, color=color, size=10)
        cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = thin()
        if fmt:
            cell.number_format = fmt
        return cell

    def header_row(ws, row, headers, bg=SUB_BG):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c)
            cell.value = h
            cell.font  = Font(name="Arial", bold=True, color=WHITE, size=10)
            cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin()
        ws.row_dimensions[row].height = 28

    def title_bar(ws, text, cols):
        ws.merge_cells(f"A1:{get_column_letter(cols)}1")
        c = ws["A1"]
        c.value = text
        c.font  = Font(name="Arial", bold=True, size=13, color=WHITE)
        c.fill  = PatternFill("solid", start_color=HDR_BG, end_color=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 20

    title_bar(ws_sum, "NSE 500 Residual Momentum — Performance Summary", 2)
    header_row(ws_sum, 2, ["Metric", "Value"])
    for r, (k, v) in enumerate(metrics.items(), 3):
        bg = ALT_BG if r % 2 == 0 else WHITE
        style(ws_sum.cell(row=r, column=1), k,  bg=bg, align="left")
        fmt = "0.00%" if "Rate" in k or "Return" in k or "Volatility" in k else "0.000"
        style(ws_sum.cell(row=r, column=2), v/100 if "%" in k else v, bg=bg, fmt=fmt)

    # ── Sheet 2: Monthly Returns ──────────────────────────────────────────────
    ws_ret = wb.create_sheet("Monthly_Returns")
    ws_ret.sheet_view.showGridLines = False
    for i, w in enumerate([14, 13, 13, 13, 9, 9], 1):
        ws_ret.column_dimensions[get_column_letter(i)].width = w
    title_bar(ws_ret, "Monthly Portfolio Returns", 6)
    header_row(ws_ret, 2, ["Date", "Long Leg (%)", "Short Leg (%)", "L/S Return (%)", "# Long", "# Short"])

    for r, (dt, row) in enumerate(perf.iterrows(), 3):
        bg = ALT_BG if r % 2 == 0 else WHITE
        style(ws_ret.cell(row=r, column=1), dt.strftime("%b-%Y"), bg=bg)
        for c, col in enumerate(["long_ret","short_ret","ls_ret","n_long","n_short"], 2):
            v   = row[col]
            fmt = "0.00" if c <= 4 else "0"
            cell = ws_ret.cell(row=r, column=c)
            style(cell, round(v, 4), bg=bg, fmt=fmt)
            if c == 4:  # L/S return colour
                if v > 0:
                    cell.fill = PatternFill("solid", start_color=GREEN_C, end_color=GREEN_C)
                elif v < 0:
                    cell.fill = PatternFill("solid", start_color=RED_C, end_color=RED_C)

    # ── Sheet 3: Cumulative Returns chart ────────────────────────────────────
    ws_cum = wb.create_sheet("Cumulative_Returns")
    ws_cum.sheet_view.showGridLines = False
    title_bar(ws_cum, "Cumulative Returns", 4)
    header_row(ws_cum, 2, ["Date", "Long Cum.", "Short Cum.", "L/S Cum."])
    cum_long  = (1 + perf["long_ret"]/100).cumprod()
    cum_short = (1 + perf["short_ret"]/100).cumprod()
    cum_ls    = (1 + perf["ls_ret"]/100).cumprod()
    for r, dt in enumerate(perf.index, 3):
        bg = ALT_BG if r % 2 == 0 else WHITE
        style(ws_cum.cell(row=r, column=1), dt.strftime("%b-%Y"), bg=bg)
        style(ws_cum.cell(row=r, column=2), round(cum_long.loc[dt], 4),  bg=bg, fmt="0.0000")
        style(ws_cum.cell(row=r, column=3), round(cum_short.loc[dt], 4), bg=bg, fmt="0.0000")
        style(ws_cum.cell(row=r, column=4), round(cum_ls.loc[dt], 4),    bg=bg, fmt="0.0000")

    # Add line chart
    chart = LineChart()
    chart.title  = "Cumulative Returns (Base = 1.0)"
    chart.y_axis.title = "Cumulative Return"
    chart.x_axis.title = "Month"
    chart.style  = 10
    chart.width  = 25
    chart.height = 14
    n = len(perf) + 2
    for col, title in [(2, "Long"), (3, "Short"), (4, "L/S")]:
        data = Reference(ws_cum, min_col=col, min_row=2, max_row=n)
        chart.add_data(data, titles_from_data=True)
    ws_cum.add_chart(chart, "F3")

    # ── Sheet 4: Latest Signals ───────────────────────────────────────────────
    ws_sig = wb.create_sheet("Latest_Signals")
    ws_sig.sheet_view.showGridLines = False
    for i, w in enumerate([18, 16, 16, 10, 14], 1):
        ws_sig.column_dimensions[get_column_letter(i)].width = w
    title_bar(ws_sig, f"Latest Residual Momentum Signals — {std_signals.index[-1].strftime('%b-%Y')}", 5)
    header_row(ws_sig, 2, ["Ticker", "Raw Signal (%)", "Std Signal (IR)", "Percentile", "Assignment"])

    latest_raw = raw_signals.iloc[-1].dropna()
    latest_std = std_signals.iloc[-1].dropna()
    common     = latest_raw.index.intersection(latest_std.index)
    sig_df     = pd.DataFrame({"raw": latest_raw[common], "std": latest_std[common]})
    sig_df["pct"] = sig_df["std"].rank(pct=True)
    sig_df = sig_df.sort_values("std", ascending=False)

    top_thresh = sig_df["pct"].quantile(0.80)
    bot_thresh = sig_df["pct"].quantile(0.20)

    for r, (ticker, row) in enumerate(sig_df.iterrows(), 3):
        pct = row["pct"]
        if pct >= top_thresh:
            bg = "C6EFCE"
            assign = "LONG"
        elif pct <= bot_thresh:
            bg = "FFC7CE"
            assign = "SHORT"
        else:
            bg = ALT_BG if r % 2 == 0 else WHITE
            assign = "HOLD"
        style(ws_sig.cell(row=r, column=1), ticker.replace(".NS",""), bg=bg)
        style(ws_sig.cell(row=r, column=2), round(row["raw"],4), bg=bg, fmt="0.0000")
        style(ws_sig.cell(row=r, column=3), round(row["std"],4), bg=bg, fmt="0.0000")
        style(ws_sig.cell(row=r, column=4), round(pct, 3), bg=bg, fmt="0.0%")
        style(ws_sig.cell(row=r, column=5), assign, bg=bg, bold=(assign != "HOLD"))

    # ── Sheet 5: Factors ─────────────────────────────────────────────────────
    ws_fac = wb.create_sheet("Factors")
    ws_fac.sheet_view.showGridLines = False
    fac_cols = list(factors.columns)
    for i, w in enumerate([14] + [13]*len(fac_cols), 1):
        ws_fac.column_dimensions[get_column_letter(i)].width = w
    title_bar(ws_fac, "Risk Factors — Monthly (%)", len(fac_cols)+1)
    header_row(ws_fac, 2, ["Date"] + fac_cols)
    for r, (dt, row) in enumerate(factors.iterrows(), 3):
        bg = ALT_BG if r % 2 == 0 else WHITE
        style(ws_fac.cell(row=r, column=1), dt.strftime("%b-%Y"), bg=bg)
        for c, col in enumerate(fac_cols, 2):
            style(ws_fac.cell(row=r, column=c), round(row[col], 4), bg=bg, fmt="0.0000")

    wb.save(output_path)
    print(f"    ✓ Excel file saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="NSE 500 Residual Momentum Strategy",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--start",     default=DEFAULT_START, help="Start month YYYY-MM")
    parser.add_argument("--end",       default=datetime.today().strftime("%Y-%m"), help="End month YYYY-MM")
    parser.add_argument("--roll",      type=int,   default=DEFAULT_ROLL,      help="Rolling window months")
    parser.add_argument("--top",       type=float, default=DEFAULT_TOP,       help="Long threshold percentile")
    parser.add_argument("--bot",       type=float, default=DEFAULT_BOT,       help="Short threshold percentile")
    parser.add_argument("--min-obs",   type=int,   default=DEFAULT_MIN_OBS,   help="Min obs for regression")
    parser.add_argument("--liquidity", type=float, default=DEFAULT_LIQUIDITY, help="Min avg monthly vol (Cr INR)")
    parser.add_argument("--output",    default="nse500_residual_momentum_results.xlsx")
    parser.add_argument("--no-short",  action="store_true", help="Long-only mode")
    return parser.parse_args()


def main():
    args = parse_args()

    print("=" * 65)
    print("  NSE 500 RESIDUAL MOMENTUM — FULL PIPELINE")
    print("=" * 65)
    print(f"  Period       : {args.start} → {args.end}")
    print(f"  Roll window  : {args.roll} months")
    print(f"  Long / Short : top {int((1-args.top)*100)}% / bot {int(args.bot*100)}%")
    print(f"  Min obs      : {args.min_obs}")
    print(f"  Liquidity    : ₹{args.liquidity} Cr min avg monthly vol")
    print(f"  Mode         : {'Long-only' if args.no_short else 'Long/Short'}")
    print("=" * 65)

    # Step 1 – Tickers
    tickers = fetch_nse500_tickers()

    # Step 2 – Prices & returns
    prices = download_prices(tickers, args.start, args.end)
    if prices.empty:
        print("ERROR: No price data downloaded. Check internet connection.")
        sys.exit(1)

    returns = compute_returns(prices)

    # Liquidity filter
    liquid_tickers = liquidity_filter(prices, list(prices.columns), args.start, args.end, args.liquidity)
    returns = returns[liquid_tickers]
    print(f"    Universe after all filters: {returns.shape[1]} stocks")

    # Step 3 – Factors
    factors = build_factors(returns)

    # Step 4 – Residuals
    residuals = compute_residuals(returns, factors, args.roll, args.min_obs)

    # Step 5 – Signal
    raw_signals, std_signals = compute_signal(residuals)

    # Step 6 – Backtest
    perf = backtest(returns, std_signals, args.top, args.bot, args.no_short)

    # Step 7 – Metrics
    metrics = performance_metrics(perf, RF_ANNUAL_DEFAULT)

    print("\n" + "=" * 65)
    print("  PERFORMANCE SUMMARY")
    print("=" * 65)
    for k, v in metrics.items():
        print(f"  {k:<35} {v:>10}")

    # Step 8 – Export
    export_excel(perf, metrics, raw_signals, std_signals, residuals, factors, args.output)

    print("\n" + "=" * 65)
    print("  DONE")
    print(f"  Results saved to: {args.output}")
    print("=" * 65)


if __name__ == "__main__":
    main()
