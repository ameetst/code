"""
ETF Momentum Backtest Engine -- v8 (Sharpe/Clenow Double-Confirmation Intersection)
=====================================================================================
EVALUATION ONLY -- does not touch etf_momentum_ranking.py.

Tests a "double confirmation" ranking rule: compute the existing Sharpe
composite (0.5*Z6+0.5*Z3) and the Clenow composite (0.5*Z6+0.5*Z3, see
../../Sharpe/momentum_lib.py::compute_clenow) completely SEPARATELY, each
independently ranked among screen-pass ETFs. An ETF only qualifies if it's
in the top-20 of BOTH rankings (the intersection). Within that intersection,
hold the top 5 by average rank (mean of RANK_SHARPE and RANK_CLENOW, lower
is better).

  Selection  : intersection of {RANK_SHARPE <= 20} and {RANK_CLENOW <= 20};
               ranked within the intersection by avg(RANK_SHARPE, RANK_CLENOW).
  Shortfall  : if the intersection has fewer than `active_slots` names,
               remaining slots sit in cash (no relaxing to single-metric picks).
  Exit rule  : a held ETF exits if it drops OUT of the intersection (either
               rank > 20), in addition to the existing 52wk-high-DD and TSL
               triggers. (The old "RANK_INVESTABLE > 20" trigger is replaced
               by "not in intersection" -- its direct analogue here.)

Screen / regime (current live rule) / hold-and-replace / position sizing /
costs are identical to v5/v6/v7, so this isolates the selection-rule change.

Modes run: CURRENT (today's live Sharpe-only ranking, for reference) vs
INTERSECTION (the new double-confirmation rule).

Usage:
  python etf_backtest_v8_intersection.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import yfinance as yf

# =========================================================
# CONFIG
# =========================================================
_BASE = Path(__file__).resolve().parent
_ETFS_DIR = _BASE.parent


class CONFIG:
    INPUT_FILE       = str(_BASE / "ETF - Backtest  - Copy.xlsx")
    SECTOR_FILE       = str(_ETFS_DIR / "ETF_SECTOR.xlsx")

    START_CAPITAL     = 1_000_000.0
    CASH_INTEREST_PA  = 0.02
    TRADE_COST_FIXED  = 20.0

    WINDOW_3M         = 63
    WINDOW_6M         = 126
    ANNUALIZE         = 252
    DAILY_RF          = 0.07 / 252

    TOP_N             = 5
    TOP_N_PARTIAL     = 3
    MAX_DRAWDOWN_FROM_HIGH = 0.25

    SHARPE_W6M        = 0.5
    SHARPE_W3M        = 0.5
    CLENOW_W6M        = 0.5
    CLENOW_W3M        = 0.5
    INTERSECTION_TOPK = 20   # top-K threshold on EACH ranking for the intersection

    TREND_FAST_EMA_WINDOW = 50
    TREND_EMA_WINDOW      = 100
    BENCHMARK_TICKER  = "^CRSLDX"
    REGIME_XLSX_FALLBACK = "MONIFTY500"

    SECTOR_CAP        = 1
    TSL_THRESHOLD     = 0.05
    EXIT_MAX_DD_FROM_HIGH = 0.25
    EXIT_MAX_RANK     = 20   # used by CURRENT mode only


# =========================================================
# SECTOR CLASSIFICATION (identical to v5/v6/v7)
# =========================================================
_SECTOR_LOOKUP: dict[str, str] = {}


def _load_sector_lookup():
    global _SECTOR_LOOKUP
    p = Path(CONFIG.SECTOR_FILE)
    if not p.exists():
        print(f"  [warn] {p.name} not found; using keyword fallback only")
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True)
        ws = wb["ETF_SECTOR"]
        for r in range(2, ws.max_row + 1):
            ticker = ws.cell(r, 1).value
            sector = ws.cell(r, 2).value
            if ticker and sector:
                _SECTOR_LOOKUP[str(ticker).strip().upper()] = str(sector).strip()
        wb.close()
        print(f"  [sectors] Loaded {len(_SECTOR_LOOKUP)} mappings from {p.name}")
    except Exception as e:
        print(f"  [warn] Could not load {p.name}: {e}")


_SECTOR_RULES = [
    ("PSU_BANK",         ["psu bank","psubnk","psubank","bse psu bank"]),
    ("PRIVATE_BANK",     ["private bank","pvt bank","pvtban","nifty pb "]),
    ("BANKING_BROAD",    ["nifty bank","bse bank"," bank ","banketf","bankbees","banknifty","nifban"]),
    ("IT_TECH",          ["nifty it","bse it"," it etf","itbees","itietf","nifit","tech etf"]),
    ("HEALTHCARE",       ["healthcare","pharma","health "," hc "," hc\\"]),
    ("METAL",            ["metal"]),
    ("ENERGY",           ["energy","oil & gas","o&g","oilietf","power etf","bse power"]),
    ("INFRA",            ["infra"]),
    ("CONSUMPTION",      ["consumption","consump","fmcg","consumer"]),
    ("REALTY",           ["realty","real estate"]),
    ("DEFENCE",          ["defence","dfnc"]),
    ("PSE",              ["pse etf","cpse","nifty pse","bharat 22","cpseetf"]),
    ("AUTO",             ["auto"]),
    ("CHEMICALS",        ["chemical"]),
    ("FIN_SERVICES",     ["fin serv","financial serv","finietf","bfsi","capital mkt",
                          "captl mkt","capital market","cptmkt","capital mrkts"]),
    ("COMMODITIES",      ["commodity","commo"]),
    ("MANUFACTURING",    ["manufactur","manu"]),
    ("EV_MOBILITY",      ["ev&new","ev new","nifty ev"]),
    ("DIGITAL_INTERNET", ["internet","digital"]),
    ("RAILWAY",          ["railway"]),
    ("TOURISM",          ["trsm","tourism"]),
    ("MNC",              ["mnc"]),
    ("GOLD",             ["gold"]),
    ("SILVER",           ["silver"]),
    ("GOVT_BONDS",       ["g-sec","gsec","gilt","bond etf","bharat bond","ebbetf"]),
    ("DIVIDEND",         ["dividend","div opp"]),
    ("IPO",              ["ipo"]),
    ("ESG",              ["esg"]),
    ("FACTOR_MOMENTUM",  ["momentum","mmt","mmntm"]),
    ("FACTOR_VALUE",     ["value 20","value 30","value 50","enhanced val","enhval"]),
    ("FACTOR_QUALITY",   ["quality","qlty"," ql "," ql30","qual30"]),
    ("FACTOR_LOW_VOL",   ["low vol","lowvol","lw- vol"]),
    ("FACTOR_ALPHA",     ["alpha"]),
    ("FACTOR_EQUAL_WT",  ["equal weight","eq weight","eqwt","eqlwgt","eqlwght","equal wt"]),
    ("INTERNATIONAL",    ["nasdaq","s&p 500","hang seng","hangseng","hngsng","msci","fang+"]),
    ("MIDCAP",           ["midcap","mid cap","mdsmc","midsmall"]),
    ("SMALLCAP",         ["smallcap","small cap","sml100","smcp","mosmall","sc 250"]),
    ("NEXT_50",          ["next 50","next50","juniorbees","jr bees"]),
    ("BROAD_MARKET",     ["nifty 50","nifty50","sensex","nifty 100","nifty100",
                          "nifty 200","nifty 500","nifty500","total market",
                          "total mrkt","bse 500","bse500","multicap","mltcp",
                          "lgmdcp","gth sectors","flexicap","flexi"]),
    ("SERVICES",         ["services","svcs"]),
]


def classify_sector(etf_name: str, ticker: str) -> str:
    t_upper = ticker.strip().upper()
    if t_upper in _SECTOR_LOOKUP:
        return _SECTOR_LOOKUP[t_upper]
    n = etf_name.lower()
    t = ticker.lower()
    for sector, keywords in _SECTOR_RULES:
        for kw in keywords:
            if kw in n or kw in t:
                return sector
    return "OTHER"


# =========================================================
# DATA LOADING (identical to v5/v6/v7)
# =========================================================
def load_data(filepath: str):
    print(f"Loading data from {filepath} ...")
    raw = pd.read_excel(filepath, sheet_name="DATA", header=None)
    header = raw.iloc[0]
    date_cols = [c for c in range(2, raw.shape[1])
                 if pd.notna(header.iloc[c]) and isinstance(header.iloc[c], (datetime, pd.Timestamp))]
    dates = pd.to_datetime([header.iloc[c] for c in date_cols])

    meta = pd.DataFrame({
        "ETF_NAME": raw.iloc[1:, 0].fillna("").astype(str).str.strip(),
        "TICKER":   raw.iloc[1:, 1].astype(str).str.strip(),
    }).reset_index(drop=True)

    price_df = raw.iloc[1:, date_cols].apply(pd.to_numeric, errors="coerce").replace(0, np.nan)
    price_df.columns = dates
    price_df.index = meta["TICKER"]
    prices = price_df.T.sort_index().ffill()
    print(f"  {len(meta)} ETFs | {len(dates)} date columns "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})\n")
    return meta, prices


def get_week_start_days(all_dates: pd.DatetimeIndex) -> list:
    week_starts = []
    prev_key = None
    for d in all_dates:
        key = (d.isocalendar()[0], d.isocalendar()[1])
        if key != prev_key:
            week_starts.append(d)
            prev_key = key
    return week_starts


# =========================================================
# SCORING
# =========================================================
def sharpe_score(series: pd.Series, window: int) -> float:
    clean = series.dropna()
    if len(clean) < window + 1:
        return np.nan
    log_ret = np.log(clean.iloc[-window - 1:] / clean.iloc[-window - 1:].shift(1)).dropna()
    excess = log_ret - CONFIG.DAILY_RF
    if excess.std() == 0:
        return np.nan
    return (excess.mean() / excess.std()) * np.sqrt(CONFIG.ANNUALIZE)


def clenow_score(series: pd.Series, window: int, trading_days: int = 252) -> float:
    px = series.dropna()
    if len(px) < window * 0.90:
        return np.nan
    n = min(len(px), window)
    px_w = px.iloc[-n:].values
    if np.any(px_w <= 0):
        return np.nan
    x = np.arange(n, dtype=float)
    y = np.log(px_w)
    if np.std(x) == 0 or np.std(y) == 0:
        return np.nan
    slope = np.cov(x, y, bias=True)[0, 1] / np.var(x)
    r = np.corrcoef(x, y)[0, 1]
    r2 = r ** 2
    ann_slope = slope * trading_days
    return ann_slope * r2


def _zscore(series: pd.Series) -> pd.Series:
    mu = series.mean()
    sig = series.std()
    if sig == 0 or np.isnan(sig):
        return pd.Series(0.0, index=series.index)
    return (series - mu) / sig


def build_ranking(meta: pd.DataFrame, prices: pd.DataFrame, as_of: pd.Timestamp) -> pd.DataFrame:
    """
    Computes Sharpe composite/rank and Clenow composite/rank INDEPENDENTLY,
    plus the intersection flag and combined (avg) rank for selection.
    """
    hist = prices.loc[:as_of]
    records = []
    for _, row in meta.iterrows():
        t = row["TICKER"]
        if t not in hist.columns:
            continue
        s = hist[t]
        close = float(s.dropna().iloc[-1]) if s.dropna().shape[0] > 0 else np.nan
        high_52w = float(s.tail(252).max()) if len(s) > 0 else np.nan

        sh6 = sharpe_score(s, CONFIG.WINDOW_6M)
        sh3 = sharpe_score(s, CONFIG.WINDOW_3M)
        cl6 = clenow_score(s, CONFIG.WINDOW_6M)
        cl3 = clenow_score(s, CONFIG.WINDOW_3M)

        if pd.notna(close) and pd.notna(high_52w) and high_52w > 0:
            pct_from_high = (high_52w - close) / high_52w
            high_pass = pct_from_high <= CONFIG.MAX_DRAWDOWN_FROM_HIGH
        else:
            pct_from_high = np.nan
            high_pass = True

        records.append({
            "TICKER": t,
            "ETF_NAME": row["ETF_NAME"],
            "SECTOR": classify_sector(row["ETF_NAME"], t),
            "CLOSE": close,
            "PCT_FROM_HIGH": pct_from_high * 100 if pd.notna(pct_from_high) else np.nan,
            "SHARPE_6M": sh6, "SHARPE_3M": sh3,
            "CLENOW_6M": cl6, "CLENOW_3M": cl3,
            "SCREEN_PASS": high_pass,
        })

    df = pd.DataFrame(records)
    if df.empty:
        return df

    inv_mask = df["SCREEN_PASS"]
    for col, zcol in [("SHARPE_6M", "_SZ6"), ("SHARPE_3M", "_SZ3"),
                      ("CLENOW_6M", "_CZ6"), ("CLENOW_3M", "_CZ3")]:
        df[zcol] = np.nan
        if inv_mask.sum() > 0:
            df.loc[inv_mask, zcol] = _zscore(df.loc[inv_mask, col])

    df["SHARPE_COMPOSITE"] = np.nan
    both_s = df["_SZ6"].notna() & df["_SZ3"].notna()
    only6s = df["_SZ6"].notna() & df["_SZ3"].isna()
    only3s = df["_SZ6"].isna() & df["_SZ3"].notna()
    df.loc[both_s, "SHARPE_COMPOSITE"] = CONFIG.SHARPE_W6M * df.loc[both_s, "_SZ6"] + CONFIG.SHARPE_W3M * df.loc[both_s, "_SZ3"]
    df.loc[only6s, "SHARPE_COMPOSITE"] = CONFIG.SHARPE_W6M * df.loc[only6s, "_SZ6"]
    df.loc[only3s, "SHARPE_COMPOSITE"] = CONFIG.SHARPE_W3M * df.loc[only3s, "_SZ3"]

    df["CLENOW_COMPOSITE"] = np.nan
    both_c = df["_CZ6"].notna() & df["_CZ3"].notna()
    only6c = df["_CZ6"].notna() & df["_CZ3"].isna()
    only3c = df["_CZ6"].isna() & df["_CZ3"].notna()
    df.loc[both_c, "CLENOW_COMPOSITE"] = CONFIG.CLENOW_W6M * df.loc[both_c, "_CZ6"] + CONFIG.CLENOW_W3M * df.loc[both_c, "_CZ3"]
    df.loc[only6c, "CLENOW_COMPOSITE"] = CONFIG.CLENOW_W6M * df.loc[only6c, "_CZ6"]
    df.loc[only3c, "CLENOW_COMPOSITE"] = CONFIG.CLENOW_W3M * df.loc[only3c, "_CZ3"]

    inv = df[df["SCREEN_PASS"]].copy()
    if len(inv) > 0:
        inv["RANK_SHARPE"] = inv["SHARPE_COMPOSITE"].rank(ascending=False, na_option="bottom").astype(int)
        inv["RANK_CLENOW"] = inv["CLENOW_COMPOSITE"].rank(ascending=False, na_option="bottom").astype(int)
        df = df.merge(inv[["TICKER", "RANK_SHARPE", "RANK_CLENOW"]], on="TICKER", how="left")
    else:
        df["RANK_SHARPE"] = np.nan
        df["RANK_CLENOW"] = np.nan

    # RANK_INVESTABLE for CURRENT (Sharpe-only) mode -- kept for compatibility
    df["RANK_INVESTABLE"] = df["RANK_SHARPE"]

    df["IN_INTERSECTION"] = (
        df["SCREEN_PASS"]
        & df["RANK_SHARPE"].notna() & (df["RANK_SHARPE"] <= CONFIG.INTERSECTION_TOPK)
        & df["RANK_CLENOW"].notna() & (df["RANK_CLENOW"] <= CONFIG.INTERSECTION_TOPK)
    )
    df["COMBINED_RANK"] = (df["RANK_SHARPE"] + df["RANK_CLENOW"]) / 2.0

    df = df.drop(columns=["_SZ6", "_SZ3", "_CZ6", "_CZ3"], errors="ignore")
    return df


# =========================================================
# REGIME -- held fixed at the CURRENT live rule for both modes
# =========================================================
def evaluate_regime(price, ema50, ema100) -> tuple:
    if pd.isna(price) or pd.isna(ema50) or pd.isna(ema100):
        return "BULL", CONFIG.TOP_N
    if ema50 > ema100 and price > ema50:
        return "BULL", CONFIG.TOP_N
    elif price > ema100:
        return "PARTIAL", CONFIG.TOP_N_PARTIAL
    else:
        return "BEAR", 0


# =========================================================
# EXIT LOGIC
# =========================================================
def should_exit_current(row: pd.Series | None, peak: float, current_price: float) -> tuple[bool, str]:
    """Live-script exit rule: 52wk-high DD, RANK_INVESTABLE (Sharpe) > 20, TSL."""
    if row is None:
        return True, "Ticker no longer in ranking universe"
    reasons = []
    pct_from_high = row.get("PCT_FROM_HIGH", np.nan)
    if pd.notna(pct_from_high) and abs(pct_from_high) > CONFIG.EXIT_MAX_DD_FROM_HIGH * 100:
        reasons.append(f"52wk high DD {pct_from_high:.1f}%")
    inv_rank = row.get("RANK_INVESTABLE", np.nan)
    if pd.notna(inv_rank) and inv_rank > CONFIG.EXIT_MAX_RANK:
        reasons.append(f"Rank {int(inv_rank)}")
    if peak and peak > 0 and current_price and current_price > 0:
        dd_from_peak = (peak - current_price) / peak
        if dd_from_peak >= CONFIG.TSL_THRESHOLD:
            reasons.append(f"TSL {dd_from_peak*100:.1f}%")
    if reasons:
        return True, " | ".join(reasons)
    return False, ""


def should_exit_intersection(row: pd.Series | None, peak: float, current_price: float) -> tuple[bool, str]:
    """New rule: 52wk-high DD, drop out of Sharpe/Clenow top-20 intersection, TSL."""
    if row is None:
        return True, "Ticker no longer in ranking universe"
    reasons = []
    pct_from_high = row.get("PCT_FROM_HIGH", np.nan)
    if pd.notna(pct_from_high) and abs(pct_from_high) > CONFIG.EXIT_MAX_DD_FROM_HIGH * 100:
        reasons.append(f"52wk high DD {pct_from_high:.1f}%")
    in_intersection = row.get("IN_INTERSECTION", False)
    if not in_intersection:
        rs = row.get("RANK_SHARPE", np.nan)
        rc = row.get("RANK_CLENOW", np.nan)
        reasons.append(f"Dropped intersection (SharpeRk={rs}, ClenowRk={rc})")
    if peak and peak > 0 and current_price and current_price > 0:
        dd_from_peak = (peak - current_price) / peak
        if dd_from_peak >= CONFIG.TSL_THRESHOLD:
            reasons.append(f"TSL {dd_from_peak*100:.1f}%")
    if reasons:
        return True, " | ".join(reasons)
    return False, ""


# =========================================================
# TRADE HELPERS (identical to v5/v6/v7)
# =========================================================
def _sell(t, slot, price, d, reason, regime_label, cash, trade_log):
    proceeds = slot["shares"] * price
    cost = CONFIG.TRADE_COST_FIXED
    pnl = proceeds - (slot["shares"] * slot["entry_price"]) - cost
    trade_log.append({
        "TYPE": "SELL", "REASON": reason, "TICKER": t, "NAME": slot.get("name", ""),
        "ENTRY_DATE": slot["entry_date"], "EXIT_DATE": d,
        "HOLDING_DAYS": (d - slot["entry_date"]).days,
        "ENTRY_PRICE": round(slot["entry_price"], 4), "EXIT_PRICE": round(price, 4),
        "SHARES": round(slot["shares"], 4),
        "GROSS_PNL": round(proceeds - slot["shares"] * slot["entry_price"], 2),
        "COSTS": cost, "NET_PNL": round(pnl, 2), "REGIME": regime_label,
    })
    return cash + proceeds - cost


def _buy(t, r, slot_size, p_entry, d, regime_label, trade_log, rank_note):
    shares = slot_size / p_entry
    slot = {
        "shares": shares, "entry_price": p_entry, "peak": p_entry,
        "entry_date": d, "name": r["ETF_NAME"], "sector": r["SECTOR"],
    }
    trade_log.append({
        "TYPE": "BUY", "REASON": f"NEW BUY ({rank_note})",
        "TICKER": t, "NAME": r["ETF_NAME"], "ENTRY_DATE": d, "EXIT_DATE": None,
        "HOLDING_DAYS": None, "ENTRY_PRICE": round(p_entry, 4), "EXIT_PRICE": None,
        "SHARES": round(shares, 4), "GROSS_PNL": None,
        "COSTS": CONFIG.TRADE_COST_FIXED, "NET_PNL": None, "REGIME": regime_label,
    })
    return slot, slot_size + CONFIG.TRADE_COST_FIXED


# =========================================================
# MAIN BACKTEST ENGINE -- Weekly Hold-and-Replace
# =========================================================
def run_backtest(meta, prices, regime_s, regime_ema50, regime_ema100, mode: str) -> dict:
    """mode: 'current' (Sharpe-only, live rule) or 'intersection' (double confirmation)."""
    all_dates = prices.index
    week_starts = set(get_week_start_days(all_dates))
    date_list = list(all_dates)
    date_index = {d: i for i, d in enumerate(date_list)}

    cash = CONFIG.START_CAPITAL
    slots: list[dict] = []
    equity_history = []
    trade_log = []
    regime_label = "BULL"

    print(f"\n{'='*70}\n  Running mode = {mode.upper()}\n{'='*70}")
    print(f"Period: {date_list[0].date()} -> {date_list[-1].date()}")

    for d in all_dates:
        idx = date_index[d]
        cash *= (1 + CONFIG.CASH_INTEREST_PA / 365.0)

        if d in week_starts:
            prev_day = date_list[idx - 1] if idx > 0 else d

            if regime_s is not None and prev_day in regime_s.index:
                price = regime_s.loc[prev_day]
                ema50 = regime_ema50.loc[prev_day]
                ema100 = regime_ema100.loc[prev_day]
                regime_label, active_slots = evaluate_regime(price, ema50, ema100)
            else:
                regime_label, active_slots = "BULL", CONFIG.TOP_N

            rank_df = build_ranking(meta, prices, prev_day)
            rank_by_ticker = {r["TICKER"]: r for _, r in rank_df.iterrows()} if not rank_df.empty else {}

            port_val = sum(
                s["shares"] * prices.loc[d, s["ticker"]]
                for s in slots
                if s["ticker"] != "CASH" and s["ticker"] in prices.columns and pd.notna(prices.loc[d, s["ticker"]])
            )
            total_equity = cash + port_val
            slot_size = total_equity / CONFIG.TOP_N

            new_slots = []
            held_tickers = set()
            sector_count: dict[str, int] = {}

            should_exit_fn = should_exit_current if mode == "current" else should_exit_intersection

            for prev_slot in slots:
                t = prev_slot["ticker"]
                if t == "CASH":
                    continue
                if len(held_tickers) >= active_slots:
                    price_now = prices.loc[d, t] if t in prices.columns else np.nan
                    if pd.notna(price_now):
                        cash = _sell(t, prev_slot, price_now, d, "REGIME_SLOT_CUT", regime_label, cash, trade_log)
                    continue

                row = rank_by_ticker.get(t)
                current_price = float(prices.loc[d, t]) if t in prices.columns and pd.notna(prices.loc[d, t]) else 0.0
                exit_flag, reason = should_exit_fn(row, prev_slot.get("peak", 0), current_price)

                if not exit_flag:
                    sector = prev_slot["sector"]
                    sc = sector_count.get(sector, 0)
                    sector_count[sector] = sc + 1
                    held_tickers.add(t)
                    new_peak = max(prev_slot.get("peak", 0) or 0, current_price)
                    new_slots.append({**prev_slot, "peak": new_peak})
                else:
                    price_now = current_price if current_price > 0 else prev_slot["entry_price"]
                    cash = _sell(t, prev_slot, price_now, d, reason, regime_label, cash, trade_log)

            open_slots = active_slots - len(new_slots)
            if open_slots > 0 and not rank_df.empty:
                if mode == "current":
                    candidates = rank_df[
                        rank_df["SCREEN_PASS"] & (rank_df["RANK_INVESTABLE"] > 0)
                    ].sort_values("RANK_INVESTABLE")
                    rank_note_fn = lambda r: f"rank={int(r['RANK_INVESTABLE'])}"
                else:
                    candidates = rank_df[rank_df["IN_INTERSECTION"]].sort_values("COMBINED_RANK")
                    rank_note_fn = lambda r: f"SharpeRk={int(r['RANK_SHARPE'])},ClenowRk={int(r['RANK_CLENOW'])}"

                filled = 0
                for _, r in candidates.iterrows():
                    if filled >= open_slots:
                        break
                    t = r["TICKER"]
                    if t in held_tickers:
                        continue
                    sector = r["SECTOR"]
                    if sector_count.get(sector, 0) >= CONFIG.SECTOR_CAP:
                        continue
                    p_entry = prices.loc[d, t] if t in prices.columns else np.nan
                    if pd.isna(p_entry) or p_entry <= 0:
                        continue

                    slot, spend = _buy(t, r, slot_size, p_entry, d, regime_label, trade_log, rank_note_fn(r))
                    cash -= spend
                    slot["ticker"] = t
                    new_slots.append(slot)
                    sector_count[sector] = sector_count.get(sector, 0) + 1
                    held_tickers.add(t)
                    filled += 1
                # if candidates exhausted before filling open_slots, remainder just sits in cash

            slots = new_slots

        port_val = sum(
            s["shares"] * prices.loc[d, s["ticker"]]
            for s in slots
            if s["ticker"] in prices.columns and pd.notna(prices.loc[d, s["ticker"]])
        )
        equity_history.append({"date": d, "equity": cash + port_val, "regime": regime_label,
                                "n_holdings": len(slots)})

    eq = pd.DataFrame(equity_history).set_index("date")
    tlog = pd.DataFrame(trade_log)
    return {"equity": eq, "trades": tlog}


def compute_metrics(eq: pd.DataFrame) -> dict:
    years = (eq.index[-1] - eq.index[0]).days / 365.25
    initial = CONFIG.START_CAPITAL
    final = eq["equity"].iloc[-1]
    cagr = (final / initial) ** (1 / years) - 1
    eq = eq.copy()
    eq["peak"] = eq["equity"].cummax()
    eq["drawdown"] = (eq["equity"] - eq["peak"]) / eq["peak"]
    max_dd = eq["drawdown"].min()
    daily_ret = eq["equity"].pct_change().dropna()
    vol = daily_ret.std() * np.sqrt(252)
    sharpe = cagr / vol if vol > 0 else 0
    regime_days = eq["regime"].value_counts(normalize=True) * 100
    avg_holdings = eq["n_holdings"].mean()
    return {
        "years": years, "initial": initial, "final": final, "cagr": cagr,
        "max_dd": max_dd, "vol": vol, "sharpe": sharpe, "regime_days": regime_days,
        "avg_holdings": avg_holdings,
    }


def print_results(mode: str, res: dict, metrics: dict):
    tlog = res["trades"]
    sells = tlog[tlog["TYPE"] == "SELL"] if len(tlog) else pd.DataFrame()
    buys = tlog[tlog["TYPE"] == "BUY"] if len(tlog) else pd.DataFrame()

    print(f"\n{'='*60}\n  v8 RESULTS -- mode={mode.upper()}\n{'='*60}")
    print(f"  Period          : {res['equity'].index[0].date()} -> {res['equity'].index[-1].date()}")
    print(f"  Start Capital   : INR {metrics['initial']:>12,.0f}")
    print(f"  End Capital     : INR {metrics['final']:>12,.0f}")
    print(f"  Total Return    : {(metrics['final']/metrics['initial'] - 1):>10.2%}")
    print(f"  CAGR            : {metrics['cagr']:>10.2%}")
    print(f"  Max Drawdown    : {metrics['max_dd']:>10.2%}")
    print(f"  Annualised Vol  : {metrics['vol']:>10.2%}")
    print(f"  Sharpe (simple) : {metrics['sharpe']:>10.2f}")
    print(f"  Avg # Holdings  : {metrics['avg_holdings']:.2f}  (of {CONFIG.TOP_N} slots)")
    print(f"  Total BUYs      : {len(buys)}  |  Total SELLs: {len(sells)}")
    if len(sells) > 0:
        wins = (sells["NET_PNL"] > 0).sum()
        print(f"  Win Rate        : {wins/len(sells):.1%}")
        print(f"  Avg Net P&L     : INR {sells['NET_PNL'].mean():,.0f}")
        reason_counts = sells["REASON"].apply(lambda r: r.split(" ")[0] if " " in r else r).value_counts()
        print(f"  Exit reasons    : {dict(reason_counts)}")
    print(f"  Regime day mix  : {metrics['regime_days'].round(1).to_dict()}")


def main():
    _load_sector_lookup()
    meta, prices = load_data(CONFIG.INPUT_FILE)
    all_dates = prices.index

    print(f"Fetching Nifty 500 regime data ({CONFIG.BENCHMARK_TICKER}) ...")
    regime_s = regime_ema50 = regime_ema100 = None
    try:
        reg_raw = yf.download(
            CONFIG.BENCHMARK_TICKER,
            start=all_dates[0].strftime("%Y-%m-%d"),
            end=(all_dates[-1] + pd.Timedelta(days=5)).strftime("%Y-%m-%d"),
            auto_adjust=True, progress=False,
        )
        regime_raw = reg_raw["Close"].squeeze().dropna()
        regime_raw.index = pd.to_datetime(regime_raw.index).tz_localize(None)
        if len(regime_raw) == 0:
            raise ValueError("empty live series")
        regime_s = regime_raw.reindex(all_dates, method="ffill")
        print(f"  Live regime source: {CONFIG.BENCHMARK_TICKER} "
              f"({regime_raw.index[0].date()} -> {regime_raw.index[-1].date()})")
    except Exception as e:
        print(f"  [warn] Live fetch failed ({e}); falling back to {CONFIG.REGIME_XLSX_FALLBACK} column")
        if CONFIG.REGIME_XLSX_FALLBACK in prices.columns:
            regime_s = prices[CONFIG.REGIME_XLSX_FALLBACK].dropna().reindex(all_dates, method="ffill")
        else:
            print("  [warn] Fallback column not found either; regime will default to BULL throughout")

    if regime_s is not None:
        regime_ema50 = regime_s.ewm(span=CONFIG.TREND_FAST_EMA_WINDOW, adjust=False).mean()
        regime_ema100 = regime_s.ewm(span=CONFIG.TREND_EMA_WINDOW, adjust=False).mean()

    results = {}
    metrics = {}
    for mode in ["current", "intersection"]:
        res = run_backtest(meta, prices, regime_s, regime_ema50, regime_ema100, mode)
        res["equity"].to_csv(_BASE / f"v8_{mode}_backtest_equity.csv")
        res["trades"].to_csv(_BASE / f"v8_{mode}_backtest_trade_log.csv", index=False)
        m = compute_metrics(res["equity"])
        print_results(mode, res, m)
        results[mode] = res
        metrics[mode] = m

    fig, ax = plt.subplots(figsize=(13, 7))
    for mode, color in [("current", "#7f8c8d"), ("intersection", "#8e44ad")]:
        eq = results[mode]["equity"]
        norm = eq["equity"] / eq["equity"].iloc[0] * 100
        ax.plot(eq.index, norm, label=f"{mode.upper()}", color=color, linewidth=1.6)
    ax.set_title("ETF Momentum v8 -- CURRENT (Sharpe) vs INTERSECTION (Sharpe+Clenow top-20 double confirm)")
    ax.set_ylabel("Growth of 100")
    ax.legend()
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    fig.tight_layout()
    fig.savefig(_BASE / "v8_comparison_equity_curve.png", dpi=140)
    print(f"\nComparison chart -> {_BASE / 'v8_comparison_equity_curve.png'}")

    print(f"\n{'='*60}\n  SUMMARY: CURRENT vs INTERSECTION\n{'='*60}")
    print(f"  {'Metric':<20}{'CURRENT':>15}{'INTERSECTION':>15}")
    for label, key, fmt in [
        ("CAGR", "cagr", "{:.2%}"), ("Max Drawdown", "max_dd", "{:.2%}"),
        ("Volatility", "vol", "{:.2%}"), ("Sharpe", "sharpe", "{:.2f}"),
        ("Avg Holdings", "avg_holdings", "{:.2f}"),
        ("Final Equity", "final", "INR {:,.0f}"),
    ]:
        c = fmt.format(metrics["current"][key])
        i = fmt.format(metrics["intersection"][key])
        print(f"  {label:<20}{c:>15}{i:>15}")


if __name__ == "__main__":
    main()
