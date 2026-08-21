"""
ETF Momentum Backtest -- v13 (CURRENT live algo vs Final V2, full history)
=====================================================================================
EVALUATION ONLY -- does not touch any live script or state file.

Full before/after comparison using the newly-updated backtest data file
("ETF - Backtest  - Copy.xlsx", now 2016-04 -> 2026-08, 329 ETFs -- up from
the earlier 2020-04 -> 2026-07 / 240-ETF snapshot used in v9-v12).

  CURRENT  : exact replica of etf_momentum_ranking.py (the live script).
             Score  = Z(Sharpe_6M)*0.5 + Z(Sharpe_3M)*0.5, live-fixed-weight
                      missing-data behaviour.
             Rebal  = Weekly hold-and-replace; exits on 25% off 52wk-high,
                      rank>20, or 5% TSL.
             Sizing = Equal weight (1/TOP_N per slot).
             Cap    = 1 ETF per sector.

  V2       : exact replica of etf_momentum_ranking_v2.py (the consolidated
             script built from the v9-v12 backtests).
             Score  = mean(Z(Sharpe_12M/6M/3M)), ALL THREE required (3W-STRICT).
             Rebal  = Monthly full-flush; no intra-month exit monitoring.
             Sizing = Inverse-volatility (63d realised vol) across the
                      month's picks, falls back to equal split if unavailable.
             Cap    = None (Top-N by rank regardless of sector).

Both share: identical screen (52wk-high proximity), identical tiered
EMA50/EMA100 regime rule, identical trade costs (INR 20/leg) and cash yield
(2% p.a.), identical start capital.

Usage:
  python etf_backtest_v13_current_vs_v2.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import yfinance as yf

_BASE = Path(__file__).resolve().parent
_ETFS_DIR = _BASE.parent


class CONFIG:
    INPUT_FILE        = str(_BASE / "ETF - Backtest  - Copy.xlsx")
    SECTOR_FILE        = str(_ETFS_DIR / "ETF_SECTOR.xlsx")

    START_CAPITAL      = 1_000_000.0
    CASH_INTEREST_PA   = 0.02
    TRADE_COST_FIXED   = 20.0

    ANNUALIZE          = 252
    DAILY_RF           = 0.07 / 252

    TOP_N              = 5
    TOP_N_PARTIAL      = 3
    MAX_DRAWDOWN_FROM_HIGH = 0.25

    # CURRENT (live) composite weights
    SHARPE_W6M         = 0.5
    SHARPE_W3M         = 0.5
    # V2 windows
    WINDOWS_V2          = {"12M": 252, "6M": 126, "3M": 63}
    VOL_WINDOW          = 63

    TREND_FAST_EMA_WINDOW = 50
    TREND_EMA_WINDOW      = 100
    BENCHMARK_TICKER   = "^CRSLDX"
    REGIME_XLSX_FALLBACK = "MONIFTY500"

    TSL_THRESHOLD      = 0.05
    EXIT_MAX_DD_FROM_HIGH = 0.25
    EXIT_MAX_RANK      = 20

    CURRENT_SECTOR_CAP = 1
    V2_SECTOR_CAP      = TOP_N   # >= TOP_N -> never binds


# =========================================================
# SECTOR CLASSIFICATION (same rules as etf_momentum_ranking.py)
# =========================================================
_SECTOR_LOOKUP: dict[str, str] = {}


def _load_sector_lookup():
    global _SECTOR_LOOKUP
    p = Path(CONFIG.SECTOR_FILE)
    if not p.exists():
        print(f"  [warn] {p.name} not found; using keyword fallback only")
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True)
        ws = wb["ETF_SECTOR"]
        for r in range(2, ws.max_row + 1):
            ticker = ws.cell(r, 1).value
            sector = ws.cell(r, 2).value
            if ticker and sector:
                _SECTOR_LOOKUP[str(ticker).strip().upper()] = str(sector).strip()
        wb.close()
        print(f"  [sectors] Loaded {len(_SECTOR_LOOKUP)} mappings from {p.name}")
    except Exception as e:
        print(f"  [warn] Could not load {p.name}: {e}")


_SECTOR_RULES = [
    ("PSU_BANK",         ["psu bank","psubnk","psubank","bse psu bank"]),
    ("PRIVATE_BANK",     ["private bank","pvt bank","pvtban","nifty pb "]),
    ("BANKING_BROAD",    ["nifty bank","bse bank"," bank ","banketf","bankbees","banknifty","nifban"]),
    ("IT_TECH",          ["nifty it","bse it"," it etf","itbees","itietf","nifit","tech etf"]),
    ("HEALTHCARE",       ["healthcare","pharma","health "," hc "," hc\\"]),
    ("METAL",            ["metal"]),
    ("ENERGY",           ["energy","oil & gas","o&g","oilietf","power etf","bse power"]),
    ("INFRA",            ["infra"]),
    ("CONSUMPTION",      ["consumption","consump","fmcg","consumer"]),
    ("REALTY",           ["realty","real estate"]),
    ("DEFENCE",          ["defence","dfnc"]),
    ("PSE",              ["pse etf","cpse","nifty pse","bharat 22","cpseetf"]),
    ("AUTO",             ["auto"]),
    ("CHEMICALS",        ["chemical"]),
    ("FIN_SERVICES",     ["fin serv","financial serv","finietf","bfsi","capital mkt",
                          "captl mkt","capital market","cptmkt","capital mrkts"]),
    ("COMMODITIES",      ["commodity","commo"]),
    ("MANUFACTURING",    ["manufactur","manu"]),
    ("EV_MOBILITY",      ["ev&new","ev new","nifty ev"]),
    ("DIGITAL_INTERNET", ["internet","digital"]),
    ("RAILWAY",          ["railway"]),
    ("TOURISM",          ["trsm","tourism"]),
    ("MNC",              ["mnc"]),
    ("GOLD",             ["gold"]),
    ("SILVER",           ["silver"]),
    ("GOVT_BONDS",       ["g-sec","gsec","gilt","bond etf","bharat bond","ebbetf"]),
    ("DIVIDEND",         ["dividend","div opp"]),
    ("IPO",              ["ipo"]),
    ("ESG",              ["esg"]),
    ("FACTOR_MOMENTUM",  ["momentum","mmt","mmntm"]),
    ("FACTOR_VALUE",     ["value 20","value 30","value 50","enhanced val","enhval"]),
    ("FACTOR_QUALITY",   ["quality","qlty"," ql "," ql30","qual30"]),
    ("FACTOR_LOW_VOL",   ["low vol","lowvol","lw- vol"]),
    ("FACTOR_ALPHA",     ["alpha"]),
    ("FACTOR_EQUAL_WT",  ["equal weight","eq weight","eqwt","eqlwgt","eqlwght","equal wt"]),
    ("INTERNATIONAL",    ["nasdaq","s&p 500","hang seng","hangseng","hngsng","msci","fang+"]),
    ("MIDCAP",           ["midcap","mid cap","mdsmc","midsmall"]),
    ("SMALLCAP",         ["smallcap","small cap","sml100","smcp","mosmall","sc 250"]),
    ("NEXT_50",          ["next 50","next50","juniorbees","jr bees"]),
    ("BROAD_MARKET",     ["nifty 50","nifty50","sensex","nifty 100","nifty100",
                          "nifty 200","nifty 500","nifty500","total market",
                          "total mrkt","bse 500","bse500","multicap","mltcp",
                          "lgmdcp","gth sectors","flexicap","flexi"]),
    ("SERVICES",         ["services","svcs"]),
]


def classify_sector(etf_name: str, ticker: str) -> str:
    t_upper = ticker.strip().upper()
    if t_upper in _SECTOR_LOOKUP:
        return _SECTOR_LOOKUP[t_upper]
    n = etf_name.lower()
    t = ticker.lower()
    for sector, keywords in _SECTOR_RULES:
        for kw in keywords:
            if kw in n or kw in t:
                return sector
    return "OTHER"


# =========================================================
# DATA LOADING
# =========================================================
def load_data(filepath: str):
    print(f"Loading data from {filepath} ...")
    raw = pd.read_excel(filepath, sheet_name="DATA", header=None)
    header = raw.iloc[0]
    date_cols = [c for c in range(2, raw.shape[1])
                 if pd.notna(header.iloc[c]) and isinstance(header.iloc[c], (datetime, pd.Timestamp))]
    dates = pd.to_datetime([header.iloc[c] for c in date_cols])

    meta = pd.DataFrame({
        "ETF_NAME": raw.iloc[1:, 0].fillna("").astype(str).str.strip(),
        "TICKER":   raw.iloc[1:, 1].astype(str).str.strip(),
    }).reset_index(drop=True)

    price_df = raw.iloc[1:, date_cols].apply(pd.to_numeric, errors="coerce").replace(0, np.nan)
    price_df.columns = dates
    price_df.index = meta["TICKER"]
    prices = price_df.T.sort_index().ffill()
    print(f"  {len(meta)} ETFs | {len(dates)} date columns "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})\n")
    return meta, prices


def get_week_start_days(all_dates: pd.DatetimeIndex) -> list:
    week_starts = []
    prev_key = None
    for d in all_dates:
        key = (d.isocalendar()[0], d.isocalendar()[1])
        if key != prev_key:
            week_starts.append(d)
            prev_key = key
    return week_starts


def get_month_start_days(all_dates: pd.DatetimeIndex) -> list:
    month_starts = []
    prev_key = None
    for d in all_dates:
        key = (d.year, d.month)
        if key != prev_key:
            month_starts.append(d)
            prev_key = key
    return month_starts


# =========================================================
# SCORING
# =========================================================
def sharpe_score(series: pd.Series, window: int) -> float:
    clean = series.dropna()
    if len(clean) < window + 1:
        return np.nan
    log_ret = np.log(clean.iloc[-window - 1:] / clean.iloc[-window - 1:].shift(1)).dropna()
    excess = log_ret - CONFIG.DAILY_RF
    if excess.std() == 0:
        return np.nan
    return (excess.mean() / excess.std()) * np.sqrt(CONFIG.ANNUALIZE)


def realized_vol(series: pd.Series, window: int) -> float:
    clean = series.dropna()
    if len(clean) < window + 1:
        return np.nan
    log_ret = np.log(clean.iloc[-window - 1:] / clean.iloc[-window - 1:].shift(1)).dropna()
    sd = log_ret.std()
    if sd <= 0 or np.isnan(sd):
        return np.nan
    return sd * np.sqrt(CONFIG.ANNUALIZE)


def _zscore(series: pd.Series) -> pd.Series:
    mu = series.mean()
    sig = series.std()
    if sig == 0 or np.isnan(sig):
        return pd.Series(0.0, index=series.index)
    return (series - mu) / sig


def build_ranking_current(meta: pd.DataFrame, prices: pd.DataFrame, as_of: pd.Timestamp) -> pd.DataFrame:
    """CURRENT (live) composite: Z(6M)*0.5 + Z(3M)*0.5, live-fixed-weight missing-data behaviour."""
    hist = prices.loc[:as_of]
    records = []
    for _, row in meta.iterrows():
        t = row["TICKER"]
        if t not in hist.columns:
            continue
        s = hist[t]
        close = float(s.dropna().iloc[-1]) if s.dropna().shape[0] > 0 else np.nan
        high_52w = float(s.tail(252).max()) if len(s) > 0 else np.nan
        sh6 = sharpe_score(s, 126)
        sh3 = sharpe_score(s, 63)
        if pd.notna(close) and pd.notna(high_52w) and high_52w > 0:
            pct_from_high = (high_52w - close) / high_52w
            high_pass = pct_from_high <= CONFIG.MAX_DRAWDOWN_FROM_HIGH
        else:
            pct_from_high = np.nan
            high_pass = True
        records.append({
            "TICKER": t, "ETF_NAME": row["ETF_NAME"], "SECTOR": classify_sector(row["ETF_NAME"], t),
            "CLOSE": close, "PCT_FROM_HIGH": pct_from_high * 100 if pd.notna(pct_from_high) else np.nan,
            "SHARPE_6M": sh6, "SHARPE_3M": sh3, "SCREEN_PASS": high_pass,
        })

    df = pd.DataFrame(records)
    if df.empty:
        return df

    inv_mask = df["SCREEN_PASS"]
    df["_Z6_INV"] = np.nan
    df["_Z3_INV"] = np.nan
    if inv_mask.sum() > 0:
        df.loc[inv_mask, "_Z6_INV"] = _zscore(df.loc[inv_mask, "SHARPE_6M"])
        df.loc[inv_mask, "_Z3_INV"] = _zscore(df.loc[inv_mask, "SHARPE_3M"])

    z6i, z3i = df["_Z6_INV"], df["_Z3_INV"]
    both = z6i.notna() & z3i.notna()
    only6 = z6i.notna() & z3i.isna()
    only3 = z6i.isna() & z3i.notna()
    df["_WTD_INV"] = np.nan
    df.loc[both, "_WTD_INV"] = CONFIG.SHARPE_W6M * z6i[both] + CONFIG.SHARPE_W3M * z3i[both]
    df.loc[only6, "_WTD_INV"] = CONFIG.SHARPE_W6M * z6i[only6]
    df.loc[only3, "_WTD_INV"] = CONFIG.SHARPE_W3M * z3i[only3]

    inv = df[df["SCREEN_PASS"]].copy()
    if len(inv) > 0:
        inv["RANK_INVESTABLE"] = inv["_WTD_INV"].rank(ascending=False, na_option="bottom").astype(int)
        df = df.merge(inv[["TICKER", "RANK_INVESTABLE"]], on="TICKER", how="left")
    else:
        df["RANK_INVESTABLE"] = np.nan

    return df.drop(columns=["_Z6_INV", "_Z3_INV", "_WTD_INV"], errors="ignore")


def build_ranking_v2(meta: pd.DataFrame, prices: pd.DataFrame, as_of: pd.Timestamp) -> pd.DataFrame:
    """V2 (3W-STRICT) composite: mean(Z(12M/6M/3M)), all three required."""
    hist = prices.loc[:as_of]
    records = []
    for _, row in meta.iterrows():
        t = row["TICKER"]
        if t not in hist.columns:
            continue
        s = hist[t]
        close = float(s.dropna().iloc[-1]) if s.dropna().shape[0] > 0 else np.nan
        high_52w = float(s.tail(252).max()) if len(s) > 0 else np.nan
        sharpes = {label: sharpe_score(s, days) for label, days in CONFIG.WINDOWS_V2.items()}
        if pd.notna(close) and pd.notna(high_52w) and high_52w > 0:
            pct_from_high = (high_52w - close) / high_52w
            high_pass = pct_from_high <= CONFIG.MAX_DRAWDOWN_FROM_HIGH
        else:
            pct_from_high = np.nan
            high_pass = True
        rec = {
            "TICKER": t, "ETF_NAME": row["ETF_NAME"], "SECTOR": classify_sector(row["ETF_NAME"], t),
            "CLOSE": close, "PCT_FROM_HIGH": pct_from_high * 100 if pd.notna(pct_from_high) else np.nan,
            "SCREEN_PASS": high_pass,
        }
        for label, val in sharpes.items():
            rec[f"SHARPE_{label}"] = val
        records.append(rec)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    inv_mask = df["SCREEN_PASS"]
    labels = list(CONFIG.WINDOWS_V2.keys())
    z_cols = {}
    for label in labels:
        z = pd.Series(np.nan, index=df.index)
        if inv_mask.sum() > 0:
            z.loc[inv_mask] = _zscore(df.loc[inv_mask, f"SHARPE_{label}"])
        z_cols[label] = z
    zdf = pd.DataFrame(z_cols)

    valid_all = zdf.notna().all(axis=1)
    composite = zdf.mean(axis=1, skipna=True)
    composite[~valid_all] = np.nan
    df["_WTD_INV"] = np.nan
    df.loc[inv_mask, "_WTD_INV"] = composite[inv_mask]

    inv = df[df["SCREEN_PASS"]].copy()
    if len(inv) > 0:
        inv["RANK_INVESTABLE"] = inv["_WTD_INV"].rank(ascending=False, na_option="bottom").astype(int)
        df = df.merge(inv[["TICKER", "RANK_INVESTABLE"]], on="TICKER", how="left")
    else:
        df["RANK_INVESTABLE"] = np.nan

    return df.drop(columns=["_WTD_INV"], errors="ignore")


def evaluate_regime(price, ema50, ema100) -> tuple:
    if pd.isna(price) or pd.isna(ema50) or pd.isna(ema100):
        return "BULL", CONFIG.TOP_N
    if ema50 > ema100 and price > ema50:
        return "BULL", CONFIG.TOP_N
    elif price > ema100:
        return "PARTIAL", CONFIG.TOP_N_PARTIAL
    else:
        return "BEAR", 0


def should_exit(row: pd.Series | None, peak: float, current_price: float) -> tuple[bool, str]:
    if row is None:
        return True, "Ticker no longer in ranking universe"
    reasons = []
    pct_from_high = row.get("PCT_FROM_HIGH", np.nan)
    if pd.notna(pct_from_high) and abs(pct_from_high) > CONFIG.EXIT_MAX_DD_FROM_HIGH * 100:
        reasons.append(f"52wk high DD {pct_from_high:.1f}%")
    inv_rank = row.get("RANK_INVESTABLE", np.nan)
    if pd.notna(inv_rank) and inv_rank > CONFIG.EXIT_MAX_RANK:
        reasons.append(f"Rank {int(inv_rank)}")
    if peak and peak > 0 and current_price and current_price > 0:
        dd_from_peak = (peak - current_price) / peak
        if dd_from_peak >= CONFIG.TSL_THRESHOLD:
            reasons.append(f"TSL {dd_from_peak*100:.1f}%")
    if reasons:
        return True, " | ".join(reasons)
    return False, ""


def _sell(t, slot, price, d, reason, regime_label, cash, trade_log):
    proceeds = slot["shares"] * price
    cost = CONFIG.TRADE_COST_FIXED
    pnl = proceeds - (slot["shares"] * slot["entry_price"]) - cost
    trade_log.append({
        "TYPE": "SELL", "REASON": reason, "TICKER": t, "NAME": slot.get("name", ""),
        "ENTRY_DATE": slot["entry_date"], "EXIT_DATE": d,
        "HOLDING_DAYS": (d - slot["entry_date"]).days,
        "ENTRY_PRICE": round(slot["entry_price"], 4), "EXIT_PRICE": round(price, 4),
        "SHARES": round(slot["shares"], 4),
        "GROSS_PNL": round(proceeds - slot["shares"] * slot["entry_price"], 2),
        "COSTS": cost, "NET_PNL": round(pnl, 2), "REGIME": regime_label,
    })
    return cash + proceeds - cost


def _buy(t, r, slot_size, p_entry, d, regime_label, trade_log):
    shares = slot_size / p_entry
    slot = {
        "shares": shares, "entry_price": p_entry, "peak": p_entry,
        "entry_date": d, "name": r["ETF_NAME"], "sector": r["SECTOR"],
    }
    trade_log.append({
        "TYPE": "BUY", "REASON": f"NEW BUY (rank={int(r['RANK_INVESTABLE'])})",
        "TICKER": t, "NAME": r["ETF_NAME"], "ENTRY_DATE": d, "EXIT_DATE": None,
        "HOLDING_DAYS": None, "ENTRY_PRICE": round(p_entry, 4), "EXIT_PRICE": None,
        "SHARES": round(shares, 4), "GROSS_PNL": None,
        "COSTS": CONFIG.TRADE_COST_FIXED, "NET_PNL": None, "REGIME": regime_label,
    })
    return slot, slot_size + CONFIG.TRADE_COST_FIXED


def compute_batch_weights_invvol(tickers: list, prices: pd.DataFrame, as_of: pd.Timestamp,
                                  slot_size_each: float) -> dict:
    n = len(tickers)
    if n == 0:
        return {}
    total_capital = n * slot_size_each
    vols = {}
    for t in tickers:
        s = prices.loc[:as_of, t] if t in prices.columns else pd.Series(dtype=float)
        vols[t] = realized_vol(s, CONFIG.VOL_WINDOW)
    if any(pd.isna(v) or v <= 0 for v in vols.values()):
        return {t: slot_size_each for t in tickers}
    inv = {t: 1.0 / v for t, v in vols.items()}
    s_inv = sum(inv.values())
    return {t: (inv[t] / s_inv) * total_capital for t in tickers}


# =========================================================
# ENGINE A: CURRENT -- weekly hold-and-replace, equal weight, sector cap=1
# =========================================================
def run_backtest_current(meta, prices, regime_s, regime_ema50, regime_ema100) -> dict:
    all_dates = prices.index
    week_starts = set(get_week_start_days(all_dates))
    date_list = list(all_dates)
    date_index = {d: i for i, d in enumerate(date_list)}

    cash = CONFIG.START_CAPITAL
    slots: list[dict] = []
    equity_history = []
    trade_log = []
    regime_label = "BULL"

    for d in all_dates:
        idx = date_index[d]
        cash *= (1 + CONFIG.CASH_INTEREST_PA / 365.0)

        if d in week_starts:
            prev_day = date_list[idx - 1] if idx > 0 else d
            if regime_s is not None and prev_day in regime_s.index:
                price = regime_s.loc[prev_day]; ema50 = regime_ema50.loc[prev_day]; ema100 = regime_ema100.loc[prev_day]
                regime_label, active_slots = evaluate_regime(price, ema50, ema100)
            else:
                regime_label, active_slots = "BULL", CONFIG.TOP_N

            rank_df = build_ranking_current(meta, prices, prev_day)
            rank_by_ticker = {r["TICKER"]: r for _, r in rank_df.iterrows()} if not rank_df.empty else {}

            port_val = sum(s["shares"] * prices.loc[d, s["ticker"]] for s in slots
                           if s["ticker"] in prices.columns and pd.notna(prices.loc[d, s["ticker"]]))
            total_equity = cash + port_val
            slot_size = total_equity / CONFIG.TOP_N

            new_slots = []
            held_tickers = set()
            sector_count: dict[str, int] = {}

            for prev_slot in slots:
                t = prev_slot["ticker"]
                if len(held_tickers) >= active_slots:
                    price_now = prices.loc[d, t] if t in prices.columns else np.nan
                    if pd.notna(price_now):
                        cash = _sell(t, prev_slot, price_now, d, "REGIME_SLOT_CUT", regime_label, cash, trade_log)
                    continue
                row = rank_by_ticker.get(t)
                current_price = float(prices.loc[d, t]) if t in prices.columns and pd.notna(prices.loc[d, t]) else 0.0
                exit_flag, reason = should_exit(row, prev_slot.get("peak", 0), current_price)
                if not exit_flag:
                    sector = prev_slot["sector"]
                    sector_count[sector] = sector_count.get(sector, 0) + 1
                    held_tickers.add(t)
                    new_peak = max(prev_slot.get("peak", 0) or 0, current_price)
                    new_slots.append({**prev_slot, "peak": new_peak})
                else:
                    price_now = current_price if current_price > 0 else prev_slot["entry_price"]
                    cash = _sell(t, prev_slot, price_now, d, reason, regime_label, cash, trade_log)

            open_slots = active_slots - len(new_slots)
            if open_slots > 0 and not rank_df.empty:
                candidates = rank_df[rank_df["SCREEN_PASS"] & (rank_df["RANK_INVESTABLE"] > 0)].sort_values("RANK_INVESTABLE")
                filled = 0
                for _, r in candidates.iterrows():
                    if filled >= open_slots:
                        break
                    t = r["TICKER"]
                    if t in held_tickers:
                        continue
                    sector = r["SECTOR"]
                    if sector_count.get(sector, 0) >= CONFIG.CURRENT_SECTOR_CAP:
                        continue
                    p_entry = prices.loc[d, t] if t in prices.columns else np.nan
                    if pd.isna(p_entry) or p_entry <= 0:
                        continue
                    slot, spend = _buy(t, r, slot_size, p_entry, d, regime_label, trade_log)
                    cash -= spend
                    slot["ticker"] = t
                    new_slots.append(slot)
                    sector_count[sector] = sector_count.get(sector, 0) + 1
                    held_tickers.add(t)
                    filled += 1
            slots = new_slots

        port_val = sum(s["shares"] * prices.loc[d, s["ticker"]] for s in slots
                       if s["ticker"] in prices.columns and pd.notna(prices.loc[d, s["ticker"]]))
        equity_history.append({"date": d, "equity": cash + port_val, "regime": regime_label, "n_holdings": len(slots)})

    eq = pd.DataFrame(equity_history).set_index("date")
    return {"equity": eq, "trades": pd.DataFrame(trade_log)}


# =========================================================
# ENGINE B: V2 -- monthly full-flush, inverse-vol sizing, no sector cap
# =========================================================
def run_backtest_v2(meta, prices, regime_s, regime_ema50, regime_ema100) -> dict:
    all_dates = prices.index
    month_starts = set(get_month_start_days(all_dates))
    date_list = list(all_dates)
    date_index = {d: i for i, d in enumerate(date_list)}

    cash = CONFIG.START_CAPITAL
    slots: list[dict] = []
    equity_history = []
    trade_log = []
    regime_label = "BULL"

    for d in all_dates:
        idx = date_index[d]
        cash *= (1 + CONFIG.CASH_INTEREST_PA / 365.0)

        if d in month_starts:
            prev_day = date_list[idx - 1] if idx > 0 else d
            if regime_s is not None and prev_day in regime_s.index:
                price = regime_s.loc[prev_day]; ema50 = regime_ema50.loc[prev_day]; ema100 = regime_ema100.loc[prev_day]
                regime_label, active_slots = evaluate_regime(price, ema50, ema100)
            else:
                regime_label, active_slots = "BULL", CONFIG.TOP_N

            for prev_slot in slots:
                t = prev_slot["ticker"]
                price_now = prices.loc[d, t] if t in prices.columns else np.nan
                if pd.isna(price_now):
                    price_now = prev_slot["entry_price"]
                cash = _sell(t, prev_slot, price_now, d, "MONTHLY_FLUSH", regime_label, cash, trade_log)
            slots = []

            total_equity = cash
            slot_size = total_equity / CONFIG.TOP_N

            rank_df = build_ranking_v2(meta, prices, prev_day)
            new_slots = []
            if active_slots > 0 and not rank_df.empty:
                candidates = rank_df[rank_df["SCREEN_PASS"] & (rank_df["RANK_INVESTABLE"] > 0)].sort_values("RANK_INVESTABLE")
                to_buy = []
                sector_count: dict[str, int] = {}
                for _, r in candidates.iterrows():
                    if len(to_buy) >= active_slots:
                        break
                    t = r["TICKER"]
                    sector = r["SECTOR"]
                    if sector_count.get(sector, 0) >= CONFIG.V2_SECTOR_CAP:
                        continue
                    p_entry = prices.loc[d, t] if t in prices.columns else np.nan
                    if pd.isna(p_entry) or p_entry <= 0:
                        continue
                    to_buy.append((t, r))
                    sector_count[sector] = sector_count.get(sector, 0) + 1

                weights = compute_batch_weights_invvol([t for t, _ in to_buy], prices, prev_day, slot_size)
                for t, r in to_buy:
                    p_entry = prices.loc[d, t]
                    slot, spend = _buy(t, r, weights[t], p_entry, d, regime_label, trade_log)
                    cash -= spend
                    slot["ticker"] = t
                    new_slots.append(slot)
            slots = new_slots

        port_val = sum(s["shares"] * prices.loc[d, s["ticker"]] for s in slots
                       if s["ticker"] in prices.columns and pd.notna(prices.loc[d, s["ticker"]]))
        equity_history.append({"date": d, "equity": cash + port_val, "regime": regime_label, "n_holdings": len(slots)})

    eq = pd.DataFrame(equity_history).set_index("date")
    return {"equity": eq, "trades": pd.DataFrame(trade_log)}


def compute_metrics(eq: pd.DataFrame) -> dict:
    years = (eq.index[-1] - eq.index[0]).days / 365.25
    initial = CONFIG.START_CAPITAL
    final = eq["equity"].iloc[-1]
    cagr = (final / initial) ** (1 / years) - 1
    eq = eq.copy()
    eq["peak"] = eq["equity"].cummax()
    eq["drawdown"] = (eq["equity"] - eq["peak"]) / eq["peak"]
    max_dd = eq["drawdown"].min()
    daily_ret = eq["equity"].pct_change().dropna()
    vol = daily_ret.std() * np.sqrt(252)
    sharpe = cagr / vol if vol > 0 else 0
    return {"years": years, "initial": initial, "final": final, "cagr": cagr,
            "max_dd": max_dd, "vol": vol, "sharpe": sharpe}


def print_results(name: str, res: dict, metrics: dict):
    tlog = res["trades"]
    sells = tlog[tlog["TYPE"] == "SELL"] if len(tlog) else pd.DataFrame()
    buys = tlog[tlog["TYPE"] == "BUY"] if len(tlog) else pd.DataFrame()
    print(f"\n{'='*60}\n  v13 RESULTS -- {name}\n{'='*60}")
    print(f"  Period          : {res['equity'].index[0].date()} -> {res['equity'].index[-1].date()}")
    print(f"  CAGR            : {metrics['cagr']:>10.2%}")
    print(f"  Max Drawdown    : {metrics['max_dd']:>10.2%}")
    print(f"  Annualised Vol  : {metrics['vol']:>10.2%}")
    print(f"  Sharpe (simple) : {metrics['sharpe']:>10.2f}")
    print(f"  Total BUYs      : {len(buys)}  |  Total SELLs: {len(sells)}")
    if len(sells) > 0:
        wins = (sells["NET_PNL"] > 0).sum()
        print(f"  Win Rate        : {wins/len(sells):.1%}")


def main():
    _load_sector_lookup()
    meta, prices = load_data(CONFIG.INPUT_FILE)
    all_dates = prices.index

    print(f"Fetching Nifty 500 regime data ({CONFIG.BENCHMARK_TICKER}) ...")
    regime_s = regime_ema50 = regime_ema100 = None
    try:
        reg_raw = yf.download(CONFIG.BENCHMARK_TICKER, start=all_dates[0].strftime("%Y-%m-%d"),
                               end=(all_dates[-1] + pd.Timedelta(days=5)).strftime("%Y-%m-%d"),
                               auto_adjust=True, progress=False)
        regime_raw = reg_raw["Close"].squeeze().dropna()
        regime_raw.index = pd.to_datetime(regime_raw.index).tz_localize(None)
        if len(regime_raw) == 0:
            raise ValueError("empty live series")
        regime_s = regime_raw.reindex(all_dates, method="ffill")
        print(f"  Live regime source: {CONFIG.BENCHMARK_TICKER} ({regime_raw.index[0].date()} -> {regime_raw.index[-1].date()})")
    except Exception as e:
        print(f"  [warn] Live fetch failed ({e}); falling back to {CONFIG.REGIME_XLSX_FALLBACK} column")
        if CONFIG.REGIME_XLSX_FALLBACK in prices.columns:
            regime_s = prices[CONFIG.REGIME_XLSX_FALLBACK].dropna().reindex(all_dates, method="ffill")

    if regime_s is not None:
        regime_ema50 = regime_s.ewm(span=CONFIG.TREND_FAST_EMA_WINDOW, adjust=False).mean()
        regime_ema100 = regime_s.ewm(span=CONFIG.TREND_EMA_WINDOW, adjust=False).mean()

    print(f"\n{'='*70}\n  Running CURRENT (live: weekly, equal-weight, sector-cap=1)\n{'='*70}")
    res_current = run_backtest_current(meta, prices, regime_s, regime_ema50, regime_ema100)
    res_current["equity"].to_csv(_BASE / "v13_CURRENT_backtest_equity.csv")
    res_current["trades"].to_csv(_BASE / "v13_CURRENT_backtest_trade_log.csv", index=False)
    m_current = compute_metrics(res_current["equity"])
    print_results("CURRENT", res_current, m_current)

    print(f"\n{'='*70}\n  Running V2 (3W-STRICT, monthly, inv-vol, no cap)\n{'='*70}")
    res_v2 = run_backtest_v2(meta, prices, regime_s, regime_ema50, regime_ema100)
    res_v2["equity"].to_csv(_BASE / "v13_V2_backtest_equity.csv")
    res_v2["trades"].to_csv(_BASE / "v13_V2_backtest_trade_log.csv", index=False)
    m_v2 = compute_metrics(res_v2["equity"])
    print_results("V2", res_v2, m_v2)

    # ---- Comparison chart ----
    fig, ax = plt.subplots(figsize=(13, 7))
    for name, res, color in [("CURRENT (live)", res_current, "#7f8c8d"), ("V2 (3W-STRICT/Monthly/InvVol/NoCap)", res_v2, "#2980b9")]:
        eq = res["equity"]
        norm = eq["equity"] / eq["equity"].iloc[0] * 100
        ax.plot(eq.index, norm, label=name, color=color, linewidth=1.4)
    ax.set_title("ETF Momentum v13 -- CURRENT (live) vs V2, full history 2016-2026")
    ax.set_ylabel("Growth of 100 (log scale)")
    ax.set_yscale("log")
    ax.legend()
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    fig.tight_layout()
    fig.savefig(_BASE / "v13_comparison_equity_curve.png", dpi=140)
    print(f"\nComparison chart -> {_BASE / 'v13_comparison_equity_curve.png'}")

    # ---- Summary table ----
    print(f"\n{'='*70}\n  SUMMARY: CURRENT (live) vs V2 -- full history {all_dates[0].date()} -> {all_dates[-1].date()}\n{'='*70}")
    print(f"  {'Metric':<20}{'CURRENT':>18}{'V2':>18}")
    for label, key, fmt in [
        ("CAGR", "cagr", "{:.2%}"), ("Total Return", None, None),
        ("Max Drawdown", "max_dd", "{:.2%}"), ("Volatility", "vol", "{:.2%}"),
        ("Sharpe (simple)", "sharpe", "{:.2f}"), ("Final Equity", "final", "INR {:,.0f}"),
    ]:
        if key is None:
            c = f"{(m_current['final']/m_current['initial'] - 1):.2%}"
            v = f"{(m_v2['final']/m_v2['initial'] - 1):.2%}"
        else:
            c = fmt.format(m_current[key]); v = fmt.format(m_v2[key])
        print(f"  {label:<20}{c:>18}{v:>18}")


if __name__ == "__main__":
    main()
