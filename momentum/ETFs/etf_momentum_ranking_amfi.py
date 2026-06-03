"""
etf_momentum_ranking_amfi.py
============================
AMFI-backed version of the ETF Dual Momentum + Weighted Sharpe ranking script.

Differences from etf_momentum_ranking.py:
  - Price data  : AMFI_NAV_History.xlsx (produced by download_amfi_nav.py)
  - ETF metadata: AMFI ETF Codes.csv
  - ETF ID      : AMFI Scheme Code (numeric string, e.g. "153744")
  - ETF Name    : AMFI Scheme Name (e.g. "Groww BSE Power ETF")
  - Regime      : Nifty 500 index (^CRSLDX) fetched via yfinance — not ETF NAV
  - TSL NAV     : Fetched from mfapi.in by AMFI code — no .NS suffix needed

Scoring pipeline (unchanged):
  Step 1 - SCREEN  : 52-week high proximity filter (< MAX_DRAWDOWN_FROM_HIGH)
  Step 2 - SCORE   : Weighted Sharpe (6M + 3M) on all ETFs
  Step 3 - REGIME  : Tiered BULL / PARTIAL / BEAR via EMA50/EMA100 on Nifty 500
  Step 4 - ALLOCATE: Top-N from screened + ranked investable universe

Usage:
  python etf_momentum_ranking_amfi.py          # monthly rebalance
  python etf_momentum_ranking_amfi.py --tsl    # daily TSL check
"""

from __future__ import annotations
import sys
import json
import time
import requests
import numpy as np
from scipy import stats
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIG  <- edit these freely
# =========================================================
class CONFIG:
    # Input files (same directory as this script)
    AMFI_NAV_FILE   = "AMFI_NAV_History.xlsx"   # produced by download_amfi_nav.py
    AMFI_CODES_FILE = "AMFI ETF Codes.csv"       # AMFI scheme codes + names
    OUTPUT_FILE     = "etf_rankings_amfi.xlsx"
    HOLDINGS_FILE   = "holdings_log_amfi.json"   # separate log from original script

    # Momentum windows (trading days)
    WINDOW_6M   = 126
    WINDOW_3M   = 63
    ANNUALIZE   = 252

    # Portfolio allocation
    TOP_N         = 5    # slots when regime = BULL
    TOP_N_PARTIAL = 3    # slots when regime = PARTIAL

    # 52-week high proximity filter
    MAX_DRAWDOWN_FROM_HIGH = 0.25   # ETF must be within 25% of its 52wk high

    # Weighted Sharpe blending weights
    SHARPE_W6M  = 0.5
    SHARPE_W3M  = 0.5

    # R-squared blend weights (display metric only)
    R2_W6M      = 0.5
    R2_W3M      = 0.5

    # Regime filter — uses Nifty 500 index via yfinance
    REGIME_YF_TICKER   = "^CRSLDX"    # Nifty 500 index
    TREND_FAST_EMA     = 50
    TREND_SLOW_EMA     = 100
    REGIME_HISTORY_DAYS = 150         # days of index history to fetch for EMA

    # Sector cap — max ETFs per sector in allocation
    SECTOR_CAP = 1

    # Daily risk-free rate (7% annual / 252)
    DAILY_RF = 0.07 / 252

    # Trailing Stop Loss threshold
    TSL_THRESHOLD = 0.10   # 10% drawdown from peak


# =========================================================
# SECTOR CLASSIFICATION
# =========================================================
_SECTOR_RULES = [
    ("PSU_BANK",         ["psu bank", "psubnk", "psubank", "bse psu bank"]),
    ("PRIVATE_BANK",     ["private bank", "pvt bank", "pvtban", "nifty pb "]),
    ("BANKING_BROAD",    ["nifty bank", "bse bank", " bank ", "banketf", "bankbees"]),
    ("IT_TECH",          ["nifty it", "bse it", " it etf", "itbees", "nifit"]),
    ("HEALTHCARE",       ["healthcare", "pharma", "health ", "hospitals"]),
    ("METAL",            ["metal"]),
    ("ENERGY",           ["energy", "oil & gas", "o&g", "power etf", "bse power"]),
    ("INFRA",            ["infra"]),
    ("CONSUMPTION",      ["consumption", "consump", "fmcg", "consumer"]),
    ("REALTY",           ["realty", "real estate"]),
    ("DEFENCE",          ["defence", "dfnc"]),
    ("PSE",              ["pse etf", "cpse", "nifty pse", "bharat 22", "cpseetf"]),
    ("AUTO",             ["auto"]),
    ("CHEMICALS",        ["chemical"]),
    ("FIN_SERVICES",     ["fin serv", "financial serv", "capital mkt", "capital market"]),
    ("COMMODITIES",      ["commodity", "commo"]),
    ("MANUFACTURING",    ["manufactur", "manu"]),
    ("EV_MOBILITY",      ["ev & new", "ev new", "nifty ev"]),
    ("DIGITAL_INTERNET", ["internet", "digital"]),
    ("RAILWAY",          ["railway"]),
    ("TOURISM",          ["tourism"]),
    ("MNC",              ["mnc"]),
    ("GOLD",             ["gold"]),
    ("SILVER",           ["silver"]),
    ("GOVT_BONDS",       ["g-sec", "gsec", "gilt", "bond etf", "bharat bond"]),
    ("DIVIDEND",         ["dividend", "div opp"]),
    ("IPO",              ["ipo"]),
    ("ESG",              ["esg"]),
    ("FACTOR_MOMENTUM",  ["momentum", "mmt"]),
    ("FACTOR_VALUE",     ["value 20", "value 30", "value 50", "enhanced val"]),
    ("FACTOR_QUALITY",   ["quality"]),
    ("FACTOR_LOW_VOL",   ["low vol", "lowvol"]),
    ("FACTOR_ALPHA",     ["alpha"]),
    ("FACTOR_EQUAL_WT",  ["equal weight", "equal wt"]),
    ("INTERNATIONAL",    ["nasdaq", "s&p 500", "hang seng", "msci", "fang+"]),
    ("MIDCAP",           ["midcap", "mid cap", "midsmall"]),
    ("SMALLCAP",         ["smallcap", "small cap"]),
    ("NEXT_50",          ["next 50", "next50"]),
    ("BROAD_MARKET",     ["nifty 50", "nifty50", "sensex", "nifty 100", "nifty 200",
                          "nifty 500", "total market", "bse 500", "multicap",
                          "flexicap", "flexi", "largemidc"]),
    ("SERVICES",         ["services"]),
    ("LIQUID",           ["liquid", "1d rate"]),
]


def classify_sector(etf_name: str) -> str:
    n = etf_name.lower()
    for sector, keywords in _SECTOR_RULES:
        for kw in keywords:
            if kw in n:
                return sector
    return "OTHER"


# =========================================================
# 1. DATA LOADING — AMFI NAV History
# =========================================================
def load_etf_data_amfi(nav_file: str, codes_file: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load ETF price history from AMFI_NAV_History.xlsx and metadata from
    AMFI ETF Codes.csv.

    Returns
    -------
    meta   : DataFrame with columns ETF_NAME, TICKER (TICKER = AMFI code string)
    prices : DataFrame (index=dates, columns=AMFI code strings), forward-filled
    """
    print(f"[load]   {Path(nav_file).name}  +  {Path(codes_file).name}")

    # Load codes CSV for metadata (Code -> Scheme Name mapping)
    codes_df = pd.read_csv(codes_file)
    codes_df.columns = [c.strip() for c in codes_df.columns]
    codes_df["Code"] = codes_df["Code"].astype(str).str.strip()
    # Build name -> code lookup (for matching columns in the NAV xlsx)
    name_to_code = dict(zip(codes_df["Scheme Name"].str.strip(), codes_df["Code"]))
    code_to_name = dict(zip(codes_df["Code"], codes_df["Scheme Name"].str.strip()))

    # Load NAV history (wide format: index=Date, columns=Scheme Names)
    nav_df = pd.read_excel(nav_file, index_col=0, sheet_name="NAV History")
    nav_df.index = pd.to_datetime(nav_df.index)
    nav_df.sort_index(inplace=True)

    # Rename columns from scheme names to AMFI codes
    rename_map = {}
    unmatched  = []
    for col in nav_df.columns:
        col_clean = str(col).strip()
        code = name_to_code.get(col_clean)
        if code:
            rename_map[col] = code
        else:
            unmatched.append(col_clean)

    if unmatched:
        print(f"  [warn] {len(unmatched)} column(s) in NAV file not matched to AMFI codes "
              f"(will be skipped): {unmatched[:5]}{'...' if len(unmatched)>5 else ''}")

    nav_df = nav_df.rename(columns=rename_map)
    # Keep only columns that were matched to an AMFI code
    matched_codes = [c for c in nav_df.columns if c in set(codes_df["Code"])]
    nav_df = nav_df[matched_codes]
    nav_df = nav_df.apply(pd.to_numeric, errors="coerce").replace(0, np.nan)

    # Forward-fill missing NAVs (weekends/holidays)
    prices = nav_df.ffill()

    # Build meta — only for ETFs present in the prices grid
    meta_rows = []
    for code in prices.columns:
        name = code_to_name.get(code, f"Unknown ({code})")
        meta_rows.append({
            "TICKER"  : code,
            "ETF_NAME": name,
        })
    meta = pd.DataFrame(meta_rows).reset_index(drop=True)

    print(f"         {len(meta)} ETFs  |  {len(prices)} date rows  "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})")
    return meta, prices


# =========================================================
# 2. SCORING (unchanged from original)
# =========================================================
def sharpe_score(series: pd.Series, window: int) -> float:
    """Annualised Sharpe over lookback window, excess of daily RF."""
    clean = series.dropna()
    if len(clean) < window + 1:
        return np.nan
    log_ret = np.log(clean.iloc[-window - 1:] / clean.iloc[-window - 1:].shift(1)).dropna()
    excess  = log_ret - CONFIG.DAILY_RF
    if excess.std() == 0:
        return np.nan
    return (excess.mean() / excess.std()) * np.sqrt(CONFIG.ANNUALIZE)


def r2_score(series: pd.Series, window: int) -> float:
    """R-squared from log-linear regression over the lookback window."""
    clean = series.dropna()
    if len(clean) < window:
        return np.nan
    y = np.log(clean.iloc[-window:].values.astype(float))
    x = np.arange(len(y))
    _, _, r, _, _ = stats.linregress(x, y)
    return r ** 2


# =========================================================
# 3. REGIME FILTER — Nifty 500 index via yfinance
# =========================================================
def regime_status() -> dict:
    """
    Fetch Nifty 500 index history from yfinance (^CRSLDX) and compute
    tiered regime state:
      BULL    - EMA50 > EMA100 AND Price > EMA50  -> TOP_N slots
      PARTIAL - Price > EMA100 (not BULL)         -> TOP_N_PARTIAL slots
      BEAR    - Price <= EMA100                   -> 0 slots, full cash
    """
    import yfinance as yf

    ticker_label = CONFIG.REGIME_YF_TICKER
    nifty_price  = np.nan
    ema_50       = np.nan
    ema_100      = np.nan
    label        = "BULL"
    active_slots = CONFIG.TOP_N

    try:
        data = yf.download(
            ticker_label,
            period=f"{CONFIG.REGIME_HISTORY_DAYS + 30}d",
            auto_adjust=True,
            progress=False,
            timeout=15,
        )
        if data.empty:
            raise ValueError("Empty response from yfinance")

        close = data["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
        close = close.dropna().sort_index()

        if len(close) >= CONFIG.TREND_SLOW_EMA:
            nifty_price = float(close.iloc[-1])
            ema_50      = float(close.ewm(span=CONFIG.TREND_FAST_EMA, adjust=False).mean().iloc[-1])
            ema_100     = float(close.ewm(span=CONFIG.TREND_SLOW_EMA, adjust=False).mean().iloc[-1])

            if ema_50 > ema_100 and nifty_price > ema_50:
                label        = "BULL"
                active_slots = CONFIG.TOP_N
            elif nifty_price > ema_100:
                label        = "PARTIAL"
                active_slots = CONFIG.TOP_N_PARTIAL
            else:
                label        = "BEAR"
                active_slots = 0
        else:
            print(f"  [warn] Insufficient index history ({len(close)} pts); defaulting to BULL")

    except Exception as e:
        print(f"  [warn] Could not fetch regime data from yfinance: {e}")
        print("         Defaulting to BULL regime.")

    return {
        "label"        : label,
        "active_slots" : active_slots,
        "nifty_price"  : nifty_price,
        "nifty_ema_50" : ema_50,
        "nifty_ema_100": ema_100,
        "trend_ticker" : ticker_label,
    }


# =========================================================
# 4. SCORING + RANKING (identical logic to original)
# =========================================================
def build_ranking(meta: pd.DataFrame, prices: pd.DataFrame) -> pd.DataFrame:
    """
    Compute Sharpe scores, R2 scores, 52-week high screen and investable
    ranks for all ETFs. Identifier = AMFI code (stored in TICKER column).
    """
    records = []

    # EOM date for MTD return calculation
    last_date    = prices.index[-1]
    prev_indices = prices.index[
        (prices.index.month != last_date.month) |
        (prices.index.year  != last_date.year)
    ]
    if len(prev_indices) > 0:
        eom_date   = prev_indices[-1]
        eom_series = prices.loc[eom_date]
        print(f"         Comp date: {eom_date.date()} (prev month eom)")
    else:
        eom_series = pd.Series(dtype=float)

    for _, row in meta.iterrows():
        code     = row["TICKER"]    # AMFI scheme code string
        etf_name = row["ETF_NAME"]

        if code not in prices.columns:
            continue

        s     = prices[code]
        close = float(s.iloc[-1]) if len(s) > 0 else np.nan

        # 52-week high from trailing 252 trading days
        high_52w = float(s.tail(252).max()) if len(s) > 0 else np.nan

        # Sharpe scores
        sh6 = sharpe_score(s, CONFIG.WINDOW_6M)
        sh3 = sharpe_score(s, CONFIG.WINDOW_3M)

        # R2 scores
        r6 = r2_score(s, CONFIG.WINDOW_6M)
        r3 = r2_score(s, CONFIG.WINDOW_3M)

        # Weighted Sharpe — fallback gracefully for newly-listed ETFs
        if not np.isnan(sh6) and not np.isnan(sh3):
            wtd_sharpe = CONFIG.SHARPE_W6M * sh6 + CONFIG.SHARPE_W3M * sh3
        elif not np.isnan(sh6):
            wtd_sharpe = sh6
        elif not np.isnan(sh3):
            wtd_sharpe = sh3
        else:
            wtd_sharpe = np.nan

        # Sharpe * R2 composite (display metric only)
        _sh6 = 0.0 if np.isnan(sh6) else sh6
        _sh3 = 0.0 if np.isnan(sh3) else sh3
        _r6  = 0.0 if (r6 is None or np.isnan(r6)) else r6
        _r3  = 0.0 if (r3 is None or np.isnan(r3)) else r3
        sr2_6m    = _sh6 * _r6
        sr2_3m    = _sh3 * _r3
        sr2_blend = (sr2_6m + sr2_3m) / 2

        # 100-day EMA
        clean_s = s.dropna()
        ema_100 = (float(clean_s.ewm(span=100, adjust=False).mean().iloc[-1])
                   if len(clean_s) >= 100 else np.nan)

        # MTD return vs EOM
        eom_price = eom_series.get(code, np.nan)
        if pd.notna(close) and pd.notna(eom_price) and eom_price > 0:
            eom_ret = (close / eom_price - 1) * 100
        else:
            eom_ret = np.nan

        # 52-week high screen
        if pd.notna(close) and pd.notna(high_52w) and high_52w > 0:
            pct_from_high = (high_52w - close) / high_52w
            high_pass     = pct_from_high <= CONFIG.MAX_DRAWDOWN_FROM_HIGH
        else:
            pct_from_high = np.nan
            high_pass     = True    # insufficient history -> don't penalise

        records.append({
            "TICKER"       : code,
            "ETF_NAME"     : etf_name,
            "SECTOR"       : classify_sector(etf_name),
            "CLOSE"        : close,
            "52WK_HIGH"    : high_52w,
            "PCT_FROM_HIGH": pct_from_high * 100 if not np.isnan(pct_from_high) else np.nan,
            "EMA_100"      : ema_100,
            "EOM_PCT"      : eom_ret,
            "SHARPE_6M"    : sh6,
            "SHARPE_3M"    : sh3,
            "WTD_SHARPE"   : wtd_sharpe,
            "R2_6M"        : r6,
            "R2_3M"        : r3,
            "SR2_6M"       : sr2_6m,
            "SR2_3M"       : sr2_3m,
            "SR2_BLEND"    : sr2_blend,
            "SCREEN_PASS"  : high_pass,
        })

    df = pd.DataFrame(records)

    # Universe rank (all ETFs, Weighted Sharpe)
    df["RANK_UNIVERSE"] = df["WTD_SHARPE"].rank(ascending=False, na_option="bottom").astype(int)
    df["RANK_SHARPE"]   = df["RANK_UNIVERSE"]

    # SR2 rank (display)
    df["RANK_SR2"] = df["SR2_BLEND"].rank(ascending=False, na_option="bottom").astype(int)

    # Investable rank (screen-pass ETFs only)
    inv = df[df["SCREEN_PASS"]].copy()
    if len(inv) > 0:
        inv["RANK_INVESTABLE"] = inv["WTD_SHARPE"].rank(ascending=False, na_option="bottom").astype(int)
        df = df.merge(inv[["TICKER", "RANK_INVESTABLE"]], on="TICKER", how="left")
    else:
        df["RANK_INVESTABLE"] = np.nan

    df["RANK_INVESTABLE"] = df["RANK_INVESTABLE"].fillna(0).astype(int)

    # Sort: screen-pass ETFs first, then by universe rank
    df["_sort"] = df["RANK_INVESTABLE"].replace(0, 9999)
    df = df.sort_values(["_sort", "RANK_UNIVERSE"]).drop(columns="_sort").reset_index(drop=True)

    return df


# =========================================================
# 5. PORTFOLIO ALLOCATION (identical to original)
# =========================================================
def build_allocation(df: pd.DataFrame, regime: dict) -> pd.DataFrame:
    """Select top ETFs from investable universe per regime state."""
    active = regime["active_slots"]
    total  = CONFIG.TOP_N
    w      = 1.0 / total

    if active == 0:
        return pd.DataFrame([{
            "SLOT"    : i + 1, "TICKER": "CASH",
            "ETF_NAME": "Cash / Money Market", "SECTOR": "CASH",
            "WEIGHT"  : w, "INV_RANK": "-",
            "REASON"  : f"Regime = {regime['label']} -> full cash",
        } for i in range(total)])

    investable = (df[df["SCREEN_PASS"] & (df["RANK_INVESTABLE"] > 0)]
                  .sort_values("RANK_INVESTABLE")
                  .reset_index(drop=True))
    is_partial   = (active == CONFIG.TOP_N_PARTIAL)
    sector_count : dict[str, int] = {}
    slots  = []
    cidx   = 0
    slotno = 1

    while slotno <= active:
        filled = False
        while cidx < len(investable):
            row    = investable.iloc[cidx]
            sector = row.get("SECTOR", "OTHER")
            cur    = sector_count.get(sector, 0)
            cidx  += 1
            if cur < CONFIG.SECTOR_CAP:
                sector_count[sector] = cur + 1
                slots.append({
                    "SLOT"    : slotno,
                    "TICKER"  : row["TICKER"],
                    "ETF_NAME": row["ETF_NAME"],
                    "SECTOR"  : sector,
                    "WEIGHT"  : w,
                    "INV_RANK": int(row["RANK_INVESTABLE"]),
                    "REASON"  : (f"WTD_SHARPE rank={int(row['RANK_INVESTABLE'])}  |  "
                                 f"Sector={sector} ({cur+1}/{CONFIG.SECTOR_CAP})  |  "
                                 f"3M Sharpe={row['SHARPE_3M']:.3f}"),
                })
                slotno += 1
                filled  = True
                break
        if not filled:
            slots.append({
                "SLOT"    : slotno, "TICKER": "CASH",
                "ETF_NAME": "Cash (universe exhausted after sector cap)",
                "SECTOR"  : "CASH", "WEIGHT": w, "INV_RANK": "-",
                "REASON"  : f"No remaining ETF after sector cap={CONFIG.SECTOR_CAP}",
            })
            slotno += 1

    for slotno in range(active + 1, total + 1):
        slots.append({
            "SLOT"    : slotno, "TICKER": "CASH",
            "ETF_NAME": "Cash / Money Market", "SECTOR": "CASH",
            "WEIGHT"  : w, "INV_RANK": "-",
            "REASON"  : (f"Regime buffer: {regime['label']} -> "
                         f"only {active} of {total} slots active"
                         if is_partial else "Universe exhausted"),
        })

    return pd.DataFrame(slots)


# =========================================================
# 6. CONSOLE SUMMARY
# =========================================================
def print_summary(df: pd.DataFrame, regime: dict, allocation: pd.DataFrame):
    W = 115
    print("\n" + "=" * W)
    print("ETF MOMENTUM RANKING (AMFI)  |  Screen -> Score -> Regime -> Allocate")
    print("=" * W)

    r = regime
    print(f"\n  REGIME: {r['label']:30s}  Active slots: {r['active_slots']} / {CONFIG.TOP_N}")
    print(f"  Index ({r['trend_ticker']}): "
          f"Price={r['nifty_price']:.2f}  |  "
          f"EMA50={r['nifty_ema_50']:.2f}  |  "
          f"EMA100={r['nifty_ema_100']:.2f}")

    # Top 5 recommendations (1 per sector)
    print("\n  " + "=" * 55)
    print("  RECOMMENDATION  (Top 5 Investable | 1 Per Sector)")
    print("  " + "=" * 55)
    inv_df = (df[df["SCREEN_PASS"] & (df["RANK_INVESTABLE"] > 0)]
              .sort_values("RANK_INVESTABLE"))
    recs, seen_sectors = [], set()
    for _, row in inv_df.iterrows():
        sec = row.get("SECTOR", "OTHER")
        if sec not in seen_sectors:
            seen_sectors.add(sec)
            recs.append(row)
        if len(recs) >= 5:
            break
    if not recs:
        print("    [None] No ETFs passed screen filter.")
    else:
        for i, rec in enumerate(recs, 1):
            print(f"    {i}. [{rec['TICKER']}]  {rec['SECTOR']:<15}  {rec['ETF_NAME'][:50]}")

    print(f"\n  ALLOCATION (Execution State — Regime: {r['label']})")
    print("  " + "-" * 110)
    for _, a in allocation.iterrows():
        is_cash = a["TICKER"] == "CASH"
        marker  = "  [CASH]" if is_cash else "  [HOLD]"
        print(f"  Slot {int(a['SLOT'])}: [{a['TICKER']:<7}]  {a['WEIGHT']:5.1%}{marker}  {a['ETF_NAME'][:55]}")

    inv_count = df["SCREEN_PASS"].sum()
    print(f"\n  RANKING  (investable rank = scored among {inv_count} ETFs passing screen)")
    print(f"  {'InvRk':>5} {'UniRk':>5} {'AMFI Code':>10} {'ETF Name':<45} "
          f"{'WtdSharpe':>10} {'Sh6M':>7} {'Sh3M':>7} {'Screen':>7}")
    print("  " + "-" * 100)
    for _, r2 in df.head(5).iterrows():
        def f(v, d=3): return f"{v:.{d}f}" if pd.notna(v) and v != 0 else "N/A"
        inv_rk = str(int(r2["RANK_INVESTABLE"])) if r2["SCREEN_PASS"] else "-"
        screen = "PASS" if r2["SCREEN_PASS"] else "FAIL"
        print(f"  {inv_rk:>5} {int(r2['RANK_UNIVERSE']):>5} {r2['TICKER']:>10} "
              f"{str(r2['ETF_NAME'])[:44]:<45} "
              f"{f(r2['WTD_SHARPE']):>10} {f(r2['SHARPE_6M']):>7} {f(r2['SHARPE_3M']):>7} "
              f"{screen:>7}")

    print(f"\n  Universe={len(df)}  Investable={inv_count}  "
          f"Screened out={len(df)-inv_count}  "
          f"Valid WtdSharpe={df['WTD_SHARPE'].notna().sum()}")
    print("=" * W)


# =========================================================
# 7. HOLDINGS LOG & REBALANCE TRACKER (same logic, separate file)
# =========================================================
HISTORY_MONTHS = 12


def _log_path(script_dir: Path) -> Path:
    return script_dir / CONFIG.HOLDINGS_FILE


def load_holdings_log(script_dir: Path) -> dict:
    p = _log_path(script_dir)
    if p.exists():
        try:
            with open(p) as fh:
                return json.load(fh)
        except Exception:
            return {}
    return {}


def save_holdings_log(script_dir: Path, log: dict):
    with open(_log_path(script_dir), "w") as fh:
        json.dump(log, fh, indent=2, default=str)


def record_to_log(allocation: pd.DataFrame, regime: dict, run_date: str) -> dict:
    slots = []
    for _, row in allocation.iterrows():
        slots.append({
            "slot"    : int(row["SLOT"]),
            "ticker"  : str(row["TICKER"]),    # AMFI code or "CASH"
            "etf_name": str(row["ETF_NAME"]),
            "sector"  : str(row.get("SECTOR", "")),
            "weight"  : float(row["WEIGHT"]),
            "inv_rank": str(row["INV_RANK"]),
        })
    return {
        "run_date"    : run_date,
        "regime"      : regime["label"],
        "active_slots": int(regime["active_slots"]),
        "allocation"  : slots,
    }


def diff_allocations(prev: dict, curr: dict) -> list[dict]:
    prev_alloc = {s["ticker"]: s for s in prev.get("allocation", [])}
    curr_alloc = {s["ticker"]: s for s in curr.get("allocation", [])}
    prev_holds = {t for t in prev_alloc if t != "CASH"}
    curr_holds = {t for t in curr_alloc if t != "CASH"}
    changes = []

    for t in sorted(curr_holds - prev_holds):
        s = curr_alloc[t]
        changes.append({"action": "BUY",  "ticker": t, "etf_name": s["etf_name"],
                        "sector": s["sector"], "prev_wt": 0.0, "curr_wt": s["weight"],
                        "prev_rk": "-", "curr_rk": s["inv_rank"], "note": "New entry"})

    for t in sorted(prev_holds - curr_holds):
        s = prev_alloc[t]
        changes.append({"action": "SELL", "ticker": t, "etf_name": s["etf_name"],
                        "sector": s["sector"], "prev_wt": s["weight"], "curr_wt": 0.0,
                        "prev_rk": s["inv_rank"], "curr_rk": "-", "note": "Exited"})

    for t in sorted(prev_holds & curr_holds):
        ps, cs = prev_alloc[t], curr_alloc[t]
        pw, cw = ps["weight"], cs["weight"]
        pr, cr = ps["inv_rank"], cs["inv_rank"]
        if abs(cw - pw) < 0.001: action, note = "HOLD", "No change"
        elif cw > pw:             action, note = "ADD",  f"Weight {pw:.1%} -> {cw:.1%}"
        else:                     action, note = "TRIM", f"Weight {pw:.1%} -> {cw:.1%}"
        try:
            rd = int(pr) - int(cr)
            if abs(rd) >= 3:
                note += f"  |  Rank: {pr} -> {cr} ({'+' if rd>0 else ''}{rd})"
        except (ValueError, TypeError):
            pass
        changes.append({"action": action, "ticker": t, "etf_name": cs["etf_name"],
                        "sector": cs["sector"], "prev_wt": pw, "curr_wt": cw,
                        "prev_rk": pr, "curr_rk": cr, "note": note})

    prev_r, curr_r = prev.get("regime", ""), curr.get("regime", "")
    if prev_r != curr_r:
        changes.insert(0, {"action": "REGIME", "ticker": "-",
                           "etf_name": f"Regime: {prev_r} -> {curr_r}", "sector": "-",
                           "prev_wt": 0.0, "curr_wt": 0.0, "prev_rk": "-", "curr_rk": "-",
                           "note": f"Slots: {prev.get('active_slots','?')} -> {curr.get('active_slots','?')}"})

    order = {"REGIME": 0, "SELL": 1, "BUY": 2, "ADD": 3, "TRIM": 4, "HOLD": 5}
    changes.sort(key=lambda x: order.get(x["action"], 9))
    return changes


def update_log(script_dir: Path, allocation: pd.DataFrame,
               regime: dict, prices: pd.DataFrame | None = None) -> tuple[dict, list[dict], dict]:
    log       = load_holdings_log(script_dir)
    month_key = datetime.today().strftime("%Y-%m")
    run_date  = datetime.today().strftime("%Y-%m-%d %H:%M")
    curr_entry = record_to_log(allocation, regime, run_date)

    sorted_keys = sorted(log.keys())
    prev_keys   = [k for k in sorted_keys if k < month_key]
    prev_entry  = log[prev_keys[-1]] if prev_keys else None

    # Enrich with entry_price / peak for TSL
    if prices is not None:
        prev_alloc = ({s["ticker"]: s for s in prev_entry.get("allocation", [])}
                      if prev_entry else {})
        for slot in curr_entry["allocation"]:
            t = slot["ticker"]
            if t == "CASH":
                slot["entry_price"] = None
                slot["peak"]        = None
                continue
            current_nav = None
            if t in prices.columns:
                s = prices[t].dropna()
                if len(s) > 0:
                    current_nav = float(s.iloc[-1])
            prev_slot = prev_alloc.get(t)
            if (prev_slot and prev_slot.get("entry_price") and prev_slot["ticker"] != "CASH"):
                slot["entry_price"] = prev_slot["entry_price"]
                old_peak = prev_slot.get("peak") or 0
                slot["peak"] = max(old_peak, current_nav) if current_nav else old_peak
            else:
                slot["entry_price"] = current_nav
                slot["peak"]        = current_nav

    changes = diff_allocations(prev_entry, curr_entry) if prev_entry else []
    log[month_key] = curr_entry
    save_holdings_log(script_dir, log)
    return prev_entry, changes, log


# =========================================================
# 8. DAILY TSL CHECK — uses mfapi.in (no yfinance .NS)
# =========================================================
def fetch_latest_nav_amfi(amfi_code: str) -> float | None:
    """
    Fetch the latest available NAV for an AMFI scheme code from mfapi.in.
    Returns None on failure.
    """
    try:
        url  = f"https://api.mfapi.in/mf/{amfi_code}"
        resp = requests.get(url, timeout=12)
        resp.raise_for_status()
        data = resp.json().get("data", [])
        if data:
            return float(data[0]["nav"])   # data[0] = most recent
    except Exception:
        pass
    return None


def check_tsl(script_dir: Path):
    """
    Daily Trailing Stop Loss check.
    Fetches latest NAV from mfapi.in for held ETFs (by AMFI code),
    compares against stored peaks, and flags any TSL breaches.
    Advisory only — does NOT auto-sell positions.
    """
    log       = load_holdings_log(script_dir)
    month_key = datetime.today().strftime("%Y-%m")

    if month_key not in log:
        print("[tsl] No allocation found for current month.")
        print("      Run monthly rebalance first: python etf_momentum_ranking_amfi.py")
        return

    entry      = log[month_key]
    allocation = entry.get("allocation", [])
    held       = [s for s in allocation if s["ticker"] != "CASH"]

    if not held:
        print("[tsl] No active positions (all cash). Nothing to check.")
        return

    print(f"[tsl] Fetching NAVs for {len(held)} position(s) via mfapi.in ...")
    live_navs: dict[str, float] = {}
    for s in held:
        code = s["ticker"]
        nav  = fetch_latest_nav_amfi(code)
        if nav is not None:
            live_navs[code] = nav
            print(f"       [{code}] {s['etf_name'][:40]:<40}  NAV: {nav:.4f}")
        else:
            print(f"       [{code}] {s['etf_name'][:40]:<40}  NAV: FETCH FAILED")
        time.sleep(0.2)

    # TSL Dashboard
    threshold = CONFIG.TSL_THRESHOLD
    W = 92
    print("\n" + "=" * W)
    print(f"  TSL CHECK (AMFI)  |  {datetime.today().strftime('%Y-%m-%d %H:%M')}"
          f"  |  Threshold: {threshold*100:.0f}%")
    print("=" * W)
    print(f"  {'Slot':>4}  {'AMFI Code':>10}  {'Entry':>8} {'Peak':>8}"
          f" {'TSL NAV':>8} {'NAV':>8} {'DD%':>7}  Status")
    print("  " + "-" * (W - 4))

    breaches = []
    for s in allocation:
        t        = s["ticker"]
        slot_num = s["slot"]

        if t == "CASH":
            print(f"  {slot_num:>4}  {'CASH':>10}  "
                  f"{'--':>8} {'--':>8} {'--':>8} {'--':>8} {'--':>7}  --")
            continue

        entry_price = s.get("entry_price")
        peak        = s.get("peak")
        nav         = live_navs.get(t)

        if nav is None:
            ep_s = f"{entry_price:.4f}" if entry_price else "--"
            pk_s = f"{peak:.4f}" if peak else "--"
            print(f"  {slot_num:>4}  {t:>10}  "
                  f"{ep_s:>8} {pk_s:>8} {'--':>8} {'ERR':>8} {'--':>7}  FETCH FAILED")
            continue

        if peak is None or nav > peak:
            peak = nav
        s["peak"] = peak

        tsl_nav = peak * (1 - threshold)
        dd      = (peak - nav) / peak if peak > 0 else 0.0
        status  = "!! TSL BREACH !!" if dd >= threshold else "OK"
        if dd >= threshold:
            breaches.append((t, s.get("etf_name", ""), dd, tsl_nav))

        ep_s = f"{entry_price:.4f}" if entry_price else "--"
        print(f"  {slot_num:>4}  {t:>10}  "
              f"{ep_s:>8} {peak:>8.4f} {tsl_nav:>8.4f} {nav:>8.4f} {dd*100:>6.1f}%  {status}")

    print("=" * W)

    if breaches:
        print("\n  !! ACTION REQUIRED !!")
        for code, name, dd, tsl_price in breaches:
            print(f"  -> SELL [{code}] {name}"
                  f" -- drawdown {dd*100:.1f}% exceeds {threshold*100:.0f}% TSL"
                  f" (trigger: {tsl_price:.4f})")
        print(f"  -> Move proceeds to cash until next monthly rebalance.\n")
    else:
        print(f"\n  All positions within TSL threshold. No action needed.\n")

    log[month_key] = entry
    save_holdings_log(script_dir, log)
    print(f"[tsl] Updated peaks saved to {CONFIG.HOLDINGS_FILE}")


def run_tsl_check():
    SCRIPT_DIR = Path(__file__).resolve().parent
    check_tsl(SCRIPT_DIR)


# =========================================================
# 9. EXCEL OUTPUT (identical styling, AMFI code as ticker)
# =========================================================
def _brd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _h(ws, row, col, val, bg="1F4E79", fg="FFFFFF", sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, size=sz, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _brd()
    return c

def _d(ws, row, col, val, bg="FFFFFF", fmt=None, bold=False, fg="000000"):
    if isinstance(val, float) and np.isnan(val):
        val = None
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _brd()
    if fmt:
        c.number_format = fmt
    return c


def _write_rebalance_sheet(wb, prev_entry, changes, log, NAVY, GREEN, DKGREEN, ORANGE, YELLOW, GREY):
    ACTION_COLORS = {"BUY": "C6EFCE", "SELL": "FFC7CE", "ADD": "DAEEF3",
                     "TRIM": "FFEB9C", "HOLD": "F2F2F2", "REGIME": "D9D9D9"}
    if "Rebalance" in [s.title for s in wb.worksheets]:
        del wb["Rebalance"]
    wb.create_sheet("Rebalance", 1)
    wr  = wb["Rebalance"]
    row = 1

    def title_row(ws, r, text, cols, bg=NAVY):
        ws.merge_cells(f"A{r}:{get_column_letter(cols)}{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20

    def hdr_row(ws, r, hdrs, widths, bg=NAVY):
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _brd()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[r].height = 28

    def data_row(ws, r, vals, fmts, bg="F2F2F2", bold=False):
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            if isinstance(v, float) and np.isnan(v): v = None
            c = ws.cell(row=r, column=ci, value=v)
            c.font      = Font(name="Arial", size=9, bold=bold)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _brd()
            if f and f != "@": c.number_format = f
        ws.row_dimensions[r].height = 15

    # Section 1 — Current Allocation
    title_row(wr, row, "SECTION 1 - CURRENT ALLOCATION", 7)
    row += 1
    hdr_row(wr, row, ["Slot", "Inv Rank", "AMFI Code", "ETF Name", "Sector", "Weight", "Action"],
            [7, 9, 12, 45, 18, 9, 9])
    row += 1
    change_map  = {c["ticker"]: c["action"] for c in (changes or [])}
    curr_month  = datetime.today().strftime("%Y-%m")
    curr_entry  = log.get(curr_month, {})
    for sl in curr_entry.get("allocation", []):
        t      = sl["ticker"]
        action = change_map.get(t, "HOLD") if t != "CASH" else "CASH"
        bg     = ACTION_COLORS.get(action, GREY)
        data_row(wr, row,
                 [sl["slot"], sl["inv_rank"], t, sl["etf_name"], sl["sector"], sl["weight"], action],
                 ["0", "@", "@", "@", "@", "0%", "@"],
                 bg=bg, bold=(t != "CASH"))
        row += 1
    row += 1

    # Section 2 — Changes
    prev_label = prev_entry.get("run_date", "N/A")[:7] if prev_entry else "N/A"
    title_row(wr, row, f"SECTION 2 - CHANGES vs PREVIOUS ({prev_label})", 7,
              bg="375623" if prev_entry else "7F6000")
    row += 1
    if not changes:
        wr.merge_cells(f"A{row}:G{row}")
        c = wr.cell(row=row, column=1, value="No previous month data - first recorded rebalance.")
        c.font      = Font(name="Arial", italic=True, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
    else:
        hdr_row(wr, row, ["Action", "AMFI Code", "ETF Name", "Sector",
                           "Prev Wt", "Curr Wt", "Note"],
                [9, 12, 45, 18, 9, 9, 50])
        row += 1
        for ch in changes:
            bg   = ACTION_COLORS.get(ch["action"], GREY)
            bold = ch["action"] in ("BUY", "SELL", "REGIME")
            data_row(wr, row,
                     [ch["action"], ch["ticker"], ch["etf_name"], ch["sector"],
                      ch["prev_wt"] or None, ch["curr_wt"] or None, ch["note"]],
                     ["@", "@", "@", "@", "0%", "0%", "@"],
                     bg=bg, bold=bold)
            row += 1
    row += 1

    # Section 3 — 12-Month History
    title_row(wr, row, f"SECTION 3 - LAST {HISTORY_MONTHS} MONTHS HISTORY", 7, bg="203864")
    row += 1
    sorted_months = sorted(log.keys())[-int(HISTORY_MONTHS):]
    all_codes: list[str] = []
    seen: set[str] = set()
    for mk in reversed(sorted_months):
        for sl in log[mk].get("allocation", []):
            t = sl["ticker"]
            if t != "CASH" and t not in seen:
                all_codes.append(t)
                seen.add(t)

    hdr_row(wr, row, ["Month", "Regime"] + all_codes, [12, 24] + [12]*len(all_codes))
    row += 1
    for mk in sorted_months:
        entry_m  = log[mk]
        regime_l = entry_m.get("regime", "")
        held     = {s["ticker"]: s["weight"] for s in entry_m.get("allocation", [])
                    if s["ticker"] != "CASH"}
        r_bg     = ("E2EFDA" if "BULL" in regime_l else
                    "FCE4D6" if "BEAR" in regime_l else "FFF2CC")
        vals = [mk, regime_l]
        fmts = ["@", "@"]
        for t in all_codes:
            w2 = held.get(t)
            vals.append(w2)
            fmts.append("0%" if w2 is not None else "@")
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            if isinstance(v, float) and np.isnan(v): v = None
            c = wr.cell(row=row, column=ci, value=v)
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _brd()
            if ci <= 2:
                c.fill = PatternFill("solid", fgColor=r_bg)
                c.font = Font(name="Arial", size=9, bold=(ci==1))
            elif v is not None:
                c.fill = PatternFill("solid", fgColor=DKGREEN)
                c.font = Font(name="Arial", size=9, bold=True)
            else:
                c.fill = PatternFill("solid", fgColor="F2F2F2")
            if f and f != "@": c.number_format = f
        wr.row_dimensions[row].height = 15
        row += 1
    wr.freeze_panes = "A2"


def save_excel(df, regime, allocation, out_path,
               prev_entry=None, changes=None, log=None):
    wb      = Workbook()
    NAVY    = "1F4E79"
    GREEN   = "E2EFDA"
    DKGREEN = "C6EFCE"
    ORANGE  = "FCE4D6"
    YELLOW  = "FFF2CC"
    GREY    = "F2F2F2"
    BULL_C  = "375623"
    BEAR_C  = "C00000"
    PART_C  = "7F6000"

    regime_color = (BULL_C if regime["label"] == "BULL" else
                    BEAR_C if "BEAR" in regime["label"] else PART_C)

    # Sheet 1: Rankings
    ws = wb.active
    ws.title = "Rankings"
    ws.merge_cells("A1:V1")
    c = ws["A1"]
    c.value     = ("ETF Momentum Ranking (AMFI)  |  "
                   "Step 1: Screen  ->  Step 2: Weighted Sharpe  ->  "
                   "Step 3: Regime  ->  Step 4: Allocate")
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:V2")
    r = regime
    rtext = (f"REGIME: {r['label']}  |  Active slots: {r['active_slots']}/{CONFIG.TOP_N}  |  "
             f"Index ({r['trend_ticker']}): Price {r['nifty_price']:.2f}  |  "
             f"EMA50: {r['nifty_ema_50']:.2f}  |  EMA100: {r['nifty_ema_100']:.2f}")
    rc = ws["A2"]
    rc.value     = rtext
    rc.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    rc.fill      = PatternFill("solid", fgColor=regime_color)
    rc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    COLS = [
        ("Investable\nRank",   10, "0"),
        ("Universe\nRank",     10, "0"),
        ("Sharpe\nRank",        9, "0"),
        ("SR2\nRank",           9, "0"),
        ("AMFI\nCode",         12, "@"),
        ("ETF Name",           48, "@"),
        ("Sector",             18, "@"),
        ("Close\nNAV",         10, "0.0000"),
        ("52Wk\nHigh",         10, "0.0000"),
        ("% From\n52Wk High",  12, "0.00"),
        ("EMA 100",            10, "0.0000"),
        ("% From\nEOM",        11, "0.00"),
        ("Wtd Sharpe\nScore",  14, "0.000"),
        ("Sharpe\n6M",         13, "0.000"),
        ("Sharpe\n3M",         13, "0.000"),
        ("R2\n6M",             10, "0.000"),
        ("R2\n3M",             10, "0.000"),
        ("SR2\n6M",            12, "0.000"),
        ("SR2\n3M",            12, "0.000"),
        ("SR2\nBlend",         12, "0.000"),
        ("Screen\nResult",     11, "@"),
    ]
    KEYS = [
        "RANK_INVESTABLE", "RANK_UNIVERSE", "RANK_SHARPE", "RANK_SR2",
        "TICKER", "ETF_NAME", "SECTOR", "CLOSE", "52WK_HIGH", "PCT_FROM_HIGH",
        "EMA_100", "EOM_PCT",
        "WTD_SHARPE", "SHARPE_6M", "SHARPE_3M",
        "R2_6M", "R2_3M", "SR2_6M", "SR2_3M", "SR2_BLEND",
        "SCREEN_PASS",
    ]

    HDR = 3
    for ci, (hdr, width, _) in enumerate(COLS, 1):
        _h(ws, HDR, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HDR].height = 32

    for ri, (_, row) in enumerate(df.iterrows(), start=HDR + 1):
        passed   = row["SCREEN_PASS"]
        inv_rk   = row["RANK_INVESTABLE"]
        in_alloc = passed and (inv_rk > 0) and (inv_rk <= regime["active_slots"])
        bg = (DKGREEN if in_alloc else GREEN if passed else ORANGE)

        close_above_ema = (pd.notna(row["CLOSE"]) and pd.notna(row["EMA_100"])
                           and row["CLOSE"] > row["EMA_100"])

        for ci, (key, (_, _, fmt)) in enumerate(zip(KEYS, COLS), 1):
            val = row[key]
            if key == "SCREEN_PASS":
                val = "PASS" if val else "FAIL"
            elif key == "RANK_INVESTABLE" and (not passed or val == 0):
                val = "-"
            cell_fg = "1A5C2E" if (key == "CLOSE" and close_above_ema) else "000000"
            _d(ws, ri, ci, val, bg=bg,
               fmt=fmt if fmt != "@" else None, bold=in_alloc, fg=cell_fg)
        ws.row_dimensions[ri].height = 14

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{HDR + len(df)}"

    # Sheet 2: Allocation
    wa = wb.create_sheet("Allocation")
    wa.merge_cells("A1:G1")
    c = wa["A1"]
    c.value     = (f"Top-{CONFIG.TOP_N} Allocation (AMFI)  |  "
                   f"Screen: <=25% from 52wk high  |  "
                   f"Regime={regime['label']}  |  Active slots={regime['active_slots']}")
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=regime_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    wa.row_dimensions[1].height = 22

    for ci, (hdr, w) in enumerate(
            zip(["Slot", "Inv Rank", "Sector", "AMFI Code", "ETF Name", "Weight", "Detail"],
                [8, 10, 18, 12, 48, 10, 60]), 1):
        _h(wa, 2, ci, hdr)
        wa.column_dimensions[get_column_letter(ci)].width = w
    wa.row_dimensions[2].height = 22

    for ri, (_, row) in enumerate(allocation.iterrows(), start=3):
        is_cash = row["TICKER"] == "CASH"
        is_buf  = is_cash and "buffer" in str(row["REASON"])
        bg = YELLOW if is_buf else (ORANGE if is_cash else DKGREEN)
        for ci, (v, f) in enumerate(
                zip([row["SLOT"], row["INV_RANK"], row.get("SECTOR", ""),
                     row["TICKER"], row["ETF_NAME"], row["WEIGHT"], row["REASON"]],
                    ["0", "@", "@", "@", "@", "0.0%", "@"]), 1):
            _d(wa, ri, ci, v, bg=bg, fmt=f if f != "@" else None, bold=True)
        wa.row_dimensions[ri].height = 18

    # Sheet 3: Regime Detail
    wr = wb.create_sheet("Regime")
    wr.column_dimensions["A"].width = 45
    wr.column_dimensions["B"].width = 25
    _h(wr, 1, 1, "Regime Parameter", bg=NAVY)
    _h(wr, 1, 2, "Value / Status",   bg=NAVY)
    r = regime
    regime_rows = [
        ("Regime Label",                             r["label"]),
        ("Active slots",                             f"{r['active_slots']} of {CONFIG.TOP_N}"),
        ("--- LOGIC ---",                            ""),
        ("BULL  (EMA50>EMA100 & Price>EMA50) -> slots", str(CONFIG.TOP_N)),
        ("PARTIAL (Price>EMA100, not BULL) -> slots",   str(CONFIG.TOP_N_PARTIAL)),
        ("BEAR  (Price<=EMA100) -> slots",           "0  (full cash)"),
        ("--- PARAMETERS ---",                       ""),
        ("Index used (yfinance)",                    r["trend_ticker"]),
        ("Current price",                            f"{r['nifty_price']:.2f}"),
        (f"{CONFIG.TREND_FAST_EMA}-day EMA",         f"{r['nifty_ema_50']:.2f}"),
        (f"{CONFIG.TREND_SLOW_EMA}-day EMA",         f"{r['nifty_ema_100']:.2f}"),
        ("Price above EMA50?",                       str(r["nifty_price"] > r["nifty_ema_50"]
                                                         if pd.notna(r["nifty_price"]) else "N/A")),
        ("EMA50 > EMA100?",                          str(r["nifty_ema_50"] > r["nifty_ema_100"]
                                                         if pd.notna(r["nifty_ema_50"]) else "N/A")),
    ]
    for ri2, (lbl, val) in enumerate(regime_rows, start=2):
        is_sec = lbl.startswith("---")
        ok_bg  = (NAVY if is_sec else
                  GREEN  if "True"  in str(val) or "BULL" in str(val) else
                  ORANGE if "False" in str(val) or "BEAR" in str(val) else GREY)
        c1 = _d(wr, ri2, 1, lbl, bg=ok_bg if not is_sec else NAVY)
        c2 = _d(wr, ri2, 2, val, bg=ok_bg if not is_sec else NAVY)
        if is_sec:
            c1.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c2.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")

    # Sheet 4: Rebalance Tracker
    if changes is not None:
        _write_rebalance_sheet(wb, prev_entry, changes, log,
                               NAVY, GREEN, DKGREEN, ORANGE, YELLOW, GREY)

    wb.save(out_path)
    print(f"\n[saved] -> {Path(out_path).resolve()}")


# =========================================================
# 10. MAIN PIPELINE
# =========================================================
def run_pipeline(nav_file=None, out=None):
    SCRIPT_DIR = Path(__file__).resolve().parent
    if nav_file is None:
        nav_file = str(SCRIPT_DIR / CONFIG.AMFI_NAV_FILE)
    codes_file = str(SCRIPT_DIR / CONFIG.AMFI_CODES_FILE)
    if out is None:
        out = str(SCRIPT_DIR / CONFIG.OUTPUT_FILE)

    meta, prices = load_etf_data_amfi(nav_file, codes_file)

    print("[regime] Fetching Nifty 500 regime from yfinance ...")
    regime = regime_status()
    print(f"         {regime['label']}  |  "
          f"Price={regime['nifty_price']:.2f}  |  "
          f"EMA50={regime['nifty_ema_50']:.2f}  |  "
          f"EMA100={regime['nifty_ema_100']:.2f}")

    print("[scores] Screening + scoring all ETFs ...")
    ranking    = build_ranking(meta, prices)

    print("[alloc]  Building allocation ...")
    allocation = build_allocation(ranking, regime)

    print_summary(ranking, regime, allocation)

    print("[log]    Updating holdings log ...")
    prev_entry, changes, log = update_log(SCRIPT_DIR, allocation, regime, prices)

    if prev_entry:
        prev_month = prev_entry.get("run_date", "?")[:7]
        print(f"         Previous month: {prev_month}  |  Changes: {len(changes)}")
        for ch in changes:
            arrow = {"BUY": "+ BUY", "SELL": "- SELL", "ADD": "^ ADD",
                     "TRIM": "v TRIM", "HOLD": "= HOLD", "REGIME": "! REGIME"}.get(ch["action"], "  ")
            print(f"           {arrow:8s} [{ch['ticker']:<8}]  {ch['note']}")
    else:
        print("         First run - no previous holdings to compare.")

    save_excel(ranking, regime, allocation, out,
               prev_entry=prev_entry, changes=changes, log=log)


if __name__ == "__main__":
    if "--tsl" in sys.argv:
        run_tsl_check()
    else:
        nav_arg = sys.argv[1] if len(sys.argv) > 1 else None
        out_arg = sys.argv[2] if len(sys.argv) > 2 else None
        run_pipeline(nav_arg, out_arg)
