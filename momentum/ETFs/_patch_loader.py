"""
Patch: Replace load_etf_data in etf_momentum_ranking.py.
The ETF.xlsx date headers are a dynamic Excel array formula based on TODAY()
that cannot be read by openpyxl (no cached values).
Fix: generate business days in Python to match the formula output.
"""
import re

FILE = "etf_momentum_ranking.py"
content = open(FILE, encoding="utf-8").read()

NEW_FUNC = r'''# =========================================================
# 1. DATA LOADING
# =========================================================
def load_etf_data(filepath: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load ETF data from ETF.xlsx.

    The date headers (row 1, col C onwards) are an Excel dynamic array formula:
        =TRANSPOSE(LET(d, SEQUENCE(365,...TODAY()...), FILTER(d, WEEKDAY(d,2)<6)))
    This formula is TODAY()-based and has no cached values readable by openpyxl.

    Fix: read only the price grid via openpyxl (formula-safe), and reconstruct
    the date index in Python by generating the last N business days (Mon-Fri)
    to match the column count exactly.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    ws = wb["DATA"]
    max_col = ws.max_column
    max_row = ws.max_row

    # Price data starts at Excel column 3 (C), 0-based index 2
    PRICE_START_COL = 3   # Excel 1-indexed
    n_price_cols = max_col - PRICE_START_COL + 1  # number of date columns

    # Reconstruct date index: last N business days (Mon-Fri) ending today
    # This mirrors the Excel formula which generates ~261 trading days per year
    today = pd.Timestamp.today().normalize()
    # Generate enough biz days; filter to exact count needed
    candidate_dates = pd.bdate_range(end=today, periods=n_price_cols)
    dates = list(candidate_dates)  # ascending order, length = n_price_cols

    # Read ETF names, tickers and price rows
    rows = list(ws.iter_rows(min_row=2, max_row=max_row, values_only=True))
    wb.close()

    etf_names, tickers, price_rows = [], [], []
    for row in rows:
        name   = str(row[0]).strip() if row[0] is not None else ""
        ticker = str(row[1]).strip() if row[1] is not None else ""
        if not ticker or ticker == "None":
            continue
        etf_names.append(name)
        tickers.append(ticker)
        # cols 2 onwards (0-based) are price columns
        price_rows.append(row[2: 2 + n_price_cols])

    meta = pd.DataFrame({"ETF_NAME": etf_names, "TICKER": tickers})

    price_raw = pd.DataFrame(price_rows, index=tickers, columns=dates)
    price_raw = price_raw.apply(pd.to_numeric, errors="coerce").replace(0, np.nan)

    prices = price_raw.T.sort_index().ffill()
    print(f"[load]   {filepath}")
    print(f"         {len(tickers)} ETFs  |  {len(dates)} date cols  "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})")
    return meta.reset_index(drop=True), prices
'''

# Locate and replace the entire function block
# Find the DATA LOADING section header
start = content.find("# 1. DATA LOADING")
if start < 0:
    print("ERROR: Could not find '# 1. DATA LOADING' in file")
    exit(1)

# Walk back to find the section comment block start (the === line)
block_start = content.rfind("# =", 0, start)

# Find the end of the function (next triple-blank-line followed by # = section)
rest = content[block_start:]
# Find next section separator after the function
end_match = re.search(r'\n\n\n# =', rest)
if not end_match:
    print("ERROR: Could not find end of function block")
    exit(1)

block_end = block_start + end_match.start()

content = content[:block_start] + NEW_FUNC + "\n\n\n" + content[block_end + 3:]
open(FILE, "w", encoding="utf-8").write(content)
print("SUCCESS: load_etf_data replaced with business-day date generator")
print(f"  Replaced chars {block_start}-{block_end}")
