"""
ETF Dual Momentum + Clenow + Weighted Sharpe + Tiered Regime Filter
====================================================================
Scoring pipeline (in correct order):
  Step 1 - SCREEN  : Apply 52-week high proximity filter.
                     Only ETFs within MAX_DRAWDOWN_FROM_HIGH of their 52wk high are investable.
  Step 2 - SCORE   : Compute Clenow (6M+3M), Weighted Sharpe, Composite
                     on ALL 224 ETFs for reference.
                     Investable rank is computed on the screened subset only.
  Step 3 - REGIME  : Two-layer tiered filter determines how many slots to fill.
  Step 4 - ALLOCATE: Select Top-N from the investable (screened) ranked list.

Regime Filter - Tiered (three states):
  BULL   (both pass) : all TOP_N slots active
  PARTIAL (one fails): TOP_N_PARTIAL slots active, rest = cash
  BEAR   (Price <= EMA100): full cash

  Layer 1 - Trend  : MONIFTY500 above its 100-day EMA

  Rebalance: Weekly (Monday / first trading day).
  Exit logic: Hold positions unless 52wk-high DD > 25%, rank > 20, or TSL > 5%.
  Layer 2 - EMA50  : Price above 50-day EMA (for BULL)

All parameters in CONFIG below.
"""

from __future__ import annotations
import sys
import json
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import yfinance as yf
    _YF_AVAILABLE = True
except ImportError:
    _YF_AVAILABLE = False


# =========================================================
# CONFIG  <- loaded from strategy_config.json if available,
#            otherwise uses hardcoded defaults below.
# =========================================================
_SCRIPT_DIR = Path(__file__).resolve().parent
_CONFIG_FILE = "strategy_config.json"


def _config_path(script_dir: Path | None = None) -> Path:
    return (script_dir or _SCRIPT_DIR) / _CONFIG_FILE


def load_config_from_json(script_dir: Path | None = None) -> dict:
    """Load strategy_config.json. Returns empty dict if file missing."""
    p = _config_path(script_dir)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config_to_json(config_dict: dict, script_dir: Path | None = None):
    """Write config values to strategy_config.json."""
    p = _config_path(script_dir)
    with open(p, "w") as f:
        json.dump(config_dict, f, indent=2, default=str)


def _apply_json_config(cls, cfg: dict):
    """Override CONFIG class attributes from a JSON dict."""
    _KEYS = [
        "INPUT_FILE", "OUTPUT_FILE",
        "WINDOW_6M", "WINDOW_3M", "ANNUALIZE",
        "TOP_N", "TOP_N_PARTIAL",
        "MAX_DRAWDOWN_FROM_HIGH",
        "SHARPE_W6M", "SHARPE_W3M",
        "REGIME_TICKER", "REGIME_FALLBACKS", "REGIME_INDEX_TICKER",
        "TREND_FAST_EMA_WINDOW", "TREND_EMA_WINDOW",
        "SECTOR_CAP",
        "TSL_THRESHOLD",
        "EXIT_MAX_DD_FROM_HIGH", "EXIT_MAX_RANK",
        "HISTORY_PERIODS",
    ]
    for key in _KEYS:
        if key in cfg:
            setattr(cls, key, cfg[key])
    # DAILY_RF is derived from the annual rate
    if "DAILY_RF_ANNUAL" in cfg:
        cls.DAILY_RF = cfg["DAILY_RF_ANNUAL"] / cls.ANNUALIZE
    return cls


def get_config_as_dict() -> dict:
    """Return current CONFIG values as a JSON-serialisable dict."""
    return {
        "INPUT_FILE": CONFIG.INPUT_FILE,
        "OUTPUT_FILE": CONFIG.OUTPUT_FILE,
        "WINDOW_6M": CONFIG.WINDOW_6M,
        "WINDOW_3M": CONFIG.WINDOW_3M,
        "ANNUALIZE": CONFIG.ANNUALIZE,
        "TOP_N": CONFIG.TOP_N,
        "TOP_N_PARTIAL": CONFIG.TOP_N_PARTIAL,
        "MAX_DRAWDOWN_FROM_HIGH": CONFIG.MAX_DRAWDOWN_FROM_HIGH,
        "SHARPE_W6M": CONFIG.SHARPE_W6M,
        "SHARPE_W3M": CONFIG.SHARPE_W3M,
        "REGIME_TICKER": CONFIG.REGIME_TICKER,
        "REGIME_FALLBACKS": CONFIG.REGIME_FALLBACKS,
        "REGIME_INDEX_TICKER": CONFIG.REGIME_INDEX_TICKER,
        "TREND_FAST_EMA_WINDOW": CONFIG.TREND_FAST_EMA_WINDOW,
        "TREND_EMA_WINDOW": CONFIG.TREND_EMA_WINDOW,
        "SECTOR_CAP": CONFIG.SECTOR_CAP,
        "DAILY_RF_ANNUAL": CONFIG.DAILY_RF * CONFIG.ANNUALIZE,
        "TSL_THRESHOLD": CONFIG.TSL_THRESHOLD,
        "EXIT_MAX_DD_FROM_HIGH": CONFIG.EXIT_MAX_DD_FROM_HIGH,
        "EXIT_MAX_RANK": CONFIG.EXIT_MAX_RANK,
        "HISTORY_PERIODS": CONFIG.HISTORY_PERIODS,
    }


class CONFIG:
    INPUT_FILE  = "ETF.xlsx"
    OUTPUT_FILE = "etf_rankings.xlsx"

    # Momentum windows (trading days)
    WINDOW_6M   = 126
    WINDOW_3M   = 63
    ANNUALIZE   = 252

    # Portfolio allocation
    TOP_N         = 5    # slots when regime = BULL (both layers pass)
    TOP_N_PARTIAL = 3    # slots when regime = PARTIAL (one layer fails)
                         # remaining slots go to cash as a buffer

    # 52-week high proximity filter — the sole screen applied BEFORE ranking
    # ETF must be trading within MAX_DRAWDOWN_FROM_HIGH of its 52-week high.
    # Removes deep-drawdown ETFs that are bouncing off a bottom rather than
    # exhibiting genuine momentum. 0.25 = must be >= 75% of 52wk high.
    MAX_DRAWDOWN_FROM_HIGH = 0.25

    # Weighted Sharpe blending (6M trend + 3M recency tilt)
    # Composite ranking is based entirely on Weighted Sharpe
    SHARPE_W6M  = 0.5
    SHARPE_W3M  = 0.5

    # Regime filter
    # Nifty 500 used (not Nifty 50) — broader coverage matches full ETF universe
    # (large + mid + small cap); mid/small roll over before large caps in India
    # Primary source: live Nifty 500 INDEX via yfinance (not an ETF proxy).
    # Falls back to the ETF.xlsx-based tickers below if the live fetch fails.
    REGIME_INDEX_TICKER = "^CRSLDX"   # Yahoo Finance symbol for Nifty 500 Index
    REGIME_TICKER      = "MONIFTY500"
    REGIME_FALLBACKS   = ["BSE500IETF", "HDFCBSE500", "NIFTYBEES"]
    TREND_FAST_EMA_WINDOW = 50     # Layer 1: fast EMA
    TREND_EMA_WINDOW   = 100       # Layer 1: slow EMA

    # Sector cap — max ETFs per sector in final allocation
    # Prevents concentration in duplicates tracking the same index
    SECTOR_CAP = 1

    # Daily risk-free rate for Sharpe (7% annual / 252)
    DAILY_RF = 0.07 / 252

    # Trailing Stop Loss — daily monitoring via --tsl flag
    TSL_THRESHOLD = 0.05   # 5% drawdown from peak triggers alert

    # Weekly rebalance: selective exit triggers
    EXIT_MAX_DD_FROM_HIGH = 0.25   # exit if >25% from 52-week high
    EXIT_MAX_RANK         = 20     # exit if investable rank > 20
    HISTORY_PERIODS       = 52     # weeks of log history to retain in Excel


# ── Apply JSON overrides on module load ──────────────────────
_apply_json_config(CONFIG, load_config_from_json())



# =========================================================
# SECTOR CLASSIFICATION
# Auto-derived from ETF name keywords. Rules are ordered
# most-specific first — first match wins.
# =========================================================
# ── Primary: direct ticker→sector lookup from ETF_SECTOR.xlsx ──────────
_SECTOR_LOOKUP: dict[str, str] = {}
_SECTOR_FILE = _SCRIPT_DIR / "ETF_SECTOR.xlsx"

def _load_sector_lookup():
    """Load ticker→sector mapping from ETF_SECTOR.xlsx (sheet 'ETF_SECTOR')."""
    global _SECTOR_LOOKUP
    if not _SECTOR_FILE.exists():
        print(f"  [warn] {_SECTOR_FILE.name} not found; using keyword fallback only")
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_SECTOR_FILE, data_only=True)
        ws = wb["ETF_SECTOR"]
        for r in range(2, ws.max_row + 1):
            ticker = ws.cell(r, 1).value
            sector = ws.cell(r, 2).value
            if ticker and sector:
                _SECTOR_LOOKUP[str(ticker).strip().upper()] = str(sector).strip()
        wb.close()
        print(f"  [sectors] Loaded {len(_SECTOR_LOOKUP)} mappings from {_SECTOR_FILE.name}")
    except Exception as e:
        print(f"  [warn] Could not load {_SECTOR_FILE.name}: {e}")

_load_sector_lookup()   # run once at module import


# ── Fallback: keyword-based rules for tickers not in ETF_SECTOR.xlsx ───
_SECTOR_RULES = [
    ("PSU_BANK",         ["psu bank","psubnk","psubank","bse psu bank"]),
    ("PRIVATE_BANK",     ["private bank","pvt bank","pvtban","nifty pb "]),
    ("BANKING_BROAD",    ["nifty bank","bse bank"," bank ","banketf","bankbees","banknifty","nifban"]),
    ("IT_TECH",          ["nifty it","bse it"," it etf","itbees","itietf","nifit","tech etf"]),
    ("HEALTHCARE",       ["healthcare","pharma","health "," hc "," hc\\"]),
    ("METAL",            ["metal"]),
    ("ENERGY",           ["energy","oil & gas","o&g","oilietf","power etf","bse power"]),
    ("INFRA",            ["infra"]),
    ("CONSUMPTION",      ["consumption","consump","fmcg","consumer"]),
    ("REALTY",           ["realty","real estate"]),
    ("DEFENCE",          ["defence","dfnc"]),
    ("PSE",              ["pse etf","cpse","nifty pse","bharat 22","cpseetf"]),
    ("AUTO",             ["auto"]),
    ("CHEMICALS",        ["chemical"]),
    ("FIN_SERVICES",     ["fin serv","financial serv","finietf","bfsi","capital mkt",
                          "captl mkt","capital market","cptmkt","capital mrkts"]),
    ("COMMODITIES",      ["commodity","commo"]),
    ("MANUFACTURING",    ["manufactur","manu"]),
    ("EV_MOBILITY",      ["ev&new","ev new","nifty ev"]),
    ("DIGITAL_INTERNET", ["internet","digital"]),
    ("RAILWAY",          ["railway"]),
    ("TOURISM",          ["trsm","tourism"]),
    ("MNC",              ["mnc"]),
    ("GOLD",             ["gold"]),
    ("SILVER",           ["silver"]),
    ("GOVT_BONDS",       ["g-sec","gsec","gilt","bond etf","bharat bond","ebbetf"]),
    ("DIVIDEND",         ["dividend","div opp"]),
    ("IPO",              ["ipo"]),
    ("ESG",              ["esg"]),
    ("FACTOR_MOMENTUM",  ["momentum","mmt","mmntm"]),
    ("FACTOR_VALUE",     ["value 20","value 30","value 50","enhanced val","enhval"]),
    ("FACTOR_QUALITY",   ["quality","qlty"," ql "," ql30","qual30"]),
    ("FACTOR_LOW_VOL",   ["low vol","lowvol","lw- vol"]),
    ("FACTOR_ALPHA",     ["alpha"]),
    ("FACTOR_EQUAL_WT",  ["equal weight","eq weight","eqwt","eqlwgt","eqlwght","equal wt"]),
    ("INTERNATIONAL",    ["nasdaq","s&p 500","hang seng","hangseng","hngsng","msci","fang+"]),
    ("MIDCAP",           ["midcap","mid cap","mdsmc","midsmall"]),
    ("SMALLCAP",         ["smallcap","small cap","sml100","smcp","mosmall","sc 250"]),
    ("NEXT_50",          ["next 50","next50","juniorbees","jr bees"]),
    ("BROAD_MARKET",     ["nifty 50","nifty50","sensex","nifty 100","nifty100",
                          "nifty 200","nifty 500","nifty500","total market",
                          "total mrkt","bse 500","bse500","multicap","mltcp",
                          "lgmdcp","gth sectors","flexicap","flexi"]),
    ("SERVICES",         ["services","svcs"]),
]

def classify_sector(etf_name: str, ticker: str) -> str:
    """
    Classify ETF sector.
    Priority: 1. Direct lookup from ETF_SECTOR.xlsx  2. Keyword rules  3. 'OTHER'
    """
    # 1. Direct lookup (exact ticker match)
    t_upper = ticker.strip().upper()
    if t_upper in _SECTOR_LOOKUP:
        return _SECTOR_LOOKUP[t_upper]

    # 2. Keyword fallback
    n = etf_name.lower()
    t = ticker.lower()
    for sector, keywords in _SECTOR_RULES:
        for kw in keywords:
            if kw in n or kw in t:
                return sector
    return "OTHER"


# =========================================================
# 1. DATA LOADING
# =========================================================
def load_etf_data(filepath: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load ETF data from ETF.xlsx.

    The date headers (row 1, col C onwards) are an Excel dynamic array formula:
        =TRANSPOSE(LET(d, SEQUENCE(365,...TODAY()...), FILTER(d, WEEKDAY(d,2)<6)))
    This formula is TODAY()-based and has no cached values readable by openpyxl.

    Fix: read only the price grid via openpyxl (formula-safe), and reconstruct
    the date index in Python by generating the last N business days (Mon-Fri)
    to match the column count exactly.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    ws = wb["DATA"]
    max_col = ws.max_column
    max_row = ws.max_row

    # Price data starts at Excel column 3 (C), 0-based index 2
    PRICE_START_COL = 3   # Excel 1-indexed
    n_price_cols = max_col - PRICE_START_COL + 1  # number of date columns

    # Reconstruct date index: last N business days (Mon-Fri) ending today
    # This mirrors the Excel formula which generates ~261 trading days per year
    today = pd.Timestamp.today().normalize()
    # Generate enough biz days; filter to exact count needed
    candidate_dates = pd.bdate_range(end=today, periods=n_price_cols)
    dates = list(candidate_dates)  # ascending order, length = n_price_cols

    # Read ETF names, tickers and price rows
    rows = list(ws.iter_rows(min_row=2, max_row=max_row, values_only=True))
    wb.close()

    etf_names, tickers, price_rows = [], [], []
    for row in rows:
        name   = str(row[0]).strip() if row[0] is not None else ""
        ticker = str(row[1]).strip() if row[1] is not None else ""
        if not ticker or ticker == "None":
            continue
        etf_names.append(name)
        tickers.append(ticker)
        # cols 2 onwards (0-based) are price columns
        price_rows.append(row[2: 2 + n_price_cols])

    meta = pd.DataFrame({"ETF_NAME": etf_names, "TICKER": tickers})

    price_raw = pd.DataFrame(price_rows, index=tickers, columns=dates)
    price_raw = price_raw.apply(pd.to_numeric, errors="coerce").replace(0, np.nan)

    prices = price_raw.T.sort_index().ffill()
    print(f"[load]   {filepath}")
    print(f"         {len(tickers)} ETFs  |  {len(dates)} date cols  "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})")
    return meta.reset_index(drop=True), prices



# =========================================================
# 2. SCORING
# =========================================================
def sharpe_score(series: pd.Series, window: int, daily_rf: float | None = None) -> float:
    """Annualised Sharpe over lookback window, excess of daily RF.
    daily_rf defaults to CONFIG.DAILY_RF; pass an explicit value (e.g. 0.0)
    to override for other risk-free-rate assumptions."""
    clean = series.dropna()
    if len(clean) < window + 1:
        return np.nan
    rf = CONFIG.DAILY_RF if daily_rf is None else daily_rf
    log_ret = np.log(clean.iloc[-window - 1:] / clean.iloc[-window - 1:].shift(1)).dropna()
    excess  = log_ret - rf
    if excess.std() == 0:
        return np.nan
    return (excess.mean() / excess.std()) * np.sqrt(CONFIG.ANNUALIZE)


def momentum_return(series: pd.Series, window: int) -> float:
    """Total % return over lookback window"""
    clean = series.dropna()
    if len(clean) < window:
        return np.nan
    return (clean.iloc[-1] / clean.iloc[-window] - 1) * 100


# =========================================================
# 3. REGIME FILTER
# =========================================================
NIFTY500_CACHE_FILE = "nifty500_cache.csv"
NIFTY500_CACHE_MAX_AGE_DAYS = 3   # tolerate weekends/holidays before refetching


def fetch_nifty500_index(n_days: int = 600, script_dir: Path | None = None) -> pd.Series | None:
    """
    Fetch the live Nifty 500 INDEX (CONFIG.REGIME_INDEX_TICKER, e.g. ^CRSLDX)
    via yfinance, with a local daily CSV cache to avoid hitting the network
    on every run.

    Cache behaviour:
      - If nifty500_cache.csv exists and was last written today (or within
        NIFTY500_CACHE_MAX_AGE_DAYS calendar days, to tolerate weekends/
        holidays), load and return it directly — no network call.
      - Otherwise (missing, stale, or unreadable), fetch fresh via yfinance,
        overwrite the cache, and return the new series.

    Returns None if yfinance isn't installed, the fetch fails, or no cache
    is usable — callers should fall back to the ETF.xlsx-based regime
    tickers in that case.
    """
    cache_path = (script_dir or _SCRIPT_DIR) / NIFTY500_CACHE_FILE

    # --- Try the cache first ---
    if cache_path.exists():
        try:
            mtime_date = datetime.fromtimestamp(cache_path.stat().st_mtime).date()
            age_days = (datetime.today().date() - mtime_date).days
            if age_days <= NIFTY500_CACHE_MAX_AGE_DAYS:
                cached = pd.read_csv(cache_path, index_col=0, parse_dates=True)["Close"]
                cached = cached.dropna()
                if len(cached) > 0:
                    print(f"  [regime] Using cached Nifty500 index ({cache_path.name}, "
                          f"{age_days}d old, {len(cached)} rows)")
                    return cached
        except Exception as e:
            print(f"  [warn] Nifty500 cache unreadable ({e}); will refetch")

    # --- Cache missing/stale/unreadable: fetch live ---
    if not _YF_AVAILABLE:
        print("  [warn] yfinance not installed; cannot fetch live Nifty500 index")
        return None

    try:
        end = datetime.today()
        start = end - timedelta(days=n_days)
        data = yf.download(CONFIG.REGIME_INDEX_TICKER,
                            start=start.strftime("%Y-%m-%d"),
                            end=end.strftime("%Y-%m-%d"),
                            progress=False, auto_adjust=True)
        if data is None or data.empty:
            print(f"  [warn] Live fetch for {CONFIG.REGIME_INDEX_TICKER} returned no data")
            return None

        close = data["Close"]
        if isinstance(close, pd.DataFrame):   # yfinance can return a 1-col DataFrame
            close = close.iloc[:, 0]
        close = close.dropna()
        if close.empty:
            print(f"  [warn] Live fetch for {CONFIG.REGIME_INDEX_TICKER} had no valid closes")
            return None

        close.to_frame(name="Close").to_csv(cache_path)
        print(f"  [regime] Fetched live Nifty500 index ({CONFIG.REGIME_INDEX_TICKER}, "
              f"{len(close)} rows) -> cached to {cache_path.name}")
        return close
    except Exception as e:
        print(f"  [warn] Live Nifty500 index fetch failed: {e}")
        return None


def regime_status(prices: pd.DataFrame, script_dir: Path | None = None) -> dict:
    """
    Returns tiered regime state:
      BULL    - both layers pass  -> invest TOP_N slots
      PARTIAL - one layer fails   -> invest TOP_N_PARTIAL slots, rest = cash
      BEAR    - both layers fail  -> full cash

    Trend source priority:
      1. Live Nifty 500 INDEX via yfinance (CONFIG.REGIME_INDEX_TICKER, e.g. ^CRSLDX)
      2. CONFIG.REGIME_TICKER (MONIFTY500) from ETF.xlsx
      3. CONFIG.REGIME_FALLBACKS, in order, from ETF.xlsx
      4. Default to BULL if nothing is available at all
    """
    trend_series = None
    trend_ticker = None

    # --- Priority 1: live Nifty 500 index ---
    live_series = fetch_nifty500_index(script_dir=script_dir)
    if live_series is not None and len(live_series) > 0:
        trend_series = live_series
        trend_ticker = "NIFTY500"

    # --- Priority 2/3: ETF.xlsx-based tickers (MONIFTY500, then fallbacks) ---
    if trend_series is None:
        xlsx_ticker = next(
            (t for t in [CONFIG.REGIME_TICKER] + CONFIG.REGIME_FALLBACKS
             if t in prices.columns), None
        )
        if xlsx_ticker is not None:
            s = prices[xlsx_ticker].dropna()
            if len(s) > 0:
                trend_series = s
                trend_ticker = xlsx_ticker

    # --- Priority 4: nothing available at all -> default BULL ---
    if trend_series is None:
        print("  [warn] No Nifty 500 source available (live or ETF.xlsx); "
              "trend layer defaulting to BULL")
        return {
            "regime_ok"   : True,
            "label"       : "BULL",
            "active_slots": CONFIG.TOP_N,
            "trend_ok"    : True,
            "nifty_price" : np.nan,
            "nifty_ema_50": np.nan,
            "nifty_ema_100": np.nan,
            "trend_ticker": "N/A",
        }

    s = trend_series
    if len(s) >= CONFIG.TREND_EMA_WINDOW:
        nifty_price = float(s.iloc[-1])
        nifty_ema_50  = float(s.ewm(span=CONFIG.TREND_FAST_EMA_WINDOW, adjust=False).mean().iloc[-1])
        nifty_ema_100 = float(s.ewm(span=CONFIG.TREND_EMA_WINDOW, adjust=False).mean().iloc[-1])

        # Regime (Run 1 - best performing config):
        #   BULL    : EMA50 > EMA100  AND  Price > EMA50  -> TOP_N slots
        #   PARTIAL : Price > EMA100  (but not BULL)       -> TOP_N_PARTIAL slots
        #   BEAR    : Price <= EMA100                       -> 0 slots, full cash
        if nifty_ema_50 > nifty_ema_100 and nifty_price > nifty_ema_50:
            label        = "BULL"
            active_slots = CONFIG.TOP_N
        elif nifty_price > nifty_ema_100:
            label        = "PARTIAL"
            active_slots = CONFIG.TOP_N_PARTIAL
        else:
            label        = "BEAR"
            active_slots = 0

        trend_ok = active_slots > 0
    else:
        trend_ok    = True
        nifty_price = float(s.iloc[-1]) if len(s) else np.nan
        nifty_ema_50 = np.nan
        nifty_ema_100 = np.nan
        label = "BULL"
        active_slots = CONFIG.TOP_N

    return {
        "regime_ok"   : active_slots == CONFIG.TOP_N,
        "label"       : label,
        "active_slots": active_slots,
        "trend_ok"    : trend_ok,
        "nifty_price" : nifty_price,
        "nifty_ema_50": nifty_ema_50,
        "nifty_ema_100": nifty_ema_100,
        "trend_ticker": trend_ticker or "N/A",
    }


# =========================================================
# 4. SCORING + RANKING
#    Abs momentum screen applied FIRST to determine investable universe,
#    then composite ranking done on that screened subset.
# =========================================================
def build_ranking(meta: pd.DataFrame, prices: pd.DataFrame) -> pd.DataFrame:
    records = []

    # Get EOM date for calculation (last day of previous month with data)
    last_date = prices.index[-1]
    prev_indices = prices.index[(prices.index.month != last_date.month) | (prices.index.year != last_date.year)]
    if len(prev_indices) > 0:
        eom_date = prev_indices[-1]
        eom_series = prices.loc[eom_date]
        print(f"         Comp date: {eom_date.date()} (prev month eom)")
    else:
        eom_series = pd.Series(dtype=float)

    for _, row in meta.iterrows():
        ticker = row["TICKER"]
        if ticker not in prices.columns:
            continue
        s = prices[ticker]
        close = float(s.iloc[-1]) if len(s) > 0 else np.nan
        # Calculate 52-week high from 1 year of historical data (252 trailing trading days)
        high_52w = float(s.tail(252).max()) if len(s) > 0 else np.nan

        sh6 = sharpe_score(s, CONFIG.WINDOW_6M)
        sh3 = sharpe_score(s, CONFIG.WINDOW_3M)
        # Raw Sharpe values stored; WTD_SHARPE computed post-loop via Z-scores

        # --- EMA calculation (100 day) ---
        clean_s = s.dropna()
        ema_100 = float(clean_s.ewm(span=100, adjust=False).mean().iloc[-1]) if len(clean_s) >= 100 else np.nan

        # --- Return from EOM ---
        eom_price = eom_series.get(ticker, np.nan)
        if pd.notna(close) and pd.notna(eom_price) and eom_price > 0:
            eom_ret = (close / eom_price - 1) * 100
        else:
            eom_ret = np.nan

        # --- Screen: 52-week high proximity filter (calculated from history) ---
        if pd.notna(close) and pd.notna(high_52w) and high_52w > 0:
            pct_from_high = (high_52w - close) / high_52w
            high_pass     = pct_from_high <= CONFIG.MAX_DRAWDOWN_FROM_HIGH
        else:
            pct_from_high = np.nan
            high_pass     = True   # no data -> don't penalise

        # --- Screen: above 100 EMA filter (calculated for reference only) ---
        ema_100 = float(clean_s.ewm(span=100, adjust=False).mean().iloc[-1]) if len(clean_s) >= 100 else np.nan

        screen_pass = high_pass

        records.append({
            "TICKER"        : ticker,
            "ETF_NAME"      : row["ETF_NAME"],
            "SECTOR"        : classify_sector(row["ETF_NAME"], ticker),
            "CLOSE"         : close,
            "52WK_HIGH"     : high_52w,
            "PCT_FROM_HIGH" : pct_from_high * 100 if not np.isnan(pct_from_high) else np.nan,
            "EMA_100"       : ema_100,
            "EOM_PCT"       : eom_ret,
            "SHARPE_6M"     : sh6,
            "SHARPE_3M"     : sh3,
            "SCREEN_PASS"   : screen_pass,
        })

    df = pd.DataFrame(records)

    # =================================================================
    # CROSS-SECTIONAL Z-SCORING OF SHARPE RATIOS
    # Z-score = (value - mean) / std, computed across ETFs with valid
    # (non-NaN) values only.  NaN ETFs are excluded from mean/std
    # calculation and remain NaN after Z-scoring.
    # =================================================================
    def _zscore(series: pd.Series) -> pd.Series:
        """Cross-sectional Z-score; NaN values stay NaN."""
        mu  = series.mean()   # pandas .mean() skips NaN by default
        sig = series.std()    # pandas .std() skips NaN by default
        if sig == 0 or np.isnan(sig):
            return pd.Series(0.0, index=series.index)
        return (series - mu) / sig

    # ── Investable Z-scores (reference pool = screen-pass ETFs) ──
    inv_mask = df["SCREEN_PASS"]
    df["_Z6_INV"] = np.nan
    df["_Z3_INV"] = np.nan
    if inv_mask.sum() > 0:
        df.loc[inv_mask, "_Z6_INV"] = _zscore(df.loc[inv_mask, "SHARPE_6M"])
        df.loc[inv_mask, "_Z3_INV"] = _zscore(df.loc[inv_mask, "SHARPE_3M"])

    # Blend Z-scored Sharpe for investable composite (fallback if only one valid)
    z6i = df["_Z6_INV"]; z3i = df["_Z3_INV"]
    both_i = z6i.notna() & z3i.notna()
    only6i = z6i.notna() & z3i.isna()
    only3i = z6i.isna()  & z3i.notna()
    df["_WTD_INV"] = np.nan
    df.loc[both_i, "_WTD_INV"] = CONFIG.SHARPE_W6M * z6i[both_i] + CONFIG.SHARPE_W3M * z3i[both_i]
    df.loc[only6i, "_WTD_INV"] = CONFIG.SHARPE_W6M * z6i[only6i]
    df.loc[only3i, "_WTD_INV"] = CONFIG.SHARPE_W3M * z3i[only3i]

    # ── Universe Z-scores (reference pool = ALL ETFs with valid Sharpe) ──
    z6u = _zscore(df["SHARPE_6M"]); z3u = _zscore(df["SHARPE_3M"])
    both_u = z6u.notna() & z3u.notna()
    only6u = z6u.notna() & z3u.isna()
    only3u = z6u.isna()  & z3u.notna()
    df["WTD_SHARPE"] = np.nan
    df.loc[both_u, "WTD_SHARPE"] = CONFIG.SHARPE_W6M * z6u[both_u] + CONFIG.SHARPE_W3M * z3u[both_u]
    df.loc[only6u, "WTD_SHARPE"] = z6u[only6u]
    df.loc[only3u, "WTD_SHARPE"] = z3u[only3u]

    # ── Rankings ─────────────────────────────────────────────────
    # Investable rank — Z-scored composite among screen-pass ETFs only.
    # ETFs that fail the 52-week-high screen get RANK_INVESTABLE = NaN
    # (blank on display), not 0 — they were never ranked at all, since
    # they're not part of the investable population being ranked.
    inv = df[df["SCREEN_PASS"]].copy()
    if len(inv) > 0:
        inv["RANK_INVESTABLE"] = inv["_WTD_INV"].rank(ascending=False, na_option="bottom").astype(int)
        df = df.merge(inv[["TICKER", "RANK_INVESTABLE"]], on="TICKER", how="left")
    else:
        df["RANK_INVESTABLE"] = np.nan

    # Drop intermediate Z-score working columns
    df = df.drop(columns=["_Z6_INV", "_Z3_INV", "_WTD_INV"], errors="ignore")

    # Sort by investable rank first (passing ETFs at top, best rank first),
    # then by raw Weighted Sharpe for the remaining (non-investable) ETFs
    df["_sort"] = df["RANK_INVESTABLE"].fillna(9999)
    df = df.sort_values(["_sort", "WTD_SHARPE"], ascending=[True, False]).drop(columns="_sort").reset_index(drop=True)

    return df


# =========================================================
# 5. PORTFOLIO ALLOCATION
# =========================================================
def should_exit(ticker: str, ranking_df: pd.DataFrame, peak: float,
                current_price: float) -> tuple[bool, str]:
    """
    Check whether a currently-held position should be exited.
    Returns (should_exit: bool, reason: str).

    Exit triggers (ANY fires → exit):
      1. >25% away from 52-week high  (EXIT_MAX_DD_FROM_HIGH)
      2. Investable rank > 20          (EXIT_MAX_RANK)
      3. Drawdown from peak > 5%       (TSL_THRESHOLD)
    """
    reasons = []
    row = ranking_df[ranking_df["TICKER"] == ticker]
    if row.empty:
        return True, "Ticker no longer in ranking universe"

    row = row.iloc[0]

    # 1. 52-week high drawdown
    pct_from_high = row.get("PCT_FROM_HIGH", 0)
    if pd.notna(pct_from_high) and abs(pct_from_high) > CONFIG.EXIT_MAX_DD_FROM_HIGH * 100:
        reasons.append(f"52wk high DD {pct_from_high:.1f}% > {CONFIG.EXIT_MAX_DD_FROM_HIGH*100:.0f}%")

    # 2. Rank degradation
    inv_rank = row.get("RANK_INVESTABLE", float("inf"))
    if pd.notna(inv_rank) and inv_rank > CONFIG.EXIT_MAX_RANK:
        reasons.append(f"Rank {int(inv_rank)} > {CONFIG.EXIT_MAX_RANK}")

    # 3. TSL breach (drawdown from stored peak)
    if peak and peak > 0 and current_price and current_price > 0:
        dd_from_peak = (peak - current_price) / peak
        if dd_from_peak >= CONFIG.TSL_THRESHOLD:
            reasons.append(f"TSL {dd_from_peak*100:.1f}% >= {CONFIG.TSL_THRESHOLD*100:.0f}%")

    if reasons:
        return True, " | ".join(reasons)
    return False, ""


def compute_holding_peak(ticker: str, first_buy_date, prices_df: pd.DataFrame,
                          current_price: float) -> float:
    """
    Highest observed price for `ticker` since `first_buy_date` (inclusive) —
    the TSL reference peak for an actual (tradelog) holding, as opposed to the
    model-slot peak tracked in holdings_log.json. Falls back to `current_price`
    when no price history is available.
    """
    peak = current_price
    if first_buy_date and ticker in prices_df.columns:
        try:
            fbd = pd.Timestamp(first_buy_date)
        except (TypeError, ValueError):
            fbd = None
        if fbd is not None:
            series = prices_df[ticker].dropna()
            series = series[series.index >= fbd]
            if not series.empty:
                peak = max(peak, float(series.max()))
    return peak


def evaluate_holdings_exit_rules(holdings_metrics: list, ranking_df: pd.DataFrame,
                                  prices_df: pd.DataFrame) -> list:
    """
    Attach exit-rule evaluation to each tradelog holding, reusing should_exit()
    (52wk-high drawdown, investable rank degradation, TSL from peak-since-entry).

    Returns a new list of dicts: each input dict plus "Peak Price", "Exit Flag",
    and "Exit Reason" ("OK" when no rule is breached).
    """
    enriched = []
    for h in holdings_metrics:
        peak = compute_holding_peak(h["Ticker"], h.get("First Buy Date"),
                                     prices_df, h["Current Price"])
        exit_flag, exit_reason = should_exit(h["Ticker"], ranking_df, peak, h["Current Price"])
        row = dict(h)
        row["Peak Price"]  = peak
        row["Exit Flag"]   = exit_flag
        row["Exit Reason"] = exit_reason if exit_flag else "OK"
        enriched.append(row)
    return enriched


def build_allocation(df: pd.DataFrame, regime: dict,
                     prev_allocation: list | None = None,
                     prices: pd.DataFrame | None = None) -> pd.DataFrame:
    """
    Weekly Hold-and-Replace (WRH) allocation.

    If prev_allocation is provided (list of slot dicts from last week):
      1. Check each held position against exit triggers (should_exit)
      2. HOLD positions that pass all checks
      3. Fill vacated + new slots from investable ranking (skip already held)

    If prev_allocation is None (first run): behaves like fresh top-N pick.

    Number of active slots determined by tiered regime state:
      BULL    -> TOP_N slots
      PARTIAL -> TOP_N_PARTIAL slots (remainder = cash buffer)
      BEAR    -> 0 slots (full cash)
    """
    active = regime["active_slots"]
    total  = CONFIG.TOP_N
    w      = 1.0 / total

    # Full cash — regime is BEAR
    if active == 0:
        return pd.DataFrame([{
            "SLOT"        : i + 1,
            "TICKER"      : "CASH",
            "ETF_NAME"    : "Cash / Money Market",
            "SECTOR"      : "CASH",
            "WEIGHT"      : w,
            "INV_RANK"    : "-",
            "REASON"      : f"Regime = {regime['label']} -> full cash"
        } for i in range(total)])

    # Investable ETFs sorted by investable rank (composite score)
    investable = df[df["SCREEN_PASS"] & (df["RANK_INVESTABLE"] > 0)].copy()
    investable = investable.sort_values("RANK_INVESTABLE").reset_index(drop=True)
    is_partial = (active == CONFIG.TOP_N_PARTIAL)

    slots = []
    held_tickers = set()      # tickers retained from previous week
    sector_count: dict[str, int] = {}

    # ── Phase 1: evaluate holds from previous allocation ─────────────
    if prev_allocation:
        for prev_slot in prev_allocation:
            t = prev_slot.get("ticker", "")
            if t == "CASH" or not t:
                continue
            peak          = prev_slot.get("peak", 0) or 0
            current_price = 0.0
            if prices is not None and t in prices.columns:
                s = prices[t].dropna()
                if len(s) > 0:
                    current_price = float(s.iloc[-1])

            exit_flag, exit_reason = should_exit(t, df, peak, current_price)

            if not exit_flag and len(held_tickers) < active:
                sector = prev_slot.get("sector", "OTHER")
                sc     = sector_count.get(sector, 0)
                # Look up current rank
                rk_row = df[df["TICKER"] == t]
                inv_rk = int(rk_row.iloc[0]["RANK_INVESTABLE"]) if not rk_row.empty and pd.notna(rk_row.iloc[0].get("RANK_INVESTABLE")) else "-"
                sector_count[sector] = sc + 1
                held_tickers.add(t)
                slots.append({
                    "SLOT"    : len(slots) + 1,
                    "TICKER"  : t,
                    "ETF_NAME": prev_slot.get("etf_name", t),
                    "SECTOR"  : sector,
                    "WEIGHT"  : w,
                    "INV_RANK": inv_rk,
                    "REASON"  : "HOLD — no exit trigger",
                })
            else:
                if exit_flag:
                    print(f"    EXIT {t}: {exit_reason}")

    # ── Phase 2: fill remaining active slots from ranking ────────────
    open_slots   = active - len(slots)
    candidate_idx = 0
    filled_new    = 0
    while filled_new < open_slots and candidate_idx < len(investable):
        row    = investable.iloc[candidate_idx]
        ticker = row["TICKER"]
        sector = row.get("SECTOR", "OTHER")
        candidate_idx += 1

        if ticker in held_tickers:
            continue   # already held from Phase 1
        sc = sector_count.get(sector, 0)
        if sc >= CONFIG.SECTOR_CAP:
            continue   # sector cap hit

        sector_count[sector] = sc + 1
        held_tickers.add(ticker)
        slots.append({
            "SLOT"    : len(slots) + 1,
            "TICKER"  : ticker,
            "ETF_NAME": row["ETF_NAME"],
            "SECTOR"  : sector,
            "WEIGHT"  : w,
            "INV_RANK": int(row["RANK_INVESTABLE"]),
            "REASON"  : (f"NEW BUY — Rank {int(row['RANK_INVESTABLE'])}  |  "
                         f"Sector={sector} ({sc+1}/{CONFIG.SECTOR_CAP})"),
        })
        filled_new += 1

    # Fill any remaining active slots with CASH (universe exhausted)
    while len(slots) < active:
        slots.append({
            "SLOT"    : len(slots) + 1,
            "TICKER"  : "CASH",
            "ETF_NAME": "Cash (sector cap / investable universe exhausted)",
            "SECTOR"  : "CASH",
            "WEIGHT"  : w,
            "INV_RANK": "-",
            "REASON"  : "No remaining qualifying ETF after cap",
        })

    # Remaining slots: cash buffer for PARTIAL regime
    for _ in range(len(slots), total):
        slots.append({
            "SLOT"    : len(slots) + 1,
            "TICKER"  : "CASH",
            "ETF_NAME": "Cash / Money Market",
            "SECTOR"  : "CASH",
            "WEIGHT"  : w,
            "INV_RANK": "-",
            "REASON"  : (f"Regime buffer: {regime['label']} -> "
                         f"only {active} of {total} slots active"
                         if is_partial else "Universe exhausted")
        })

    return pd.DataFrame(slots)


# =========================================================
# 6. CONSOLE SUMMARY
# =========================================================
def print_summary(df, regime, allocation):
    W = 110
    print("\n" + "=" * W)
    print("ETF MOMENTUM RANKING  |  Screen -> Score -> Regime -> Allocate")
    print("=" * W)

    r = regime
    print(f"\n  REGIME: {r['label']:30s}"
          f"  Active slots: {r['active_slots']} / {CONFIG.TOP_N}")
    print(f"  Trend Signal     ({r['trend_ticker']}): "
          f"Price={r['nifty_price']:.2f} | 50 EMA={r['nifty_ema_50']:.2f} | 100 EMA={r['nifty_ema_100']:.2f}")

    # Recommendations: Top 5 (1 per sector) from investable universe (regardless of regime)
    print("\n  " + "=" * 52)
    print("  RECOMMENDATION (Top 5 Investable | 1 Per Sector)")
    print("  " + "=" * 52)
    inv_df = df[df["SCREEN_PASS"] & (df["RANK_INVESTABLE"] > 0)].sort_values("RANK_INVESTABLE")
    recs = []
    seen_sectors = set()
    for _, row in inv_df.iterrows():
        sector = row.get("SECTOR", "OTHER")
        if sector not in seen_sectors:
            seen_sectors.add(sector)
            recs.append(row)
        if len(recs) >= 5:
            break
    
    if not recs:
        print("    [None] No ETFs passed screen filter.")
    else:
        for i, rec in enumerate(recs, 1):
            print(f"    {i}. {rec['TICKER']:<12} {rec['SECTOR']:<15} {rec['ETF_NAME'][:45]}")

    print(f"\n  ALLOCATION (Current Execution State)")
    print("  " + "-" * 105)
    for _, a in allocation.iterrows():
        is_cash = a["TICKER"] == "CASH"
        marker  = "  [CASH]" if is_cash else "  [HOLD]"
        print(f"  Slot {int(a['SLOT'])}: {a['TICKER']:<14} {a['WEIGHT']:5.1%}{marker}  {a['ETF_NAME'][:48]}")

    inv_count = df["SCREEN_PASS"].sum()
    print(f"\n  RANKING  (investable rank = scored among {inv_count} ETFs passing abs filter)")
    print(f"  {'InvRk':>5} {'Ticker':<14} {'ETF Name':<36} "
          f"{'WtdSharpe':>10} {'Sharpe6M':>9} {'Sharpe3M':>9} "
          f"{'Screen':>7}")
    print("  " + "-" * 95)

    for _, r2 in df.head(5).iterrows():
        def f(v, d=3): return f"{v:.{d}f}" if pd.notna(v) and v != 0 else "N/A"
        inv_rk = str(int(r2["RANK_INVESTABLE"])) if r2["SCREEN_PASS"] else "-"
        screen = "PASS" if r2["SCREEN_PASS"] else "FAIL"
        print(f"  {inv_rk:>5} {r2['TICKER']:<14} "
              f"{str(r2['ETF_NAME'])[:35]:<36} "
              f"{f(r2['WTD_SHARPE']):>10} {f(r2['SHARPE_6M']):>9} {f(r2['SHARPE_3M']):>9} "
              f"{screen:>7}")

    print(f"\n  Universe={len(df)}  Investable (both screens pass)={inv_count}  "
          f"Screened out={len(df)-inv_count}  Valid Wtd Sharpe={df['WTD_SHARPE'].notna().sum()}")
    print("=" * W)


# =========================================================
# 7b. HOLDINGS LOG & REBALANCE TRACKER
# =========================================================
import json
from datetime import datetime

HOLDINGS_LOG_FILE = "holdings_log.json"


def _week_key(dt=None):
    """Return ISO week key like '2026-W30' for the given date (default=today)."""
    if dt is None:
        dt = datetime.today()
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"

def _log_path(script_dir: Path) -> Path:
    return script_dir / HOLDINGS_LOG_FILE


def load_holdings_log(script_dir: Path) -> dict:
    """Load existing log; return empty dict if none exists yet."""
    p = _log_path(script_dir)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_holdings_log(script_dir: Path, log: dict):
    """Persist the updated log to disk."""
    with open(_log_path(script_dir), "w") as f:
        json.dump(log, f, indent=2, default=str)


def record_to_log(allocation: pd.DataFrame, regime: dict, run_date: str) -> dict:
    """Serialise current allocation + regime into a log entry."""
    slots = []
    for _, row in allocation.iterrows():
        slots.append({
            "slot"    : int(row["SLOT"]),
            "ticker"  : str(row["TICKER"]),
            "etf_name": str(row["ETF_NAME"]),
            "sector"  : str(row.get("SECTOR", "")),
            "weight"  : float(row["WEIGHT"]),
            "inv_rank": str(row["INV_RANK"]),
        })
    return {
        "run_date"    : run_date,
        "regime"      : regime["label"],
        "active_slots": int(regime["active_slots"]),
        "allocation"  : slots,
    }


def diff_allocations(prev: dict, curr: dict) -> list[dict]:
    """
    Compare previous and current allocation dicts.
    Returns a list of change records, one per affected ticker.
    """
    prev_alloc = {s["ticker"]: s for s in prev.get("allocation", [])}
    curr_alloc = {s["ticker"]: s for s in curr.get("allocation", [])}

    prev_holds = {t for t, s in prev_alloc.items() if t != "CASH"}
    curr_holds = {t for t, s in curr_alloc.items() if t != "CASH"}

    changes = []

    # BUY — new entrants
    for t in sorted(curr_holds - prev_holds):
        s = curr_alloc[t]
        changes.append({
            "action"  : "BUY",
            "ticker"  : t,
            "etf_name": s["etf_name"],
            "sector"  : s["sector"],
            "prev_wt" : 0.0,
            "curr_wt" : s["weight"],
            "prev_rk" : "-",
            "curr_rk" : s["inv_rank"],
            "note"    : "New entry",
        })

    # SELL — exits
    for t in sorted(prev_holds - curr_holds):
        s = prev_alloc[t]
        # Check if it moved to cash or just dropped off
        in_curr_as_cash = any(
            sl["ticker"] == "CASH" for sl in curr["allocation"]
        )
        changes.append({
            "action"  : "SELL",
            "ticker"  : t,
            "etf_name": s["etf_name"],
            "sector"  : s["sector"],
            "prev_wt" : s["weight"],
            "curr_wt" : 0.0,
            "prev_rk" : s["inv_rank"],
            "curr_rk" : "-",
            "note"    : "Exited",
        })

    # HOLD / ADD / TRIM — existing positions
    for t in sorted(prev_holds & curr_holds):
        ps = prev_alloc[t]
        cs = curr_alloc[t]
        pw = ps["weight"]
        cw = cs["weight"]
        pr = ps["inv_rank"]
        cr = cs["inv_rank"]

        if abs(cw - pw) < 0.001:
            action = "HOLD"
            note   = "No change"
        elif cw > pw:
            action = "ADD"
            note   = f"Weight increased {pw:.1%} -> {cw:.1%}"
        else:
            action = "TRIM"
            note   = f"Weight reduced {pw:.1%} -> {cw:.1%}"

        # Flag rank drift even on holds
        try:
            rk_drift = int(pr) - int(cr)
            if abs(rk_drift) >= 3:
                note += f"  |  Rank: {pr} -> {cr} ({'+' if rk_drift>0 else ''}{rk_drift})"
        except (ValueError, TypeError):
            pass

        changes.append({
            "action"  : action,
            "ticker"  : t,
            "etf_name": cs["etf_name"],
            "sector"  : cs["sector"],
            "prev_wt" : pw,
            "curr_wt" : cw,
            "prev_rk" : pr,
            "curr_rk" : cr,
            "note"    : note,
        })

    # REGIME CHANGE note
    prev_regime = prev.get("regime", "")
    curr_regime = curr.get("regime", "")
    if prev_regime != curr_regime:
        changes.insert(0, {
            "action"  : "REGIME",
            "ticker"  : "—",
            "etf_name": f"Regime changed: {prev_regime} -> {curr_regime}",
            "sector"  : "—",
            "prev_wt" : 0.0,
            "curr_wt" : 0.0,
            "prev_rk" : "-",
            "curr_rk" : "-",
            "note"    : f"Active slots: {prev.get('active_slots','?')} -> {curr.get('active_slots','?')}",
        })

    # Sort: REGIME first, then SELL, BUY, ADD, TRIM, HOLD
    order = {"REGIME":0,"SELL":1,"BUY":2,"ADD":3,"TRIM":4,"HOLD":5}
    changes.sort(key=lambda x: order.get(x["action"], 9))
    return changes


def update_log(script_dir: Path, allocation: pd.DataFrame,
               regime: dict, prices: pd.DataFrame = None) -> tuple[dict, list[dict]]:
    """
    Load log, diff vs previous month, save updated log.
    When prices is provided, enriches each slot with entry_price and peak
    for trailing stop loss tracking.
    Returns (prev_entry_or_None, list_of_changes, full_log).
    """
    log      = load_holdings_log(script_dir)
    month_key = _week_key()   # weekly key, e.g. "2026-W30"
    run_date  = datetime.today().strftime("%Y-%m-%d %H:%M")

    curr_entry = record_to_log(allocation, regime, run_date)

    # Find most recent previous month entry
    sorted_keys = sorted(log.keys())
    prev_keys   = [k for k in sorted_keys if k < month_key]
    prev_entry  = log[prev_keys[-1]] if prev_keys else None

    # ── Enrich slots with entry_price / peak for TSL tracking ──────────
    if prices is not None:
        prev_alloc = (
            {s["ticker"]: s for s in prev_entry.get("allocation", [])}
            if prev_entry else {}
        )
        for slot in curr_entry["allocation"]:
            t = slot["ticker"]
            if t == "CASH":
                slot["entry_price"] = None
                slot["peak"]        = None
                continue

            # Current NAV from the loaded price grid
            current_nav = None
            if t in prices.columns:
                s = prices[t].dropna()
                if len(s) > 0:
                    current_nav = float(s.iloc[-1])

            prev_slot = prev_alloc.get(t)
            if (prev_slot is not None
                    and prev_slot.get("entry_price") is not None
                    and prev_slot["ticker"] != "CASH"):
                # HOLD — carry forward entry_price, update peak
                slot["entry_price"] = prev_slot["entry_price"]
                old_peak = prev_slot.get("peak") or 0
                slot["peak"] = max(old_peak, current_nav) if current_nav else old_peak
            else:
                # New BUY
                slot["entry_price"] = current_nav
                slot["peak"]        = current_nav

    # Compute diff
    changes = diff_allocations(prev_entry, curr_entry) if prev_entry else []

    # Save current week (overwrites if same week run again — latest wins)
    log[month_key] = curr_entry
    save_holdings_log(script_dir, log)

    return prev_entry, changes, log



# =========================================================
# 7c. DAILY TSL CHECK (via --tsl flag)
# =========================================================
def check_tsl(script_dir: Path) -> dict | None:
    """
    Daily Trailing Stop Loss check.
    Fetches live NAVs via yfinance for held ETFs (max 5),
    compares against stored peaks, and flags any TSL breaches.
    Advisory only — does NOT auto-sell positions.

    Returns a dict with:
      - "rows": list of dicts (one per slot) with Slot, Ticker, Entry, Peak, etc.
      - "breaches": list of dicts for slots exceeding TSL threshold
      - "message": status message string
    Returns None if no allocation found or fetch failed.
    """
    if not _YF_AVAILABLE:
        msg = "[tsl] yfinance not installed; cannot fetch live NAVs for TSL check."
        print(msg)
        return {"rows": [], "breaches": [], "message": msg}

    log = load_holdings_log(script_dir)
    month_key = _week_key()   # weekly key

    if month_key not in log:
        msg = "[tsl] No allocation found for current week."
        print(msg)
        print("      Run the weekly rebalance first: python etf_momentum_ranking.py")
        return {"rows": [], "breaches": [], "message": msg}

    entry = log[month_key]
    allocation = entry.get("allocation", [])

    # Get held tickers (non-CASH)
    held = [s for s in allocation if s["ticker"] != "CASH"]
    if not held:
        msg = "[tsl] No active positions (all cash). Nothing to check."
        print(msg)
        return {"rows": [], "breaches": [], "message": msg}

    # Fetch live prices via yfinance (NSE tickers need .NS suffix)
    tickers_nse = [s["ticker"] + ".NS" for s in held]
    print(f"[tsl] Fetching live NAVs for {len(held)} position(s) via yfinance ...")

    live_prices: dict[str, float] = {}
    try:
        data = yf.download(tickers_nse, period="5d", auto_adjust=True, progress=False)
        close = data["Close"]
        if isinstance(close, pd.Series):
            close = close.to_frame(tickers_nse[0])
        for s in held:
            yf_t = s["ticker"] + ".NS"
            if yf_t in close.columns:
                vals = close[yf_t].dropna()
                if len(vals) > 0:
                    live_prices[s["ticker"]] = float(vals.iloc[-1])
    except Exception as e:
        msg = f"[tsl] ERROR fetching prices: {e}"
        print(msg)
        return {"rows": [], "breaches": [], "message": msg}

    # ── TSL Dashboard ────────────────────────────────────────────
    threshold = CONFIG.TSL_THRESHOLD
    W = 86
    print("\n" + "=" * W)
    print(f"  TSL CHECK  |  {datetime.today().strftime('%Y-%m-%d %H:%M')}"
          f"  |  Threshold: {threshold*100:.0f}%")
    print("=" * W)
    print(f"  {'Slot':>4}  {'Ticker':<14} {'Entry':>8} {'Peak':>8}"
          f" {'TSL NAV':>8} {'NAV':>8} {'DD%':>7}  Status")
    print("  " + "-" * (W - 4))

    result_rows = []
    breaches = []
    for s in allocation:
        t        = s["ticker"]
        slot_num = s["slot"]

        if t == "CASH":
            print(f"  {slot_num:>4}  {'CASH':<14}"
                  f" {'--':>8} {'--':>8} {'--':>8} {'--':>8} {'--':>7}  --")
            result_rows.append({
                "Slot": slot_num, "Ticker": "CASH",
                "ETF Name": "Cash / Money Market",
                "Entry": None, "Peak": None, "TSL NAV": None,
                "Current NAV": None, "DD%": None, "Status": "—",
            })
            continue

        entry_price = s.get("entry_price")
        peak        = s.get("peak")
        nav         = live_prices.get(t)

        if nav is None:
            ep_str = f"{entry_price:.2f}" if entry_price else "--"
            pk_str = f"{peak:.2f}"        if peak        else "--"
            print(f"  {slot_num:>4}  {t:<14}"
                  f" {ep_str:>8} {pk_str:>8} {'--':>8} {'ERR':>8} {'--':>7}  FETCH FAILED")
            result_rows.append({
                "Slot": slot_num, "Ticker": t,
                "ETF Name": s.get("etf_name", ""),
                "Entry": entry_price, "Peak": peak, "TSL NAV": None,
                "Current NAV": None, "DD%": None, "Status": "FETCH FAILED",
            })
            continue

        # Update peak if current NAV is a new high
        if peak is None or nav > peak:
            peak = nav
        s["peak"] = peak

        # TSL NAV = the exact price at which the stop loss triggers
        tsl_nav = peak * (1 - threshold)

        # Compute drawdown
        dd = (peak - nav) / peak if peak > 0 else 0.0

        status_cli = "!! TSL BREACH !!" if dd >= threshold else "OK"
        status_ui  = "⚠️ BREACH" if dd >= threshold else "✅ OK"
        if dd >= threshold:
            breaches.append({
                "Ticker": t, "ETF Name": s.get("etf_name", ""),
                "DD%": round(dd * 100, 1),
                "TSL NAV": round(tsl_nav, 2),
            })

        ep_str = f"{entry_price:.2f}" if entry_price else "--"
        print(f"  {slot_num:>4}  {t:<14}"
              f" {ep_str:>8} {peak:>8.2f} {tsl_nav:>8.2f} {nav:>8.2f} {dd*100:>6.1f}%  {status_cli}")

        result_rows.append({
            "Slot": slot_num, "Ticker": t,
            "ETF Name": s.get("etf_name", ""),
            "Entry": round(entry_price, 2) if entry_price else None,
            "Peak": round(peak, 2),
            "TSL NAV": round(tsl_nav, 2),
            "Current NAV": round(nav, 2),
            "DD%": round(dd * 100, 1),
            "Status": status_ui,
        })

    print("=" * W)

    if breaches:
        print("\n  !! ACTION REQUIRED !!")
        for b in breaches:
            print(f"  -> SELL {b['Ticker']} ({b['ETF Name']})"
                  f" -- drawdown {b['DD%']:.1f}% exceeds {threshold*100:.0f}% TSL"
                  f" (trigger: {b['TSL NAV']:.2f})")
        print(f"  -> Move proceeds to cash until next weekly rebalance.\n")
        msg = f"{len(breaches)} TSL BREACH(ES) DETECTED"
    else:
        print(f"\n  All positions within TSL threshold. No action needed.\n")
        msg = "All positions within TSL threshold. No action needed."

    # Save updated peaks back to log
    log[month_key] = entry
    save_holdings_log(script_dir, log)
    print(f"[tsl] Updated peaks saved to {HOLDINGS_LOG_FILE}")

    return {"rows": result_rows, "breaches": breaches, "message": msg}


def run_tsl_check():
    """Entry point for daily TSL monitoring via --tsl flag."""
    SCRIPT_DIR = Path(__file__).resolve().parent
    check_tsl(SCRIPT_DIR)


# =========================================================
# 7. EXCEL OUTPUT
# =========================================================
def _brd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _h(ws, row, col, val, bg="1F4E79", fg="FFFFFF", sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, size=sz, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _brd()
    return c

def _d(ws, row, col, val, bg="FFFFFF", fmt=None, bold=False, fg="000000"):
    if isinstance(val, float) and np.isnan(val):
        val = None
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _brd()
    if fmt:
        c.number_format = fmt
    return c


def _write_rebalance_sheet(wb, prev_entry, changes, log,
                            NAVY, GREEN, DKGREEN, ORANGE, YELLOW, GREY):
    """Write the Rebalance sheet with 3 sections:
       1. Current allocation
       2. Changes vs last month
       3. Last 12 months history
    """
    ACTION_COLORS = {
        "BUY"   : "C6EFCE",   # green
        "SELL"  : "FFC7CE",   # red
        "ADD"   : "DAEEF3",   # light blue
        "TRIM"  : "FFEB9C",   # amber
        "HOLD"  : "F2F2F2",   # grey
        "REGIME": "D9D9D9",   # dark grey
    }
    NAVY2  = "1F4E79"

    wb_sheets = [s.title for s in wb.worksheets]
    if "Rebalance" in wb_sheets:
        del wb["Rebalance"]
    wb.create_sheet("Rebalance", 1)   # insert as second sheet
    wr = wb["Rebalance"]

    row: int = 1

    def title_row(ws, r, text, cols, bg=NAVY2):
        ws.merge_cells(f"A{r}:{get_column_letter(cols)}{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20

    def hdr_row(ws, r, hdrs, widths, bg=NAVY2):
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _brd()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[r].height = 28

    def data_row(ws, r, vals, fmts, bg="F2F2F2", bold=False):
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            if isinstance(v, float) and np.isnan(v): v = None
            c = ws.cell(row=r, column=ci, value=v)
            c.font      = Font(name="Arial", size=9, bold=bold)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _brd()
            if f and f != "@": c.number_format = f
        ws.row_dimensions[r].height = 15

    # ── Section 1: Current Allocation ─────────────────────────────
    title_row(wr, row, "SECTION 1 — CURRENT ALLOCATION", 7)
    row += 1
    hdr_row(wr, row,
            ["Slot","Inv Rank","Ticker","ETF Name","Sector","Weight","Action"],
            [7, 9, 14, 40, 18, 9, 9])
    row += 1

    curr_tickers = set()
    prev_tickers = set(
        s["ticker"] for s in (prev_entry or {}).get("allocation", [])
        if s["ticker"] != "CASH"
    ) if prev_entry else set()

    # Build quick lookup from changes
    change_map = {c["ticker"]: c["action"] for c in (changes or [])}

    from datetime import datetime
    curr_month = _week_key()
    curr_entry = log.get(curr_month, {})
    for sl in curr_entry.get("allocation", []):
        t      = sl["ticker"]
        action = change_map.get(t, "HOLD") if t != "CASH" else "CASH"
        bg     = ACTION_COLORS.get(action, GREY)
        if t != "CASH": curr_tickers.add(t)
        data_row(wr, row,
                 [sl["slot"], sl["inv_rank"], t,
                  sl["etf_name"], sl["sector"],
                  sl["weight"], action],
                 ["0","@","@","@","@","0%","@"],
                 bg=bg, bold=(t != "CASH"))
        row += 1

    row += 1  # spacer

    # ── Section 2: Changes vs Previous Month ──────────────────────
    prev_month_label = prev_entry.get("run_date","N/A")[:7] if prev_entry else "N/A"
    title_row(wr, row,
              f"SECTION 2 — CHANGES vs PREVIOUS ({prev_month_label})", 7,
              bg="375623" if prev_entry else "7F6000")
    row += 1

    if not changes:
        wr.merge_cells(f"A{row}:G{row}")
        c = wr.cell(row=row, column=1,
                    value="No previous month data — this is the first recorded rebalance.")
        c.font      = Font(name="Arial", italic=True, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
    else:
        hdr_row(wr, row,
                ["Action","Ticker","ETF Name","Sector",
                 "Prev Wt","Curr Wt","Note"],
                [9, 14, 38, 18, 9, 9, 48])
        row += 1
        for ch in changes:
            bg   = ACTION_COLORS.get(ch["action"], GREY)
            bold = ch["action"] in ("BUY","SELL","REGIME")
            data_row(wr, row,
                     [ch["action"], ch["ticker"], ch["etf_name"],
                      ch["sector"], ch["prev_wt"] or None,
                      ch["curr_wt"] or None, ch["note"]],
                     ["@","@","@","@","0%","0%","@"],
                     bg=bg, bold=bold)
            row += 1

    row += 1  # spacer

    # ── Section 3: 12-Month History ───────────────────────────────
    title_row(wr, row, f"SECTION 3 — LAST {CONFIG.HISTORY_PERIODS} PERIODS HISTORY", 7,
              bg="203864")
    row += 1

    sorted_months: list[str] = list(sorted(log.keys()))[-int(CONFIG.HISTORY_PERIODS):]  # pyre-ignore[16]

    # Collect all unique tickers ever held (excluding CASH)
    all_tickers: list[str] = []
    seen: set[str] = set()
    for mk in reversed(sorted_months):
        for sl in log[mk].get("allocation", []):
            t = sl["ticker"]
            if t != "CASH" and t not in seen:
                all_tickers.append(t)
                seen.add(t)

    # Header: Month | Regime | Ticker1 | Ticker2 | ...
    hdrs   = ["Month", "Regime"] + all_tickers
    widths = [12, 24] + [12] * len(all_tickers)
    hdr_row(wr, row, hdrs, widths)
    row += 1  # pyre-ignore[58]

    for mk in sorted_months:
        entry    = log[mk]
        regime_l = entry.get("regime", "")
        held     = {s["ticker"]: s["weight"]
                    for s in entry.get("allocation", [])
                    if s["ticker"] != "CASH"}

        regime_bg = ("E2EFDA" if "BULL" in regime_l else
                     "FCE4D6" if "BEAR" in regime_l else "FFF2CC")

        vals = [mk, regime_l]
        fmts = ["@", "@"]
        for t in all_tickers:
            w = held.get(t)
            vals.append(w)
            fmts.append("0%" if w is not None else "@")

        # Write row cell by cell for per-cell colouring
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            if isinstance(v, float) and np.isnan(v): v = None
            c = wr.cell(row=row, column=ci, value=v)
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _brd()
            if ci <= 2:
                c.fill = PatternFill("solid", fgColor=regime_bg)
                c.font = Font(name="Arial", size=9, bold=(ci==1))
            elif v is not None:
                c.fill = PatternFill("solid", fgColor=DKGREEN)
                c.font = Font(name="Arial", size=9, bold=True)
            else:
                c.fill = PatternFill("solid", fgColor="F2F2F2")
            if f and f != "@": c.number_format = f
        wr.row_dimensions[row].height = 15
        row += 1  # pyre-ignore[58]

    wr.freeze_panes = "A2"


def save_excel(df, regime, allocation, out_path, prev_entry=None, changes=None, log=None):
    wb = Workbook()
    NAVY    = "1F4E79"
    GREEN   = "E2EFDA"
    DKGREEN = "C6EFCE"
    ORANGE  = "FCE4D6"
    YELLOW  = "FFF2CC"
    GREY    = "F2F2F2"
    BULL_C  = "375623"
    BEAR_C  = "C00000"
    PART_C  = "7F6000"

    regime_color = (BULL_C if regime["label"] == "BULL" else
                    BEAR_C if "BEAR" in regime["label"] else PART_C)

    # ── Sheet 1: Rankings ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Rankings"

    # Title
    ws.merge_cells("A1:Y1")
    c = ws["A1"]
    c.value     = ("ETF Momentum Ranking  |  "
                   "Step 1: Screen (abs filter)  ->  "
                   "Step 2: Score (Weighted Sharpe)  ->  "
                   "Step 3: Regime  ->  Step 4: Allocate")
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Regime row
    ws.merge_cells("A2:Y2")
    r = regime
    rtext = (f"REGIME: {r['label']}  |  Active slots: {r['active_slots']}/{CONFIG.TOP_N}  |  "
             f"Trend Signal ({r['trend_ticker']}): Price {r['nifty_price']:.2f}  |  "
             f"50 EMA: {r['nifty_ema_50']:.2f}  |  100 EMA: {r['nifty_ema_100']:.2f}")
    rc = ws["A2"]
    rc.value     = rtext
    rc.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    rc.fill      = PatternFill("solid", fgColor=regime_color)
    rc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    COLS = [
        ("Investable\nRank",    10, "0"),
        ("Ticker",              14, "@"),
        ("ETF Name",            40, "@"),
        ("Sector",              18, "@"),
        ("Close",               10, "0.00"),
        ("52Wk\nHigh",          10, "0.00"),
        ("% From\n52Wk High",   12, "0.00"),
        ("EMA 100",             10, "0.00"),
        ("% From\nEOM",         11, "0.00"),
        ("Wtd Sharpe\nScore",   14, "0.000"),
        ("Sharpe\n6M",          13, "0.000"),
        ("Sharpe\n3M",          13, "0.000"),
        ("Screen\nResult",      11, "@"),
    ]
    KEYS = [
        "RANK_INVESTABLE",
        "TICKER", "ETF_NAME", "SECTOR", "CLOSE", "52WK_HIGH", "PCT_FROM_HIGH",
        "EMA_100",
        "EOM_PCT",
        "WTD_SHARPE", "SHARPE_6M", "SHARPE_3M",
        "SCREEN_PASS",
    ]

    HDR = 3
    for ci, (hdr, width, _) in enumerate(COLS, 1):
        _h(ws, HDR, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HDR].height = 32

    for ri, (_, row) in enumerate(df.iterrows(), start=HDR + 1):
        passed  = row["SCREEN_PASS"]
        inv_rk  = row["RANK_INVESTABLE"]
        in_alloc = passed and (inv_rk > 0) and (inv_rk <= regime["active_slots"])
        bg = (DKGREEN if in_alloc else
              GREEN   if passed else
              ORANGE)

        # Pre-compute NAV vs EMA condition for this row
        # CLOSE > EMA_100: keep row background, set font to dark green
        close_above_ema100 = (
            pd.notna(row["CLOSE"]) and
            pd.notna(row["EMA_100"]) and
            row["CLOSE"] > row["EMA_100"]
        )

        for ci, (key, (_, _, fmt)) in enumerate(zip(KEYS, COLS), 1):
            val = row[key]
            if key == "SCREEN_PASS":
                val = "PASS" if val else "FAIL"
            elif key == "RANK_INVESTABLE" and (not passed or pd.isna(val)):
                val = None
            # Close cell: dark green font when NAV > 100 EMA, background unchanged
            cell_fg = "1A5C2E" if (key == "CLOSE" and close_above_ema100) else "000000"
            _d(ws, ri, ci, val, bg=bg,
               fmt=fmt if fmt != "@" else None, bold=in_alloc, fg=cell_fg)
        ws.row_dimensions[ri].height = 14

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{HDR + len(df)}"

    # ── Sheet 2: Allocation ────────────────────────────────────────
    wa = wb.create_sheet("Allocation")
    wa.merge_cells("A1:F1")
    c = wa["A1"]
    c.value     = (f"Top-{CONFIG.TOP_N} Allocation  |  Screen: ≤{CONFIG.MAX_DRAWDOWN_FROM_HIGH*100:.0f}% from 52wk high  |  "
                   f"Regime={regime['label']}  |  Active slots={regime['active_slots']}")
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=regime_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    wa.row_dimensions[1].height = 22

    for ci, (hdr, w) in enumerate(
            zip(["Slot", "Inv Rank", "Sector", "Ticker", "ETF Name", "Weight", "Detail"],
                [8, 10, 18, 14, 40, 10, 55]), 1):
        _h(wa, 2, ci, hdr)
        wa.column_dimensions[get_column_letter(ci)].width = w
    wa.row_dimensions[2].height = 22
    wa.merge_cells("A1:G1")

    for ri, (_, row) in enumerate(allocation.iterrows(), start=3):
        is_cash = row["TICKER"] == "CASH"
        is_buf  = is_cash and "buffer" in str(row["REASON"])
        bg = YELLOW if is_buf else (ORANGE if is_cash else DKGREEN)
        for ci, (v, f) in enumerate(
                zip([row["SLOT"], row["INV_RANK"], row.get("SECTOR",""),
                     row["TICKER"], row["ETF_NAME"], row["WEIGHT"], row["REASON"]],
                    ["0", "@", "@", "@", "@", "0.0%", "@"]), 1):
            _d(wa, ri, ci, v, bg=bg, fmt=f if f != "@" else None, bold=True)
        wa.row_dimensions[ri].height = 18

    # ── Sheet 3: Regime Detail ─────────────────────────────────────
    wr = wb.create_sheet("Regime")
    wr.column_dimensions["A"].width = 40
    wr.column_dimensions["B"].width = 25
    _h(wr, 1, 1, "Regime Parameter", bg=NAVY)
    _h(wr, 1, 2, "Value / Status", bg=NAVY)

    regime_rows = [
        ("Regime Label",                                  r["label"]),
        ("Active slots",                                  f"{r['active_slots']} of {CONFIG.TOP_N}"),
        ("--- LOGIC ---",                                 ""),
        ("BULL (50EMA > 100EMA & Price > 50EMA) -> slots",str(CONFIG.TOP_N)),
        ("PARTIAL (50EMA <= 100EMA & Price > 50EMA) -> slots", str(CONFIG.TOP_N_PARTIAL)),
        ("BEAR  (Price <= 50EMA)  -> slots",              "0  (full cash)"),
        ("--- PARAMETERS ---",                            ""),
        ("Index used",                                    r["trend_ticker"]),
        ("Current price",                                 f"{r['nifty_price']:.2f}"),
        (f"{CONFIG.TREND_FAST_EMA_WINDOW}-day EMA",       f"{r['nifty_ema_50']:.2f}"),
        (f"{CONFIG.TREND_EMA_WINDOW}-day EMA",            f"{r['nifty_ema_100']:.2f}"),
        ("Price above 50-day EMA?",                       str(r["nifty_price"] > r["nifty_ema_50"] if pd.notna(r["nifty_price"]) else "N/A")),
        ("50-day EMA > 100-day EMA?",                     str(r["nifty_ema_50"] > r["nifty_ema_100"] if pd.notna(r["nifty_ema_50"]) else "N/A")),
    ]
    for ri2, (lbl, val) in enumerate(regime_rows, start=2):
        is_section = lbl.startswith("---")
        ok_bg = (NAVY   if is_section else
                 GREEN  if "True"  in str(val) or "BULL" in str(val) else
                 ORANGE if "False" in str(val) or "BEAR" in str(val) else
                 GREY)
        fg = "FFFFFF" if is_section else "000000"
        c1 = _d(wr, ri2, 1, lbl, bg=ok_bg if not is_section else NAVY)
        c2 = _d(wr, ri2, 2, val, bg=ok_bg if not is_section else NAVY)
        if is_section:
            c1.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c2.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")

    # ── Sheet 4: Rebalance Tracker ───────────────────────────
    if changes is not None:
        _write_rebalance_sheet(wb, prev_entry, changes, log,
                               NAVY, GREEN, DKGREEN, ORANGE, YELLOW, GREY)

    wb.save(out_path)
    print(f"\n[saved] -> {Path(out_path).resolve()}")


# =========================================================
# 8. MAIN
# =========================================================
def run_pipeline(fp=None, out=None):
    SCRIPT_DIR = Path(__file__).resolve().parent
    if fp is None:
        fp = str(SCRIPT_DIR / CONFIG.INPUT_FILE)
    if out is None:
        out = str(SCRIPT_DIR / CONFIG.OUTPUT_FILE)

    print(f"[load]   {fp}")
    meta, prices = load_etf_data(fp)
    print(f"         {len(meta)} ETFs | {len(prices)} days "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})")

    print("[regime] Computing tiered regime filter ...")
    regime = regime_status(prices, script_dir=SCRIPT_DIR)

    print("[scores] Screening + scoring all ETFs ...")
    ranking = build_ranking(meta, prices)

    # Load previous week's allocation for hold-and-replace
    prev_log   = load_holdings_log(SCRIPT_DIR)
    prev_wk    = _week_key()
    sorted_wks = sorted(prev_log.keys())
    prev_wks   = [k for k in sorted_wks if k < prev_wk]
    prev_alloc = prev_log[prev_wks[-1]].get("allocation", []) if prev_wks else None

    print("[alloc]  Building allocation (WRH — hold & replace) ...")
    allocation = build_allocation(ranking, regime,
                                  prev_allocation=prev_alloc,
                                  prices=prices)

    print_summary(ranking, regime, allocation)

    print("[log]    Updating holdings log ...")
    prev_entry, changes, log = update_log(SCRIPT_DIR, allocation, regime, prices)

    if prev_entry:
        prev_period = prev_entry.get("run_date","?")[:10]
        print(f"         Previous: {prev_period}  |  Changes: {len(changes)}")
        for ch in changes:
            arrow = {"BUY":"+ BUY","SELL":"- SELL","ADD":"^ ADD",
                     "TRIM":"v TRIM","HOLD":"= HOLD","REGIME":"! REGIME"}.get(ch["action"],"  ")
            print(f"           {arrow:8s} {ch['ticker']:<14} {ch['note']}")
    else:
        print("         First run — no previous holdings to compare.")

    save_excel(ranking, regime, allocation, out,
               prev_entry=prev_entry, changes=changes, log=log)


if __name__ == "__main__":
    if "--tsl" in sys.argv:
        run_tsl_check()
    else:
        fp_arg  = sys.argv[1] if len(sys.argv) > 1 else None
        out_arg = sys.argv[2] if len(sys.argv) > 2 else None
        run_pipeline(fp_arg, out_arg)