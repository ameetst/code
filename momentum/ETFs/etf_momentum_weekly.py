"""
ETF Weekly Momentum Rebalance Strategy
========================================
Rebalances every Monday with:
1. MOM_ACCEL > 0 filter (0.4*(Z_1m+Z_3m) - 0.2*Z_6m)
2. Within 25% of 52-week high filter
3. Ranked by Weighted Sharpe (50% 6M + 50% 3M)
4. Exit rules: 5% TSL or rank outside top 20
5. Weekly cadence for regime filter

Exit timing: Monday only (batch processing)
Position tracking: JSON with entry/peak/current metrics
"""

from __future__ import annotations
import sys
import json
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl

_SCRIPT_DIR = Path(__file__).resolve().parent
_CONFIG_FILE = "strategy_config.json"

def _config_path(script_dir: Path | None = None) -> Path:
    return (script_dir or _SCRIPT_DIR) / _CONFIG_FILE

def load_config_from_json(script_dir: Path | None = None) -> dict:
    """Load strategy_config.json. Returns empty dict if file missing."""
    p = _config_path(script_dir)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

class CONFIG:
    INPUT_FILE  = "ETF.xlsx"
    OUTPUT_FILE = "etf_rankings_weekly.xlsx"

    # Momentum windows (trading days)
    WINDOW_6M   = 126
    WINDOW_3M   = 63
    WINDOW_1M   = 21
    ANNUALIZE   = 252

    # Portfolio allocation
    TOP_N         = 5
    TOP_N_PARTIAL = 3

    # Filters
    MAX_DRAWDOWN_FROM_HIGH = 0.25

    # Weighted Sharpe blending
    SHARPE_W6M  = 0.5
    SHARPE_W3M  = 0.5

    # Regime filter
    REGIME_TICKER      = "MONIFTY500"
    REGIME_FALLBACKS   = ["BSE500IETF", "HDFCBSE500", "NIFTYBEES"]
    TREND_FAST_EMA_WINDOW = 50
    TREND_EMA_WINDOW   = 100

    # Sector cap
    SECTOR_CAP = 1

    # Risk-free rate
    DAILY_RF = 0.07 / 252

    # Exit rules
    TSL_THRESHOLD = 0.05  # 5% trailing stop loss
    RANK_RETENTION = 20   # Keep if rank <= 20

    # Weekly rebalance
    REBALANCE_FREQUENCY = "weekly"

def _apply_json_config(cls, cfg: dict):
    """Override CONFIG class attributes from a JSON dict."""
    _KEYS = [
        "INPUT_FILE", "OUTPUT_FILE",
        "WINDOW_6M", "WINDOW_3M", "WINDOW_1M", "ANNUALIZE",
        "TOP_N", "TOP_N_PARTIAL",
        "MAX_DRAWDOWN_FROM_HIGH",
        "SHARPE_W6M", "SHARPE_W3M",
        "REGIME_TICKER", "REGIME_FALLBACKS",
        "TREND_FAST_EMA_WINDOW", "TREND_EMA_WINDOW",
        "SECTOR_CAP",
        "TSL_THRESHOLD",
        "RANK_RETENTION_THRESHOLD",
    ]
    for key in _KEYS:
        if key in cfg:
            setattr(cls, key, cfg[key])
    if "DAILY_RF_ANNUAL" in cfg:
        cls.DAILY_RF = cfg["DAILY_RF_ANNUAL"] / cls.ANNUALIZE
    return cls

# Apply JSON overrides
_apply_json_config(CONFIG, load_config_from_json())


# =========================================================
# SECTOR CLASSIFICATION (same as monthly version)
# =========================================================
_SECTOR_RULES = [
    ("PSU_BANK",         ["psu bank","psubnk","psubank","bse psu bank"]),
    ("PRIVATE_BANK",     ["private bank","pvt bank","pvtban","nifty pb "]),
    ("BANKING_BROAD",    ["nifty bank","bse bank"," bank ","banketf","bankbees","banknifty","nifban"]),
    ("IT_TECH",          ["nifty it","bse it"," it etf","itbees","itietf","nifit","tech etf"]),
    ("HEALTHCARE",       ["healthcare","pharma","health "," hc "," hc\\"]),
    ("METAL",            ["metal"]),
    ("ENERGY",           ["energy","oil & gas","o&g","oilietf","power etf","bse power"]),
    ("INFRA",            ["infra"]),
    ("CONSUMPTION",      ["consumption","consump","fmcg","consumer"]),
    ("REALTY",           ["realty","real estate"]),
    ("DEFENCE",          ["defence","dfnc"]),
    ("PSE",              ["pse etf","cpse","nifty pse","bharat 22","cpseetf"]),
    ("AUTO",             ["auto"]),
    ("CHEMICALS",        ["chemical"]),
    ("FIN_SERVICES",     ["fin serv","financial serv","finietf","bfsi","capital mkt","captl mkt","capital market","cptmkt","capital mrkts"]),
    ("GOLD",             ["gold"]),
    ("SILVER",           ["silver"]),
    ("GOVT_BONDS",       ["g-sec","gsec","gilt","bond etf","bharat bond","ebbetf"]),
    ("FACTOR_MOMENTUM",  ["momentum","mmt","mmntm"]),
    ("FACTOR_VALUE",     ["value 20","value 30","value 50","enhanced val","enhval"]),
    ("FACTOR_QUALITY",   ["quality","qlty"," ql "," ql30","qual30"]),
    ("FACTOR_LOW_VOL",   ["low vol","lowvol","lw- vol"]),
    ("INTERNATIONAL",    ["nasdaq","s&p 500","hang seng","hangseng","hngsng","msci","fang+"]),
    ("MIDCAP",           ["midcap","mid cap","mdsmc","midsmall"]),
    ("SMALLCAP",         ["smallcap","small cap","sml100","smcp"]),
    ("NEXT_50",          ["next 50","next50","juniorbees","jr bees"]),
    ("BROAD_MARKET",     ["nifty 50","nifty50","sensex","nifty 100","nifty100","nifty 200","nifty 500","nifty500","total market","total mrkt","bse 500","bse500","multicap","mltcp","lgmdcp","gth sectors","flexicap","flexi"]),
]

def classify_sector(etf_name: str, ticker: str) -> str:
    n = etf_name.lower()
    t = ticker.lower()
    for sector, keywords in _SECTOR_RULES:
        for kw in keywords:
            if kw in n or kw in t:
                return sector
    return "OTHER"


# =========================================================
# 1. DATA LOADING
# =========================================================
def load_etf_data(filepath: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load ETF price data from ETF.xlsx"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    ws = wb["DATA"]
    max_col = ws.max_column
    max_row = ws.max_row

    PRICE_START_COL = 3
    n_price_cols = max_col - PRICE_START_COL + 1

    today = pd.Timestamp.today().normalize()
    candidate_dates = pd.bdate_range(end=today, periods=n_price_cols)
    dates = list(candidate_dates)

    rows = list(ws.iter_rows(min_row=2, max_row=max_row, values_only=True))
    wb.close()

    etf_names, tickers, price_rows = [], [], []
    for row in rows:
        name   = str(row[0]).strip() if row[0] is not None else ""
        ticker = str(row[1]).strip() if row[1] is not None else ""
        if not ticker or ticker == "None":
            continue
        etf_names.append(name)
        tickers.append(ticker)
        price_rows.append(row[2: 2 + n_price_cols])

    meta = pd.DataFrame({"ETF_NAME": etf_names, "TICKER": tickers})
    price_raw = pd.DataFrame(price_rows, index=tickers, columns=dates)
    price_raw = price_raw.apply(pd.to_numeric, errors="coerce").replace(0, np.nan)
    prices = price_raw.T.sort_index().ffill()

    print(f"[load]   {filepath}")
    print(f"         {len(tickers)} ETFs  |  {len(dates)} date cols  "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})")
    return meta.reset_index(drop=True), prices


# =========================================================
# 2. SCORING
# =========================================================
def sharpe_score(series: pd.Series, window: int, daily_rf: float | None = None) -> float:
    """Annualised Sharpe over lookback window"""
    clean = series.dropna()
    if len(clean) < window + 1:
        return np.nan
    rf = CONFIG.DAILY_RF if daily_rf is None else daily_rf
    log_ret = np.log(clean.iloc[-window - 1:] / clean.iloc[-window - 1:].shift(1)).dropna()
    excess  = log_ret - rf
    if excess.std() == 0:
        return np.nan
    return (excess.mean() / excess.std()) * np.sqrt(CONFIG.ANNUALIZE)


def _zscore(series: pd.Series) -> pd.Series:
    """Cross-sectional Z-score; NaN values stay NaN."""
    mu  = series.mean()
    sig = series.std()
    if sig == 0 or np.isnan(sig):
        return pd.Series(0.0, index=series.index)
    return (series - mu) / sig


# =========================================================
# 3. REGIME FILTER (Weekly Cadence)
# =========================================================
def regime_status(prices: pd.DataFrame) -> dict:
    """
    Returns tiered regime state for weekly cadence:
      BULL    - EMA50 > EMA100 AND Price > EMA50 -> invest TOP_N slots
      PARTIAL - Price > EMA100 (but not BULL)     -> invest TOP_N_PARTIAL slots
      BEAR    - Price <= EMA100                    -> 0 slots, no new entries
    """
    trend_ticker = next(
        (t for t in [CONFIG.REGIME_TICKER] + CONFIG.REGIME_FALLBACKS
         if t in prices.columns), None
    )

    if trend_ticker is None:
        print("  [warn] No Nifty 500 proxy found; defaulting to BULL regime")
        return {
            "regime_ok"   : True,
            "label"       : "BULL",
            "active_slots": CONFIG.TOP_N,
            "trend_ok"    : True,
            "nifty_price" : np.nan,
            "nifty_ema_50": np.nan,
            "nifty_ema_100": np.nan,
            "trend_ticker": "N/A",
        }

    s = prices[trend_ticker].dropna()
    if len(s) < CONFIG.TREND_EMA_WINDOW:
        return {
            "regime_ok"   : True,
            "label"       : "BULL",
            "active_slots": CONFIG.TOP_N,
            "trend_ok"    : True,
            "nifty_price" : float(s.iloc[-1]),
            "nifty_ema_50": np.nan,
            "nifty_ema_100": np.nan,
            "trend_ticker": trend_ticker,
        }

    nifty_price = float(s.iloc[-1])
    nifty_ema_50  = float(s.ewm(span=CONFIG.TREND_FAST_EMA_WINDOW, adjust=False).mean().iloc[-1])
    nifty_ema_100 = float(s.ewm(span=CONFIG.TREND_EMA_WINDOW, adjust=False).mean().iloc[-1])

    if nifty_ema_50 > nifty_ema_100 and nifty_price > nifty_ema_50:
        label        = "BULL"
        active_slots = CONFIG.TOP_N
    elif nifty_price > nifty_ema_100:
        label        = "PARTIAL"
        active_slots = CONFIG.TOP_N_PARTIAL
    else:
        label        = "BEAR"
        active_slots = 0

    return {
        "regime_ok"   : active_slots == CONFIG.TOP_N,
        "label"       : label,
        "active_slots": active_slots,
        "trend_ok"    : active_slots > 0,
        "nifty_price" : nifty_price,
        "nifty_ema_50": nifty_ema_50,
        "nifty_ema_100": nifty_ema_100,
        "trend_ticker": trend_ticker,
    }


# =========================================================
# 4. WEEKLY RANKING WITH MOM_ACCEL FILTERS
# =========================================================
def build_weekly_ranking(meta: pd.DataFrame, prices: pd.DataFrame) -> pd.DataFrame:
    """
    Build ranking with:
    1. 52-week high filter (<=25% drawdown)
    2. MOM_ACCEL > 0 filter
    3. Ranked by Weighted Sharpe (50% 6M + 50% 3M)
    """
    records = []

    for _, row in meta.iterrows():
        ticker = row["TICKER"]
        if ticker not in prices.columns:
            continue

        s = prices[ticker]
        close = float(s.iloc[-1]) if len(s) > 0 else np.nan
        high_52w = float(s.tail(252).max()) if len(s) > 0 else np.nan

        # Compute Sharpe ratios
        sh6 = sharpe_score(s, CONFIG.WINDOW_6M)
        sh3 = sharpe_score(s, CONFIG.WINDOW_3M)
        sh1 = sharpe_score(s, CONFIG.WINDOW_1M)

        # Compute MOM_ACCEL: 0.4*(Z_1m+Z_3m) - 0.2*Z_6m (Z-scores computed later)
        accel_sh1 = sharpe_score(s, CONFIG.WINDOW_1M, daily_rf=0.0)
        accel_sh3 = sharpe_score(s, CONFIG.WINDOW_3M, daily_rf=0.0)
        accel_sh6 = sharpe_score(s, CONFIG.WINDOW_6M, daily_rf=0.0)

        if np.isnan(accel_sh1) or np.isnan(accel_sh3):
            mom_accel_raw = np.nan
        else:
            accel_sh6_eff = 0.0 if np.isnan(accel_sh6) else accel_sh6
            mom_accel_raw = (accel_sh1 + accel_sh3) - accel_sh6_eff

        # Filter 1: 52-week high proximity
        if pd.notna(close) and pd.notna(high_52w) and high_52w > 0:
            pct_from_high = (high_52w - close) / high_52w
            high_pass = pct_from_high <= CONFIG.MAX_DRAWDOWN_FROM_HIGH
        else:
            pct_from_high = np.nan
            high_pass = True

        records.append({
            "TICKER"        : ticker,
            "ETF_NAME"      : row["ETF_NAME"],
            "SECTOR"        : classify_sector(row["ETF_NAME"], ticker),
            "CLOSE"         : close,
            "52WK_HIGH"     : high_52w,
            "PCT_FROM_HIGH" : pct_from_high * 100 if not np.isnan(pct_from_high) else np.nan,
            "SHARPE_6M"     : sh6,
            "SHARPE_3M"     : sh3,
            "SHARPE_1M"     : sh1,
            "MOM_ACCEL_RAW" : mom_accel_raw,
            "SCREEN_PASS_HIGH": high_pass,
        })

    df = pd.DataFrame(records)

    # Z-score Sharpe ratios cross-sectionally
    def _zscore(series: pd.Series) -> pd.Series:
        mu  = series.mean()
        sig = series.std()
        if sig == 0 or np.isnan(sig):
            return pd.Series(0.0, index=series.index)
        return (series - mu) / sig

    # Z-score for investable (pass 52-week high filter)
    inv_mask = df["SCREEN_PASS_HIGH"]
    df["_Z6_INV"] = np.nan
    df["_Z3_INV"] = np.nan
    df["_Z1_INV"] = np.nan
    df["_MOM_ACCEL_INV"] = np.nan

    if inv_mask.sum() > 0:
        df.loc[inv_mask, "_Z6_INV"] = _zscore(df.loc[inv_mask, "SHARPE_6M"])
        df.loc[inv_mask, "_Z3_INV"] = _zscore(df.loc[inv_mask, "SHARPE_3M"])
        df.loc[inv_mask, "_Z1_INV"] = _zscore(df.loc[inv_mask, "SHARPE_1M"])
        df.loc[inv_mask, "_MOM_ACCEL_INV"] = _zscore(df.loc[inv_mask, "MOM_ACCEL_RAW"])

    # Weighted Sharpe for ranking (50% 6M + 50% 3M)
    z6i = df["_Z6_INV"]; z3i = df["_Z3_INV"]
    both_i = z6i.notna() & z3i.notna()
    only6i = z6i.notna() & z3i.isna()
    only3i = z6i.isna() & z3i.notna()

    df["_WTD_SHARPE_INV"] = np.nan
    df.loc[both_i, "_WTD_SHARPE_INV"] = CONFIG.SHARPE_W6M * z6i[both_i] + CONFIG.SHARPE_W3M * z3i[both_i]
    df.loc[only6i, "_WTD_SHARPE_INV"] = z6i[only6i]
    df.loc[only3i, "_WTD_SHARPE_INV"] = z3i[only3i]

    # Ranking by Weighted Sharpe
    df["RANK_INV_WEIGHTED"] = df["_WTD_SHARPE_INV"].rank(ascending=False, na_option="bottom").astype(int)

    # Apply MOM_ACCEL filter: must be > 0
    df["MOM_ACCEL_PASS"] = df["_MOM_ACCEL_INV"] > 0

    # Combined screen: both filters must pass
    df["SCREEN_PASS"] = df["SCREEN_PASS_HIGH"] & df["MOM_ACCEL_PASS"]

    # Recompute rank on final screened ETFs
    screened = df[df["SCREEN_PASS"]].copy()
    if len(screened) > 0:
        screened["RANK_FINAL"] = screened["_WTD_SHARPE_INV"].rank(ascending=False).astype(int)
        df = df.merge(screened[["TICKER", "RANK_FINAL"]], on="TICKER", how="left")
    else:
        df["RANK_FINAL"] = np.nan

    df["RANK_FINAL"] = df["RANK_FINAL"].fillna(0).astype(int)

    # Drop working columns
    df = df.drop(columns=["_Z6_INV", "_Z3_INV", "_Z1_INV", "_MOM_ACCEL_INV", "_WTD_SHARPE_INV", "SCREEN_PASS_HIGH", "MOM_ACCEL_PASS", "MOM_ACCEL_RAW"], errors="ignore")

    # Sort by final rank
    df["_sort"] = df["RANK_FINAL"].replace(0, 9999)
    df = df.sort_values(["_sort", "RANK_INV_WEIGHTED"]).drop(columns="_sort").reset_index(drop=True)

    return df


# =========================================================
# 5. HOLDINGS & EXIT LOGIC
# =========================================================
def load_holdings(filepath: str = "holdings_log_weekly.json") -> dict:
    """Load weekly holdings log"""
    try:
        with open(filepath) as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_holdings(holdings: dict, filepath: str = "holdings_log_weekly.json"):
    """Save weekly holdings log"""
    with open(filepath, "w") as f:
        json.dump(holdings, f, indent=2)

def check_exit_conditions(holdings: dict, ranking_df: pd.DataFrame, current_date: str) -> dict:
    """
    Check exit conditions for each holding:
    1. 5% TSL from running peak since entry
    2. Rank outside top 20 across investable ETFs
    Returns dict of exits: {ticker: reason}
    """
    exits = {}

    for ticker, position in holdings.items():
        if position["status"] != "active":
            continue

        # Get current rank
        current_rank_row = ranking_df[ranking_df["TICKER"] == ticker]
        if current_rank_row.empty:
            exits[ticker] = "No longer in ranking universe"
            continue

        current_rank = int(current_rank_row["RANK_FINAL"].iloc[0])

        # Check 1: Rank outside top 20
        if current_rank > CONFIG.RANK_RETENTION:
            exits[ticker] = f"Rank {current_rank} outside top {CONFIG.RANK_RETENTION}"
            continue

        # Check 2: 5% TSL
        entry_price = position["entry_price"]
        running_peak = position.get("running_peak", entry_price)
        current_close = current_rank_row["CLOSE"].iloc[0]

        if current_close < running_peak * (1 - CONFIG.TSL_THRESHOLD):
            drawdown_pct = (running_peak - current_close) / running_peak * 100
            exits[ticker] = f"5% TSL hit: {drawdown_pct:.2f}% drawdown from peak {running_peak:.2f}"
            continue

        # Update running peak if current close is higher
        if current_close > running_peak:
            position["running_peak"] = current_close

    return exits


# =========================================================
# 6. MAIN WEEKLY REBALANCE
# =========================================================
def weekly_rebalance(current_date: str = None) -> dict:
    """
    Execute weekly rebalance on Monday:
    1. Load data and compute scores
    2. Check exit conditions
    3. Execute exits
    4. Check regime
    5. Select top 5 (or top 3 in PARTIAL) with sector cap
    6. Update holdings
    """
    if current_date is None:
        current_date = datetime.now().strftime("%Y-%m-%d")

    # Check if Monday
    day_of_week = pd.Timestamp(current_date).dayofweek
    if day_of_week != 0:  # 0 = Monday
        print(f"[info] {current_date} is not a Monday. Skipping rebalance.")
        return {"status": "skipped", "reason": "Not a Monday"}

    print(f"\n{'='*70}")
    print(f"WEEKLY REBALANCE: {current_date}")
    print(f"{'='*70}\n")

    # Load data
    meta, prices = load_etf_data(CONFIG.INPUT_FILE)

    # Build ranking
    print("Building ranking with MOM_ACCEL filters...")
    ranking_df = build_weekly_ranking(meta, prices)

    # Check regime
    regime = regime_status(prices)
    print(f"\nREGIME: {regime['label']} ({regime['active_slots']} active slots)")

    # Load current holdings
    holdings = load_holdings()
    week_key = current_date

    # Check exits
    print(f"\nChecking exit conditions...")
    exits = check_exit_conditions(holdings, ranking_df, current_date)

    for ticker, reason in exits.items():
        print(f"  EXIT: {ticker} - {reason}")
        if ticker in holdings:
            holdings[ticker]["status"] = "exited"
            holdings[ticker]["exit_date"] = current_date
            holdings[ticker]["exit_reason"] = reason

    # If BEAR regime, no new entries
    active_slots = regime["active_slots"]
    if active_slots == 0:
        print(f"\nREGIME = BEAR: No new entries (only monitoring existing)")
        save_holdings(holdings)
        return {"status": "bear_regime", "exits": exits}

    # Select new positions
    active_holdings = {t: p for t, p in holdings.items() if p["status"] == "active"}
    candidates = ranking_df[ranking_df["SCREEN_PASS"]].copy()

    # Keep existing if rank <= 20
    new_positions = {}
    sector_count = {}

    # First, add existing active positions if still ranked
    for ticker in active_holdings:
        ticker_row = candidates[candidates["TICKER"] == ticker]
        if not ticker_row.empty:
            rank = int(ticker_row["RANK_FINAL"].iloc[0])
            if rank <= CONFIG.RANK_RETENTION:
                sector = ticker_row["SECTOR"].iloc[0]
                if sector_count.get(sector, 0) < CONFIG.SECTOR_CAP:
                    new_positions[ticker] = holdings[ticker]
                    sector_count[sector] = sector_count.get(sector, 0) + 1

    # Add new candidates to reach TOP_N
    for _, row in candidates.iterrows():
        if len(new_positions) >= active_slots:
            break

        ticker = row["TICKER"]
        if ticker in new_positions:
            continue

        sector = row["SECTOR"]
        if sector_count.get(sector, 0) >= CONFIG.SECTOR_CAP:
            continue

        entry_price = row["CLOSE"]
        new_positions[ticker] = {
            "ticker": ticker,
            "entry_date": current_date,
            "entry_price": entry_price,
            "running_peak": entry_price,
            "current_price": entry_price,
            "status": "active",
            "sector": sector,
        }
        sector_count[sector] = sector_count.get(sector, 0) + 1

    # Fill remaining with cash if needed
    for ticker in active_holdings:
        if len(new_positions) < active_slots and ticker not in new_positions:
            new_positions["CASH"] = {"ticker": "CASH", "status": "cash", "weight": (active_slots - len(new_positions)) / active_slots}

    # Update holdings
    holdings = new_positions
    save_holdings(holdings)

    print(f"\nNew allocation ({len(new_positions)} positions):")
    for ticker, pos in new_positions.items():
        print(f"  {ticker}: sector={pos.get('sector', 'CASH')}, entry={pos.get('entry_price', 'N/A')}")

    return {"status": "success", "exits": exits, "new_positions": len(new_positions)}


if __name__ == "__main__":
    result = weekly_rebalance()
    print(f"\n[result] {result}")
