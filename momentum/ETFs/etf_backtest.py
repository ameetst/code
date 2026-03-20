"""
ETF Dual Momentum Strategy — Backtest Engine
=============================================
Walk-forward backtest: at each month-end rebalance date, scores and regime
are computed using ONLY price data available up to that date (no lookahead).
Returns are measured from that month-end close to the next month-end close.

Usage:
    python etf_backtest.py                    # uses ETF.xlsx in same folder
    python etf_backtest.py my_data.xlsx       # custom input file

Output:
    etf_backtest_results.xlsx                 # results workbook

Modes:
    Simplified  — runs on any data length, even 1 year (few data points)
    Full        — same engine, automatically uses all available history

Requires etf_momentum_ranking.py in the same folder (imports scoring logic).
"""

from __future__ import annotations
import sys
import importlib.util
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Import scoring functions from main script ─────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
_spec = importlib.util.spec_from_file_location(
    "etf_main", SCRIPT_DIR / "etf_momentum_ranking.py"
)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

load_etf_data    = _mod.load_etf_data
build_ranking    = _mod.build_ranking
build_allocation = _mod.build_allocation
regime_status    = _mod.regime_status
CONFIG           = _mod.CONFIG


# =========================================================
# BACKTEST CONFIG
# =========================================================
class BT_CONFIG:
    INPUT_FILE    = "ETF.xlsx"
    OUTPUT_FILE   = "etf_backtest_results.xlsx"

    # Transaction cost per trade (one-way): brokerage + STT + impact
    # 0.1% = 10 bps is a reasonable estimate for liquid Indian ETFs
    TRANSACTION_COST = 0.001

    # Risk-free rate used for cash slots and Sharpe (annualised)
    RISK_FREE_ANNUAL  = 0.07

    # Benchmark ticker — must exist in the price data
    BENCHMARK_TICKER  = "MONIFTY500"
    BENCHMARK_FALLBACKS = ["BSE500IETF", "NIFTYBEES"]

    # Minimum trading days of warmup before first signal
    # Must be >= WINDOW_6M (126) to have valid 6M scores
    WARMUP_DAYS = 130


# =========================================================
# 1. REBALANCE DATE GENERATION
# =========================================================
def get_rebalance_dates(price_index: pd.DatetimeIndex) -> list[pd.Timestamp]:
    """
    Return the last available trading day of each calendar month
    in the price series, excluding the warmup period.
    """
    dummy = pd.Series(1, index=price_index)
    all_month_ends = dummy.resample("ME").last().dropna().index.tolist()

    # Drop the very last month-end — need a subsequent period to measure returns
    return all_month_ends[:-1]


# =========================================================
# 2. SINGLE REBALANCE STEP
# =========================================================
def score_at_date(
    meta: pd.DataFrame,
    prices: pd.DataFrame,
    as_of: pd.Timestamp,
) -> tuple[pd.DataFrame, dict, pd.DataFrame]:
    """
    Slice prices up to `as_of`, run full scoring + regime + allocation.
    Returns (ranking_df, regime_dict, allocation_df).
    """
    prices_slice = prices.loc[:as_of].copy()
    regime  = regime_status(prices_slice)
    ranking = build_ranking(meta, prices_slice)
    alloc   = build_allocation(ranking, regime)
    return ranking, regime, alloc


# =========================================================
# 3. PERIOD RETURN CALCULATION
# =========================================================
def period_return(
    allocation: pd.DataFrame,
    prices: pd.DataFrame,
    date_start: pd.Timestamp,
    date_end: pd.Timestamp,
    prev_allocation: pd.DataFrame | None,
) -> tuple[float, float]:
    """
    Compute weighted portfolio return from date_start to date_end.
    Cash slots earn daily risk-free rate.
    Transaction costs deducted on changed positions.

    Returns (gross_return, net_return) as decimal fractions.
    """
    daily_rf = BT_CONFIG.RISK_FREE_ANNUAL / 252

    # Price slice for the period
    period_prices = prices.loc[date_start:date_end]
    if len(period_prices) < 2:
        return 0.0, 0.0

    trading_days = len(period_prices) - 1

    # Portfolio return
    gross_ret = 0.0
    for _, row in allocation.iterrows():
        ticker = row["TICKER"]
        weight = float(row["WEIGHT"])

        if ticker == "CASH":
            # Cash earns risk-free for the period
            slot_ret = (1 + daily_rf) ** trading_days - 1
        else:
            if ticker not in period_prices.columns:
                slot_ret = 0.0
            else:
                p_start = period_prices[ticker].dropna().iloc[0]
                p_end   = period_prices[ticker].dropna().iloc[-1]
                slot_ret = (p_end / p_start) - 1 if p_start > 0 else 0.0

        gross_ret += weight * slot_ret

    # Transaction costs — charged on positions that changed
    tc = 0.0
    if prev_allocation is not None:
        prev_tickers = set(
            r["TICKER"] for _, r in prev_allocation.iterrows()
            if r["TICKER"] != "CASH"
        )
        curr_tickers = set(
            r["TICKER"] for _, r in allocation.iterrows()
            if r["TICKER"] != "CASH"
        )
        # Buys + sells (each position change costs one-way)
        changed = prev_tickers.symmetric_difference(curr_tickers)
        tc = len(changed) * CONFIG.TOP_N**-1 * BT_CONFIG.TRANSACTION_COST

    net_ret = gross_ret - tc
    return gross_ret, net_ret


# =========================================================
# 4. BENCHMARK RETURN
# =========================================================
def benchmark_return(
    prices: pd.DataFrame,
    date_start: pd.Timestamp,
    date_end: pd.Timestamp,
) -> float:
    """Buy-and-hold return of the benchmark over the period."""
    ticker = next(
        (t for t in [BT_CONFIG.BENCHMARK_TICKER] + BT_CONFIG.BENCHMARK_FALLBACKS
         if t in prices.columns), None
    )
    if ticker is None:
        return np.nan

    p = prices[ticker].dropna()
    p_start = p.loc[:date_start].iloc[-1] if len(p.loc[:date_start]) else np.nan
    p_end   = p.loc[:date_end].iloc[-1]   if len(p.loc[:date_end])   else np.nan

    if np.isnan(p_start) or np.isnan(p_end) or p_start == 0:
        return np.nan
    return (p_end / p_start) - 1


# =========================================================
# 5. PERFORMANCE METRICS
# =========================================================
def compute_metrics(monthly_returns: list[float], label: str) -> dict:
    """Compute CAGR and annualised Sharpe from a list of monthly returns."""
    r = np.array(monthly_returns, dtype=float)
    r = r[~np.isnan(r)]

    if len(r) == 0:
        return {"label": label, "months": 0, "cagr": np.nan,
                "sharpe": np.nan, "total_return": np.nan}

    total_ret  = np.prod(1 + r) - 1
    months     = len(r)
    years      = months / 12
    cagr       = (1 + total_ret) ** (1 / years) - 1 if years > 0 else np.nan

    monthly_rf = (1 + BT_CONFIG.RISK_FREE_ANNUAL) ** (1 / 12) - 1
    excess     = r - monthly_rf
    sharpe     = (excess.mean() / excess.std() * np.sqrt(12)
                  if excess.std() > 0 else np.nan)

    return {
        "label"       : label,
        "months"      : months,
        "total_return": total_ret,
        "cagr"        : cagr,
        "sharpe"      : sharpe,
    }


# =========================================================
# 6. MAIN BACKTEST LOOP
# =========================================================
def run_backtest(filepath: str) -> pd.DataFrame:
    print(f"[load]  {filepath}")
    meta, prices = load_etf_data(filepath)
    print(f"        {len(meta)} ETFs | {len(prices)} trading days "
          f"({prices.index[0].date()} -> {prices.index[-1].date()})")

    rebal_dates = get_rebalance_dates(prices.index)
    print(f"        {len(rebal_dates)} potential rebalance dates")

    # Filter to dates with sufficient warmup
    first_valid = prices.index[BT_CONFIG.WARMUP_DAYS - 1]
    rebal_dates = [d for d in rebal_dates if d >= first_valid]
    print(f"        {len(rebal_dates)} rebalance dates after {BT_CONFIG.WARMUP_DAYS}-day warmup\n")

    if len(rebal_dates) < 2:
        print("[warn]  Fewer than 2 valid rebalance dates — not enough data.")
        print("        Need at least WARMUP_DAYS + 1 month of data.")
        return pd.DataFrame()

    records       = []
    prev_alloc    = None
    prev_regime   = None

    for i, date in enumerate(rebal_dates):
        next_date = rebal_dates[i + 1] if i + 1 < len(rebal_dates) else None
        if next_date is None:
            break

        print(f"  [{i+1:>2}/{len(rebal_dates)-1}] {date.date()} -> {next_date.date()} ...",
              end=" ", flush=True)

        # Score at this date (no lookahead)
        ranking, regime, alloc = score_at_date(meta, prices, date)

        # Tickers held this month
        held = [r["TICKER"] for _, r in alloc.iterrows() if r["TICKER"] != "CASH"]

        # Period returns
        gross, net = period_return(alloc, prices, date, next_date, prev_alloc)
        bench      = benchmark_return(prices, date, next_date)

        print(f"Portfolio {net:+.2%}  Benchmark {bench:+.2%}  Regime: {regime['label'][:12]}")

        records.append({
            "DATE_START"    : date,
            "DATE_END"      : next_date,
            "REGIME"        : regime["label"],
            "ACTIVE_SLOTS"  : regime["active_slots"],
            "HOLDINGS"      : ", ".join(held) if held else "CASH",
            "GROSS_RETURN"  : gross,
            "NET_RETURN"    : net,
            "BENCHMARK_RET" : bench,
            "EXCESS_RETURN" : net - bench if not np.isnan(bench) else np.nan,
        })

        prev_alloc  = alloc
        prev_regime = regime

    return pd.DataFrame(records)


# =========================================================
# 7. EXCEL OUTPUT
# =========================================================
def _brd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _h(ws, row, col, val, bg="1F4E79", fg="FFFFFF", sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, size=sz, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _brd()
    return c

def _d(ws, row, col, val, bg="FFFFFF", fmt=None, bold=False, align="center"):
    if isinstance(val, float) and np.isnan(val): val = None
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9, bold=bold)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _brd()
    if fmt: c.number_format = fmt
    return c


def save_results(results: pd.DataFrame, out_path: str):
    wb   = Workbook()
    NAVY  = "1F4E79"
    GREEN = "E2EFDA"
    RED   = "FCE4D6"
    GREY  = "F2F2F2"
    YELLOW= "FFF2CC"
    DKGRN = "375623"
    DKRED = "C00000"

    # ── Summary metrics ───────────────────────────────────────────
    port_metrics  = compute_metrics(results["NET_RETURN"].tolist(),   "Strategy (Net)")
    gross_metrics = compute_metrics(results["GROSS_RETURN"].tolist(), "Strategy (Gross)")
    bench_metrics = compute_metrics(results["BENCHMARK_RET"].tolist(),"Nifty 500 B&H")

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "ETF Dual Momentum Strategy — Backtest Results"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Meta info
    meta_rows = [
        ("Data range",        f"{results['DATE_START'].min().date()} → {results['DATE_END'].max().date()}"),
        ("Rebalance periods", f"{len(results)} months"),
        ("Benchmark",         BT_CONFIG.BENCHMARK_TICKER),
        ("Transaction cost",  f"{BT_CONFIG.TRANSACTION_COST*100:.1f}% per one-way trade"),
        ("Risk-free rate",    f"{BT_CONFIG.RISK_FREE_ANNUAL*100:.1f}% p.a."),
        ("Warmup required",   f"{BT_CONFIG.WARMUP_DAYS} trading days"),
    ]
    for ri, (lbl, val) in enumerate(meta_rows, start=2):
        _d(ws, ri, 1, lbl, bg=GREY, bold=True, align="left")
        _d(ws, ri, 2, val, bg=GREY, align="left")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35

    row = len(meta_rows) + 3

    # Performance table
    _h(ws, row,   1, "Metric",            bg=NAVY)
    _h(ws, row,   2, "Strategy\n(Gross)", bg=NAVY)
    _h(ws, row,   3, "Strategy\n(Net)",   bg=NAVY)
    _h(ws, row,   4, "Nifty 500\nB&H",   bg=NAVY)
    ws.row_dimensions[row].height = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    row += 1

    perf_rows = [
        ("Total Return",  "total_return", "0.0%"),
        ("CAGR",          "cagr",         "0.0%"),
        ("Annualised Sharpe", "sharpe",   "0.00"),
    ]

    for lbl, key, fmt in perf_rows:
        gv = gross_metrics.get(key, np.nan)
        nv = port_metrics.get(key, np.nan)
        bv = bench_metrics.get(key, np.nan)

        # Colour net vs benchmark
        n_bg = GREEN if (not np.isnan(nv) and not np.isnan(bv) and nv > bv) else RED

        _d(ws, row, 1, lbl,  bg=GREY, bold=True, align="left")
        _d(ws, row, 2, gv,   bg=GREY, fmt=fmt)
        _d(ws, row, 3, nv,   bg=n_bg, fmt=fmt, bold=True)
        _d(ws, row, 4, bv,   bg=GREY, fmt=fmt)
        row += 1

    # Excess return
    net_cagr   = port_metrics.get("cagr",  np.nan)
    bench_cagr = bench_metrics.get("cagr", np.nan)
    if not np.isnan(net_cagr) and not np.isnan(bench_cagr):
        excess = net_cagr - bench_cagr
        ex_bg  = GREEN if excess > 0 else RED
        _d(ws, row, 1, "Excess CAGR (vs Nifty 500)", bg=GREY, bold=True, align="left")
        _d(ws, row, 2, "",    bg=GREY)
        _d(ws, row, 3, excess, bg=ex_bg, fmt="+0.0%;-0.0%", bold=True)
        _d(ws, row, 4, "",    bg=GREY)
        row += 1

    # Regime breakdown
    row += 1
    _h(ws, row, 1, "Regime Distribution", bg=NAVY)
    _h(ws, row, 2, "Months",              bg=NAVY)
    _h(ws, row, 3, "Avg Net Return",      bg=NAVY)
    _h(ws, row, 4, "vs Benchmark",        bg=NAVY)
    ws.row_dimensions[row].height = 22
    row += 1

    for regime_label, grp in results.groupby("REGIME"):
        avg_net   = grp["NET_RETURN"].mean()
        avg_bench = grp["BENCHMARK_RET"].mean()
        r_bg      = ("E2EFDA" if "BULL" in regime_label else
                     "FCE4D6" if "BEAR" in regime_label else "FFF2CC")
        _d(ws, row, 1, regime_label, bg=r_bg, align="left")
        _d(ws, row, 2, len(grp),     bg=r_bg, fmt="0")
        _d(ws, row, 3, avg_net,      bg=r_bg, fmt="0.0%")
        _d(ws, row, 4, avg_net - avg_bench if not np.isnan(avg_bench) else None,
           bg=r_bg, fmt="+0.0%;-0.0%")
        row += 1

    # ── Sheet 2: Monthly Returns ──────────────────────────────────
    wr = wb.create_sheet("Monthly Returns")

    COLS = [
        ("Period Start",    14, "@"),
        ("Period End",      14, "@"),
        ("Regime",          22, "@"),
        ("Active\nSlots",    9, "0"),
        ("Holdings",        45, "@"),
        ("Gross\nReturn",   12, "0.0%"),
        ("Net\nReturn",     12, "0.0%"),
        ("Nifty500\nReturn",12, "0.0%"),
        ("Excess\nReturn",  12, "+0.0%;-0.0%"),
    ]

    wr.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = wr["A1"]
    c.value     = "Monthly Returns — Strategy vs Nifty 500 Benchmark"
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    wr.row_dimensions[1].height = 20

    for ci, (hdr, width, _) in enumerate(COLS, 1):
        _h(wr, 2, ci, hdr)
        wr.column_dimensions[get_column_letter(ci)].width = width
    wr.row_dimensions[2].height = 28

    # Running cumulative returns
    port_cum  = 1.0
    bench_cum = 1.0

    for ri, (_, row_data) in enumerate(results.iterrows(), start=3):
        net   = row_data["NET_RETURN"]
        bench = row_data["BENCHMARK_RET"]
        gross = row_data["GROSS_RETURN"]
        excss = row_data["EXCESS_RETURN"]
        reg   = row_data["REGIME"]

        port_cum  *= (1 + net)   if not np.isnan(net)   else 1
        bench_cum *= (1 + bench) if not np.isnan(bench) else 1

        row_bg = ("E2EFDA" if "BULL" in reg else
                  "FCE4D6" if "BEAR" in reg else "FFF2CC")

        # Colour the return cells based on outperformance
        ret_bg = (GREEN if not np.isnan(net) and not np.isnan(bench) and net > bench
                  else RED if not np.isnan(net) and not np.isnan(bench) and net < bench
                  else GREY)

        vals = [
            row_data["DATE_START"].strftime("%Y-%m-%d"),
            row_data["DATE_END"].strftime("%Y-%m-%d"),
            reg,
            int(row_data["ACTIVE_SLOTS"]),
            row_data["HOLDINGS"],
            gross, net, bench, excss
        ]
        fmts = [None, None, None, "0", None, "0.0%", "0.0%", "0.0%", "+0.0%;-0.0%"]
        bgs  = [row_bg, row_bg, row_bg, row_bg, row_bg,
                row_bg, ret_bg, row_bg, ret_bg]

        for ci, (v, f, bg) in enumerate(zip(vals, fmts, bgs), 1):
            _d(wr, ri, ci, v, bg=bg, fmt=f,
               bold=(ci in [6,7,8,9]), align="left" if ci == 5 else "center")
        wr.row_dimensions[ri].height = 14

    # Totals row
    tr = len(results) + 3
    _d(wr, tr, 1, "CUMULATIVE", bg=NAVY, bold=True, align="left")
    wr.cell(row=tr, column=1).font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    _d(wr, tr, 5, "", bg=NAVY)
    _d(wr, tr, 6, port_cum  - 1, bg=GREEN if port_cum > bench_cum else RED,
       fmt="0.0%", bold=True)
    _d(wr, tr, 7, port_cum  - 1, bg=GREEN if port_cum > bench_cum else RED,
       fmt="0.0%", bold=True)
    _d(wr, tr, 8, bench_cum - 1, bg=GREY, fmt="0.0%", bold=True)
    _d(wr, tr, 9, (port_cum - bench_cum), bg=GREY, fmt="+0.0%;-0.0%", bold=True)
    wr.row_dimensions[tr].height = 18
    wr.freeze_panes = "A3"

    wb.save(out_path)
    print(f"\n[saved] -> {Path(out_path).resolve()}")

    # Print summary to console
    print("\n" + "=" * 60)
    print("BACKTEST SUMMARY")
    print("=" * 60)
    print(f"  Period : {results['DATE_START'].min().date()} → {results['DATE_END'].max().date()}")
    print(f"  Months : {len(results)}")
    print(f"\n  {'Metric':<22} {'Strategy':>12} {'Nifty 500':>12} {'Excess':>10}")
    print("  " + "-" * 58)
    for lbl, key, fmt_str in perf_rows:
        nv = port_metrics.get(key, np.nan)
        bv = bench_metrics.get(key, np.nan)
        ex = nv - bv if not np.isnan(nv) and not np.isnan(bv) else np.nan
        nv_s = f"{nv:{fmt_str}}" if not np.isnan(nv) else "N/A"
        bv_s = f"{bv:{fmt_str}}" if not np.isnan(bv) else "N/A"
        ex_s = f"{ex:+{fmt_str}}" if not np.isnan(ex) else "N/A"
        print(f"  {lbl:<22} {nv_s:>12} {bv_s:>12} {ex_s:>10}")
    print("=" * 60)


# =========================================================
# 8. MAIN
# =========================================================
if __name__ == "__main__":
    fp  = sys.argv[1] if len(sys.argv) > 1 else str(SCRIPT_DIR / BT_CONFIG.INPUT_FILE)
    out = sys.argv[2] if len(sys.argv) > 2 else str(SCRIPT_DIR / BT_CONFIG.OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("ETF DUAL MOMENTUM — BACKTEST ENGINE")
    print("=" * 60)
    n_months = None
    try:
        import pandas as pd
        _, px = load_etf_data(fp)
        n_months = len(pd.Series(1, index=px.index).resample("ME").last().dropna()) - 1
    except Exception:
        pass

    if n_months and n_months < 8:
        print(f"\n[warn]  Only ~{n_months} months of data detected.")
        print("        Backtest will run but results will have low statistical power.")
        print("        For meaningful results, use 3-5 years of data.\n")
    elif n_months:
        print(f"\n[info]  ~{n_months} months of data detected — proceeding.\n")

    results = run_backtest(fp)

    if results.empty:
        print("\n[error] Backtest produced no results. Check data length and format.")
        sys.exit(1)

    save_results(results, out)