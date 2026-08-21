"""
etf_update_prices_dhan.py
===========================
Dhan HQ v2 fetcher for the ETF Momentum strategy's price data, using the
shared `dhandata` library. Reads the ETF universe (ETF_NAME/TICKER) from
ETF.xlsx and writes a NEW workbook (ETF_dhan.xlsx by default) in the same
single-sheet "DATA" shape etf_momentum_ranking_v2.py already expects
(col A = ETF_NAME, col B = TICKER, col C onward = daily close price grid)
-- so you can point the ranking script at it directly. No changes needed
to etf_momentum_ranking_v2.py itself -- it already accepts a file path as
its first CLI argument.

This does NOT touch ETF.xlsx or the AMFI NAV pipeline (download_amfi_nav.py /
AMFI_NAV_History.xlsx) -- those keep working exactly as before. This is a
parallel script, same non-destructive pattern as milt_update_prices_dhan.py:
it only ever *reads* ETF.xlsx (for the ticker list) and writes to a
differently-named output file.

IMPORTANT -- date-header quirk this script deliberately replicates:
etf_momentum_ranking_v2.py's load_etf_data() does NOT read date values from
row 1 -- it reconstructs the date index as "the last N business days
(Mon-Fri) ending TODAY", where N = however many price columns are present
in the file. This script builds its price grid with that exact same
"last N business days ending today" logic, so the columns line up
correctly PROVIDED you run this fetcher and etf_momentum_ranking_v2.py on
the same calendar day. (This constraint already exists for ETF.xlsx today,
via its own TODAY()-based header formula -- it is not something the Dhan
migration introduces.) Cells for calendar business days that weren't
actual NSE trading days (holidays) are left blank; the reader's own
.ffill() already carries the prior close forward across those, same as it
does today for the existing AMFI-sourced file.

Setup (one-time, same as MILT's dhandata setup)
-------------------------------------------------
    cd <path-to>\\dhan_datahq
    pip install -e .
Then make sure dhan_datahq\\.dhan_token_cache.json has a usable token.

Usage
-----
    python etf_update_prices_dhan.py [--tickers-from ETF.xlsx]
        [--output ETF_dhan.xlsx] [--lookback-years 2]

Then, to compare against the live AMFI-sourced file WITHOUT touching any
production state (safe as long as this calendar month already has an entry
in holdings_log_v2.json -- check that first; a brand-new month would
trigger a real full-flush write):

    python etf_momentum_ranking_v2.py ETF_dhan.xlsx etf_rankings_v2_dhan.xlsx

That passes the Dhan file as input and a *different* output report path,
so it never touches etf_rankings_v2.xlsx / holdings_log_v2.json /
ETF_positions_ledger_v2.json / ETF_tradelog_v2.json.

Known differences vs. the AMFI-NAV-sourced ETF.xlsx
------------------------------------------------------
* Price basis: Dhan gives the ETF's actual traded exchange close; ETF.xlsx
  today carries AMFI's published NAV. These are related but not identical
  (NAV is struck once/day off the fund's underlying basket plus a small
  timing/expense-ratio lag; exchange close is whatever it last traded at
  on NSE). Expect small, usually sub-1%, day-to-day differences -- a
  different cause from MILT's unadjusted-price gap, but the same kind of
  "two legitimate sources, not identical" situation.
* 9 Axis Mutual Fund ETFs are listed on Dhan under a "reversed" symbol
  (e.g. ETF.xlsx's AXISBNKETF -> Dhan's BNKETFAXIS) -- pre-populated in
  ALIAS_MAP below. Verified match rate: 283/292 direct, 9/9 via alias ->
  292/292 (100%) of ETF.xlsx's current universe resolves on Dhan.
* Regime index (^CRSLDX via yfinance, cached in nifty500_cache.csv) and its
  ETF.xlsx-based fallback tickers (MONIFTY500, BSE500IETF, HDFCBSE500,
  NIFTYBEES) are untouched by this script -- that pipeline stays exactly
  as it is today. MONIFTY500 happens to also be one of the 292 ETFs in the
  main universe, so it gets fetched normally here too, but this script
  does not change how the regime filter sources its index level.
"""

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

import dhandata as dh

# ── Config ────────────────────────────────────────────────────────────────
DEFAULT_TICKERS_FROM = "ETF.xlsx"
DEFAULT_OUTPUT = "ETF_dhan.xlsx"
DEFAULT_LOOKBACK_YEARS = 2
BDAYS_PER_YEAR = 261   # matches ETF.xlsx's own ~261-column-per-year convention

# ETF.xlsx ticker -> Dhan UNDERLYING_SYMBOL, for the 9 Axis Mutual Fund ETFs
# listed on Dhan under a differently-ordered symbol. Confirmed against the
# real api-scrip-master-detailed.csv (Aug 2026).
ALIAS_MAP = {
    "AXSENSEX":   "SENSEXAXIS",
    "AXISGOLD":   "GOLDAXIS",
    "AXISNIFTY":  "NIFTYAXIS",
    "AXISBNKETF": "BNKETFAXIS",
    "AXISHCETF":  "HEALTHAXIS",
    "AXISCETF":   "CONSUMAXIS",
    "AXISTECETF": "ITAXIS",
    "AXISVALUE":  "VALUEAXIS",
    "AXISILVER":  "SILVERAXIS",
}


def read_universe(etf_xlsx_path: str) -> list:
    """Read (ETF_NAME, TICKER) pairs from ETF.xlsx's DATA sheet, cols A/B."""
    wb = openpyxl.load_workbook(etf_xlsx_path, data_only=True, read_only=True)
    ws = wb["DATA"]
    pairs = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
        name, ticker = row[0], row[1]
        if ticker:
            pairs.append((str(name).strip(), str(ticker).strip()))
    wb.close()
    return pairs


def main(etf_xlsx_path: str, output_file: str, lookback_years: float):
    print(f"Reading ETF universe from: {etf_xlsx_path}")
    universe = read_universe(etf_xlsx_path)
    tickers = [t for _, t in universe]
    name_by_ticker = {t: n for n, t in universe}
    print(f"  {len(tickers)} ETFs\n")

    today = pd.Timestamp.today().normalize()
    n_cols = int(round(lookback_years * BDAYS_PER_YEAR))
    all_bdays = pd.bdate_range(end=today, periods=n_cols)
    from_date = all_bdays[0].date().isoformat()
    # Dhan's toDate is non-inclusive -- pad by a day so today's bar (once
    # the market's closed) is actually included.
    to_date = (today + pd.Timedelta(days=1)).date().isoformat()
    print(f"Date range: {from_date} -> {today.date().isoformat()} "
          f"({n_cols} business-day columns, matching etf_momentum_ranking_v2.py's "
          f"own date-reconstruction convention)\n")

    print(f"Fetching {len(tickers)} ETFs' daily closes via Dhan "
          f"(single-symbol calls, rate-limited to ~5/sec -- this will take a couple of minutes) ...")
    data, unmatched, failed = dh.get_daily_history_bulk(
        tickers, from_date=from_date, to_date=to_date,
        fields=["close"], alias_map=ALIAS_MAP, progress=True,
    )
    print()

    if unmatched:
        print(f"{len(unmatched)} ticker(s) had no match in the Dhan instrument master:")
        for t in unmatched:
            print(f"   {t}  ({name_by_ticker.get(t, '?')})")
        print("  -> add a mapping to ALIAS_MAP at the top of this script if you find the new symbol.\n")

    if failed:
        print(f"{len(failed)} ticker(s) matched but the API call failed:")
        for t, err in failed.items():
            print(f"   {t}: {err}")
        print()

    if not data:
        print("No data fetched -- aborting before writing an empty workbook.")
        sys.exit(1)

    print(f"Writing {output_file} ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA"
    ws.cell(row=1, column=1, value="ETF_NAME")
    ws.cell(row=1, column=2, value="TICKER")
    for c, d in enumerate(all_bdays, start=3):
        cell = ws.cell(row=1, column=c, value=d.date())
        cell.number_format = "dd-mmm-yy"

    bday_dates = all_bdays.date  # numpy array of datetime.date, same order as columns
    n_ok = 0
    for r, (name, ticker) in enumerate(universe, start=2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=ticker)
        df = data.get(ticker)
        if df is None or "close" not in df.columns or df.empty:
            continue
        n_ok += 1
        series = df["close"].reindex(bday_dates)
        for c, v in enumerate(series.values, start=3):
            if pd.notna(v):
                ws.cell(row=r, column=c, value=round(float(v), 4))

    wb.save(output_file)
    print("Done.")
    print(f"\n{n_ok}/{len(tickers)} ETFs fetched successfully "
          f"({len(unmatched)} unmatched, {len(failed)} API failures).")

    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Fetch daily close prices for the ETF Momentum universe via Dhan HQ v2.")
    parser.add_argument("--tickers-from", type=str, default=DEFAULT_TICKERS_FROM,
                        help=f"Existing ETF.xlsx to read ETF_NAME/TICKER from (default: {DEFAULT_TICKERS_FROM})")
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT,
                        help=f"Output workbook path (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--lookback-years", type=float, default=DEFAULT_LOOKBACK_YEARS,
                        help=f"Years of daily history to fetch (default: {DEFAULT_LOOKBACK_YEARS})")
    args = parser.parse_args()

    if not Path(args.tickers_from).exists():
        print(f"ERROR: {args.tickers_from} not found.")
        sys.exit(1)

    main(args.tickers_from, args.output, args.lookback_years)
