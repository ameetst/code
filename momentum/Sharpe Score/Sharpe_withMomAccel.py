"""
N500 Sharpe Z-Score + Clenow Z-Score + Residual Momentum Ranking
=================================================================
  • 12M/9M/6M/3M annualised Sharpe ratios → Z-score → SHARPE_ALL composite
  • 12M/9M/6M/3M Clenow (AnnSlope×R²)    → Z-score → CLENOW_Z composite
  • 12M/9M/6M/3M residual Sharpe (OLS)   → Z-score → RES_MOM composite
  • Eligibility filter: PCT_FROM_52H >= -25%
  • Ranked by SHARPE_ALL

Usage:  python Sharpe.py path/to/n500.xlsx
"""

import sys, datetime, warnings
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy import stats

warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
FILE         = "n500.xlsx" if len(sys.argv) < 2 else sys.argv[1]
OUTPUT_FILE  = "n500_rankings.xlsx"
RFR_ANNUAL   = 0.07
TRADING_DAYS = 252
TOP_N        = 20
WINDOWS         = {"12M": 252, "9M": 189, "6M": 126, "3M": 63}   # Clenow + ResMom
SHARPE_WINDOWS  = {"12M": 252, "9M": 189, "6M": 126, "3M": 63, "1M": 21}  # Sharpe only

# ── LOAD ──────────────────────────────────────────────────────────────────────
print(f"Loading {FILE} ...")
wb_in    = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
ws       = wb_in["DATA"]
all_rows = list(ws.iter_rows(values_only=True))
wb_in.close()

header       = all_rows[0]
date_indices = [i for i, h in enumerate(header)
                if isinstance(h, (datetime.datetime, datetime.date))]
dates        = [header[i] for i in date_indices]

tickers, price_matrix = [], []
for row in all_rows[1:]:
    if row[0] is None: continue
    px = []
    for i in date_indices:
        v = row[i]
        try:    px.append(float(v) if v and float(v) > 0 else np.nan)
        except: px.append(np.nan)
    tickers.append(str(row[0]).strip())
    price_matrix.append(px)

prices_df = pd.DataFrame(price_matrix, index=tickers, columns=dates)

# ── SEPARATE NIFTY500 BENCHMARK ───────────────────────────────────────────────
nifty_series  = prices_df.loc["NIFTY500"].copy()
stock_tickers = [t for t in tickers if t != "NIFTY500"]
prices_df     = prices_df.loc[stock_tickers]

valid_days = sum(
    1 for di in date_indices
    if any(row[di] and row[di] != 0 for row in all_rows[1:] if row[0])
)
print(f"  {len(prices_df)} stocks  |  {len(dates)} date columns  "
      f"({valid_days} actual trading days)  "
      f"|  {dates[0].strftime('%d-%b-%Y')} → {dates[-1].strftime('%d-%b-%Y')}\n")

# ── SHARPE RATIOS ─────────────────────────────────────────────────────────────
rfr_daily = RFR_ANNUAL / TRADING_DAYS

def sharpe_ratio(series: pd.Series, window: int) -> float:
    px = series.dropna()
    if len(px) < window * 0.90: return np.nan
    px_w     = px if len(px) < window + 1 else px.iloc[-(window + 1):]
    log_rets = np.diff(np.log(px_w.values))
    excess   = log_rets - rfr_daily
    sd       = excess.std(ddof=1)
    if sd < 1e-12: return np.nan
    return (excess.mean() / sd) * np.sqrt(TRADING_DAYS)

print("Computing Sharpe ratios ...")
sharpe_data = {}
for label, window in SHARPE_WINDOWS.items():
    col = [sharpe_ratio(prices_df.loc[t], window) for t in stock_tickers]
    sharpe_data[label] = col
    valid = sum(1 for v in col if not np.isnan(v))
    print(f"  {label} ({window}d): {valid}/{len(stock_tickers)} valid")
sharpe_df = pd.DataFrame(sharpe_data, index=stock_tickers)

# ── SHARPE Z-SCORES → SHARPE_ALL COMPOSITE ─────────────────────────────────────
print("\nZ-scoring Sharpe cross-sectionally ...")
z_df = pd.DataFrame(index=stock_tickers)
for label in SHARPE_WINDOWS:
    s = sharpe_df[label]
    mu, sd = s.mean(), s.std(ddof=1)
    z_df[f"Z_{label}"] = (s - mu) / sd if sd > 0 else 0.0

z_cols            = [f"Z_{l}" for l in WINDOWS]   # COMPOSITE uses 4 core windows only
z_df["COMPOSITE"] = z_df[z_cols].mean(axis=1)
# SHARPE_3: short-term composite (Z_3M + Z_6M) — reflects recent momentum
# MOM_ACCEL: short-term composite minus long-term composite, Z-scored
#   short = mean(Z_1M, Z_3M, Z_6M) — last 1–6 months
#   long  = mean(Z_9M, Z_12M)      — last 9–12 months
#   raw   = short − long  →  positive = momentum accelerating recently
z_df["SHARPE_ST"] = z_df[["Z_1M", "Z_3M", "Z_6M"]].mean(axis=1)   # short-term
z_df["SHARPE_LT"] = z_df[["Z_9M", "Z_12M"]].mean(axis=1)           # long-term
z_df["SHARPE_3"]  = z_df[["Z_12M", "Z_6M", "Z_3M"]].mean(axis=1)  # 3-window EW (excl. 9M, 1M)
accel_raw_series  = z_df["SHARPE_ST"] - z_df["SHARPE_LT"]
mu_acc = accel_raw_series.mean()
sd_acc = accel_raw_series.std(ddof=1)
z_df["MOM_ACCEL"] = (accel_raw_series - mu_acc) / sd_acc if sd_acc > 0 else 0.0
print(f"  MOM_ACCEL  — accelerating (>0): {(z_df['MOM_ACCEL'] > 0).sum()} stocks  |  "
      f"decelerating (<0): {(z_df['MOM_ACCEL'] < 0).sum()} stocks")

# ── CLENOW SCORES — multi-window composite ────────────────────────────────────
# Per window: log(price) = a + b*t  →  ann_slope × R² = raw Clenow score
# Z-score each window cross-sectionally → CLENOW_Z (same structure as SHARPE_ALL)
print("\nComputing multi-window Clenow scores ...")

def clenow_window(series: pd.Series, window: int) -> tuple:
    px = series.dropna()
    if len(px) < window * 0.90: return np.nan, np.nan, np.nan
    n         = min(len(px), window)
    px_w      = px.iloc[-n:].values
    x         = np.arange(n)
    y         = np.log(px_w)
    slope, _, r, _, _ = stats.linregress(x, y)
    r2        = r ** 2
    ann_slope = slope * TRADING_DAYS
    return ann_slope, r2, ann_slope * r2

clenow_slope_data, clenow_r2_data, clenow_raw_data = {}, {}, {}
for label, window in WINDOWS.items():
    slopes, r2s, raws = [], [], []
    for t in stock_tickers:
        sl, r2, raw = clenow_window(prices_df.loc[t], window)
        slopes.append(sl); r2s.append(r2); raws.append(raw)
    clenow_slope_data[f"CL_{label}"] = slopes
    clenow_r2_data[f"CR_{label}"]    = r2s
    clenow_raw_data[f"CS_{label}"]   = raws
    valid = sum(1 for v in raws if not np.isnan(v))
    print(f"  {label} ({window}d): {valid}/{len(stock_tickers)} valid")

clenow_slope_df = pd.DataFrame(clenow_slope_data, index=stock_tickers)
clenow_r2_df    = pd.DataFrame(clenow_r2_data,    index=stock_tickers)
clenow_raw_df   = pd.DataFrame(clenow_raw_data,   index=stock_tickers)

cz_df = pd.DataFrame(index=stock_tickers)
for label in WINDOWS:
    s = clenow_raw_df[f"CS_{label}"]
    mu, sd = s.mean(), s.std(ddof=1)
    cz_df[f"CZ_{label}"] = (s - mu) / sd if sd > 0 else 0.0

cz_cols           = [f"CZ_{l}" for l in WINDOWS]
cz_df["CLENOW_Z"] = cz_df[cz_cols].mean(axis=1)

# ── COMBINE ───────────────────────────────────────────────────────────────────
result = z_df.join(sharpe_df.rename(columns={l: f"S_{l}" for l in SHARPE_WINDOWS}))
result = result.join(clenow_slope_df)
result = result.join(clenow_r2_df)
result = result.join(clenow_raw_df)
result = result.join(cz_df)

# ── NORMALISE COMPOSITE (SHARPE_ALL) ───────────────────────────────────────────
def normalise_composite(v):
    if pd.isna(v):  return np.nan
    if v > 1:       return v + 1
    if v < 0:       return 1.0 / (1.0 - v)
    return v

result["COMPOSITE"]  = result["COMPOSITE"].map(normalise_composite)
result["SHARPE_ALL"]  = result["COMPOSITE"]   # display alias
result["SHARPE_3"]   = result["SHARPE_3"].map(normalise_composite)
result["SHARPE_ST"]  = result["SHARPE_ST"].map(normalise_composite)
result["SHARPE_LT"]  = result["SHARPE_LT"].map(normalise_composite)
result["RANK"]      = result["COMPOSITE"].rank(ascending=False, method="first",
                                                na_option="bottom")
result = result.sort_values("COMPOSITE", ascending=False)

# ── RETURN CONTEXT ────────────────────────────────────────────────────────────
def safe_ret(series, n):
    px = series.dropna()
    return (px.iloc[-1] / px.iloc[-n] - 1) * 100 if len(px) > n else np.nan

ret_data = {t: {
    "1M%":  safe_ret(prices_df.loc[t], 22),
    "3M%":  safe_ret(prices_df.loc[t], 63),
    "12M%": safe_ret(prices_df.loc[t], 245),
} for t in stock_tickers}
result = result.join(pd.DataFrame(ret_data).T)

# ── 52-WEEK HIGH ──────────────────────────────────────────────────────────────
def pct_from_52w_high(series: pd.Series, window: int = 252) -> float:
    px = series.dropna()
    if len(px) < 2: return np.nan
    px_w     = px.iloc[-window:] if len(px) >= window else px
    high_52w = px_w.max()
    last_px  = px.iloc[-1]
    if high_52w <= 0: return np.nan
    return (last_px / high_52w - 1) * 100

print("\nComputing 52-week high proximity ...")
pct_52h = {t: pct_from_52w_high(prices_df.loc[t]) for t in stock_tickers}
result["PCT_FROM_52H"] = pd.Series(pct_52h)

# ── ELIGIBILITY FILTER: PCT_FROM_52H >= -25% only ────────────────────────────
eligible = (result["PCT_FROM_52H"] >= -25)
result["RANK"] = np.nan
result.loc[eligible, "RANK"] = (
    result.loc[eligible, "COMPOSITE"]
    .rank(ascending=False, method="first", na_option="bottom")
)
result = result.sort_values(["RANK", "COMPOSITE"], ascending=[True, False])
print(f"  {eligible.sum()} / {len(result)} stocks eligible (PCT_FROM_52H >= -25%)")

# ── RESIDUAL MOMENTUM ─────────────────────────────────────────────────────────
print("\nComputing residual momentum scores ...")
nifty_log_rets = np.diff(np.log(nifty_series.dropna().values))

def residual_sharpe(stock_series: pd.Series, mkt_rets: np.ndarray,
                    window: int) -> float:
    px = stock_series.dropna()
    if len(px) < window * 0.90: return np.nan
    n        = min(len(px) - 1, window)
    s_rets   = np.diff(np.log(px.iloc[-n-1:].values))
    m_rets   = mkt_rets[-n:]
    if len(s_rets) != len(m_rets) or len(s_rets) < 10: return np.nan
    X        = np.column_stack([np.ones(len(m_rets)), m_rets])
    try:    coeffs, _, _, _ = np.linalg.lstsq(X, s_rets, rcond=None)
    except: return np.nan
    residuals = s_rets - X @ coeffs
    sd = residuals.std(ddof=1)
    if sd < 1e-12: return np.nan
    return (residuals.mean() / sd) * np.sqrt(TRADING_DAYS)

resmom_data = {}
for label, window in WINDOWS.items():
    col   = [residual_sharpe(prices_df.loc[t], nifty_log_rets, window)
             for t in stock_tickers]
    valid = sum(1 for v in col if not np.isnan(v))
    resmom_data[f"RS_{label}"] = col
    print(f"  {label} ({window}d): {valid}/{len(stock_tickers)} valid")

resmom_df = pd.DataFrame(resmom_data, index=stock_tickers)
rs_z_df   = pd.DataFrame(index=stock_tickers)
for label in WINDOWS:
    s = resmom_df[f"RS_{label}"]
    mu, sd = s.mean(), s.std(ddof=1)
    rs_z_df[f"RZ_{label}"] = (s - mu) / sd if sd > 0 else 0.0

rz_cols            = [f"RZ_{l}" for l in WINDOWS]
rs_z_df["RES_MOM"] = rs_z_df[rz_cols].mean(axis=1)

result = result.join(resmom_df)
result = result.join(rs_z_df)

# ── MARKET REGIME ─────────────────────────────────────────────────────────────
nifty_px = nifty_series.dropna()
if len(nifty_px) >= 63:
    n50    = nifty_px.ewm(span=50, adjust=False).mean().iloc[-1]
    n21    = nifty_px.ewm(span=21, adjust=False).mean().iloc[-1]
    n63    = nifty_px.ewm(span=63, adjust=False).mean().iloc[-1]
    last_n = nifty_px.iloc[-1]
    is_buy = (last_n > n50) and (n21 > n63)
    regime_flag = "BUY" if is_buy else "NOT BUY (Risk Off)"
else:
    regime_flag = "UNKNOWN"

# ── CONSOLE OUTPUT ────────────────────────────────────────────────────────────
SEP  = "─" * 100
HEAD = (f"{'RNK':>4}  {'TICKER':<12}  {'SHARPE_ALL':>10}  {'CLENOW_Z':>9}  "
        f"{'RES_MOM':>9}  {'MOM_ACCEL':>10}  {'SHARPE_3':>9}  "
        f"{'1M%':>7}  {'3M%':>7}  {'12M%':>7}")

print(f"\n{'':=<100}")
print(f"  N500 MOMENTUM — TOP {TOP_N}  ·  Sharpe Z + Clenow Z + Residual + Accel")
print(f"  MARKET REGIME : {regime_flag}  (NIFTY500)")
print(f"  Windows: 12M/9M/6M/3M  |  RFR={RFR_ANNUAL*100:.1f}%  |  Filter: PCT_FROM_52H >= -25%")
print(f"{'':=<100}")
print(HEAD); print(SEP)

def fs(v, w=7): return f"{v:>{w}.3f}" if pd.notna(v) else f"{'—':>{w}}"
def fp(v, w=7): return f"{v:>{w}.1f}" if pd.notna(v) else f"{'—':>{w}}"

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 1):
    print(f"{i:>4}  {ticker:<12}  "
          f"{fs(row['COMPOSITE'],10)}  {fs(row['CLENOW_Z'],9)}  "
          f"{fs(row['RES_MOM'],9)}  {fs(row['MOM_ACCEL'],10)}  {fs(row['SHARPE_3'],9)}  "
          f"{fp(row['1M%'])}  {fp(row['3M%'])}  {fp(row['12M%'])}")

print(SEP)
print(f"\n  SHARPE_ALL = mean(Z_12M..Z_3M)  |  CLENOW_Z = mean(CZ_12M..CZ_3M)  |  "
      f"RES_MOM = residual Sharpe  |  MOM_ACCEL = Z(ST−LT Sharpe)  |  "
      f"SHARPE_3 = mean(Z_12M,Z_6M,Z_3M)\n")

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
print(f"Writing {OUTPUT_FILE} ...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border_all():
    s = Side(style="thin", color="333355")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = fill("1E2A4A")
ALT_FILL   = fill("13131F")
POS_FILL   = fill("003322")
NEG_FILL   = fill("330011")
GOLD_FONT  = Font(name="Calibri", color="FFC840", bold=True, size=11)
CYAN_FONT  = Font(name="Calibri", color="00CCFF", bold=True, size=11)
TEXT_FONT  = Font(name="Calibri", color="E0E0F0", size=11)
MUTED_FONT = Font(name="Calibri", color="7070A0", size=11)
HDR_FONT   = Font(name="Calibri", color="00CCFF", bold=True, size=11)
GREEN_FONT = Font(name="Calibri", color="00E5A0", bold=True, size=11)
RED_FONT   = Font(name="Calibri", color="FF4466", bold=True, size=11)

def set_hdr(cell, value):
    cell.value     = value
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border_all()

def set_cell(cell, value, font=None, bg=None, num_fmt=None, align="right"):
    cell.value     = value
    cell.font      = font or TEXT_FONT
    cell.fill      = bg or fill("0D0D14")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border_all()
    if num_fmt:
        cell.number_format = num_fmt

# ── SHEET 1 — TOP20 ───────────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("TOP20")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C3"

ws1.merge_cells("A1:L1")
tc           = ws1["A1"]
tc.value     = (f"N500 MOMENTUM  ·  Top {TOP_N} by SHARPE_ALL  ·  Filter: PCT_FROM_52H ≥ −25%  ·  "
                f"RFR={RFR_ANNUAL*100:.1f}%  ·  "
                f"{dates[0].strftime('%d-%b-%Y')} → {dates[-1].strftime('%d-%b-%Y')}  ·  "
                f"Regime: {regime_flag}")
tc.font      = Font(name="Calibri", color="FFC840", bold=True, size=11)
tc.fill      = fill("0A0A18")
tc.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

top20_cols = [
    ("RNK",        5), ("TICKER",    12), ("SHARPE_ALL",10),
    ("CLENOW_Z",  10), ("RES_MOM",   10), ("MOM_ACCEL", 10),
    ("SHARPE_3",  10), ("SHARPE_ST",  9), ("SHARPE_LT",  9),
    ("1M%",        8), ("3M%",        8), ("12M%",        8),
]
for c, (col_name, col_w) in enumerate(top20_cols, 1):
    set_hdr(ws1.cell(row=2, column=c), col_name)
    ws1.column_dimensions[get_column_letter(c)].width = col_w
ws1.row_dimensions[2].height = 18

def pnl_fnt(v): return GREEN_FONT if pd.notna(v) and v >= 0 else RED_FONT
def pnl_bg(v):  return POS_FILL   if pd.notna(v) and v >= 0 else NEG_FILL

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 3):
    bg     = ALT_FILL if i % 2 == 0 else fill("0D0D14")
    rank_v = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    values = [
        (rank_v,              GOLD_FONT,            bg,                   None),
        (ticker,              GOLD_FONT,            bg,                   None),
        (row["COMPOSITE"],    CYAN_FONT,            bg,                   "0.000"),
        (row["CLENOW_Z"],     CYAN_FONT,            bg,                   "0.000"),
        (row["RES_MOM"],      TEXT_FONT,            bg,                   "0.000"),
        (row["MOM_ACCEL"],    CYAN_FONT,            bg,                   "0.000"),
        (row["SHARPE_3"],     TEXT_FONT,            bg,                   "0.000"),
        (row["SHARPE_ST"],    MUTED_FONT,           bg,                   "0.000"),
        (row["SHARPE_LT"],    MUTED_FONT,           bg,                   "0.000"),
        (row["1M%"],          pnl_fnt(row["1M%"]),  pnl_bg(row["1M%"]),  "0.0"),
        (row["3M%"],          pnl_fnt(row["3M%"]),  pnl_bg(row["3M%"]),  "0.0"),
        (row["12M%"],         pnl_fnt(row["12M%"]), pnl_bg(row["12M%"]), "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws1.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws1.row_dimensions[i].height = 16

# ── SHEET 2 — CALCS ───────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("CALCS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C3"

ws2.merge_cells("A1:AP1")
t2           = ws2["A1"]
t2.value     = (f"N500  ·  Full Calculations  ·  All {len(stock_tickers)} stocks  ·  "
                f"{dates[0].strftime('%d-%b-%Y')} → {dates[-1].strftime('%d-%b-%Y')}")
t2.font      = Font(name="Calibri", color="FFC840", bold=True, size=11)
t2.fill      = fill("0A0A18")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

calcs_cols = [
    ("RANK",      6), ("TICKER",    12),
    ("S_12M",     9), ("S_9M",       9), ("S_6M",    9), ("S_3M",    9), ("S_1M",   9),
    ("Z_12M",     9), ("Z_9M",       9), ("Z_6M",    9), ("Z_3M",    9), ("Z_1M",   9),
    ("SHARPE_ALL",10), ("SHARPE_ST",  9), ("SHARPE_LT", 9),
    ("SHARPE_3", 10), ("MOM_ACCEL", 10),
    ("CS_12M",    9), ("CS_9M",      9), ("CS_6M",   9), ("CS_3M",   9),
    ("CZ_12M",    9), ("CZ_9M",      9), ("CZ_6M",   9), ("CZ_3M",   9),
    ("CLENOW_Z", 10),
    ("CL_12M",    9), ("CL_9M",      9), ("CL_6M",   9), ("CL_3M",   9),
    ("CR_12M",    9), ("CR_9M",      9), ("CR_6M",   9), ("CR_3M",   9),
    ("RS_12M",    9), ("RS_9M",      9), ("RS_6M",   9), ("RS_3M",   9),
    ("RZ_12M",    9), ("RZ_9M",      9), ("RZ_6M",   9), ("RZ_3M",   9),
    ("RES_MOM",  10),
    ("1M%",       8), ("3M%",        8), ("12M%",    8),
    ("52H%",     10),
]

for c, (col_name, col_w) in enumerate(calcs_cols, 1):
    set_hdr(ws2.cell(row=2, column=c), col_name)
    ws2.column_dimensions[get_column_letter(c)].width = col_w
ws2.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.iterrows(), 3):
    bg         = ALT_FILL if i % 2 == 0 else fill("0D0D14")
    rank_v     = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h     = row["PCT_FROM_52H"]
    pct52h_ok  = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt = GREEN_FONT if pct52h_ok else MUTED_FONT
    pct52h_bg  = fill("003322") if pct52h_ok else fill("1A0A0A")

    values = [
        (rank_v,           GOLD_FONT,  bg,        None),
        (ticker,           GOLD_FONT,  bg,        None),
        (row["S_12M"],     TEXT_FONT,  bg,        "0.000"),
        (row["S_9M"],      TEXT_FONT,  bg,        "0.000"),
        (row["S_6M"],      TEXT_FONT,  bg,        "0.000"),
        (row["S_3M"],      TEXT_FONT,  bg,        "0.000"),
        (row["S_1M"],      TEXT_FONT,  bg,        "0.000"),
        (row["Z_12M"],     TEXT_FONT,  bg,        "0.000"),
        (row["Z_9M"],      TEXT_FONT,  bg,        "0.000"),
        (row["Z_6M"],      TEXT_FONT,  bg,        "0.000"),
        (row["Z_3M"],      TEXT_FONT,  bg,        "0.000"),
        (row["Z_1M"],      TEXT_FONT,  bg,        "0.000"),
        (row["COMPOSITE"],   CYAN_FONT,  bg,        "0.000"),
        (row["SHARPE_ALL"],  CYAN_FONT,  bg,        "0.000"),
        (row["SHARPE_ST"],   MUTED_FONT, bg,        "0.000"),
        (row["SHARPE_LT"],   MUTED_FONT, bg,        "0.000"),
        (row["SHARPE_3"],    TEXT_FONT,  bg,        "0.000"),
        (row["MOM_ACCEL"],   CYAN_FONT,  bg,        "0.000"),
        (row["CS_12M"],      TEXT_FONT,  bg,        "0.000"),
        (row["CS_9M"],     TEXT_FONT,  bg,        "0.000"),
        (row["CS_6M"],     TEXT_FONT,  bg,        "0.000"),
        (row["CS_3M"],     TEXT_FONT,  bg,        "0.000"),
        (row["CZ_12M"],    TEXT_FONT,  bg,        "0.000"),
        (row["CZ_9M"],     TEXT_FONT,  bg,        "0.000"),
        (row["CZ_6M"],     TEXT_FONT,  bg,        "0.000"),
        (row["CZ_3M"],     TEXT_FONT,  bg,        "0.000"),
        (row["CLENOW_Z"],  CYAN_FONT,  bg,        "0.000"),
        (row["CL_12M"],    MUTED_FONT, bg,        "0.000"),
        (row["CL_9M"],     MUTED_FONT, bg,        "0.000"),
        (row["CL_6M"],     MUTED_FONT, bg,        "0.000"),
        (row["CL_3M"],     MUTED_FONT, bg,        "0.000"),
        (row["CR_12M"],    MUTED_FONT, bg,        "0.000"),
        (row["CR_9M"],     MUTED_FONT, bg,        "0.000"),
        (row["CR_6M"],     MUTED_FONT, bg,        "0.000"),
        (row["CR_3M"],     MUTED_FONT, bg,        "0.000"),
        (row["RS_12M"],    MUTED_FONT, bg,        "0.000"),
        (row["RS_9M"],     MUTED_FONT, bg,        "0.000"),
        (row["RS_6M"],     MUTED_FONT, bg,        "0.000"),
        (row["RS_3M"],     MUTED_FONT, bg,        "0.000"),
        (row["RZ_12M"],    MUTED_FONT, bg,        "0.000"),
        (row["RZ_9M"],     MUTED_FONT, bg,        "0.000"),
        (row["RZ_6M"],     MUTED_FONT, bg,        "0.000"),
        (row["RZ_3M"],     MUTED_FONT, bg,        "0.000"),
        (row["RES_MOM"],   CYAN_FONT,  bg,        "0.000"),
        (row["1M%"],       TEXT_FONT,  bg,        "0.0"),
        (row["3M%"],       TEXT_FONT,  bg,        "0.0"),
        (row["12M%"],      TEXT_FONT,  bg,        "0.0"),
        (pct52h,           pct52h_fnt, pct52h_bg, "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws2.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws2.row_dimensions[i].height = 15

wb_out.save(OUTPUT_FILE)
print(f"  ✓  Saved → {OUTPUT_FILE}")
print(f"     Sheet 'TOP20' : top {TOP_N} stocks")
print(f"     Sheet 'CALCS' : all {len(stock_tickers)} stocks, 47 columns")