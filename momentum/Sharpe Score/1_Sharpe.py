"""
N500 Momentum Ranking
=====================
Ranks NSE N500 stocks by a multi-window momentum composite.

Scores computed (via momentum_lib):
  SHARPE_ALL  — equal-weighted Z-score of 12M/9M/6M/3M Sharpe ratios
  SHARPE_3    — equal-weighted Z-score of 12M/6M/3M Sharpe ratios
  RES_MOM     — equal-weighted Z-score of 12M/9M/6M/3M residual Sharpe
                (display only — not used in ranking or exit logic)

Eligibility filter : PCT_FROM_52H >= -25%
Ranking            : SHARPE_ALL (COMPOSITE)

Exit Logic
----------
Two distinct exit triggers are evaluated every Monday rebalance.
They are intentionally separate — the 52H exit overrides the hold lock,
the rank exit respects it.

  1. 52H DISQUALIFICATION (overrides 28-day hold lock)
       If PCT_FROM_52H < -25%, the stock is removed from the ranking
       universe entirely (RANK = NaN). Any held stock with NaN rank
       is flagged EXIT_52H = True and must be sold immediately.

  2. RANK-BASED EXIT (respects 28-day hold lock)
       If a held stock's rank drops to > HOLD_RANK_BUFFER (default 40)
       AND it has been held for >= MIN_HOLD_DAYS (default 28 calendar days),
       it is flagged EXIT_RANK = True.

  3. REGIME GATE on new entries
       New buys are only permitted when regime = BUY.
       In NOT BUY regime, existing positions are monitored for exit
       but no new entries are made.

Position Ledger
---------------
Positions are tracked in a JSON file (LEDGER_FILE) with the structure:
  {
    "TICKER": {
      "entry_date": "YYYY-MM-DD",
      "entry_price": float
    },
    ...
  }

The ledger is loaded at the start of each run, used to evaluate exit
conditions, and updated with new entries / removals at the end of the run.

Usage:  python Sharpe.py path/to/n500.xlsx [path/to/ledger.json]
"""

import sys
import json
import datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

import momentum_lib as ml

# ── CONFIG ────────────────────────────────────────────────────────────────────
FILE              = "NSEAll_updated.xlsx" if len(sys.argv) < 2 else sys.argv[1]
OUTPUT_FILE       = "NSEAll_rankings.xlsx"
LEDGER_FILE       = sys.argv[2] if len(sys.argv) >= 3 else "positions_ledger.json"

PORTFOLIO_CAPITAL = 2_000_000   # INR — baseline for allocation display
RFR_ANNUAL        = 0.07
TRADING_DAYS      = 252
TOP_N             = 20
HOLD_RANK_BUFFER  = 40          # exit rank threshold (rank > this → eligible for exit)
MIN_HOLD_DAYS     = 28          # calendar days before rank-based exit is permitted
LIQUID_YIELD_PA   = 0.06        # 6% p.a. on idle cash (Blueprint spec)

WINDOWS        = {"12M": 252, "9M": 189, "6M": 126, "3M": 63}
SHARPE_WINDOWS = {"12M": 252, "9M": 189, "6M": 126, "3M": 63}

rfr_daily = RFR_ANNUAL / TRADING_DAYS

TODAY = datetime.date.today()

# ── POSITION LEDGER ───────────────────────────────────────────────────────────
def load_ledger(path: str) -> dict:
    """
    Load the position ledger from JSON.
    Returns a dict of { ticker: { entry_date, entry_price } }.
    Creates an empty ledger if the file does not exist.
    """
    p = Path(path)
    if not p.exists():
        print(f"  Ledger not found at '{path}' — starting with empty ledger.")
        return {}
    with open(p, "r") as f:
        raw = json.load(f)
    # Validate and normalise
    ledger = {}
    for ticker, rec in raw.items():
        try:
            ledger[ticker] = {
                "entry_date":  datetime.date.fromisoformat(rec["entry_date"]),
                "entry_price": float(rec["entry_price"]),
            }
        except (KeyError, ValueError) as e:
            print(f"  Warning: skipping malformed ledger entry for {ticker}: {e}")
    print(f"  Ledger loaded: {len(ledger)} open position(s) from '{path}'")
    return ledger


def save_ledger(ledger: dict, path: str):
    """Persist the updated ledger back to JSON."""
    serialisable = {
        ticker: {
            "entry_date":  rec["entry_date"].isoformat(),
            "entry_price": rec["entry_price"],
        }
        for ticker, rec in ledger.items()
    }
    with open(path, "w") as f:
        json.dump(serialisable, f, indent=2)
    print(f"  Ledger saved: {len(ledger)} position(s) → '{path}'")


def days_held(ticker: str, ledger: dict) -> int:
    """Return calendar days since entry for a held ticker. -1 if not in ledger."""
    if ticker not in ledger:
        return -1
    return (TODAY - ledger[ticker]["entry_date"]).days


# ── LOAD PRICES ───────────────────────────────────────────────────────────────
print(f"Loading {FILE} ...")
prices_df, nifty_series, stock_tickers, dates = ml.load_prices(FILE)

valid_days = sum(1 for d in prices_df.columns
                 if prices_df[d].notna().any() and (prices_df[d] != 0).any())
print(f"  {len(prices_df)} stocks  |  {len(dates)} date columns  "
      f"({valid_days} actual trading days)  "
      f"|  {dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}\n")

# ── LOAD LEDGER ───────────────────────────────────────────────────────────────
print(f"Loading position ledger ...")
ledger = load_ledger(LEDGER_FILE)

# ── COMPUTE SCORES ────────────────────────────────────────────────────────────
sharpe_df, z_df = ml.compute_sharpe(prices_df, stock_tickers,
                                     SHARPE_WINDOWS, rfr_daily, TRADING_DAYS)

ret_df   = ml.compute_returns(prices_df, stock_tickers)
pct_52h  = ml.compute_pct_from_52h(prices_df, stock_tickers)

# ── COMBINE ───────────────────────────────────────────────────────────────────
result = z_df.join(sharpe_df.rename(columns={l: f"S_{l}" for l in SHARPE_WINDOWS}))

for col in ["COMPOSITE", "SHARPE_3"]:
    result[col] = result[col].map(ml.normalise_composite)
result["SHARPE_ALL"] = result["COMPOSITE"]

result["RANK"] = result["COMPOSITE"].rank(ascending=False, method="first",
                                           na_option="bottom")
result = result.sort_values("COMPOSITE", ascending=False)
result = result.join(ret_df)

# ── 52H FILTER + RE-RANK ──────────────────────────────────────────────────────
print("\nComputing 52-week high proximity ...")
result["PCT_FROM_52H"] = pct_52h

eligible = result["PCT_FROM_52H"] >= -25
result["RANK"] = np.nan
result.loc[eligible, "RANK"] = (
    result.loc[eligible, "COMPOSITE"]
    .rank(ascending=False, method="first", na_option="bottom")
)
result = result.sort_values(["RANK", "COMPOSITE"], ascending=[True, False])
print(f"  {eligible.sum()} / {len(result)} stocks eligible (PCT_FROM_52H >= -25%)")

# ── RESIDUAL MOMENTUM ─────────────────────────────────────────────────────────
resmom_df, rs_z_df = ml.compute_residual_momentum(prices_df, stock_tickers,
                                                    nifty_series, WINDOWS, TRADING_DAYS)
result = result.join(resmom_df)
result = result.join(rs_z_df)

# ── MARKET REGIME ─────────────────────────────────────────────────────────────
regime_flag   = ml.compute_market_regime(nifty_series)
regime_is_buy = regime_flag.startswith("BUY")

# ── EXIT EVALUATION ───────────────────────────────────────────────────────────
#
# Two exit triggers — evaluated independently for every held position.
#
# EXIT_52H  — 52H disqualification.
#   The stock has RANK = NaN because PCT_FROM_52H breached -25%.
#   This overrides the 28-day hold lock unconditionally.
#
# EXIT_RANK — Rank-based exit.
#   The stock's rank has fallen beyond HOLD_RANK_BUFFER (40)
#   AND the stock has been held for at least MIN_HOLD_DAYS (28 days).
#   The hold lock protects against rank whipsaw for recently bought stocks.
#
# NOT BUY regime — no new entries. Existing positions continue to be
#   evaluated for both exit conditions normally.
#
print(f"\n{'─'*60}")
print(f"  EXIT EVALUATION  ({TODAY.strftime('%d-%b-%Y')})")
print(f"  Held positions   : {len(ledger)}")
print(f"{'─'*60}")

exit_52h_list   = []   # immediate exits — 52H breach (lock overridden)
exit_rank_list  = []   # rank exits — rank > 40 AND hold >= 28 days
hold_list       = []   # retained positions

for ticker, rec in ledger.items():
    held_days  = days_held(ticker, ledger)
    rank_val   = result.loc[ticker, "RANK"] if ticker in result.index else np.nan
    pct52h_val = result.loc[ticker, "PCT_FROM_52H"] if ticker in result.index else np.nan

    # ── Trigger 1: 52H disqualification (NaN rank means failed the 52H gate)
    if pd.isna(rank_val):
        exit_52h_list.append({
            "ticker":       ticker,
            "held_days":    held_days,
            "rank":         None,
            "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
            "entry_date":   rec["entry_date"].isoformat(),
            "entry_price":  rec["entry_price"],
            "exit_trigger": "52H_BREACH",
        })

    # ── Trigger 2: Rank exit (respects 28-day lock)
    elif rank_val > HOLD_RANK_BUFFER:
        if held_days >= MIN_HOLD_DAYS:
            exit_rank_list.append({
                "ticker":       ticker,
                "held_days":    held_days,
                "rank":         int(rank_val),
                "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
                "entry_date":   rec["entry_date"].isoformat(),
                "entry_price":  rec["entry_price"],
                "exit_trigger": "RANK_EXIT",
            })
        else:
            # Rank has dropped but hold lock still active — hold and note
            hold_list.append({
                "ticker":       ticker,
                "held_days":    held_days,
                "rank":         int(rank_val),
                "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
                "note":         f"rank {int(rank_val)} > {HOLD_RANK_BUFFER} but lock active "
                                f"({held_days}/{MIN_HOLD_DAYS}d)",
            })
    else:
        # Healthy — rank within buffer, no 52H breach
        hold_list.append({
            "ticker":    ticker,
            "held_days": held_days,
            "rank":      int(rank_val),
            "pct_52h":   round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
            "note":      "HOLD",
        })

# ── PRINT EXIT SUMMARY ────────────────────────────────────────────────────────
all_exits = exit_52h_list + exit_rank_list

if exit_52h_list:
    print(f"\n  [EXIT — 52H BREACH]  Sell immediately. Hold lock overridden.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'RANK':>6}  {'52H%':>7}  {'ENTRY':>10}  {'@ PRICE':>10}")
    print(f"  {'─'*62}")
    for e in exit_52h_list:
        print(f"  {e['ticker']:<14} {e['held_days']:>4}d  "
              f"  {'NaN':>6}  {str(e['pct_52h']):>7}  "
              f"{e['entry_date']:>10}  {e['entry_price']:>10,.2f}")

if exit_rank_list:
    print(f"\n  [EXIT — RANK DROP]  Rank > {HOLD_RANK_BUFFER} and hold >= {MIN_HOLD_DAYS} days.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'RANK':>6}  {'52H%':>7}  {'ENTRY':>10}  {'@ PRICE':>10}")
    print(f"  {'─'*62}")
    for e in exit_rank_list:
        print(f"  {e['ticker']:<14} {e['held_days']:>4}d  "
              f"  {e['rank']:>6}  {str(e['pct_52h']):>7}  "
              f"{e['entry_date']:>10}  {e['entry_price']:>10,.2f}")

if hold_list:
    print(f"\n  [HOLD]  {len(hold_list)} position(s) retained.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'RANK':>6}  {'52H%':>7}  NOTE")
    print(f"  {'─'*62}")
    for h in hold_list:
        print(f"  {h['ticker']:<14} {h['held_days']:>4}d  "
              f"  {str(h['rank']):>6}  {str(h['pct_52h']):>7}  {h['note']}")

if not ledger:
    print("  No open positions in ledger — nothing to evaluate.")

print(f"\n  Summary: {len(exit_52h_list)} 52H exit(s)  |  "
      f"{len(exit_rank_list)} rank exit(s)  |  {len(hold_list)} hold(s)")
print(f"{'─'*60}")

# ── ENTRY CANDIDATES ──────────────────────────────────────────────────────────
#
# Stocks eligible for new entry this week:
#   - In Top 20 by SHARPE_ALL
#   - Not already in the ledger (already held)
#   - Regime must be BUY (no new entries in NOT BUY)
#
currently_held  = set(ledger.keys()) - {e["ticker"] for e in all_exits}
top_n_tickers   = result.head(TOP_N).index.tolist()

if regime_is_buy:
    entry_candidates = [t for t in top_n_tickers if t not in currently_held]
    if entry_candidates:
        print(f"\n  [NEW ENTRIES — BUY REGIME]  {len(entry_candidates)} candidate(s):")
        for t in entry_candidates:
            px = prices_df.loc[t].dropna()
            last_px = px.iloc[-1] if len(px) > 0 else np.nan
            print(f"    {t:<14}  rank {int(result.loc[t,'RANK']):>3}  "
                  f"last price: {last_px:,.2f}")
    else:
        print(f"\n  [NEW ENTRIES — BUY REGIME]  All Top {TOP_N} positions already held.")
else:
    entry_candidates = []
    print(f"\n  [NEW ENTRIES BLOCKED]  Regime = {regime_flag}")
    print(f"  No new buys this week. Existing positions monitored for exit only.")

# ── UPDATE LEDGER ─────────────────────────────────────────────────────────────
#
# Remove exits from ledger, add new entries with today's price.
# The caller is responsible for confirming execution before saving —
# in a live system you'd only save after trades are confirmed filled.
#
for e in all_exits:
    ledger.pop(e["ticker"], None)

for ticker in entry_candidates:
    px = prices_df.loc[ticker].dropna()
    last_px = float(px.iloc[-1]) if len(px) > 0 else 0.0
    ledger[ticker] = {
        "entry_date":  TODAY,
        "entry_price": last_px,
    }

save_ledger(ledger, LEDGER_FILE)

# ── CAPITAL ALLOCATION (VOLATILITY WEIGHTING) ─────────────────────────────────
print("\nCalculating Dynamic Volatility Weights for Top N Portfolio ...")

result["TARGET_WT"] = np.nan
result["ALLOC_INR"] = np.nan

raw_weights = {}
for ticker in top_n_tickers:
    comp_score = result.loc[ticker, "COMPOSITE"]
    px = prices_df.loc[ticker].dropna()

    if len(px) > 10:
        vols = []
        for w in [252, 189, 126, 63]:
            px_w  = px.iloc[-w:] if len(px) >= w else px
            log_r = np.diff(np.log(px_w.values))
            if len(log_r) > 5:
                vols.append(np.std(log_r, ddof=1) * np.sqrt(252))
        if vols and np.mean(vols) > 0:
            raw_weights[ticker] = comp_score / np.mean(vols)
        else:
            raw_weights[ticker] = comp_score
    else:
        raw_weights[ticker] = comp_score

total_raw = sum(raw_weights.values())
for ticker in top_n_tickers:
    norm_w   = raw_weights[ticker] / total_raw if total_raw > 0 else 1.0 / len(top_n_tickers)
    capped_w = min(0.05, norm_w)
    result.loc[ticker, "TARGET_WT"] = capped_w
    result.loc[ticker, "ALLOC_INR"] = capped_w * PORTFOLIO_CAPITAL

total_equity_weight = result.head(TOP_N)["TARGET_WT"].sum()
total_cash_weight   = max(0.0, 1.0 - total_equity_weight)
total_cash_inr      = total_cash_weight * PORTFOLIO_CAPITAL

# ── CONSOLE OUTPUT ────────────────────────────────────────────────────────────
SEP  = "-" * 100
HEAD = (f"{'RNK':>4}  {'TICKER':<12}  {'STATUS':<10}  {'TARGET_WT':>9}  {'ALLOC_INR':>11}  "
        f"{'SHARPE_ALL':>10}  {'RES_MOM':>9}  {'SHARPE_3':>9}  {'52H%':>8}")

# Determine per-ticker display status
exit_tickers  = {e["ticker"] for e in exit_52h_list}
rank_exit_set = {e["ticker"] for e in exit_rank_list}
new_entry_set = set(entry_candidates)

def ticker_status(ticker):
    if ticker in exit_tickers:   return "EXIT-52H"
    if ticker in rank_exit_set:  return "EXIT-RANK"
    if ticker in new_entry_set:  return "NEW BUY"
    if ticker in currently_held: return "HOLD"
    return ""

print(f"\n{'':=<100}")
print(f"  N500 MOMENTUM - TOP {TOP_N}  .  Sharpe Z + Sharpe 3W + Residual")
print(f"  MARKET REGIME  : {regime_flag}")
print(f"  Checks         : (1) NIFTY500 EMA50 > EMA200  (2) NIFTY500 price > EMA50")
print(f"  Windows        : 12M/9M/6M/3M (Overlapping)  |  RFR={RFR_ANNUAL*100:.1f}%")
print(f"  Policies       : Weekly Executions | {MIN_HOLD_DAYS}-Day Min Hold Lock | "
      f"52H% >= -25% | Rank buffer = {HOLD_RANK_BUFFER}")
print(f"  Capital Model  : {PORTFOLIO_CAPITAL:,.0f} INR Base | 5% Capped Volatility Sizing | "
      f"Cash yield {LIQUID_YIELD_PA*100:.0f}% p.a.")
print(f"{'':=<100}")

print(HEAD); print(SEP)

def fs(v, w=7): return f"{v:>{w}.3f}" if pd.notna(v) else f"{'--':>{w}}"
def fp(v, w=7): return f"{v:>{w}.1f}" if pd.notna(v) else f"{'--':>{w}}"
def fw(v, w=7): return f"{v*100:>{w}.1f}%" if pd.notna(v) else f"{'--':>{w}}"

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 1):
    status = ticker_status(ticker)
    print(f"{i:>4}  {ticker:<12}  {status:<10}  "
          f"{fw(row['TARGET_WT'],8)}     {row['ALLOC_INR']:9,.0f}  "
          f"{fs(row['COMPOSITE'],10)}  "
          f"{fs(row['RES_MOM'],9)}  {fs(row['SHARPE_3'],9)}  "
          f"{fp(row['PCT_FROM_52H'], 8)}")

print(SEP)
print(f" {'-':>4}  {'CASH (LIQUID)':<12}  {'':10}  {total_cash_weight*100:8.1f}%     "
      f"{total_cash_inr:9,.0f}")
print(SEP)
print(f"\n  SHARPE_ALL = mean(Z_12M..Z_3M)  |  "
      f"SHARPE_3 = mean(Z_12M,Z_6M,Z_3M)  |  "
      f"RES_MOM = residual Sharpe composite (display only)\n")

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
print(f"Writing {OUTPUT_FILE} ...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border_all():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL    = fill("E8EEF2")
ALT_FILL    = fill("F8F9FA")
POS_FILL    = fill("E6F4EA")
NEG_FILL    = fill("FCE8E6")
EXIT52_FILL = fill("FFF3E0")   # amber — 52H exit
EXITRK_FILL = fill("FCE8E6")   # red-tint — rank exit
NEWBY_FILL  = fill("E8F5E9")   # green-tint — new buy
HOLD_FILL   = fill("FFFFFF")

GOLD_FONT   = Font(name="Calibri", color="1A365D", bold=True,  size=11)
CYAN_FONT   = Font(name="Calibri", color="0055CC", bold=True,  size=11)
TEXT_FONT   = Font(name="Calibri", color="111111",             size=11)
MUTED_FONT  = Font(name="Calibri", color="707070",             size=11)
HDR_FONT    = Font(name="Calibri", color="1A365D", bold=True,  size=11)
GREEN_FONT  = Font(name="Calibri", color="137333", bold=True,  size=11)
RED_FONT    = Font(name="Calibri", color="C5221F", bold=True,  size=11)
AMBER_FONT  = Font(name="Calibri", color="E65100", bold=True,  size=11)

def set_hdr(cell, value):
    cell.value     = value
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border_all()

def set_cell(cell, value, font=None, bg=None, num_fmt=None, align="right"):
    cell.value     = value
    cell.font      = font or TEXT_FONT
    cell.fill      = bg or fill("FFFFFF")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border_all()
    if num_fmt:
        cell.number_format = num_fmt

def row_style(ticker):
    """Return (font, bg_fill) based on exit / entry status."""
    if ticker in exit_tickers:   return AMBER_FONT, EXIT52_FILL
    if ticker in rank_exit_set:  return RED_FONT,   EXITRK_FILL
    if ticker in new_entry_set:  return GREEN_FONT, NEWBY_FILL
    return GOLD_FONT, HOLD_FILL

# ── SHEET 1 — TOP20 ───────────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("TOP20")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C3"

ws1.merge_cells("A1:G1")
tc           = ws1["A1"]
tc.value     = (f"N500 MOMENTUM  .  Top {TOP_N} by SHARPE_ALL  .  "
                f"Filter: PCT_FROM_52H >= -25%  .  RFR={RFR_ANNUAL*100:.1f}%  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}  .  "
                f"Regime: {regime_flag}  .  Run: {TODAY.strftime('%d-%b-%Y')}")
tc.font      = Font(name="Calibri",
                    color="FF2222" if "NOT BUY" in regime_flag else "1A365D",
                    bold=True, size=11)
tc.fill      = fill("2A0000") if "NOT BUY" in regime_flag else fill("F0F4F8")
tc.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

top20_cols = [
    ("RNK",        5), ("TICKER",    12), ("STATUS",    11),
    ("TARGET_WT%", 11), ("ALLOC_INR", 13),
    ("SHARPE_ALL", 12), ("RES_MOM",   10), ("SHARPE_3", 10), ("52H%", 10),
]
for c, (col_name, col_w) in enumerate(top20_cols, 1):
    set_hdr(ws1.cell(row=2, column=c), col_name)
    ws1.column_dimensions[get_column_letter(c)].width = col_w
ws1.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 3):
    row_fnt, row_bg = row_style(ticker)
    rank_v          = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h          = row["PCT_FROM_52H"]
    pct52h_ok       = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt      = GREEN_FONT if pct52h_ok else RED_FONT
    status          = ticker_status(ticker)

    values = [
        (rank_v,           row_fnt,   row_bg, None),
        (ticker,           row_fnt,   row_bg, None),
        (status,           row_fnt,   row_bg, None),
        (row["TARGET_WT"], row_fnt,   row_bg, "0.00%"),
        (row["ALLOC_INR"], row_fnt,   row_bg, "₹_ * #,##0_ ;_ * -#,##0_ ;_ * \"-\"_ ;_ @_ "),
        (row["COMPOSITE"], CYAN_FONT, row_bg, "0.000"),
        (row["RES_MOM"],   TEXT_FONT, row_bg, "0.000"),
        (row["SHARPE_3"],  CYAN_FONT, row_bg, "0.000"),
        (pct52h,           pct52h_fnt,row_bg, "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws1.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws1.row_dimensions[i].height = 16

# Cash summary row
summary_row = TOP_N + 3
set_cell(ws1.cell(row=summary_row, column=1), "-",             GOLD_FONT, fill("FFFFFF"), None)
set_cell(ws1.cell(row=summary_row, column=2), "CASH (LIQUID)", GOLD_FONT, fill("FFFFFF"), None, align="left")
set_cell(ws1.cell(row=summary_row, column=3), "",              MUTED_FONT,fill("FFFFFF"), None)
set_cell(ws1.cell(row=summary_row, column=4), total_cash_weight, GOLD_FONT, fill("FFFFFF"), "0.00%")
set_cell(ws1.cell(row=summary_row, column=5), total_cash_inr,  GOLD_FONT, fill("FFFFFF"),
         "₹_ * #,##0_ ;_ * -#,##0_ ;_ * \"-\"_ ;_ @_ ")
for extra_col in range(6, 10):
    set_cell(ws1.cell(row=summary_row, column=extra_col), "-", MUTED_FONT, fill("FFFFFF"), None)

# ── SHEET 2 — EXITS ───────────────────────────────────────────────────────────
ws_exit = wb_out.create_sheet("EXITS")
ws_exit.sheet_view.showGridLines = False

ws_exit.merge_cells("A1:H1")
te           = ws_exit["A1"]
te.value     = (f"Exit Actions  .  {TODAY.strftime('%d-%b-%Y')}  .  "
                f"{len(all_exits)} exit(s) this rebalance")
te.font      = Font(name="Calibri", color="C5221F", bold=True, size=11)
te.fill      = fill("FFF3E0")
te.alignment = Alignment(horizontal="center", vertical="center")
ws_exit.row_dimensions[1].height = 22

exit_cols = [
    ("TICKER", 14), ("TRIGGER", 14), ("RANK", 8), ("52H%", 8),
    ("HELD_DAYS", 10), ("ENTRY_DATE", 12), ("ENTRY_PRICE", 13), ("NOTE", 30),
]
for c, (col_name, col_w) in enumerate(exit_cols, 1):
    set_hdr(ws_exit.cell(row=2, column=c), col_name)
    ws_exit.column_dimensions[get_column_letter(c)].width = col_w
ws_exit.row_dimensions[2].height = 18

for i, e in enumerate(all_exits, 3):
    is_52h = e["exit_trigger"] == "52H_BREACH"
    row_fnt = AMBER_FONT if is_52h else RED_FONT
    row_bg  = EXIT52_FILL if is_52h else EXITRK_FILL
    note    = "Lock overridden — 52H breach" if is_52h else f"Rank > {HOLD_RANK_BUFFER}, held >= {MIN_HOLD_DAYS}d"
    vals = [
        (e["ticker"],      row_fnt, row_bg),
        (e["exit_trigger"],row_fnt, row_bg),
        (e.get("rank"),    row_fnt, row_bg),
        (e["pct_52h"],     row_fnt, row_bg),
        (e["held_days"],   row_fnt, row_bg),
        (e["entry_date"],  row_fnt, row_bg),
        (e["entry_price"], row_fnt, row_bg),
        (note,             TEXT_FONT, row_bg),
    ]
    for c, (val, fnt, bg_c) in enumerate(vals, 1):
        set_cell(ws_exit.cell(row=i, column=c), val, fnt, bg_c,
                 align="left" if c in (1, 2, 8) else "right")
    ws_exit.row_dimensions[i].height = 16

if not all_exits:
    ws_exit.merge_cells("A3:H3")
    nc = ws_exit["A3"]
    nc.value = "No exits this rebalance."
    nc.font  = MUTED_FONT
    nc.alignment = Alignment(horizontal="center", vertical="center")

# ── SHEET 3 — CALCS ───────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("CALCS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C3"

ws2.merge_cells("A1:AD1")
t2           = ws2["A1"]
t2.value     = (f"N500  .  Full Calculations  .  All {len(stock_tickers)} stocks  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}  .  "
                f"Regime: {regime_flag}")
t2.font      = Font(name="Calibri",
                    color="FF2222" if "NOT BUY" in regime_flag else "1A365D",
                    bold=True, size=11)
t2.fill      = fill("2A0000") if "NOT BUY" in regime_flag else fill("F0F4F8")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

calcs_cols = [
    ("RANK",       6), ("TICKER",    12),
    ("S_12M",      9), ("S_9M",       9), ("S_6M",     9), ("S_3M",    9),
    ("Z_12M",      9), ("Z_9M",       9), ("Z_6M",     9), ("Z_3M",    9),
    ("SHARPE_ALL",10), ("SHARPE_3",  10),
    ("RS_12M",     9), ("RS_9M",      9), ("RS_6M",    9), ("RS_3M",   9),
    ("RZ_12M",     9), ("RZ_9M",      9), ("RZ_6M",    9), ("RZ_3M",   9),
    ("RES_MOM",   10),
    ("1M%",        8), ("3M%",        8), ("12M%",     8),
    ("52H%",      10),
]
for c, (col_name, col_w) in enumerate(calcs_cols, 1):
    set_hdr(ws2.cell(row=2, column=c), col_name)
    ws2.column_dimensions[get_column_letter(c)].width = col_w
ws2.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.iterrows(), 3):
    bg         = ALT_FILL if i % 2 == 0 else fill("FFFFFF")
    rank_v     = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h     = row["PCT_FROM_52H"]
    pct52h_ok  = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt = GREEN_FONT if pct52h_ok else MUTED_FONT
    pct52h_bg  = fill("E6F4EA") if pct52h_ok else fill("F0F0F0")

    values = [
        (rank_v,             GOLD_FONT,  bg,        None),
        (ticker,             GOLD_FONT,  bg,        None),
        (row["S_12M"],       TEXT_FONT,  bg,        "0.000"),
        (row["S_9M"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_6M"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_3M"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_12M"],       TEXT_FONT,  bg,        "0.000"),
        (row["Z_9M"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_6M"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_3M"],        TEXT_FONT,  bg,        "0.000"),
        (row["COMPOSITE"],   CYAN_FONT,  bg,        "0.000"),
        (row["SHARPE_3"],    TEXT_FONT,  bg,        "0.000"),
        (row["RS_12M"],      MUTED_FONT, bg,        "0.000"),
        (row["RS_9M"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_6M"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_3M"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_12M"],      MUTED_FONT, bg,        "0.000"),
        (row["RZ_9M"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_6M"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_3M"],       MUTED_FONT, bg,        "0.000"),
        (row["RES_MOM"],     CYAN_FONT,  bg,        "0.000"),
        (row["1M%"],         TEXT_FONT,  bg,        "0.0"),
        (row["3M%"],         TEXT_FONT,  bg,        "0.0"),
        (row["12M%"],        TEXT_FONT,  bg,        "0.0"),
        (pct52h,             pct52h_fnt, pct52h_bg, "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws2.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws2.row_dimensions[i].height = 15

wb_out.save(OUTPUT_FILE)
print(f"  +  Saved -> {OUTPUT_FILE}")
print(f"     Sheet 'TOP20' : top {TOP_N} stocks with status (NEW BUY / HOLD / EXIT)")
print(f"     Sheet 'EXITS' : {len(all_exits)} exit action(s) this rebalance")
print(f"     Sheet 'CALCS' : all {len(stock_tickers)} stocks, full calculations")