"""
N500 Momentum Ranking - 4-Block Non-Overlapping Strategy
=========================================================
Ranks NSE N500 stocks by a 4-Block non-overlapping momentum composite.

Blocks:
B1: 0 - 63 days
B2: 64 - 126 days
B3: 127 - 189 days
B4: 190 - 252 days

Scores computed:
  SHARPE_ALL  — equal-weighted Z-score of B1, B2, B3, B4 Sharpe ratios
  RES_MOM     — equal-weighted Z-score of B1, B2, B3, B4 residual Sharpe

Eligibility filter : PCT_FROM_52H >= -25%
Ranking            : SHARPE_ALL (COMPOSITE)

Usage:  python Sharpe_4Block.py path/to/n500.xlsx
"""

import sys
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import momentum_lib as ml

# ── CONFIG ────────────────────────────────────────────────────────────────────
FILE         = "n500.xlsx" if len(sys.argv) < 2 else sys.argv[1]
OUTPUT_FILE  = "n500_rankings_4block.xlsx"
RFR_ANNUAL   = 0.07
TRADING_DAYS = 252
TOP_N        = 20

rfr_daily = RFR_ANNUAL / TRADING_DAYS

# ── CALCULATIONS: 4-BLOCK LOGIC ───────────────────────────────────────────────

def compute_block_sharpe_4b(prices_df, stock_tickers):
    num_days = 63
    blocks = [("B1", 0, 63), ("B2", 63, 126), ("B3", 126, 189), ("B4", 189, 252)]
    sharpe_data = {}
    
    for t in stock_tickers:
        px = prices_df.loc[t].dropna()
        if len(px) < 63:
            sharpe_data[t] = {l: np.nan for l,_,_ in blocks}
            continue
            
        res = {}
        for label, start, end in blocks:
            if len(px) <= start:
                res[label] = np.nan
                continue
                
            px_block = px.iloc[-end:] if start == 0 else px.iloc[-end:-start]
            if len(px_block) < 10:
                res[label] = np.nan
                continue
            ret = (px_block.iloc[-1] / px_block.iloc[0]) - 1.0
            ann_ret = (1 + ret) ** (252 / num_days) - 1.0
            daily_rets = px_block.pct_change().dropna()
            vol = daily_rets.std() * np.sqrt(252)
            
            if vol > 0.0:
                res[label] = (ann_ret - 0.07) / vol
            else:
                res[label] = np.nan
        sharpe_data[t] = res
        
    df = pd.DataFrame.from_dict(sharpe_data, orient='index')
    z_df = pd.DataFrame(index=df.index)
    for col in df.columns:
        s = df[col]
        z_df[f"Z_{col}"] = (s - s.mean()) / s.std()
        
    return df, z_df

def _residual_sharpe(s_rets, m_rets):
    if len(s_rets) != len(m_rets) or len(s_rets) < 10:
        return np.nan
    X = np.column_stack([np.ones(len(m_rets)), m_rets])
    try:
        coeffs, _, _, _ = np.linalg.lstsq(X, s_rets, rcond=None)
    except np.linalg.LinAlgError:
        return np.nan
    residuals = s_rets - X @ coeffs
    sd = residuals.std(ddof=1)
    if sd < 1e-12:
        return np.nan
    return (residuals.mean() / sd) * np.sqrt(252)

def compute_block_residual_momentum_4b(prices_df, stock_tickers, nifty_series):
    blocks = [("B1", 0, 63), ("B2", 63, 126), ("B3", 126, 189), ("B4", 189, 252)]
    resmom_data = {t: {} for t in stock_tickers}
    nifty_px = nifty_series.dropna()
    
    for t in stock_tickers:
        px = prices_df.loc[t].dropna()
        if len(px) < 63 or len(nifty_px) < 63:
            resmom_data[t] = {l: np.nan for l,_,_ in blocks}
            continue
            
        for label, start, end in blocks:
            if len(px) <= start or len(nifty_px) <= start:
                resmom_data[t][label] = np.nan
                continue
                
            px_block = px.iloc[-end:] if start == 0 else px.iloc[-end:-start]
            n_block = nifty_px.iloc[-end:] if start == 0 else nifty_px.iloc[-end:-start]
            
            s_rets = np.diff(np.log(px_block.values))
            m_rets = np.diff(np.log(n_block.values))
            resmom_data[t][label] = _residual_sharpe(s_rets, m_rets)
            
    df = pd.DataFrame.from_dict(resmom_data, orient='index')
    z_df = pd.DataFrame(index=df.index)
    for col in df.columns:
        s = df[col]
        z_df[f"RZ_{col}"] = (s - s.mean()) / s.std()
    
    rz_cols = [f"RZ_{l}" for l,_,_ in blocks]
    z_df["RES_MOM"] = z_df[rz_cols].mean(axis=1)
    return df, z_df

# ── LOAD ──────────────────────────────────────────────────────────────────────
print(f"Loading {FILE} ...")
prices_df, nifty_series, stock_tickers, dates = ml.load_prices(FILE)

valid_days = sum(1 for d in prices_df.columns
                 if prices_df[d].notna().any() and (prices_df[d] != 0).any())
print(f"  {len(prices_df)} stocks  |  {len(dates)} date columns  "
      f"({valid_days} actual trading days)  "
      f"|  {dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}\n")

# ── COMPUTE SCORES ────────────────────────────────────────────────────────────
print("Computing 4-Block Non-Overlapping Sharpe ...")
sharpe_df, z_df = compute_block_sharpe_4b(prices_df, stock_tickers)

print("Computing 4-Block Residual Momentum ...")
resmom_df, rs_z_df = compute_block_residual_momentum_4b(prices_df, stock_tickers, nifty_series)

ret_df  = ml.compute_returns(prices_df, stock_tickers)
pct_52h = ml.compute_pct_from_52h(prices_df, stock_tickers)

# ── COMBINE ───────────────────────────────────────────────────────────────────
result = z_df.join(sharpe_df.rename(columns={l: f"S_{l}" for l in ["B1", "B2", "B3", "B4"]}))

z_cols = [f"Z_{l}" for l in ["B1", "B2", "B3", "B4"]]
result["COMPOSITE"] = result[z_cols].mean(axis=1)
result["COMPOSITE"] = result["COMPOSITE"].map(ml.normalise_composite)
result["SHARPE_ALL"] = result["COMPOSITE"]

result["RANK"] = result["COMPOSITE"].rank(ascending=False, method="first", na_option="bottom")
result = result.sort_values("COMPOSITE", ascending=False)
result = result.join(ret_df)

# ── 52H FILTER + RE-RANK ──────────────────────────────────────────────────────
print("Computing 52-week high proximity ...")
result["PCT_FROM_52H"] = pct_52h

eligible = result["PCT_FROM_52H"] >= -25
result["RANK"] = np.nan
result.loc[eligible, "RANK"] = (
    result.loc[eligible, "COMPOSITE"]
    .rank(ascending=False, method="first", na_option="bottom")
)
result = result.sort_values(["RANK", "COMPOSITE"], ascending=[True, False])
print(f"  {eligible.sum()} / {len(result)} stocks eligible (PCT_FROM_52H >= -25%)")

# ── MERGE RESIDUAL ────────────────────────────────────────────────────────────
resmom_df = resmom_df.rename(columns={l: f"RS_{l}" for l in ["B1", "B2", "B3", "B4"]})
result = result.join(resmom_df)
result = result.join(rs_z_df)

# ── MARKET REGIME ─────────────────────────────────────────────────────────────
regime_flag, is_cash = ml.compute_market_regime(nifty_series)

# ── CONSOLE OUTPUT ────────────────────────────────────────────────────────────
SEP  = "-" * 60
HEAD = (f"{'RNK':>4}  {'TICKER':<12}  {'SHARPE_ALL':>10}  "
        f"{'RES_MOM':>9}  {'52H%':>8}")

print(f"\n{'':=<70}")
print(f"  N500 MOMENTUM - TOP {TOP_N}  .  4-Block Non-Overlapping Sharpe")
print(f"  MARKET REGIME : {regime_flag}")
print(f"  Checks        : (1) EMA50 > EMA200  (2) price > EMA50")
print(f"  Windows: B1/B2/B3/B4   |  RFR={RFR_ANNUAL*100:.1f}%  |  Filter: PCT_FROM_52H >= -25%")
print(f"{'':=<70}")

print(HEAD); print(SEP)

def fs(v, w=7): return f"{v:>{w}.3f}" if pd.notna(v) else f"{'--':>{w}}"
def fp(v, w=7): return f"{v:>{w}.1f}" if pd.notna(v) else f"{'--':>{w}}"

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 1):
    print(f"{i:>4}  {ticker:<12}  "
          f"{fs(row['COMPOSITE'],10)}  "
          f"{fs(row['RES_MOM'],9)}  "
          f"{fp(row['PCT_FROM_52H'], 8)}")

print(SEP)
print(f"\n  SHARPE_ALL = mean(Z_B1..Z_B4)  |  "
      f"RES_MOM = residual Sharpe composite\n")

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
print(f"Writing {OUTPUT_FILE} ...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border_all():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = fill("E8EEF2")
ALT_FILL   = fill("F8F9FA")
POS_FILL   = fill("E6F4EA")
NEG_FILL   = fill("FCE8E6")
GOLD_FONT  = Font(name="Calibri", color="1A365D", bold=True,  size=11)
CYAN_FONT  = Font(name="Calibri", color="0055CC", bold=True,  size=11)
TEXT_FONT  = Font(name="Calibri", color="111111",             size=11)
MUTED_FONT = Font(name="Calibri", color="707070",             size=11)
HDR_FONT   = Font(name="Calibri", color="1A365D", bold=True,  size=11)
GREEN_FONT = Font(name="Calibri", color="137333", bold=True,  size=11)
RED_FONT   = Font(name="Calibri", color="C5221F", bold=True,  size=11)

def set_hdr(cell, value):
    cell.value     = value
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border_all()

def set_cell(cell, value, font=None, bg=None, num_fmt=None, align="right"):
    cell.value     = value
    cell.font      = font or TEXT_FONT
    cell.fill      = bg or fill("FFFFFF")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border_all()
    if num_fmt:
        cell.number_format = num_fmt

# ── SHEET 1 — TOP20 ───────────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("TOP20")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C3"

ws1.merge_cells("A1:E1")
tc           = ws1["A1"]
tc.value     = (f"N500 MOMENTUM 4-BLOCK  .  Top {TOP_N}  .  Filter: PCT_FROM_52H >= -25%  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}  .  "
                f"Regime: {regime_flag}")
tc.font      = Font(name="Calibri", color="FF2222" if "NOT BUY" in regime_flag else "1A365D", bold=True, size=11)
tc.fill      = fill("2A0000") if "NOT BUY" in regime_flag else fill("F0F4F8")
tc.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

top20_cols = [
    ("RNK",       5), ("TICKER",   12), ("SHARPE_ALL", 10),
    ("RES_MOM",  10), ("52H%",       10),
]
for c, (col_name, col_w) in enumerate(top20_cols, 1):
    set_hdr(ws1.cell(row=2, column=c), col_name)
    ws1.column_dimensions[get_column_letter(c)].width = col_w
ws1.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 3):
    bg     = ALT_FILL if i % 2 == 0 else fill("FFFFFF")
    rank_v = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h = row["PCT_FROM_52H"]
    pct52h_ok  = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt = GREEN_FONT if pct52h_ok else MUTED_FONT

    values = [
        (rank_v,              GOLD_FONT,            bg,                   None),
        (ticker,              GOLD_FONT,            bg,                   None),
        (row["COMPOSITE"],    CYAN_FONT,            bg,                   "0.000"),
        (row["RES_MOM"],      TEXT_FONT,            bg,                   "0.000"),
        (pct52h,              pct52h_fnt,           bg,                   "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws1.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws1.row_dimensions[i].height = 16

# ── SHEET 2 — CALCS ───────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("CALCS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C3"

ws2.merge_cells("A1:W1")
t2           = ws2["A1"]
t2.value     = (f"N500 4-Block Calculations  .  All {len(stock_tickers)} stocks  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}  .  "
                f"Regime: {regime_flag}")
t2.font      = Font(name="Calibri", color="FF2222" if "NOT BUY" in regime_flag else "1A365D", bold=True, size=11)
t2.fill      = fill("2A0000") if "NOT BUY" in regime_flag else fill("F0F4F8")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

calcs_cols = [
    ("RANK",       6), ("TICKER",    12),
    ("S_B1",      9), ("S_B2",       9), ("S_B3",     9), ("S_B4",    9),
    ("Z_B1",      9), ("Z_B2",       9), ("Z_B3",     9), ("Z_B4",    9),
    ("SHARPE_ALL",10),
    ("RS_B1",     9), ("RS_B2",      9), ("RS_B3",    9), ("RS_B4",   9),
    ("RZ_B1",     9), ("RZ_B2",      9), ("RZ_B3",    9), ("RZ_B4",   9),
    ("RES_MOM",   10),
    ("1M%",        8), ("3M%",        8), ("12M%",     8),
    ("52H%",      10),
]

for c, (col_name, col_w) in enumerate(calcs_cols, 1):
    set_hdr(ws2.cell(row=2, column=c), col_name)
    ws2.column_dimensions[get_column_letter(c)].width = col_w
ws2.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.iterrows(), 3):
    bg         = ALT_FILL if i % 2 == 0 else fill("FFFFFF")
    rank_v     = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h     = row["PCT_FROM_52H"]
    pct52h_ok  = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt = GREEN_FONT if pct52h_ok else MUTED_FONT
    pct52h_bg  = fill("E6F4EA") if pct52h_ok else fill("F0F0F0")

    values = [
        (rank_v,             GOLD_FONT,  bg,        None),
        (ticker,             GOLD_FONT,  bg,        None),
        (row["S_B1"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_B2"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_B3"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_B4"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_B1"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_B2"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_B3"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_B4"],        TEXT_FONT,  bg,        "0.000"),
        (row["COMPOSITE"],   CYAN_FONT,  bg,        "0.000"),

        (row["RS_B1"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_B2"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_B3"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_B4"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_B1"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_B2"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_B3"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_B4"],       MUTED_FONT, bg,        "0.000"),
        (row["RES_MOM"],     CYAN_FONT,  bg,        "0.000"),
        (row["1M%"],         TEXT_FONT,  bg,        "0.0"),
        (row["3M%"],         TEXT_FONT,  bg,        "0.0"),
        (row["12M%"],        TEXT_FONT,  bg,        "0.0"),
        (pct52h,             pct52h_fnt, pct52h_bg, "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws2.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws2.row_dimensions[i].height = 15

wb_out.save(OUTPUT_FILE)
print(f"  +  Saved -> {OUTPUT_FILE}")
print(f"     Sheet 'TOP20' : top {TOP_N} stocks")
print(f"     Sheet 'CALCS' : all {len(stock_tickers)} stocks, 24 columns")