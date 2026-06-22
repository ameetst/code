"""
N500 Momentum Ranking
=====================
Ranks NSE N500 stocks by a multi-window momentum composite.

Scores computed (via momentum_lib):
  SHARPE_ALL  — equal-weighted Z-score of 12M/9M/6M/3M Sharpe ratios
  SHARPE_3    — equal-weighted Z-score of 12M/6M/3M Sharpe ratios
  RES_MOM     — equal-weighted Z-score of 12M/9M/6M/3M residual Sharpe
                (display only — not used in ranking or exit logic)

Eligibility filter : PCT_FROM_52H >= -25%
Ranking            : SHARPE_ALL (COMPOSITE)

Exit Logic
----------
Two distinct exit triggers are evaluated every Monday rebalance.
They are intentionally separate — the 52H exit overrides the hold lock,
the rank exit respects it.

  1. 52H DISQUALIFICATION (overrides 28-day hold lock)
       If PCT_FROM_52H < -25%, the stock is removed from the ranking
       universe entirely (RANK = NaN). Any held stock with NaN rank
       is flagged EXIT_52H = True and must be sold immediately.

  2. RANK-BASED EXIT (respects 28-day hold lock)
       If a held stock's rank drops to > HOLD_RANK_BUFFER (default 40)
       AND it has been held for >= MIN_HOLD_DAYS (default 28 calendar days),
       it is flagged EXIT_RANK = True.

  3. REGIME GATE on new entries
       New buys are only permitted when regime = BUY.
       In NOT BUY regime, existing positions are monitored for exit
       but no new entries are made.

Position Ledger
---------------
Positions are tracked in a JSON file (LEDGER_FILE) with the structure:
  {
    "TICKER": {
      "entry_date": "YYYY-MM-DD",
      "entry_price": float
    },
    ...
  }

The ledger is loaded at the start of each run, used to evaluate exit
conditions, and updated with new entries / removals at the end of the run.

Usage:  python Sharpe.py <UNIVERSE> [path/to/ledger.json]
          UNIVERSE examples: N500, N750, NSEAll
          Derives: <UNIVERSE>_updated.xlsx  ->  <UNIVERSE>_rankings.xlsx
                   <UNIVERSE>_positions_ledger.json (overridden by 2nd arg)
          Use --dry-run to generate recommendations without saving ledger changes.
"""

import sys
import json
import datetime
import argparse
import subprocess
import shutil
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

import momentum_lib as ml

# -- ARGUMENT PARSING ----------------------------------------------------------
_parser = argparse.ArgumentParser(description="Sharpe Momentum Ranking")
_parser.add_argument("universe", nargs="?", default="N500",
                     help="Universe name (e.g., N500, N750, NSEAll)")
_parser.add_argument("ledger", nargs="?", default=None,
                     help="Path to positions ledger JSON (optional)")
_parser.add_argument("--update", action="store_true",
                     help="Run update_stock_price.py to refresh data before computing rankings")
_parser.add_argument("--min-turnover", type=float, default=1.0,
                     help="Minimum median daily turnover in Rs Cr (default: 1.0)")
_parser.add_argument("--dry-run", action="store_true",
                     help="Generate rankings/exits without saving changes to the positions ledger")
_args = _parser.parse_args()

# -- CONFIG --------------------------------------------------------------------
UNIVERSE          = _args.universe
FILE              = f"{UNIVERSE}_updated.xlsx"
OUTPUT_FILE       = f"{UNIVERSE}_rankings.xlsx"
LEDGER_FILE       = _args.ledger or f"{UNIVERSE}_positions_ledger.json"

PORTFOLIO_CAPITAL = 1_000_000   # INR — baseline for allocation display
RFR_ANNUAL        = 0.07
TRADING_DAYS      = 252
TOP_N             = 20         # used for Excel sheet label only; actual N is dynamic
HOLD_RANK_BUFFER  = 40          # exit rank threshold
MIN_HOLD_DAYS     = 28          # calendar days before rank-based exit is permitted
LIQUID_YIELD_PA   = 0.06        # 6% p.a. on idle cash
MIN_TURNOVER_CR   = _args.min_turnover   # Minimum median daily turnover (₹ Cr)
DRY_RUN           = _args.dry_run

# -- DYNAMIC REGIME PARAMETERS -------------------------------------------------
MIN_N               = 5      # minimum holdings at lowest regime score
MAX_N               = 25     # maximum holdings at highest regime score
NEW_ENTRY_THRESHOLD = 0.40   # regime score below this — no new buys
SIGNAL_WEIGHTS      = {      # must sum to 1.0
    "ema50_breadth":     0.35,   # % stocks > own EMA50
    "ema_trend_breadth": 0.25,   # % stocks with EMA50 > EMA200
    "breadth":           0.25,
    "momentum":          0.15,
}

WINDOWS        = {"12M": 252, "9M": 189, "6M": 126, "3M": 63}
SHARPE_WINDOWS = {"12M": 252, "9M": 189, "6M": 126, "3M": 63}

rfr_daily = RFR_ANNUAL / TRADING_DAYS

TODAY = datetime.date.today()

# -- POSITION LEDGER -----------------------------------------------------------


def compute_regime_score(nifty_s: pd.Series,
                         eligible_mask: pd.Series,
                         composite_series: pd.Series,
                         prices_df: pd.DataFrame = None) -> tuple:
    """
    Compute a continuous Regime Strength Score (0.0 to 1.0) from 4 signals:
      Signal 1 (35%): EMA50 breadth     — % of universe stocks trading above their own EMA50
      Signal 2 (25%): EMA trend breadth — % of universe stocks with their EMA50 > their EMA200
      Signal 3 (25%): 52H breadth       — % stocks within -25% of 52-week high
      Signal 4 (15%): Momentum breadth  — % eligible stocks with COMPOSITE > 1.5

    Returns (regime_score: float, detail: dict)
    """
    px = nifty_s.dropna()
    if len(px) < 200:
        return 0.5, {"regime_score": 0.5, "dynamic_n": 15, "note": "insufficient data"}

    # -- Signal 1 & 2: Universe EMA breadth --
    if prices_df is not None and len(prices_df.columns) >= 200:
        ema50_all  = prices_df.ewm(span=50,  adjust=False, axis=1).mean()
        ema200_all = prices_df.ewm(span=200, adjust=False, axis=1).mean()
        last_px    = prices_df.iloc[:, -1]
        last_ema50 = ema50_all.iloc[:, -1]
        last_ema200 = ema200_all.iloc[:, -1]
        valid      = last_px.notna() & last_ema200.notna()
        n_valid    = int(valid.sum())
        if n_valid > 0:
            ema50_breadth_score = float((last_px[valid] > last_ema50[valid]).sum()) / n_valid
            ema_trend_breadth_score = float((last_ema50[valid] > last_ema200[valid]).sum()) / n_valid
        else:
            ema50_breadth_score = 0.5
            ema_trend_breadth_score = 0.5
    else:
        # Fallback: use NIFTY500 index if prices_df not available
        price  = px.iloc[-1]
        ema50  = px.ewm(span=50,  adjust=False).mean().iloc[-1]
        ema200 = px.ewm(span=200, adjust=False).mean().iloc[-1]
        ema50_breadth_score     = 1.0 if price > ema50 else 0.0
        ema_trend_breadth_score = 1.0 if ema50 > ema200 else 0.0

    # -- Signal 3: 52H breadth --
    total_stocks  = len(eligible_mask)
    elig_count    = int(eligible_mask.sum())
    breadth_score = elig_count / total_stocks if total_stocks > 0 else 0.5

    # -- Signal 4: Momentum breadth --
    elig_comp     = composite_series[eligible_mask]
    pos_mom       = int((elig_comp > 1.5).sum())
    momentum_score = pos_mom / max(1, elig_count)

    regime_score = (
        ema50_breadth_score     * SIGNAL_WEIGHTS["ema50_breadth"]     +
        ema_trend_breadth_score * SIGNAL_WEIGHTS["ema_trend_breadth"] +
        breadth_score           * SIGNAL_WEIGHTS["breadth"]           +
        momentum_score          * SIGNAL_WEIGHTS["momentum"]
    )
    dynamic_n = int(MIN_N + regime_score * (MAX_N - MIN_N))

    detail = {
        "ema50_score":     round(ema50_breadth_score, 3),
        "ema_trend_score": round(ema_trend_breadth_score, 3),
        "breadth_score":   round(breadth_score, 3),
        "momentum_score":  round(momentum_score, 3),
        "regime_score":    round(regime_score, 3),
        "dynamic_n":       dynamic_n,
        "eligible":        elig_count,
        "allow_new":       regime_score >= NEW_ENTRY_THRESHOLD,
    }
    return regime_score, detail


def load_ledger(path: str) -> dict:
    """
    Load the position ledger from JSON.
    Returns a dict of { ticker: { entry_date, entry_price } }.
    Creates an empty ledger if the file does not exist.
    """
    p = Path(path)
    if not p.exists():
        print(f"  Ledger not found at '{path}' — starting with empty ledger.")
        return {}
    with open(p, "r") as f:
        raw = json.load(f)
    # Validate and normalise
    ledger = {}
    for ticker, rec in raw.items():
        try:
            ledger[ticker] = {
                "entry_date":  datetime.date.fromisoformat(rec["entry_date"]),
                "entry_price": float(rec["entry_price"]),
            }
        except (KeyError, ValueError) as e:
            print(f"  Warning: skipping malformed ledger entry for {ticker}: {e}")
    print(f"  Ledger loaded: {len(ledger)} open position(s) from '{path}'")
    return ledger


def save_ledger(ledger: dict, path: str):
    """Persist the updated ledger back to JSON using an atomic replace."""
    serialisable = {
        ticker: {
            "entry_date":  rec["entry_date"].isoformat(),
            "entry_price": rec["entry_price"],
        }
        for ticker, rec in ledger.items()
    }
    p = Path(path)
    tmp_path = p.with_suffix(".tmp")
    bak_path = p.with_suffix(".bak")

    try:
        with open(tmp_path, "w") as f:
            json.dump(serialisable, f, indent=2)
        if p.exists():
            shutil.copy2(p, bak_path)
        shutil.move(str(tmp_path), str(p))
    except Exception as e:
        if tmp_path.exists():
            tmp_path.unlink()
        raise e

    print(f"  Ledger saved: {len(ledger)} position(s) -> '{path}'")
    if bak_path.exists():
        print(f"  Previous ledger backup -> '{bak_path}'")


def days_held(ticker: str, ledger: dict) -> int:
    """Return calendar days since entry for a held ticker. -1 if not in ledger."""
    if ticker not in ledger:
        return -1
    return (TODAY - ledger[ticker]["entry_date"]).days


# -- RUN DATA UPDATE (optional) ------------------------------------------------
if _args.update:
    UPDATE_SCRIPT = Path(__file__).resolve().parent.parent.parent / "yfinance data scripts" / "update_stock_price.py"
    TEMPLATE_DIR  = Path(__file__).resolve().parent.parent.parent / "yfinance data scripts"
    SHARPE_DIR    = Path(__file__).resolve().parent

    if not UPDATE_SCRIPT.exists():
        print(f"ERROR: update_stock_price.py not found at {UPDATE_SCRIPT}")
        sys.exit(1)

    template_file = TEMPLATE_DIR / f"{UNIVERSE}.xlsx"
    if not template_file.exists():
        print(f"ERROR: Template {template_file} not found.")
        sys.exit(1)

    print(f"\n{'='*70}")
    print(f"STEP 1: Refreshing price + volume data via update_stock_price.py")
    print(f"{'='*70}\n")

    ret = subprocess.run(
        [sys.executable, str(UPDATE_SCRIPT), UNIVERSE,
         "--output-dir", str(SHARPE_DIR)],
        cwd=str(TEMPLATE_DIR)
    )
    if ret.returncode != 0:
        print(f"\nERROR: update_stock_price.py failed (exit code {ret.returncode})")
        sys.exit(1)

    print(f"\n{'='*70}")
    print(f"STEP 2: Computing Sharpe rankings")
    print(f"{'='*70}\n")

# -- LOAD PRICES ---------------------------------------------------------------
print(f"Loading {FILE} ...")
prices_df, nifty_series, stock_tickers, dates = ml.load_prices(FILE)

valid_days = sum(1 for d in prices_df.columns
                 if prices_df[d].notna().any() and (prices_df[d] != 0).any())
print(f"  {len(prices_df)} stocks  |  {len(dates)} date columns  "
      f"({valid_days} actual trading days)  "
      f"|  {dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}\n")

# -- LOAD VOLUME (optional) ----------------------------------------------------
print(f"Loading volume data ...")
volume_df = ml.load_volume(FILE)
if volume_df is not None:
    print(f"  VOLUME sheet loaded: {len(volume_df)} tickers")
    turnover_df = ml.compute_turnover(prices_df, volume_df, stock_tickers)
    print(f"  Turnover computed for {turnover_df['TURNOVER_12M'].notna().sum()} (12M) / "
          f"{turnover_df['TURNOVER_6M'].notna().sum()} (6M) tickers")
else:
    print("  ⚠  No VOLUME sheet found — ADTV filter will be skipped.")
    print("    Run: python update_stock_price.py {universe} to generate volume data.")
    turnover_df = None

# -- LOAD LEDGER ---------------------------------------------------------------
print(f"Loading position ledger ...")
ledger = load_ledger(LEDGER_FILE)

# -- COMPUTE SCORES ------------------------------------------------------------
# ACTIVE: Raw Sharpe (production baseline: CAGR 38.3% / MDD -16.3%)
# DORMANT: Adjusted Sharpe (Skew + Kurtosis penalty: CAGR 42.2% / MDD -20.2%)
#   To activate, replace the line below with:
#       sharpe_df, z_df = ml.compute_adjusted_sharpe(prices_df, stock_tickers,
#                                                     SHARPE_WINDOWS, rfr_daily, TRADING_DAYS)
sharpe_df, z_df = ml.compute_sharpe(prices_df, stock_tickers,
                                     SHARPE_WINDOWS, rfr_daily, TRADING_DAYS)

ret_df   = ml.compute_returns(prices_df, stock_tickers)
pct_52h  = ml.compute_pct_from_52h(prices_df, stock_tickers)

# -- COMBINE -------------------------------------------------------------------
result = z_df.join(sharpe_df.rename(columns={l: f"S_{l}" for l in SHARPE_WINDOWS}))

for col in ["COMPOSITE", "SHARPE_3"]:
    result[col] = result[col].map(ml.normalise_composite)
result["SHARPE_ALL"] = result["COMPOSITE"]

result["RANK"] = result["COMPOSITE"].rank(ascending=False, method="first",
                                           na_option="bottom")
result = result.sort_values("COMPOSITE", ascending=False)
result = result.join(ret_df)

# -- 52H FILTER + ADTV FILTER + RE-RANK ----------------------------------------
print("\nComputing 52-week high proximity ...")
result["PCT_FROM_52H"] = pct_52h

# ADTV turnover filter
if turnover_df is not None:
    result = result.join(turnover_df)
    # Either 12M OR 6M median turnover must be >= threshold
    adtv_12m_ok = result["TURNOVER_12M"] >= MIN_TURNOVER_CR
    adtv_6m_ok  = result["TURNOVER_6M"]  >= MIN_TURNOVER_CR
    adtv_ok     = adtv_12m_ok | adtv_6m_ok
    # Stocks with no volume data at all are treated as ineligible
    adtv_ok     = adtv_ok.fillna(False)
    result["ADTV_ELIGIBLE"] = adtv_ok
    n_adtv_fail = (~adtv_ok).sum()
    print(f"  ADTV filter (>= {MIN_TURNOVER_CR} Cr): {adtv_ok.sum()} pass, {n_adtv_fail} fail")
else:
    adtv_ok = pd.Series(True, index=result.index)
    result["ADTV_ELIGIBLE"] = True

eligible = (result["PCT_FROM_52H"] >= -25) & adtv_ok
result["RANK"] = np.nan
result.loc[eligible, "RANK"] = (
    result.loc[eligible, "COMPOSITE"]
    .rank(ascending=False, method="first", na_option="bottom")
)
result = result.sort_values(["RANK", "COMPOSITE"], ascending=[True, False])
print(f"  {eligible.sum()} / {len(result)} stocks eligible "
      f"(PCT_FROM_52H >= -25% AND ADTV >= {MIN_TURNOVER_CR} Cr)")

# -- RESIDUAL MOMENTUM ---------------------------------------------------------
resmom_df, rs_z_df = ml.compute_residual_momentum(prices_df, stock_tickers,
                                                    nifty_series, WINDOWS, TRADING_DAYS)
result = result.join(resmom_df)
result = result.join(rs_z_df)

# -- MARKET REGIME (DYNAMIC SCORE) --------------------------------------------
print("\nComputing Dynamic Regime Score ...")
eligible = result["PCT_FROM_52H"] >= -25
regime_score, regime_detail = compute_regime_score(
    nifty_series, eligible, result["COMPOSITE"], prices_df=prices_df)
dynamic_n  = regime_detail["dynamic_n"]
allow_new  = regime_detail["allow_new"]

print(f"  Regime Score  : {regime_score:.2f}  "
      f"(EMA50Brdth={regime_detail['ema50_score']:.2f}  "
      f"TrendBrdth={regime_detail['ema_trend_score']:.2f}  "
      f"Breadth={regime_detail['breadth_score']:.2f}  "
      f"Mom={regime_detail['momentum_score']:.2f})")
print(f"  Dynamic N     : {dynamic_n}  "
      f"({'NEW BUYS ALLOWED' if allow_new else 'NO NEW BUYS  (score < ' + str(NEW_ENTRY_THRESHOLD) + ')'}  )")

# -- EXIT EVALUATION -----------------------------------------------------------
#
# Two exit triggers — evaluated independently for every held position.
#
# EXIT_52H  — 52H disqualification.
#   The stock has PCT_FROM_52H < -25%.
#   This overrides the 28-day hold lock unconditionally.
#
# EXIT_FILTER — Non-rank eligibility disqualification.
#   The stock has no RANK for a non-52H reason, e.g. ADTV failure or missing data.
#   This is kept separate so liquidity/data issues are not mislabeled as 52H breaches.
#
# EXIT_RANK — Rank-based exit.
#   The stock's rank has fallen beyond HOLD_RANK_BUFFER (40)
#   AND the stock has been held for at least MIN_HOLD_DAYS (28 days).
#   The hold lock protects against rank whipsaw for recently bought stocks.
#
# DYNAMIC REGIME — new entries gated by regime_score >= NEW_ENTRY_THRESHOLD.
#   Portfolio size = dynamic_n. Below threshold: exits only, no new buys.
#
print(f"\n{'-'*60}")
print(f"  EXIT EVALUATION  ({TODAY.strftime('%d-%b-%Y')})")
print(f"  Held positions   : {len(ledger)}")
print(f"{'-'*60}")

exit_52h_list    = []   # immediate exits — 52H breach (lock overridden)
exit_filter_list = []   # non-52H eligibility exits — e.g. ADTV failure / missing rank
exit_rank_list   = []   # rank exits — rank > 40 AND hold >= 28 days
hold_list        = []   # retained positions

for ticker, rec in ledger.items():
    held_days  = days_held(ticker, ledger)
    rank_val   = result.loc[ticker, "RANK"] if ticker in result.index else np.nan
    pct52h_val = result.loc[ticker, "PCT_FROM_52H"] if ticker in result.index else np.nan
    adtv_val   = result.loc[ticker, "ADTV_ELIGIBLE"] if (
        ticker in result.index and "ADTV_ELIGIBLE" in result.columns
    ) else True

    # -- Trigger 1: 52H disqualification (direct 52H test)
    if pd.notna(pct52h_val) and pct52h_val < -25:
        exit_52h_list.append({
            "ticker":       ticker,
            "held_days":    held_days,
            "rank":         None,
            "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
            "entry_date":   rec["entry_date"].isoformat(),
            "entry_price":  rec["entry_price"],
            "exit_trigger": "52H_BREACH",
        })

    # -- Trigger 1b: non-52H eligibility failure (e.g. ADTV / missing data)
    elif pd.isna(rank_val):
        trigger = "ADTV_FAIL" if not bool(adtv_val) else "NO_RANK"
        exit_filter_list.append({
            "ticker":       ticker,
            "held_days":    held_days,
            "rank":         None,
            "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
            "entry_date":   rec["entry_date"].isoformat(),
            "entry_price":  rec["entry_price"],
            "exit_trigger": trigger,
        })

    # -- Trigger 2: Rank exit (respects 28-day lock)
    elif rank_val > HOLD_RANK_BUFFER:
        if held_days >= MIN_HOLD_DAYS:
            exit_rank_list.append({
                "ticker":       ticker,
                "held_days":    held_days,
                "rank":         int(rank_val),
                "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
                "entry_date":   rec["entry_date"].isoformat(),
                "entry_price":  rec["entry_price"],
                "exit_trigger": "RANK_EXIT",
            })
        else:
            # Rank has dropped but hold lock still active — hold and note
            hold_list.append({
                "ticker":       ticker,
                "held_days":    held_days,
                "rank":         int(rank_val),
                "pct_52h":      round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
                "note":         f"rank {int(rank_val)} > {HOLD_RANK_BUFFER} but lock active "
                                f"({held_days}/{MIN_HOLD_DAYS}d)",
            })
    else:
        # Healthy — rank within buffer, no 52H breach
        hold_list.append({
            "ticker":    ticker,
            "held_days": held_days,
            "rank":      int(rank_val),
            "pct_52h":   round(pct52h_val, 1) if pd.notna(pct52h_val) else None,
            "note":      "HOLD",
        })

# -- PRINT EXIT SUMMARY --------------------------------------------------------
all_exits = exit_52h_list + exit_filter_list + exit_rank_list

if exit_52h_list:
    print(f"\n  [EXIT — 52H BREACH]  Sell immediately. Hold lock overridden.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'RANK':>6}  {'52H%':>7}  {'ENTRY':>10}  {'@ PRICE':>10}")
    print(f"  {'-'*62}")
    for e in exit_52h_list:
        print(f"  {e['ticker']:<14} {e['held_days']:>4}d  "
              f"  {'NaN':>6}  {str(e['pct_52h']):>7}  "
              f"{e['entry_date']:>10}  {e['entry_price']:>10,.2f}")

if exit_filter_list:
    print(f"\n  [EXIT — ELIGIBILITY]  Failed non-52H eligibility filter.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'TRIGGER':>10}  {'52H%':>7}  {'ENTRY':>10}  {'@ PRICE':>10}")
    print(f"  {'-'*70}")
    for e in exit_filter_list:
        print(f"  {e['ticker']:<14} {e['held_days']:>4}d  "
              f"  {e['exit_trigger']:>10}  {str(e['pct_52h']):>7}  "
              f"{e['entry_date']:>10}  {e['entry_price']:>10,.2f}")

if exit_rank_list:
    print(f"\n  [EXIT — RANK DROP]  Rank > {HOLD_RANK_BUFFER} and hold >= {MIN_HOLD_DAYS} days.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'RANK':>6}  {'52H%':>7}  {'ENTRY':>10}  {'@ PRICE':>10}")
    print(f"  {'-'*62}")
    for e in exit_rank_list:
        print(f"  {e['ticker']:<14} {e['held_days']:>4}d  "
              f"  {e['rank']:>6}  {str(e['pct_52h']):>7}  "
              f"{e['entry_date']:>10}  {e['entry_price']:>10,.2f}")

if hold_list:
    print(f"\n  [HOLD]  {len(hold_list)} position(s) retained.")
    print(f"  {'TICKER':<14} {'HELD':>5}  {'RANK':>6}  {'52H%':>7}  NOTE")
    print(f"  {'-'*62}")
    for h in hold_list:
        print(f"  {h['ticker']:<14} {h['held_days']:>4}d  "
              f"  {str(h['rank']):>6}  {str(h['pct_52h']):>7}  {h['note']}")

if not ledger:
    print("  No open positions in ledger — nothing to evaluate.")

print(f"\n  Summary: {len(exit_52h_list)} 52H exit(s)  |  "
      f"{len(exit_filter_list)} eligibility exit(s)  |  "
      f"{len(exit_rank_list)} rank exit(s)  |  {len(hold_list)} hold(s)")
print(f"{'-'*60}")

# -- ENTRY CANDIDATES ----------------------------------------------------------
#
# New entries this week:
#   - Top dynamic_n stocks by SHARPE_ALL (not already held)
#   - Only when regime_score >= NEW_ENTRY_THRESHOLD
#
currently_held  = set(ledger.keys()) - {e["ticker"] for e in all_exits}
top_n_tickers   = result.head(dynamic_n).index.tolist()

if allow_new:
    entry_candidates = [t for t in top_n_tickers if t not in currently_held]
    if entry_candidates:
        print(f"\n  [NEW ENTRIES — REGIME SCORE {regime_score:.2f}]  {len(entry_candidates)} candidate(s) "
              f"(Top {dynamic_n} slots):")
        for t in entry_candidates:
            px = prices_df.loc[t].dropna()
            last_px = px.iloc[-1] if len(px) > 0 else np.nan
            rank_val = result.loc[t, 'RANK']
            rank_str = f"{int(rank_val):>3}" if pd.notna(rank_val) else "  —"
            print(f"    {t:<14}  rank {rank_str}  "
                  f"last price: {last_px:,.2f}")
    else:
        print(f"\n  [NEW ENTRIES]  All Top {dynamic_n} positions already held.")
else:
    entry_candidates = []
    print(f"\n  [NEW ENTRIES BLOCKED]  Regime score {regime_score:.2f} < threshold {NEW_ENTRY_THRESHOLD}")
    print(f"  No new buys this week. Existing positions monitored for exit only.")

# -- UPDATE LEDGER -------------------------------------------------------------
#
# Remove exits from ledger, add new entries with today's price.
# The caller is responsible for confirming execution before saving —
# in a live system you'd only save after trades are confirmed filled.
#
if DRY_RUN:
    print(f"\n  [DRY RUN] Ledger not updated. Recommendations only -> '{LEDGER_FILE}' unchanged.")
else:
    for e in all_exits:
        ledger.pop(e["ticker"], None)

    for ticker in entry_candidates:
        px = prices_df.loc[ticker].dropna()
        last_px = float(px.iloc[-1]) if len(px) > 0 else 0.0
        ledger[ticker] = {
            "entry_date":  TODAY,
            "entry_price": last_px,
        }

    save_ledger(ledger, LEDGER_FILE)

# -- CAPITAL ALLOCATION (VOLATILITY WEIGHTING) ---------------------------------
print("\nCalculating Dynamic Volatility Weights for Top N Portfolio ...")

result["TARGET_WT"] = np.nan
result["ALLOC_INR"] = np.nan

raw_weights = {}
for ticker in top_n_tickers:
    comp_score = result.loc[ticker, "COMPOSITE"]
    px = prices_df.loc[ticker].dropna()

    if len(px) > 10:
        vols = []
        for w in [252, 189, 126, 63]:
            px_w  = px.iloc[-w:] if len(px) >= w else px
            log_r = np.diff(np.log(px_w.values))
            if len(log_r) > 5:
                vols.append(np.std(log_r, ddof=1) * np.sqrt(252))
        if vols and np.mean(vols) > 0:
            raw_weights[ticker] = comp_score / np.mean(vols)
        else:
            raw_weights[ticker] = comp_score
    else:
        raw_weights[ticker] = comp_score

total_raw = sum(raw_weights.values())
for ticker in top_n_tickers:
    norm_w   = raw_weights[ticker] / total_raw if total_raw > 0 else 1.0 / len(top_n_tickers)
    capped_w = min(0.05, norm_w)
    result.loc[ticker, "TARGET_WT"] = capped_w
    result.loc[ticker, "ALLOC_INR"] = capped_w * PORTFOLIO_CAPITAL

total_equity_weight = result.head(TOP_N)["TARGET_WT"].sum()
total_cash_weight   = max(0.0, 1.0 - total_equity_weight)
total_cash_inr      = total_cash_weight * PORTFOLIO_CAPITAL

# -- CONSOLE OUTPUT ------------------------------------------------------------
SEP  = "-" * 100
HEAD = (f"{'RNK':>4}  {'TICKER':<12}  {'STATUS':<10}  {'TARGET_WT':>9}  {'ALLOC_INR':>11}  "
        f"{'SHARPE_ALL':>10}  {'RES_MOM':>9}  {'SHARPE_3':>9}  {'52H%':>8}")

# Determine per-ticker display status
exit_tickers  = {e["ticker"] for e in exit_52h_list}
filter_exit_set = {e["ticker"] for e in exit_filter_list}
rank_exit_set = {e["ticker"] for e in exit_rank_list}
new_entry_set = set(entry_candidates)

def ticker_status(ticker):
    if ticker in exit_tickers:   return "EXIT-52H"
    if ticker in filter_exit_set:return "EXIT-FLTR"
    if ticker in rank_exit_set:  return "EXIT-RANK"
    if ticker in new_entry_set:  return "NEW BUY"
    if ticker in currently_held: return "HOLD"
    return "WATCH"

print(f"\n{'':=<100}")
print(f"  {UNIVERSE} MOMENTUM - TOP {dynamic_n} (Dynamic N)  .  Sharpe Z + Sharpe 3W + Residual")
print(f"  REGIME SCORE   : {regime_score:.2f}  "
      f"(EMA50Brdth={regime_detail['ema50_score']:.2f}  "
      f"TrendBrdth={regime_detail['ema_trend_score']:.2f}  "
      f"Breadth={regime_detail['breadth_score']:.2f}  "
      f"Mom={regime_detail['momentum_score']:.2f})  "
      f"{'| NEW BUYS ALLOWED' if allow_new else '| NEW BUYS BLOCKED'}")
print(f"  Windows        : 12M/9M/6M/3M (Overlapping)  |  RFR={RFR_ANNUAL*100:.1f}%")
print(f"  Policies       : Weekly | {MIN_HOLD_DAYS}-Day Hold Lock | "
      f"52H% >= -25% | Rank buffer = {HOLD_RANK_BUFFER}")
print(f"  Capital Model  : {PORTFOLIO_CAPITAL:,.0f} INR | 5% Cap Vol Sizing | "
      f"Cash yield {LIQUID_YIELD_PA*100:.0f}% p.a. | Entry threshold {NEW_ENTRY_THRESHOLD}")
print(f"{'':=<100}")

print(HEAD); print(SEP)

def fs(v, w=7): return f"{v:>{w}.3f}" if pd.notna(v) else f"{'--':>{w}}"
def fp(v, w=7): return f"{v:>{w}.1f}" if pd.notna(v) else f"{'--':>{w}}"
def fw(v, w=7): return f"{v*100:>{w}.1f}%" if pd.notna(v) else f"{'--':>{w}}"

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 1):
    status = ticker_status(ticker)
    print(f"{i:>4}  {ticker:<12}  {status:<10}  "
          f"{fw(row['TARGET_WT'],8)}     {row['ALLOC_INR']:9,.0f}  "
          f"{fs(row['COMPOSITE'],10)}  "
          f"{fs(row['RES_MOM'],9)}  {fs(row['SHARPE_3'],9)}  "
          f"{fp(row['PCT_FROM_52H'], 8)}")

print(SEP)
print(f" {'-':>4}  {'CASH (LIQUID)':<12}  {'':10}  {total_cash_weight*100:8.1f}%     "
      f"{total_cash_inr:9,.0f}")
print(SEP)
print(f"\n  SHARPE_ALL = mean(Z_12M..Z_3M)  |  "
      f"SHARPE_3 = mean(Z_12M,Z_6M,Z_3M)  |  "
      f"RES_MOM = residual Sharpe composite (display only)\n")

# -- EXCEL OUTPUT --------------------------------------------------------------
print(f"Writing {OUTPUT_FILE} ...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border_all():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL    = fill("E8EEF2")
ALT_FILL    = fill("F8F9FA")
POS_FILL    = fill("E6F4EA")
NEG_FILL    = fill("FCE8E6")
EXIT52_FILL = fill("FFF3E0")   # amber — 52H exit
EXITRK_FILL = fill("FCE8E6")   # red-tint — rank exit
NEWBY_FILL  = fill("E8F5E9")   # green-tint — new buy
HOLD_FILL   = fill("FFFFFF")

GOLD_FONT   = Font(name="Calibri", color="1A365D", bold=True,  size=11)
CYAN_FONT   = Font(name="Calibri", color="0055CC", bold=True,  size=11)
TEXT_FONT   = Font(name="Calibri", color="111111",             size=11)
MUTED_FONT  = Font(name="Calibri", color="707070",             size=11)
HDR_FONT    = Font(name="Calibri", color="1A365D", bold=True,  size=11)
GREEN_FONT  = Font(name="Calibri", color="137333", bold=True,  size=11)
RED_FONT    = Font(name="Calibri", color="C5221F", bold=True,  size=11)
AMBER_FONT  = Font(name="Calibri", color="E65100", bold=True,  size=11)

def set_hdr(cell, value):
    cell.value     = value
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border_all()

def set_cell(cell, value, font=None, bg=None, num_fmt=None, align="right"):
    cell.value     = value
    cell.font      = font or TEXT_FONT
    cell.fill      = bg or fill("FFFFFF")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border_all()
    if num_fmt:
        cell.number_format = num_fmt

def row_style(ticker):
    """Return (font, bg_fill) based on exit / entry status."""
    if ticker in exit_tickers:   return AMBER_FONT, EXIT52_FILL
    if ticker in filter_exit_set:return RED_FONT,   EXITRK_FILL
    if ticker in rank_exit_set:  return RED_FONT,   EXITRK_FILL
    if ticker in new_entry_set:  return GREEN_FONT, NEWBY_FILL
    return GOLD_FONT, HOLD_FILL

# -- SHEET 1 — TOP20 -----------------------------------------------------------
ws1 = wb_out.create_sheet("TOP20")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C3"

ws1.merge_cells("A1:G1")
tc           = ws1["A1"]
tc.value     = (f"{UNIVERSE} MOMENTUM  .  Top {dynamic_n} (Dynamic N)  .  "
                f"Filter: PCT_FROM_52H >= -25%  .  RFR={RFR_ANNUAL*100:.1f}%  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}  .  "
                f"Regime Score: {regime_score:.2f} ({'BUY' if allow_new else 'NO NEW BUYS'})  .  Run: {TODAY.strftime('%d-%b-%Y')}")
tc.font      = Font(name="Calibri",
                    color="FF2222" if not allow_new else "1A365D",
                    bold=True, size=11)
tc.fill      = fill("2A0000") if not allow_new else fill("F0F4F8")
tc.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

top20_cols = [
    ("RNK",        5), ("TICKER",    12), ("STATUS",    11),
    ("TARGET_WT%", 11), ("ALLOC_INR", 13),
    ("SHARPE_ALL", 12), ("RES_MOM",   10), ("SHARPE_3", 10), ("52H%", 10),
]
for c, (col_name, col_w) in enumerate(top20_cols, 1):
    set_hdr(ws1.cell(row=2, column=c), col_name)
    ws1.column_dimensions[get_column_letter(c)].width = col_w
ws1.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 3):
    row_fnt, row_bg = row_style(ticker)
    rank_v          = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h          = row["PCT_FROM_52H"]
    pct52h_ok       = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt      = GREEN_FONT if pct52h_ok else RED_FONT
    status          = ticker_status(ticker)

    values = [
        (rank_v,           row_fnt,   row_bg, None),
        (ticker,           row_fnt,   row_bg, None),
        (status,           row_fnt,   row_bg, None),
        (row["TARGET_WT"], row_fnt,   row_bg, "0.00%"),
        (row["ALLOC_INR"], row_fnt,   row_bg, "₹_ * #,##0_ ;_ * -#,##0_ ;_ * \"-\"_ ;_ @_ "),
        (row["COMPOSITE"], CYAN_FONT, row_bg, "0.000"),
        (row["RES_MOM"],   TEXT_FONT, row_bg, "0.000"),
        (row["SHARPE_3"],  CYAN_FONT, row_bg, "0.000"),
        (pct52h,           pct52h_fnt,row_bg, "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws1.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws1.row_dimensions[i].height = 16

# Cash summary row
summary_row = TOP_N + 3
set_cell(ws1.cell(row=summary_row, column=1), "-",             GOLD_FONT, fill("FFFFFF"), None)
set_cell(ws1.cell(row=summary_row, column=2), "CASH (LIQUID)", GOLD_FONT, fill("FFFFFF"), None, align="left")
set_cell(ws1.cell(row=summary_row, column=3), "",              MUTED_FONT,fill("FFFFFF"), None)
set_cell(ws1.cell(row=summary_row, column=4), total_cash_weight, GOLD_FONT, fill("FFFFFF"), "0.00%")
set_cell(ws1.cell(row=summary_row, column=5), total_cash_inr,  GOLD_FONT, fill("FFFFFF"),
         "₹_ * #,##0_ ;_ * -#,##0_ ;_ * \"-\"_ ;_ @_ ")
for extra_col in range(6, 10):
    set_cell(ws1.cell(row=summary_row, column=extra_col), "-", MUTED_FONT, fill("FFFFFF"), None)

# -- SHEET 2 — EXITS -----------------------------------------------------------
ws_exit = wb_out.create_sheet("EXITS")
ws_exit.sheet_view.showGridLines = False

ws_exit.merge_cells("A1:H1")
te           = ws_exit["A1"]
te.value     = (f"Exit Actions  .  {TODAY.strftime('%d-%b-%Y')}  .  "
                f"{len(all_exits)} exit(s) this rebalance")
te.font      = Font(name="Calibri", color="C5221F", bold=True, size=11)
te.fill      = fill("FFF3E0")
te.alignment = Alignment(horizontal="center", vertical="center")
ws_exit.row_dimensions[1].height = 22

exit_cols = [
    ("TICKER", 14), ("TRIGGER", 14), ("RANK", 8), ("52H%", 8),
    ("HELD_DAYS", 10), ("ENTRY_DATE", 12), ("ENTRY_PRICE", 13), ("NOTE", 30),
]
for c, (col_name, col_w) in enumerate(exit_cols, 1):
    set_hdr(ws_exit.cell(row=2, column=c), col_name)
    ws_exit.column_dimensions[get_column_letter(c)].width = col_w
ws_exit.row_dimensions[2].height = 18

for i, e in enumerate(all_exits, 3):
    is_52h = e["exit_trigger"] == "52H_BREACH"
    row_fnt = AMBER_FONT if is_52h else RED_FONT
    row_bg  = EXIT52_FILL if is_52h else EXITRK_FILL
    note    = (
        "Lock overridden — 52H breach" if is_52h
        else f"Rank > {HOLD_RANK_BUFFER}, held >= {MIN_HOLD_DAYS}d" if e["exit_trigger"] == "RANK_EXIT"
        else "Failed non-52H eligibility filter"
    )
    vals = [
        (e["ticker"],      row_fnt, row_bg),
        (e["exit_trigger"],row_fnt, row_bg),
        (e.get("rank"),    row_fnt, row_bg),
        (e["pct_52h"],     row_fnt, row_bg),
        (e["held_days"],   row_fnt, row_bg),
        (e["entry_date"],  row_fnt, row_bg),
        (e["entry_price"], row_fnt, row_bg),
        (note,             TEXT_FONT, row_bg),
    ]
    for c, (val, fnt, bg_c) in enumerate(vals, 1):
        set_cell(ws_exit.cell(row=i, column=c), val, fnt, bg_c,
                 align="left" if c in (1, 2, 8) else "right")
    ws_exit.row_dimensions[i].height = 16

if not all_exits:
    ws_exit.merge_cells("A3:H3")
    nc = ws_exit["A3"]
    nc.value = "No exits this rebalance."
    nc.font  = MUTED_FONT
    nc.alignment = Alignment(horizontal="center", vertical="center")

# -- SHEET 3 — CALCS -----------------------------------------------------------
ws2 = wb_out.create_sheet("CALCS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C3"

ws2.merge_cells("A1:AD1")
t2           = ws2["A1"]
t2.value     = (f"{UNIVERSE}  .  Full Calculations  .  All {len(stock_tickers)} stocks  .  "
                f"{dates[0].strftime('%d-%b-%Y')} -> {dates[-1].strftime('%d-%b-%Y')}  .  "
                f"Regime Score: {regime_score:.2f}")
t2.font      = Font(name="Calibri",
                    color="FF2222" if not allow_new else "1A365D",
                    bold=True, size=11)
t2.fill      = fill("2A0000") if not allow_new else fill("F0F4F8")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

calcs_cols = [
    ("RANK",       6), ("TICKER",    12),
    ("S_12M",      9), ("S_9M",       9), ("S_6M",     9), ("S_3M",    9),
    ("Z_12M",      9), ("Z_9M",       9), ("Z_6M",     9), ("Z_3M",    9),
    ("SHARPE_ALL",10), ("SHARPE_3",  10),
    ("RS_12M",     9), ("RS_9M",      9), ("RS_6M",    9), ("RS_3M",   9),
    ("RZ_12M",     9), ("RZ_9M",      9), ("RZ_6M",    9), ("RZ_3M",   9),
    ("RES_MOM",   10),
    ("1M%",        8), ("3M%",        8), ("12M%",     8),
    ("52H%",      10),
]
for c, (col_name, col_w) in enumerate(calcs_cols, 1):
    set_hdr(ws2.cell(row=2, column=c), col_name)
    ws2.column_dimensions[get_column_letter(c)].width = col_w
ws2.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.iterrows(), 3):
    bg         = ALT_FILL if i % 2 == 0 else fill("FFFFFF")
    rank_v     = int(row["RANK"]) if pd.notna(row["RANK"]) else None
    pct52h     = row["PCT_FROM_52H"]
    pct52h_ok  = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt = GREEN_FONT if pct52h_ok else MUTED_FONT
    pct52h_bg  = fill("E6F4EA") if pct52h_ok else fill("F0F0F0")

    values = [
        (rank_v,             GOLD_FONT,  bg,        None),
        (ticker,             GOLD_FONT,  bg,        None),
        (row["S_12M"],       TEXT_FONT,  bg,        "0.000"),
        (row["S_9M"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_6M"],        TEXT_FONT,  bg,        "0.000"),
        (row["S_3M"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_12M"],       TEXT_FONT,  bg,        "0.000"),
        (row["Z_9M"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_6M"],        TEXT_FONT,  bg,        "0.000"),
        (row["Z_3M"],        TEXT_FONT,  bg,        "0.000"),
        (row["COMPOSITE"],   CYAN_FONT,  bg,        "0.000"),
        (row["SHARPE_3"],    TEXT_FONT,  bg,        "0.000"),
        (row["RS_12M"],      MUTED_FONT, bg,        "0.000"),
        (row["RS_9M"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_6M"],       MUTED_FONT, bg,        "0.000"),
        (row["RS_3M"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_12M"],      MUTED_FONT, bg,        "0.000"),
        (row["RZ_9M"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_6M"],       MUTED_FONT, bg,        "0.000"),
        (row["RZ_3M"],       MUTED_FONT, bg,        "0.000"),
        (row["RES_MOM"],     CYAN_FONT,  bg,        "0.000"),
        (row["1M%"],         TEXT_FONT,  bg,        "0.0"),
        (row["3M%"],         TEXT_FONT,  bg,        "0.0"),
        (row["12M%"],        TEXT_FONT,  bg,        "0.0"),
        (pct52h,             pct52h_fnt, pct52h_bg, "0.0"),
    ]
    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        set_cell(ws2.cell(row=i, column=c), v, fnt, bg_c, nfmt,
                 align="left" if c == 2 else "right")
    ws2.row_dimensions[i].height = 15

wb_out.save(OUTPUT_FILE)
print(f"  +  Saved -> {OUTPUT_FILE}")
print(f"     Sheet 'TOP20' : top {TOP_N} stocks with status (NEW BUY / HOLD / EXIT)")
print(f"     Sheet 'EXITS' : {len(all_exits)} exit action(s) this rebalance")
print(f"     Sheet 'CALCS' : all {len(stock_tickers)} stocks, full calculations")
