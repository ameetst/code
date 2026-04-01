"""Validate n500_bt.xlsx for backtest suitability."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import datetime
import numpy as np
import pandas as pd
import openpyxl

FILE = r"C:\Users\ameet\Documents\Github\code\momentum\Sharpe Score\n500_bt.xlsx"

print("=" * 68)
print("  BACKTEST DATA VALIDATION")
print("=" * 68)

# ── Load ──────────────────────────────────────────────────────────────
print("\nLoading file...")
wb       = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
sheets   = wb.sheetnames
print(f"  Sheets found : {sheets}")

if "DATA" not in sheets:
    print("  ERROR: No sheet named 'DATA' — Sharpe.py requires sheet name = DATA")
    sys.exit(1)

ws       = wb["DATA"]
all_rows = list(ws.iter_rows(values_only=True))
wb.close()

header = all_rows[0]

# ── Date columns ──────────────────────────────────────────────────────
date_indices = [i for i, h in enumerate(header)
                if isinstance(h, (datetime.datetime, datetime.date))]
text_date_indices = [i for i, h in enumerate(header)
                     if isinstance(h, str) and len(h) >= 6 and h.replace("-","").replace("/","").isdigit()]

print(f"\n[1] DATE COLUMNS")
print(f"  Proper date columns  : {len(date_indices)}")
print(f"  Text-as-date columns : {len(text_date_indices)}  "
      f"{'(WARNING: must be real date values, not text)' if text_date_indices else ''}")

if not date_indices:
    print("  ERROR: No date columns detected. Sharpe.py will find 0 price columns.")
    sys.exit(1)

dates = [header[i] for i in date_indices]
first_date = dates[0]  if hasattr(dates[0],  "strftime") else pd.Timestamp(dates[0])
last_date  = dates[-1] if hasattr(dates[-1], "strftime") else pd.Timestamp(dates[-1])
n_dates    = len(dates)

print(f"  First date : {first_date.strftime('%d-%b-%Y')}")
print(f"  Last date  : {last_date.strftime('%d-%b-%Y')}")
print(f"  Total date columns : {n_dates}")

# ── Date range check ──────────────────────────────────────────────────
REQUIRED_START = datetime.date(2020, 4, 1)
if isinstance(first_date, datetime.datetime):
    first_d = first_date.date()
elif isinstance(first_date, pd.Timestamp):
    first_d = first_date.date()
else:
    first_d = first_date

years_of_data = (last_date.date() if hasattr(last_date, "date") else last_date -
                 first_d).days / 365.25 if hasattr(last_date, "days") else (
    (last_date.date() if hasattr(last_date,"date") else last_date) - first_d).days / 365.25

print(f"\n[2] DATE RANGE")
print(f"  Years of data : {years_of_data:.1f}")
if first_d > REQUIRED_START:
    print(f"  WARNING: Data starts {first_d.strftime('%d-%b-%Y')} — "
          f"need data from ~Apr-2020 for a full 5-year backtest.")
    print(f"           Current data allows backtest from ~{(first_d.year + 1)}-{first_d.month:02d} onward.")
else:
    print(f"  OK: Sufficient history for 5-year backtest.")

if n_dates < 1200:
    print(f"  WARNING: Only {n_dates} date columns — expected ~1500 for 6 years of trading days.")
else:
    print(f"  OK: {n_dates} trading days ({n_dates/252:.1f} years).")

# ── Stock rows ────────────────────────────────────────────────────────
tickers = []
for row in all_rows[1:]:
    if row[0] is not None:
        tickers.append(str(row[0]).strip())

print(f"\n[3] STOCK UNIVERSE")
print(f"  Total rows (incl NIFTY500) : {len(tickers)}")
has_nifty = "NIFTY500" in tickers
print(f"  NIFTY500 row present       : {'YES' if has_nifty else 'NO -- REQUIRED for regime filter + residual momentum'}")
stock_tickers = [t for t in tickers if t != "NIFTY500"]
print(f"  Stock count                : {len(stock_tickers)}")
if len(stock_tickers) < 400:
    print(f"  WARNING: Only {len(stock_tickers)} stocks — expected ~450-510 for N500 universe.")
else:
    print(f"  OK: Reasonable N500 universe size.")

# ── Price data quality ────────────────────────────────────────────────
print(f"\n[4] PRICE DATA QUALITY")

# Build a small stats pass
total_cells = 0
nan_cells   = 0
zero_cells  = 0
neg_cells   = 0
coverage_by_ticker = {}

for row in all_rows[1:]:
    if row[0] is None:
        continue
    ticker = str(row[0]).strip()
    vals   = []
    for i in date_indices:
        v = row[i]
        total_cells += 1
        if v is None or v == "":
            nan_cells += 1
            vals.append(np.nan)
        else:
            try:
                f = float(v)
                if f <= 0:
                    zero_cells  += 1 if f == 0 else 0
                    neg_cells   += 1 if f < 0 else 0
                    vals.append(np.nan)
                else:
                    vals.append(f)
            except:
                nan_cells += 1
                vals.append(np.nan)
    # valid count
    valid = sum(1 for x in vals if not np.isnan(x))
    coverage_by_ticker[ticker] = valid

nan_pct = nan_cells / total_cells * 100 if total_cells else 0
print(f"  Total price cells : {total_cells:,}")
print(f"  Missing / NaN     : {nan_cells:,}  ({nan_pct:.1f}%)")
print(f"  Zero prices       : {zero_cells:,}")
print(f"  Negative prices   : {neg_cells:,}  {'(ERROR: negative prices are invalid)' if neg_cells else ''}")

# Coverage distribution
coverages   = list(coverage_by_ticker.values())
full_cover  = sum(1 for c in coverages if c >= n_dates * 0.95)
half_cover  = sum(1 for c in coverages if c >= n_dates * 0.50)
low_cover   = sum(1 for c in coverages if c < n_dates * 0.25)

print(f"\n  Coverage per stock (out of {n_dates} date columns):")
print(f"    >= 95% coverage : {full_cover} stocks  (long-history stocks)")
print(f"    >= 50% coverage : {half_cover} stocks")
print(f"    <  25% coverage : {low_cover} stocks  (newly listed / IPOs)")

# Bottom 10 by coverage
low_stocks = sorted(coverage_by_ticker.items(), key=lambda x: x[1])[:10]
print(f"\n  10 stocks with fewest valid prices:")
for t, c in low_stocks:
    pct = c / n_dates * 100
    print(f"    {t:<18} {c:>5} days  ({pct:.0f}%)")

# ── Split-adjustment spot check ───────────────────────────────────────
print(f"\n[5] SPLIT-ADJUSTMENT SPOT CHECK")
print(f"  Checking for suspicious single-day price jumps (>50% or <-33%) ...")

suspicious = []
for row in all_rows[1:]:
    if row[0] is None:
        continue
    ticker = str(row[0]).strip()
    if ticker == "NIFTY500":
        continue
    vals = []
    for i in date_indices:
        v = row[i]
        try:
            f = float(v)
            vals.append(f if f > 0 else np.nan)
        except:
            vals.append(np.nan)
    arr = np.array(vals, dtype=float)
    arr = arr[~np.isnan(arr)]
    if len(arr) < 5:
        continue
    rets = np.diff(arr) / arr[:-1]
    big  = np.where(np.abs(rets) > 0.50)[0]
    if len(big):
        suspicious.append((ticker, len(big), rets[big[0]]*100))

if suspicious:
    print(f"  WARNING: {len(suspicious)} stocks have single-day moves > 50% "
          f"(possible unadjusted splits/bonuses):")
    for t, n, pct in suspicious[:15]:
        print(f"    {t:<18}  {n} occurrence(s)  worst: {pct:+.0f}%")
    if len(suspicious) > 15:
        print(f"    ... and {len(suspicious)-15} more")
else:
    print(f"  OK: No suspicious single-day jumps found.")

# ── Final verdict ─────────────────────────────────────────────────────
print(f"\n{'=' * 68}")
print(f"  VERDICT")
print(f"{'=' * 68}")
issues = []
if not has_nifty:           issues.append("NIFTY500 row missing")
if nan_pct > 30:            issues.append(f"High missing data ({nan_pct:.0f}%)")
if neg_cells > 0:           issues.append("Negative prices present")
if years_of_data < 5:       issues.append(f"Only {years_of_data:.1f} years of data (need ~6)")
if n_dates < 1200:          issues.append(f"Only {n_dates} date columns (need ~1500)")
if len(suspicious) > 20:    issues.append(f"{len(suspicious)} stocks with likely unadjusted splits")

if not issues:
    print("  READY FOR BACKTEST — no critical issues found.")
else:
    print("  ISSUES TO RESOLVE BEFORE BACKTESTING:")
    for iss in issues:
        print(f"    - {iss}")
print("=" * 68)
