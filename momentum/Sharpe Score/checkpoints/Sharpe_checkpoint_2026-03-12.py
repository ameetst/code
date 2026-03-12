"""
N500 Sharpe Z-Score + Clenow Score Ranking
===========================================
  • Computes 12M/9M/6M/3M annualised Sharpe ratios
  • Z-scores each window cross-sectionally
  • Equal-weighted composite Sharpe score
  • Clenow Score = Annualised_Slope × R²  (90-day exponential regression)
  • Outputs:
      - Console: top 20 ranked stocks
      - Excel:   Sheet "TOP20"   — ranked results with all scores
                 Sheet "CALCS"   — full 503-stock Sharpe + Z-score detail

Usage
-----
    python n500_sharpe.py                     # expects n500.xlsx in same folder
    python n500_sharpe.py path/to/n500.xlsx
"""

import sys, datetime, warnings
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy import stats

warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
FILE            = "n500.xlsx" if len(sys.argv) < 2 else sys.argv[1]
OUTPUT_FILE     = "n500_rankings.xlsx"
RFR_ANNUAL      = 0.07
TRADING_DAYS    = 252
TOP_N           = 20
CLENOW_WINDOW   = 90    # trading days for exponential regression
WINDOWS         = {"12M": 252, "9M": 189, "6M": 126, "3M": 63}

# ── LOAD ──────────────────────────────────────────────────────────────────────
print(f"Loading {FILE} ...")
wb_in = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
ws    = wb_in["DATA"]
all_rows = list(ws.iter_rows(values_only=True))
wb_in.close()

header       = all_rows[0]
date_indices = [i for i,h in enumerate(header)
                if isinstance(h, (datetime.datetime, datetime.date))]
dates        = [header[i] for i in date_indices]

tickers, price_matrix = [], []
for row in all_rows[1:]:
    if row[0] is None: continue
    px = []
    for i in date_indices:
        v = row[i]
        try:    px.append(float(v) if v and float(v) > 0 else np.nan)
        except: px.append(np.nan)
    tickers.append(str(row[0]).strip())
    price_matrix.append(px)

prices_df = pd.DataFrame(price_matrix, index=tickers, columns=dates)

# ── SEPARATE NIFTY500 BENCHMARK ───────────────────────────────────────────────
nifty_series  = prices_df.loc["NIFTY500"].copy()
stock_tickers = [t for t in tickers if t != "NIFTY500"]
prices_df     = prices_df.loc[stock_tickers]   # stocks only from here on

# Count real trading days (exclude all-zero/holiday columns)
valid_days = sum(
    1 for di in date_indices
    if any(row[di] and row[di] != 0 for row in all_rows[1:] if row[0])
)
print(f"  {len(prices_df)} stocks  |  {len(dates)} date columns  "
      f"({valid_days} actual trading days)  "
      f"|  {dates[0].strftime('%d-%b-%Y')} → {dates[-1].strftime('%d-%b-%Y')}\n")

# ── SHARPE ────────────────────────────────────────────────────────────────────
rfr_daily = RFR_ANNUAL / TRADING_DAYS

def sharpe_ratio(series: pd.Series, window: int) -> float:
    px = series.dropna()
    if len(px) < window * 0.90: return np.nan
    px_w     = px if len(px) < window + 1 else px.iloc[-(window + 1):]
    log_rets = np.diff(np.log(px_w.values))
    excess   = log_rets - rfr_daily
    sd       = excess.std(ddof=1)
    if sd < 1e-12: return np.nan
    return (excess.mean() / sd) * np.sqrt(TRADING_DAYS)

print("Computing Sharpe ratios ...")
sharpe_data = {}
for label, window in WINDOWS.items():
    col = [sharpe_ratio(prices_df.loc[t], window) for t in stock_tickers]
    sharpe_data[label] = col
    valid = sum(1 for v in col if not np.isnan(v))
    print(f"  {label} ({window}d): {valid}/{len(tickers)} valid")
sharpe_df = pd.DataFrame(sharpe_data, index=stock_tickers)

# ── Z-SCORES ──────────────────────────────────────────────────────────────────
print("\nZ-scoring cross-sectionally ...")
z_df = pd.DataFrame(index=stock_tickers)
for label in WINDOWS:
    s = sharpe_df[label]
    mu, sd = s.mean(), s.std(ddof=1)
    z_df[f"Z_{label}"] = (s - mu) / sd if sd > 0 else 0.0

z_cols            = [f"Z_{l}" for l in WINDOWS]
z_df["COMPOSITE"] = z_df[z_cols].mean(axis=1)

# ── CLENOW SCORE ──────────────────────────────────────────────────────────────
print(f"\nComputing Clenow scores ({CLENOW_WINDOW}-day exp regression) ...")

def clenow_score(series: pd.Series, window: int = CLENOW_WINDOW) -> tuple:
    """
    Fit log-linear (exponential) regression over last `window` trading days.
    Returns (annualised_slope, r_squared, clenow_score).
    Clenow Score = annualised_slope * R²
    """
    px = series.dropna()
    if len(px) < window: return np.nan, np.nan, np.nan
    px_w  = px.iloc[-window:].values
    x     = np.arange(window)
    y     = np.log(px_w)
    slope, intercept, r, p, se = stats.linregress(x, y)
    r2    = r ** 2
    ann_slope = slope * TRADING_DAYS   # annualise daily log slope
    return ann_slope, r2, ann_slope * r2

clenow_rows = [clenow_score(prices_df.loc[t]) for t in stock_tickers]
clenow_df   = pd.DataFrame(clenow_rows,
                            columns=["CLENOW_SLOPE", "CLENOW_R2", "CLENOW_SCORE"],
                            index=stock_tickers)
valid_c = clenow_df["CLENOW_SCORE"].notna().sum()
print(f"  Clenow: {valid_c}/{len(tickers)} valid")

# ── COMBINE ───────────────────────────────────────────────────────────────────
result = z_df.join(sharpe_df.rename(columns={l: f"S_{l}" for l in WINDOWS}))
result = result.join(clenow_df)

# ── NORMALISE COMPOSITE (SHARPE_Z) ───────────────────────────────────────────
# Rules:  v > 1  → v + 1
#         v < 0  → 1 / (1 − v)   [maps negatives into (0, 1])
#         0 ≤ v ≤ 1 → unchanged
def normalise_composite(v):
    if pd.isna(v):  return np.nan
    if v > 1:       return v + 1
    if v < 0:       return 1.0 / (1.0 - v)
    return v

result["COMPOSITE"] = result["COMPOSITE"].map(normalise_composite)

result["RANK"] = result["COMPOSITE"].rank(ascending=False, method="first",
                                           na_option="bottom")
result = result.sort_values("COMPOSITE", ascending=False)

# Attach return context
def safe_ret(series, n):
    px = series.dropna()
    return (px.iloc[-1] / px.iloc[-n] - 1) * 100 if len(px) > n else np.nan

ret_data = {t: {
    "1M%":  safe_ret(prices_df.loc[t], 22),
    "3M%":  safe_ret(prices_df.loc[t], 63),
    "12M%": safe_ret(prices_df.loc[t], 245),
} for t in stock_tickers}
result = result.join(pd.DataFrame(ret_data).T)

# ── 52-WEEK HIGH FILTER ───────────────────────────────────────────────────────
def pct_from_52w_high(series: pd.Series, window: int = 252) -> float:
    """Return % distance from 52W high: negative means price is below high."""
    px = series.dropna()
    if len(px) < 2: return np.nan
    px_w     = px.iloc[-window:] if len(px) >= window else px
    high_52w = px_w.max()
    last_px  = px.iloc[-1]
    if high_52w <= 0: return np.nan
    return (last_px / high_52w - 1) * 100   # negative = below high

print("\nComputing 52-week high proximity ...")
pct_52h = {t: pct_from_52w_high(prices_df.loc[t]) for t in stock_tickers}
result["PCT_FROM_52H"] = pd.Series(pct_52h)

# Re-rank: only stocks within 25% of 52W high qualify (PCT_FROM_52H >= -25)
eligible = result["PCT_FROM_52H"] >= -25
result["RANK"] = np.nan
result.loc[eligible, "RANK"] = (
    result.loc[eligible, "COMPOSITE"]
    .rank(ascending=False, method="first", na_option="bottom")
)
result = result.sort_values("COMPOSITE", ascending=False)
print(f"  {eligible.sum()} / {len(result)} stocks within 25% of 52W high (eligible for ranking)")


# ── RESIDUAL MOMENTUM ─────────────────────────────────────────────────────────
# For each stock and each window:
#   1. Compute daily log-returns for stock and NIFTY500
#   2. OLS: r_stock = alpha + beta * r_nifty + epsilon
#   3. Compute Sharpe ratio on epsilon (residuals)
#   4. Z-score cross-sectionally → equal-weighted composite = RES_MOM

print("\nComputing residual momentum scores ...")

nifty_log_rets = np.diff(np.log(nifty_series.dropna().values))

def residual_sharpe(stock_series: pd.Series, mkt_rets: np.ndarray,
                    window: int) -> float:
    """
    Regress last `window` days of stock log-returns on market log-returns.
    Return annualised Sharpe on the OLS residuals.
    """
    px = stock_series.dropna()
    if len(px) < window * 0.90:
        return np.nan
    # align lengths — take the last min(available, window) observations
    n        = min(len(px) - 1, window)
    s_rets   = np.diff(np.log(px.iloc[-n-1:].values))   # n returns
    m_rets   = mkt_rets[-n:]                              # last n mkt returns
    if len(s_rets) != len(m_rets) or len(s_rets) < 10:
        return np.nan
    # OLS via numpy (faster than scipy for many calls)
    X        = np.column_stack([np.ones(len(m_rets)), m_rets])
    try:
        coeffs, _, _, _ = np.linalg.lstsq(X, s_rets, rcond=None)
    except np.linalg.LinAlgError:
        return np.nan
    residuals = s_rets - X @ coeffs
    sd = residuals.std(ddof=1)
    if sd < 1e-12:
        return np.nan
    return (residuals.mean() / sd) * np.sqrt(TRADING_DAYS)

resmom_data = {}
for label, window in WINDOWS.items():
    col   = [residual_sharpe(prices_df.loc[t], nifty_log_rets, window)
             for t in stock_tickers]
    valid = sum(1 for v in col if not np.isnan(v))
    resmom_data[f"RS_{label}"] = col
    print(f"  {label} ({window}d): {valid}/{len(stock_tickers)} valid")

resmom_df = pd.DataFrame(resmom_data, index=stock_tickers)

# Z-score each window cross-sectionally
rs_z_df = pd.DataFrame(index=stock_tickers)
for label in WINDOWS:
    s = resmom_df[f"RS_{label}"]
    mu, sd = s.mean(), s.std(ddof=1)
    rs_z_df[f"RZ_{label}"] = (s - mu) / sd if sd > 0 else 0.0

rz_cols              = [f"RZ_{l}" for l in WINDOWS]
rs_z_df["RES_MOM"]   = rs_z_df[rz_cols].mean(axis=1)

result = result.join(resmom_df)
result = result.join(rs_z_df)

# ── CONSOLE OUTPUT ────────────────────────────────────────────────────────────
SEP  = "─" * 48
HEAD = f"{'RNK':>4}  {'TICKER':<12}  {'SHARPE_Z':>9}  {'CLENOW':>9}"

print(f"\n{'':=<48}")
print(f"  N500 MOMENTUM — TOP {TOP_N}  ·  Sharpe Z + Clenow")
print(f"  Clenow: {CLENOW_WINDOW}d exp-reg  |  RFR={RFR_ANNUAL*100:.1f}%")
print(f"{'':=<48}")
print(HEAD); print(SEP)

def fs(v, w=7): return f"{v:>{w}.3f}" if pd.notna(v) else f"{'—':>{w}}"

for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 1):
    print(f"{i:>4}  {ticker:<12}  "
          f"{fs(row['COMPOSITE'],9)}  {fs(row['CLENOW_SCORE'],9)}")

print(SEP)
print(f"\n  SHARPE_Z = mean(Z_12M, Z_9M, Z_6M, Z_3M)"
      f"  |  CLENOW = AnnSlope × R²\n")

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
print(f"Writing {OUTPUT_FILE} ...")

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)  # remove default sheet

# ── colour helpers ────────────────────────────────────────────────────────────
def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border_all():
    s = Side(style="thin", color="333355")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = fill("1E2A4A")
HDR2_FILL  = fill("0D0D20")
ALT_FILL   = fill("13131F")
BASE_FILL  = fill("0D0D14")
POS_FILL   = fill("003322")
NEG_FILL   = fill("330011")
GOLD_FONT  = Font(name="Calibri", color="FFC840", bold=True, size=11)
CYAN_FONT  = Font(name="Calibri", color="00CCFF", bold=True, size=11)
TEXT_FONT  = Font(name="Calibri", color="E0E0F0", size=11)
MUTED_FONT = Font(name="Calibri", color="7070A0", size=11)
HDR_FONT   = Font(name="Calibri", color="00CCFF", bold=True, size=11)
GREEN_FONT = Font(name="Calibri", color="00E5A0", bold=True, size=11)
RED_FONT   = Font(name="Calibri", color="FF4466", bold=True, size=11)

def set_hdr(cell, value, font=None):
    cell.value     = value
    cell.font      = font or HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border_all()

def set_cell(cell, value, font=None, bg=None, num_fmt=None, align="right"):
    cell.value     = value
    cell.font      = font or TEXT_FONT
    cell.fill      = bg or fill("0D0D14")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border_all()
    if num_fmt:
        cell.number_format = num_fmt

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — TOP20
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("TOP20")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C3"

# Title row
ws1.merge_cells("A1:I1")
title_cell = ws1["A1"]
title_cell.value     = f"N500 MOMENTUM  ·  Top {TOP_N} by Equal-Weighted Sharpe Z-Score  ·  RFR={RFR_ANNUAL*100:.1f}%  ·  Data: {dates[0].strftime('%d-%b-%Y')} → {dates[-1].strftime('%d-%b-%Y')}"
title_cell.font      = Font(name="Calibri", color="FFC840", bold=True, size=11)
title_cell.fill      = fill("0A0A18")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

# Column headers
top20_cols = [
    ("RNK",        5),  ("TICKER",   12), ("SHARPE_Z",  10),
    ("CLENOW",    10),  ("1M%",       8), ("3M%",        8),
    ("12M%",       8),  ("CLN_SLOPE",12), ("CLN_R2",     9),
]

for c, (col_name, col_w) in enumerate(top20_cols, 1):
    cell = ws1.cell(row=2, column=c)
    set_hdr(cell, col_name)
    ws1.column_dimensions[get_column_letter(c)].width = col_w

ws1.row_dimensions[2].height = 18

# Data rows
for i, (ticker, row) in enumerate(result.head(TOP_N).iterrows(), 3):
    bg       = ALT_FILL if i % 2 == 0 else fill("0D0D14")
    pnl_bg_1 = POS_FILL if pd.notna(row["1M%"]) and row["1M%"] >= 0 else NEG_FILL
    pnl_bg_3 = POS_FILL if pd.notna(row["3M%"]) and row["3M%"] >= 0 else NEG_FILL
    pnl_bg_12= POS_FILL if pd.notna(row["12M%"]) and row["12M%"] >= 0 else NEG_FILL

    rank_v = int(row["RANK"]) if pd.notna(row["RANK"]) else None

    values = [
        (rank_v,                  GOLD_FONT,  bg,       None),
        (ticker,                  GOLD_FONT,  bg,       None),
        (row["COMPOSITE"],        CYAN_FONT,  bg,       "0.000"),
        (row["CLENOW_SCORE"],     TEXT_FONT,  bg,       "0.000"),
        (row["1M%"],              GREEN_FONT if pd.notna(row["1M%"]) and row["1M%"]>=0 else RED_FONT, pnl_bg_1,  "0.0"),
        (row["3M%"],              GREEN_FONT if pd.notna(row["3M%"]) and row["3M%"]>=0 else RED_FONT, pnl_bg_3,  "0.0"),
        (row["12M%"],             GREEN_FONT if pd.notna(row["12M%"]) and row["12M%"]>=0 else RED_FONT, pnl_bg_12,"0.0"),
        (row["CLENOW_SLOPE"],     MUTED_FONT, bg,       "0.000"),
        (row["CLENOW_R2"],        MUTED_FONT, bg,       "0.000"),
    ]

    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        align = "left" if c == 2 else "right"
        set_cell(ws1.cell(row=i, column=c), v, fnt, bg_c, nfmt, align)
    ws1.row_dimensions[i].height = 16

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — CALCS (all 503 stocks)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("CALCS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C3"

# Title
ws2.merge_cells("A1:X1")
t2 = ws2["A1"]
t2.value     = f"N500  ·  Full Sharpe & Z-Score Calculations  ·  All {len(tickers)} stocks  ·  {dates[0].strftime('%d-%b-%Y')} → {dates[-1].strftime('%d-%b-%Y')}"
t2.font      = Font(name="Calibri", color="FFC840", bold=True, size=11)
t2.fill      = fill("0A0A18")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

calcs_cols = [
    ("RANK",        6),  ("TICKER",     12),
    ("S_12M",       9),  ("S_9M",        9),  ("S_6M",       9),  ("S_3M",       9),
    ("Z_12M",       9),  ("Z_9M",        9),  ("Z_6M",       9),  ("Z_3M",       9),
    ("SHARPE_Z",   10),
    ("CLN_SLOPE",  12),  ("CLN_R2",      9),  ("CLENOW",     10),
    ("RS_12M",      9),  ("RS_9M",       9),  ("RS_6M",       9),  ("RS_3M",      9),
    ("RZ_12M",      9),  ("RZ_9M",       9),  ("RZ_6M",       9),  ("RZ_3M",      9),
    ("RES_MOM",    10),
    ("1M%",         8),  ("3M%",         8),  ("12M%",        8),
    ("52H%",       10),
]

for c, (col_name, col_w) in enumerate(calcs_cols, 1):
    cell = ws2.cell(row=2, column=c)
    set_hdr(cell, col_name)
    ws2.column_dimensions[get_column_letter(c)].width = col_w
ws2.row_dimensions[2].height = 18

for i, (ticker, row) in enumerate(result.iterrows(), 3):
    bg = ALT_FILL if i % 2 == 0 else fill("0D0D14")

    rank_val = row["RANK"]
    rank_v   = int(rank_val) if pd.notna(rank_val) else None

    pct52h     = row["PCT_FROM_52H"]
    pct52h_ok  = pd.notna(pct52h) and pct52h >= -25
    pct52h_fnt = GREEN_FONT if pct52h_ok else MUTED_FONT
    pct52h_bg  = fill("003322") if pct52h_ok else fill("1A0A0A")

    values = [
        (rank_v,              GOLD_FONT,  bg,         None),
        (ticker,              GOLD_FONT,  bg,         None),
        (row["S_12M"],        TEXT_FONT,  bg,         "0.000"),
        (row["S_9M"],         TEXT_FONT,  bg,         "0.000"),
        (row["S_6M"],         TEXT_FONT,  bg,         "0.000"),
        (row["S_3M"],         TEXT_FONT,  bg,         "0.000"),
        (row["Z_12M"],        TEXT_FONT,  bg,         "0.000"),
        (row["Z_9M"],         TEXT_FONT,  bg,         "0.000"),
        (row["Z_6M"],         TEXT_FONT,  bg,         "0.000"),
        (row["Z_3M"],         TEXT_FONT,  bg,         "0.000"),
        (row["COMPOSITE"],    CYAN_FONT,  bg,         "0.000"),
        (row["CLENOW_SLOPE"], MUTED_FONT, bg,         "0.000"),
        (row["CLENOW_R2"],    MUTED_FONT, bg,         "0.000"),
        (row["CLENOW_SCORE"], TEXT_FONT,  bg,         "0.000"),
        (row["RS_12M"],        MUTED_FONT, bg,         "0.000"),
        (row["RS_9M"],         MUTED_FONT, bg,         "0.000"),
        (row["RS_6M"],         MUTED_FONT, bg,         "0.000"),
        (row["RS_3M"],         MUTED_FONT, bg,         "0.000"),
        (row["RZ_12M"],        MUTED_FONT, bg,         "0.000"),
        (row["RZ_9M"],         MUTED_FONT, bg,         "0.000"),
        (row["RZ_6M"],         MUTED_FONT, bg,         "0.000"),
        (row["RZ_3M"],         MUTED_FONT, bg,         "0.000"),
        (row["RES_MOM"],       CYAN_FONT,  bg,         "0.000"),
        (row["1M%"],          TEXT_FONT,  bg,         "0.0"),
        (row["3M%"],          TEXT_FONT,  bg,         "0.0"),
        (row["12M%"],         TEXT_FONT,  bg,         "0.0"),
        (pct52h,              pct52h_fnt, pct52h_bg,  "0.0"),
    ]

    for c, (val, fnt, bg_c, nfmt) in enumerate(values, 1):
        v = None if (isinstance(val, float) and np.isnan(val)) else val
        align = "left" if c == 2 else "right"
        set_cell(ws2.cell(row=i, column=c), v, fnt, bg_c, nfmt, align)
    ws2.row_dimensions[i].height = 15

wb_out.save(OUTPUT_FILE)
print(f"  ✓  Saved → {OUTPUT_FILE}")
print(f"     Sheet 'TOP20' : top {TOP_N} stocks with all scores")
print(f"     Sheet 'CALCS' : full {len(tickers)}-stock Sharpe + Z-score + Clenow detail")
