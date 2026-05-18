# -*- coding: utf-8 -*-
"""
fetch_nav.py
============
Fetches historical NSE ETF closing prices using yfinance and overwrites
the price data in 'ETF - Backtest Copy.xlsx'.

Input  : ETF - Backtest Copy.xlsx  (DATA sheet, same format as live file)
Output : ETF - Backtest Copy.xlsx  (same file, price columns overwritten)

Ticker convention: NSE symbols in column B are suffixed with .NS for yfinance.
Usage  : python fetch_nav.py
"""

import pandas as pd
import numpy as np
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────
EXCEL_FILE   = r"C:\Users\ameet\Documents\Github\code\momentum\ETFs\backtest results\ETF - Backtest  - Copy.xlsx"
SHEET_NAME   = "DATA"
START_DATE   = "2020-04-01"           # Fetch from this date
END_DATE     = datetime.today().strftime("%Y-%m-%d")
BATCH_SIZE   = 20                     # Tickers per yfinance batch call
PRICE_COL    = "Close"                # Use adjusted close or Close

# ─────────────────────────────────────────────────
# STEP 1: Read existing Excel for ticker/ETF list
# ─────────────────────────────────────────────────
print(f"\n{'='*55}")
print("  NSE ETF NAV Updater  |  yfinance  |  .NS suffix")
print(f"{'='*55}")
print(f"  File      : {EXCEL_FILE}")
print(f"  Date Range: {START_DATE}  ->  {END_DATE}")
print(f"{'='*55}\n")

file_path = Path(EXCEL_FILE)
if not file_path.exists():
    raise FileNotFoundError(f"Cannot find '{EXCEL_FILE}' in {Path.cwd()}")

raw = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, header=None, dtype={0: str, 1: str})
header_row = raw.iloc[0]

# Detect layout: does col 0 say "ETF Name"/"ETF_NAME" or "TICKER"?
col0_val = str(header_row.iloc[0]).strip().lower()
if "ticker" in col0_val:
    # Layout: Col A = TICKER, dates from Col B onwards
    ticker_col  = 0
    name_col    = None
    date_start  = 1
    print("  Layout detected: Single-column (TICKER only)")
else:
    # Layout: Col A = ETF_NAME, Col B = TICKER, dates from Col C onwards
    ticker_col  = 1
    name_col    = 0
    date_start  = 2
    print("  Layout detected: Two-column (ETF_NAME + TICKER)")

meta_rows = raw.iloc[1:].copy().reset_index(drop=True)
tickers   = meta_rows.iloc[:, ticker_col].astype(str).str.strip()
etf_names = meta_rows.iloc[:, name_col].fillna("").astype(str).str.strip() if name_col is not None else tickers

meta = pd.DataFrame({"ETF_NAME": etf_names, "TICKER": tickers})
meta = meta[meta["TICKER"].notna() & (meta["TICKER"] != "") & (meta["TICKER"] != "nan") & (meta["TICKER"] != "NaT")]
print(f"  Found {len(meta)} ETFs in Excel.\n")

# ─────────────────────────────────────────────────
# STEP 2: Batch-fetch from yfinance
# ─────────────────────────────────────────────────
yf_tickers  = [f"{t}.NS" for t in meta["TICKER"]]
all_prices  = {}      # ticker (no suffix) -> pd.Series of daily closes
failed      = []

print(f"  Fetching in batches of {BATCH_SIZE}...\n")

for i in range(0, len(yf_tickers), BATCH_SIZE):
    batch     = yf_tickers[i : i + BATCH_SIZE]
    orig_tick = meta["TICKER"].tolist()[i : i + BATCH_SIZE]

    print(f"  Batch {i//BATCH_SIZE + 1}: {', '.join(orig_tick)}")
    try:
        raw_data = yf.download(
            batch,
            start=START_DATE,
            end=END_DATE,
            auto_adjust=True,
            progress=False,
            threads=True,
        )

        # yfinance returns MultiIndex columns when >1 ticker
        if isinstance(raw_data.columns, pd.MultiIndex):
            close_df = raw_data[PRICE_COL]
        else:
            close_df = raw_data[[PRICE_COL]].rename(columns={PRICE_COL: batch[0]})

        for yf_t, orig_t in zip(batch, orig_tick):
            if yf_t in close_df.columns:
                s = close_df[yf_t].dropna()
                if len(s) > 0:
                    all_prices[orig_t] = s
                    print(f"    [OK]   {orig_t:20s}  {len(s)} rows  "
                          f"{s.index[0].date()} -> {s.index[-1].date()}")
                else:
                    failed.append(orig_t)
                    print(f"    [EMPTY] {orig_t}")
            else:
                failed.append(orig_t)
                print(f"    [MISS]  {orig_t}")

    except Exception as e:
        print(f"    [ERR] Batch failed: {e}")
        failed.extend(orig_tick)

print(f"\n  Fetched: {len(all_prices)}  |  Failed/Missing: {len(failed)}")
if failed:
    print(f"  Failed tickers: {', '.join(failed)}")

# ─────────────────────────────────────────────────
# STEP 3: Build unified price DataFrame
# ─────────────────────────────────────────────────
# Combine all individual series into one DataFrame (dates as rows, tickers as cols)
if not all_prices:
    raise RuntimeError("No price data fetched. Check internet connection or ticker symbols.")

price_df = pd.DataFrame(all_prices)   # index = date, columns = ticker
price_df = price_df.sort_index()
price_df.index = pd.to_datetime(price_df.index)

# Drop timezone info if present (Excel can't store tz-aware datetimes)
price_df.index = price_df.index.tz_localize(None)

# Replace 0s with NaN, then forward-fill gaps
price_df = price_df.replace(0, np.nan)
print(f"\n  Price matrix: {len(price_df)} dates  x  {len(price_df.columns)} tickers")

# ─────────────────────────────────────────────────
# STEP 4: Write back to Excel (openpyxl)
# ─────────────────────────────────────────────────
print(f"\n  Writing to '{EXCEL_FILE}'...")

wb = load_workbook(EXCEL_FILE)
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME, 0)   # Insert at position 0

# ── Header row ──────────────────────────────────
def hdr_cell(ws, r, c, val):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell

hdr_cell(ws, 1, 1, "ETF Name")
hdr_cell(ws, 1, 2, "Ticker")

dates = price_df.index.tolist()
for col_idx, dt in enumerate(dates, start=3):
    cell = ws.cell(row=1, column=col_idx, value=dt)
    cell.number_format = "DD-MMM-YY"
    cell.font          = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    cell.fill          = PatternFill("solid", fgColor="1F4E79")
    cell.alignment     = Alignment(horizontal="center", vertical="center")

# ── Data rows ────────────────────────────────────
rows_written = 0
for row_idx, (_, m_row) in enumerate(meta.iterrows(), start=2):
    ticker   = m_row["TICKER"]
    etf_name = m_row["ETF_NAME"]

    ws.cell(row=row_idx, column=1, value=etf_name)
    ws.cell(row=row_idx, column=2, value=ticker)

    if ticker in price_df.columns:
        series = price_df[ticker]
        for col_idx, (dt, val) in enumerate(series.items(), start=3):
            if pd.notna(val):
                ws.cell(row=row_idx, column=col_idx, value=round(float(val), 4))
        rows_written += 1

# Freeze header row
ws.freeze_panes = "C2"

wb.save(EXCEL_FILE)
print(f"  Done. {rows_written} ETFs written with price data.")
print(f"\n{'='*55}")
print("  Update complete!")
print(f"  File saved: {file_path.resolve()}")
if failed:
    print(f"  WARNING: {len(failed)} tickers had no data: {', '.join(failed)}")
print(f"{'='*55}\n")
