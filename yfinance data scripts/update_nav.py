# -*- coding: utf-8 -*-
"""
fetch_nav.py (Grid Filler Version - Formula Safe)
=================================================
Reads dynamic date formulas from Row 1, fetches historical prices, 
and fills the grid. Uses a two-pass system to prevent destroying 
your Excel formulas.

Input  : ETF.xlsx  (DATA sheet)
Output : ETF.xlsx  (Prices filled, missing = 0, formulas preserved)
"""

import pandas as pd
import numpy as np
import yfinance as yf
from openpyxl import load_workbook
from datetime import datetime, timedelta
from pathlib import Path
import sys

# ─────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────
EXCEL_FILE   = "N750.xlsx"             
SHEET_NAME   = "DATA"
BATCH_SIZE   = 20                     
PRICE_COL    = "Close"                

print(f"\n{'='*55}")
print("  NSE ETF NAV Grid Filler  |  yfinance (Formula Safe)")
print(f"{'='*55}")

file_path = Path(EXCEL_FILE)
if not file_path.exists():
    raise FileNotFoundError(f"Cannot find '{EXCEL_FILE}' in the current folder.")

# ─────────────────────────────────────────────────
# STEP 1: Pass 1 - Read Data (Evaluate Formulas)
# ─────────────────────────────────────────────────
print(f"  Reading evaluated dates from '{EXCEL_FILE}'...")
# data_only=True forces Python to read the calculated value, not the formula
wb_data = load_workbook(EXCEL_FILE, data_only=True)
if SHEET_NAME not in wb_data.sheetnames:
    raise ValueError(f"Sheet '{SHEET_NAME}' not found in {EXCEL_FILE}")
ws_data = wb_data[SHEET_NAME]

max_col = ws_data.max_column
max_row = ws_data.max_row

date_cols = {}
min_date = None
max_date = None

for c in range(3, max_col + 1):
    val = ws_data.cell(row=1, column=c).value
    if val:
        try:
            if isinstance(val, datetime):
                dt = pd.to_datetime(val).normalize()
            else:
                dt = pd.to_datetime(str(val).strip()).normalize()
                
            date_cols[c] = dt
            if min_date is None or dt < min_date: min_date = dt
            if max_date is None or dt > max_date: max_date = dt
        except Exception:
            pass # Skip if it still can't be parsed

wb_data.close() # Close the data-only version immediately so we don't accidentally save it

if not date_cols:
    print("\n  [!] No valid dates found even after evaluating formulas. Exiting.")
    print("      Note: If you just added formulas, open and save the file in MS Excel first!")
    sys.exit()

print(f"  Found {len(date_cols)} valid dates to fill (from {min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')})")

# Read Tickers (also from data_only in case they are formulas too)
ticker_row_map = {}
for r in range(2, max_row + 1):
    t_val = ws_data.cell(row=r, column=2).value
    if t_val and str(t_val).strip() != "":
        ticker_row_map[str(t_val).strip()] = r

tickers = list(ticker_row_map.keys())
print(f"  Found {len(tickers)} ETFs to process.\n")


# ─────────────────────────────────────────────────
# STEP 2: Batch-fetch from yfinance
# ─────────────────────────────────────────────────
fetch_start = min_date.strftime("%Y-%m-%d")
fetch_end   = (max_date + timedelta(days=1)).strftime("%Y-%m-%d")

yf_tickers  = [f"{t}.NS" for t in tickers]
all_prices  = {}
failed      = []

print(f"  Fetching prices from Yahoo Finance...")

for i in range(0, len(yf_tickers), BATCH_SIZE):
    batch     = yf_tickers[i : i + BATCH_SIZE]
    orig_tick = tickers[i : i + BATCH_SIZE]

    print(f"  Batch {i//BATCH_SIZE + 1}: {', '.join(orig_tick)}")
    try:
        raw_data = yf.download(
            batch,
            start=fetch_start,
            end=fetch_end,
            auto_adjust=True,
            progress=False,
            threads=True,
        )

        if raw_data.empty:
            continue

        if isinstance(raw_data.columns, pd.MultiIndex):
            close_df = raw_data[PRICE_COL]
        else:
            close_df = raw_data[[PRICE_COL]].rename(columns={PRICE_COL: batch[0]})

        for yf_t, orig_t in zip(batch, orig_tick):
            if yf_t in close_df.columns:
                s = close_df[yf_t].dropna()
                s.index = pd.to_datetime(s.index).tz_localize(None).normalize()
                all_prices[orig_t] = s
            else:
                failed.append(orig_t)

    except Exception as e:
        print(f"    [ERR] Batch failed: {e}")
        failed.extend(orig_tick)

# ─────────────────────────────────────────────────
# STEP 3: Pass 2 - Write back to Excel (Formula Safe)
# ─────────────────────────────────────────────────
print(f"\n  Writing data into the grid... (Missing NAV will be set to 0)")

# Reopen the workbook in NORMAL mode. This ensures when we save, formulas are kept.
wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

prices_written = 0
zeroes_written = 0

for ticker, row_idx in ticker_row_map.items():
    series = all_prices.get(ticker)
    
    for col_idx, target_date in date_cols.items():
        if series is not None and target_date in series.index:
            val = series.at[target_date]
            if isinstance(val, pd.Series): 
                val = val.iloc[0]
                
            if pd.notna(val) and val > 0:
                ws.cell(row=row_idx, column=col_idx, value=round(float(val), 4))
                prices_written += 1
            else:
                ws.cell(row=row_idx, column=col_idx, value=0)
                zeroes_written += 1
        else:
            ws.cell(row=row_idx, column=col_idx, value=0)
            zeroes_written += 1

wb.save(EXCEL_FILE)
print(f"\n  Done! File saved successfully with formulas protected.")
print(f"  Valid prices filled : {prices_written}")
print(f"  Zeroes filled       : {zeroes_written}")
if failed:
    print(f"  WARNING: {len(failed)} tickers had no data fetched at all: {', '.join(failed)}")
print(f"{'='*55}\n")