"""
Excel Updater — fetches historical daily close prices AND daily volume for all
date columns in the template.

Usage:
    pip install yfinance openpyxl pandas
    python update_stock_price.py {NSEAll|N750|N500} [--output-dir PATH]

Input/Output:  <universe>.xlsx -> <universe>_updated.xlsx
               Also copies to --output-dir if specified.

The template must contain two sheets:
  DATA   — tickers in col A, date headers in row 1, price cells empty
  VOLUME — same layout, volume cells empty

Both sheets are populated from a single yf.download() call.
"""

import argparse
import datetime
import time
import sys
import shutil
from pathlib import Path

import pandas as pd
import yfinance as yf
import openpyxl


# ── Config ────────────────────────────────────────────────────────────────────
BATCH_SIZE  = 50          # tickers per yfinance batch download
SLEEP_SEC   = 2           # pause between batches (avoid rate-limiting)
PERIOD      = "1y"        # history period for date columns

# Tickers whose Yahoo Finance symbol differs from the value stored in the sheet.
# Key = value as it appears in column A  →  Value = correct Yahoo Finance symbol.
TICKER_OVERRIDES = {
    "NIFTY500": "^CRSLDX",   # Nifty 500 Total Return Index
}
# ─────────────────────────────────────────────────────────────────────────────
def _read_sheet_metadata(ws):
    """Extract tickers (with row numbers) and date-column map from a sheet."""
    tickers = []
    ticker_rows = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        val = row[0].value
        if val:
            tickers.append(str(val).strip())
            ticker_rows[str(val).strip()] = row[0].row

    date_cols = {}
    for cell in ws[1]:
        if isinstance(cell.value, datetime.datetime):
            date_cols[cell.value.date()] = cell.column
    return tickers, ticker_rows, date_cols


def load_template(path: str):
    """Return workbook, both worksheets, tickers, and date-column maps.

    Uses two separate loads:
      - data_only=True  -> read cached formula results (dates in row 1, ticker names)
      - normal load     -> the writable workbook that preserves formulas on save

    Returns
    -------
    wb         : writable Workbook
    ws_data    : writable DATA worksheet
    ws_volume  : writable VOLUME worksheet (or None)
    tickers    : list[str]
    ticker_rows: dict[str, int]       — from DATA sheet
    date_cols  : dict[date, int]      — from DATA sheet
    vol_ticker_rows : dict[str, int]  — from VOLUME sheet
    vol_date_cols   : dict[date, int] — from VOLUME sheet
    """
    # Read-only load to extract cached values
    wb_read = openpyxl.load_workbook(path, data_only=True)

    # DATA sheet
    ws_data_read = wb_read["DATA"]
    tickers, ticker_rows, date_cols = _read_sheet_metadata(ws_data_read)

    # VOLUME sheet
    vol_ticker_rows, vol_date_cols = {}, {}
    if "VOLUME" in wb_read.sheetnames:
        ws_vol_read = wb_read["VOLUME"]
        _, vol_ticker_rows, vol_date_cols = _read_sheet_metadata(ws_vol_read)

    wb_read.close()

    # Writable load — keeps all formulas intact
    wb_write  = openpyxl.load_workbook(path)
    ws_write  = wb_write["DATA"]
    ws_vol_write = wb_write["VOLUME"] if "VOLUME" in wb_write.sheetnames else None

    return (wb_write, ws_write, ws_vol_write,
            tickers, ticker_rows, date_cols,
            vol_ticker_rows, vol_date_cols)
def ns(ticker: str) -> str:
    """Return the correct Yahoo Finance symbol for a ticker.

    Checks TICKER_OVERRIDES first (e.g. NIFTY500 -> ^CRSLDX), then appends
    .NS for plain NSE equity symbols that don't already carry a suffix.
    """
    if ticker in TICKER_OVERRIDES:
        return TICKER_OVERRIDES[ticker]
    t = ticker.upper()
    if not t.endswith(".NS") and not t.endswith(".BO") and not t.startswith("^"):
        return t + ".NS"
    return t
def batch_download(tickers_ns: list[str]):
    """Download 1-year daily Close AND Volume for a batch.

    Returns (close_df, volume_df) — both indexed by date.
    """
    raw = yf.download(
        tickers_ns,
        period=PERIOD,
        auto_adjust=True,
        progress=False,
        threads=True,
    )
    if raw.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Normalise: extract Close and Volume, flatten multi-index if needed
    if isinstance(raw.columns, pd.MultiIndex):
        close  = raw["Close"]
        volume = raw["Volume"] if "Volume" in raw.columns.get_level_values(0) else pd.DataFrame()
    else:
        close  = raw[["Close"]]
        volume = raw[["Volume"]] if "Volume" in raw.columns else pd.DataFrame()

    close.index = close.index.date   # datetime → date
    if not volume.empty:
        volume.index = volume.index.date
    return close, volume

def main(input_file: str, output_file: str, extra_copy_dir: str = None):
    print(f"Loading template: {input_file}")
    (wb, ws_data, ws_vol,
     tickers, ticker_rows, date_cols,
     vol_ticker_rows, vol_date_cols) = load_template(input_file)

    total = len(tickers)
    print(f"  {total} tickers | {len(date_cols)} date columns "
          f"({min(date_cols)} → {max(date_cols)})")
    if ws_vol is not None:
        print(f"  VOLUME sheet: {len(vol_ticker_rows)} tickers | "
              f"{len(vol_date_cols)} date columns")
    else:
        print("  ⚠  No VOLUME sheet in template — volume data will not be written.")
    print()

    # ── Batch-download historical close prices AND volume ─────────────────────
    all_close:  dict[str, dict] = {t: {} for t in tickers}  # ticker → {date: price}
    all_volume: dict[str, dict] = {t: {} for t in tickers}  # ticker → {date: volume}

    # Build symbol → original ticker map; overridden tickers use their Yahoo symbol as key
    ns_to_original = {ns(t): t for t in tickers}
    ns_tickers = list(ns_to_original.keys())
    # Separate out index / override tickers — they must NOT be batched with equity .NS tickers
    index_tickers  = [t for t in ns_tickers if t.startswith("^")]
    equity_tickers = [t for t in ns_tickers if not t.startswith("^")]

    # ── Download index tickers individually first ─────────────────────────────
    if index_tickers:
        print(f"Downloading {len(index_tickers)} index ticker(s): {index_tickers}")
        for sym in index_tickers:
            orig = ns_to_original[sym]
            print(f"  {sym} ({orig})", end="", flush=True)
            try:
                close_df, vol_df = batch_download([sym])
                if not close_df.empty:
                    col_name = close_df.columns[0] if isinstance(close_df.columns[0], str) else sym
                    series = close_df.iloc[:, 0].dropna()
                    all_close[orig] = series.to_dict()
                    if not vol_df.empty:
                        vol_series = vol_df.iloc[:, 0].dropna()
                        all_volume[orig] = vol_series.to_dict()
                    print(f"  ✓ ({len(series)} rows)")
                else:
                    print("  ✗ no data")
            except Exception as e:
                print(f"  ✗ ERROR: {e}")
            time.sleep(1)

    # ── Download equity tickers in batches ────────────────────────────────────
    batches = [equity_tickers[i:i + BATCH_SIZE] for i in range(0, len(equity_tickers), BATCH_SIZE)]
    print(f"\nDownloading equity prices + volume in {len(batches)} batches of ≤{BATCH_SIZE}…")

    for idx, batch in enumerate(batches, 1):
        print(f"  Batch {idx}/{len(batches)}: {batch[0]} … {batch[-1]}", end="", flush=True)
        try:
            close_df, vol_df = batch_download(batch)
            if not close_df.empty:
                for col in close_df.columns:
                    orig = ns_to_original.get(col, col.replace(".NS", ""))
                    if orig in all_close:
                        series = close_df[col].dropna()
                        all_close[orig] = series.to_dict()   # {date: float}
                        if not vol_df.empty and col in vol_df.columns:
                            vol_series = vol_df[col].dropna()
                            all_volume[orig] = vol_series.to_dict()
            print(f"  ✓ ({len(close_df)} rows)")
        except Exception as e:
            print(f"  ✗ ERROR: {e}")

        if idx < len(batches):
            time.sleep(SLEEP_SEC)

    # ── Write historical prices into DATA sheet ───────────────────────────────
    print("\nWriting historical prices to DATA sheet…")
    for ticker, price_map in all_close.items():
        row = ticker_rows.get(ticker)
        if row is None:
            continue
        for date, col in date_cols.items():
            price = price_map.get(date)
            if price is not None:
                ws_data.cell(row=row, column=col).value = round(float(price), 2)

    # ── Write volume data into VOLUME sheet ───────────────────────────────────
    if ws_vol is not None and vol_date_cols:
        print("Writing volume data to VOLUME sheet…")
        vol_tickers_written = 0
        for ticker, vol_map in all_volume.items():
            row = vol_ticker_rows.get(ticker)
            if row is None:
                continue
            if vol_map:
                vol_tickers_written += 1
            for date, col in vol_date_cols.items():
                v = vol_map.get(date)
                if v is not None:
                    ws_vol.cell(row=row, column=col).value = int(float(v))
        print(f"  {vol_tickers_written} tickers with volume data, "
              f"{len(vol_date_cols)} date columns")

    # ── Tickers with no data at all ──────────────────────────────────────────
    errors = [t for t in tickers if not all_close.get(t)]

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"\nSaving → {output_file}")
    wb.save(output_file)
    print("Done ✓")

    # ── Copy to extra directory if specified ──────────────────────────────────
    if extra_copy_dir:
        dest = Path(extra_copy_dir) / Path(output_file).name
        try:
            shutil.copy2(output_file, dest)
            print(f"Copied → {dest}")
        except Exception as e:
            print(f"⚠  Could not copy to {dest}: {e}")

    if errors:
        print(f"\n⚠  {len(errors)} tickers had missing data:")
        for t in errors:
            print(f"   {t}")
    else:
        print("All tickers fetched successfully.")

    return output_file

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update stock prices in an Excel template.")
    parser.add_argument("universe", choices=["NSEAll", "N750", "N500"], help="Universe name (e.g., N750)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Additional directory to copy the output file to (e.g., Sharpe Score folder)")
    args = parser.parse_args()

    input_file = f"{args.universe}.xlsx"
    output_file = f"{args.universe}_updated.xlsx"

    if not Path(input_file).exists():
        print(f"ERROR: {input_file} not found. Place it in the same folder as this script.")
        sys.exit(1)
        
    main(input_file, output_file, extra_copy_dir=args.output_dir)